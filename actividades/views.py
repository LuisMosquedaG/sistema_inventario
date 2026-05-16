from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.forms.models import model_to_dict
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.db import transaction
import openpyxl
from datetime import datetime, time
from io import BytesIO

from .models import Actividad
from clientes.models import Cliente, ContactoCliente
from cotizaciones.models import Cotizacion
from panel.models import Empresa
from preferencias.permissions import require_sales_permission

# --- 1. FUNCIÓN AYUDANTE ESTÁNDAR ---
def get_empresa_actual(request):
    username = request.user.username
    if '@' in username:
        subdominio = username.split('@')[1]
        try:
            return Empresa.objects.get(subdominio=subdominio)
        except Empresa.DoesNotExist:
            return None
    return None

@login_required(login_url='/login/')
@require_sales_permission('actividades', 'crear', json_response=True)
def importar_actividades_ajax(request):
    if request.method == 'POST' and request.FILES.get('archivo_actividades'):
        empresa_actual = get_empresa_actual(request)
        excel_file = request.FILES['archivo_actividades']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return JsonResponse({'success': False, 'error': 'El archivo está vacío o no tiene datos.'})
            
            data_rows = rows[1:]
            creados = 0
            errores = []
            
            with transaction.atomic():
                for idx, row in enumerate(data_rows, start=2):
                    if not any(row): continue # Saltar filas vacías
                    
                    try:
                        nombre = str(row[0] or '').strip()
                        fecha_raw = row[1]
                        hora_inicio_raw = row[2]
                        hora_fin_raw = row[3]
                        tipo = str(row[4] or 'llamada').strip().lower()
                        prioridad = str(row[5] or 'media').strip().lower()
                        cliente_input = str(row[6] or '').strip()
                        descripcion = str(row[7] or '').strip()

                        if not nombre or not fecha_raw or not hora_inicio_raw or not cliente_input:
                            errores.append(f"Fila {idx}: Actividad, Fecha, Hora Inicio y Cliente son obligatorios.")
                            continue

                        # Procesar Fecha
                        if isinstance(fecha_raw, datetime):
                            fecha = fecha_raw.date()
                        elif isinstance(fecha_raw, str):
                            fecha = datetime.strptime(fecha_raw, '%Y-%m-%d').date()
                        else:
                            fecha = fecha_raw

                        # Procesar Horas
                        def parse_time(t):
                            if isinstance(t, time): return t
                            if isinstance(t, datetime): return t.time()
                            if isinstance(t, str): return datetime.strptime(t, '%H:%M').time()
                            return t

                        hora_inicio = parse_time(hora_inicio_raw)
                        hora_fin = parse_time(hora_fin_raw) if hora_fin_raw else None

                        # Buscar Cliente
                        cliente = None
                        if cliente_input.isdigit():
                            cliente = Cliente.objects.filter(id=int(cliente_input), empresa=empresa_actual).first()
                        
                        if not cliente:
                            cliente = Cliente.objects.filter(
                                Q(nombre__icontains=cliente_input) | 
                                Q(apellidos__icontains=cliente_input) | 
                                Q(razon_social__icontains=cliente_input),
                                empresa=empresa_actual
                            ).first()

                        if not cliente:
                            errores.append(f"Fila {idx}: No se encontró el cliente '{cliente_input}'.")
                            continue

                        # Crear Actividad
                        Actividad.objects.create(
                            empresa=empresa_actual,
                            nombre=nombre,
                            fecha=fecha,
                            hora_inicio=hora_inicio,
                            hora_fin=hora_fin,
                            tipo=tipo,
                            prioridad=prioridad,
                            cliente=cliente,
                            descripcion=descripcion,
                            estado='pendiente' # Por defecto al importar
                        )
                        creados += 1

                    except Exception as row_err:
                        errores.append(f"Fila {idx}: {str(row_err)}")
            
            if creados == 0 and errores:
                return JsonResponse({'success': False, 'error': f"No se pudo importar ninguna actividad. Errores: {', '.join(errores[:3])}..."})

            msg = f'Importación exitosa. {creados} actividades creadas.'
            if errores:
                msg += f' ({len(errores)} filas con error)'
                
            return JsonResponse({'success': True, 'message': msg})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error al procesar el archivo: {str(e)}'})
            
    return JsonResponse({'success': False, 'error': 'Solicitud inválida.'})

@login_required(login_url='/login/')
@require_sales_permission('actividades', 'crear')
def descargar_plantilla_actividades(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Actividades"
    
    # Encabezados basados en los campos del modelo Actividad
    headers = [
        'Actividad', 'Fecha (YYYY-MM-DD)', 'Hora Inicio (HH:MM)', 'Hora Fin (HH:MM)', 
        'Tipo (llamada/visita/reunion/envio/seguimiento)', 
        'Prioridad (baja/media/alta/urgente)', 
        'Cliente (Nombre o ID)', 'Descripcion'
    ]
    
    ws.append(headers)
    
    # Ajustar ancho de columnas
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Actividades.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
@require_sales_permission('actividades', 'ver')
def exportar_actividades_excel(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return HttpResponse("No autorizado", status=403)

    # --- LÓGICA DE FILTRADO (Igual que lista_actividades) ---
    q = request.GET.get('q', '')
    cliente_id = request.GET.get('cliente_id', '')
    actividad_nombre = request.GET.get('actividad', '')
    fecha = request.GET.get('fecha', '')
    tipo = request.GET.get('tipo', '')
    prioridad = request.GET.get('prioridad', '')
    estado = request.GET.get('estado', '')
    sucursal_id_filtro = request.GET.get('sucursal', '')

    actividades = Actividad.objects.filter(empresa=empresa_actual).select_related('cliente', 'contacto', 'cotizacion', 'sucursal').order_by('-fecha', '-hora_inicio')

    if q:
        actividades = actividades.filter(
            Q(nombre__icontains=q) |
            Q(cliente__razon_social__icontains=q) |
            Q(cliente__nombre__icontains=q) |
            Q(cliente__apellidos__icontains=q) |
            Q(tipo__icontains=q) |
            Q(prioridad__icontains=q) |
            Q(estado__icontains=q)
        )
    if cliente_id and cliente_id != 'all':
        actividades = actividades.filter(cliente_id=cliente_id)
    if actividad_nombre:
        actividades = actividades.filter(nombre__icontains=actividad_nombre)
    if fecha:
        try:
            actividades = actividades.filter(fecha=fecha)
        except:
            pass
    if tipo:
        actividades = actividades.filter(tipo=tipo)
    if prioridad:
        actividades = actividades.filter(prioridad=prioridad)
    if estado:
        actividades = actividades.filter(estado=estado)
    if sucursal_id_filtro:
        actividades = actividades.filter(sucursal_id=sucursal_id_filtro)

    # --- Generación del Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Actividades"

    headers = [
        'ID', 'Actividad', 'Fecha', 'Hora Inicio', 'Hora Fin', 'Tipo', 'Prioridad',
        'Cliente', 'Contacto', 'Cotización', 'Estado', 'Descripción', 'Sucursal', 'Fecha Creación'
    ]
    ws.append(headers)

    for a in actividades:
        cliente_str = a.cliente.razon_social if a.cliente.razon_social else f"{a.cliente.nombre} {a.cliente.apellidos}"
        contacto_str = a.contacto.nombre_completo if a.contacto else "N/A"
        cotizacion_str = a.cotizacion.folio_completo if a.cotizacion else "N/A"
        sucursal_str = a.sucursal.nombre if a.sucursal else "N/A"
        
        ws.append([
            a.id, a.nombre, a.fecha.strftime('%Y-%m-%d'), a.hora_inicio.strftime('%H:%M'),
            a.hora_fin.strftime('%H:%M') if a.hora_fin else "N/A", a.get_tipo_display(),
            a.get_prioridad_display(), cliente_str, contacto_str, cotizacion_str,
            a.get_estado_display(), a.descripcion, sucursal_str, a.creado_en.strftime('%Y-%m-%d %H:%M')
        ])

    # Ajustar ancho de columnas
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Actividades_{empresa_actual.nombre}.xlsx"'
    wb.save(response)
    return response

# 1. VISTA PRINCIPAL (LISTA)
@login_required(login_url='/login/')
@require_sales_permission('actividades', 'ver')
def lista_actividades(request):
    empresa_actual = get_empresa_actual(request)
    
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)
    
    # --- LÓGICA DE FILTRADO ---
    q = request.GET.get('q', '')
    cliente_id = request.GET.get('cliente_id', '')
    actividad_nombre = request.GET.get('actividad', '')
    fecha = request.GET.get('fecha', '')
    tipo = request.GET.get('tipo', '')
    prioridad = request.GET.get('prioridad', '')
    estado = request.GET.get('estado', '')
    sucursal_id_filtro = request.GET.get('sucursal', '')

    actividades = Actividad.objects.filter(empresa=empresa_actual).select_related('cliente', 'contacto', 'cotizacion', 'sucursal').order_by('-fecha', '-hora_inicio')

    if q:
        actividades = actividades.filter(
            Q(nombre__icontains=q) |
            Q(cliente__razon_social__icontains=q) |
            Q(cliente__nombre__icontains=q) |
            Q(cliente__apellidos__icontains=q) |
            Q(tipo__icontains=q) |
            Q(prioridad__icontains=q) |
            Q(estado__icontains=q)
        )
    if cliente_id and cliente_id != 'all':
        actividades = actividades.filter(cliente_id=cliente_id)
    if actividad_nombre:
        actividades = actividades.filter(nombre__icontains=actividad_nombre)
    if fecha:
        try:
            actividades = actividades.filter(fecha=fecha)
        except:
            pass
    if tipo:
        actividades = actividades.filter(tipo=tipo)
    if prioridad:
        actividades = actividades.filter(prioridad=prioridad)
    if estado:
        actividades = actividades.filter(estado=estado)
    if sucursal_id_filtro:
        actividades = actividades.filter(sucursal_id=sucursal_id_filtro)

    # Para el buscador visual
    cliente_nombre_display = ""
    if cliente_id and cliente_id != 'all':
        try:
            c_obj = Cliente.objects.get(id=cliente_id, empresa=empresa_actual)
            cliente_nombre_display = c_obj.razon_social if c_obj.razon_social else f"{c_obj.nombre} {c_obj.apellidos}"
        except:
            pass

    filtros = {
        'q': q,
        'cliente_id': cliente_id,
        'cliente_nombre': cliente_nombre_display,
        'actividad': actividad_nombre,
        'fecha': fecha,
        'tipo': tipo,
        'prioridad': prioridad,
        'estado': estado,
        'sucursal': sucursal_id_filtro
    }
    # --- FIN LÓGICA DE FILTRADO ---
    from preferencias.models import Sucursal
    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    
    clientes = Cliente.objects.filter(empresa=empresa_actual)
    
    return render(request, 'dashboard_actividades.html', {
        'actividades': actividades,
        'clientes': clientes,
        'sucursales': sucursales,
        'filtros': filtros
    })

@login_required(login_url='/login/')
@require_sales_permission('actividades', 'crear', json_response=True)
def crear_actividad(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'Empresa no detectada'})

    if request.method == 'POST':
        try:
            # Capturamos datos
            nombre = request.POST.get('nombre')
            fecha = request.POST.get('fecha')
            hora_inicio = request.POST.get('hora_inicio')
            hora_fin = request.POST.get('hora_fin')
            tipo = request.POST.get('tipo')
            prioridad = request.POST.get('prioridad')
            cliente_id = request.POST.get('cliente')
            contacto_id = request.POST.get('contacto')
            cotizacion_id = request.POST.get('cotizacion')
            correo = request.POST.get('correo')
            direccion = request.POST.get('direccion')
            descripcion = request.POST.get('descripcion')
            estado = request.POST.get('estado', 'borrador') 

            if not cliente_id:
                return JsonResponse({'success': False, 'error': 'El cliente es obligatorio.'})

            # --- SEGURIDAD: Verificar que el Cliente pertenezca a la empresa ---
            cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa_actual)

            # --- SEGURIDAD: Verificar Contacto (si se envía) ---
            contacto = None
            if contacto_id:
                # Como Contacto depende de Cliente, si el Cliente es de la empresa, sus contactos lo son.
                # Pero verificamos que exista.
                contacto = get_object_or_404(ContactoCliente, id=contacto_id, cliente=cliente)

            # --- SEGURIDAD: Verificar Cotización (si se envía) ---
            # Asumimos que Cotización tiene empresa, si no, verificar por cliente.
            cotizacion = None
            if cotizacion_id:
                # Nota: Si tu modelo Cotizacion no tiene campo empresa directo, verifica por cliente
                try:
                    cotizacion = Cotizacion.objects.get(id=cotizacion_id, cliente=cliente)
                except Cotizacion.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'La cotización no pertenece a este cliente.'})

            # Asignar sucursal desde la sesión
            sucursal_obj = None
            sucursal_id = request.session.get('sucursal_id')
            if sucursal_id:
                from preferencias.models import Sucursal
                try:
                    sucursal_obj = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
                except Sucursal.DoesNotExist:
                    pass

            Actividad.objects.create(
                nombre=nombre, 
                fecha=fecha, 
                hora_inicio=hora_inicio, 
                hora_fin=hora_fin,
                tipo=tipo, 
                prioridad=prioridad,
                cliente=cliente, # <--- Pasamos el objeto validado
                contacto=contacto, # <--- Pasamos el objeto validado
                cotizacion=cotizacion, # <--- Pasamos el objeto validado
                correo=correo, 
                direccion=direccion, 
                descripcion=descripcion,
                estado=estado,
                empresa=empresa_actual,
                sucursal=sucursal_obj
            )
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# 3. VISTA: EDITAR ACTIVIDAD
@login_required(login_url='/login/')
@require_sales_permission('actividades', 'editar', json_response=True)
def editar_actividad(request, actividad_id):
    empresa_actual = get_empresa_actual(request)
    
    if request.method == 'POST':
        try:
            # --- SEGURIDAD: Verificar que la actividad pertenezca a la empresa ---
            actividad = get_object_or_404(Actividad, id=actividad_id, empresa=empresa_actual)
            
            if actividad.estado != 'borrador':
                return JsonResponse({'success': False, 'error': 'Solo se pueden editar actividades en borrador.'})

            # Actualizar campos simples
            actividad.nombre = request.POST.get('nombre')
            actividad.fecha = request.POST.get('fecha')
            actividad.hora_inicio = request.POST.get('hora_inicio')
            actividad.hora_fin = request.POST.get('hora_fin')
            actividad.tipo = request.POST.get('tipo')
            actividad.prioridad = request.POST.get('prioridad')
            actividad.descripcion = request.POST.get('descripcion')
            actividad.correo = request.POST.get('correo')
            actividad.direccion = request.POST.get('direccion')

            # ACTUALIZAR RELACIONES CON VALIDACIÓN
            cliente_id = request.POST.get('cliente')
            contacto_id = request.POST.get('contacto')
            cotizacion_id = request.POST.get('cotizacion')

            if cliente_id:
                # Validar nuevo cliente
                nuevo_cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa_actual)
                actividad.cliente = nuevo_cliente
            
            if contacto_id:
                # Validar contacto (perteneciente al cliente seleccionado arriba)
                actividad.contacto = get_object_or_404(ContactoCliente, id=contacto_id, cliente=actividad.cliente)
            else:
                actividad.contacto = None

            if cotizacion_id:
                # Validar cotización (perteneciente al cliente)
                actividad.cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id, cliente=actividad.cliente)
            else:
                actividad.cotizacion = None
            
            actividad.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# 4. VISTA: CAMBIAR ESTADO
@login_required(login_url='/login/')
@csrf_exempt
@require_sales_permission('actividades', 'aprobar', json_response=True)
def cambiar_estado_actividad(request, actividad_id):
    empresa_actual = get_empresa_actual(request)
    
    if request.method == 'POST':
        try:
            # --- SEGURIDAD ---
            actividad = get_object_or_404(Actividad, id=actividad_id, empresa=empresa_actual)
            
            nuevo_estado = request.POST.get('nuevo_estado')
            
            if actividad.estado == 'borrador' and nuevo_estado == 'pendiente':
                actividad.estado = 'pendiente'
                actividad.save()
                return JsonResponse({'success': True, 'mensaje': 'Actividad aprobada y enviada a pendiente.'})
            
            elif actividad.estado == 'pendiente' and nuevo_estado == 'completada':
                actividad.estado = 'completada'
                actividad.save()
                return JsonResponse({'success': True, 'mensaje': 'Actividad completada exitosamente.'})
            
            else:
                return JsonResponse({'success': False, 'error': 'Transición de estado no permitida.'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# 5. VISTA: REPROGRAMAR
@login_required(login_url='/login/')
@require_sales_permission('actividades', 'editar', json_response=True)
def reprogramar_actividad(request, actividad_id):
    empresa_actual = get_empresa_actual(request)
    
    if request.method == 'POST':
        try:
            # --- SEGURIDAD ---
            actividad = get_object_or_404(Actividad, id=actividad_id, empresa=empresa_actual)
            
            if actividad.estado != 'pendiente':
                return JsonResponse({'success': False, 'error': 'Solo se pueden reprogramar actividades pendientes.'})

            actividad.fecha = request.POST.get('fecha')
            actividad.hora_inicio = request.POST.get('hora_inicio')
            actividad.hora_fin = request.POST.get('hora_fin')
            actividad.descripcion = request.POST.get('descripcion')
            
            actividad.save()
            return JsonResponse({'success': True, 'mensaje': 'Actividad reprogramada.'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# 6. VISTA: CANCELAR
@login_required(login_url='/login/')
@require_sales_permission('actividades', 'eliminar', json_response=True)
def cancelar_actividad(request, actividad_id):
    empresa_actual = get_empresa_actual(request)
    
    if request.method == 'POST':
        try:
            # --- SEGURIDAD ---
            actividad = get_object_or_404(Actividad, id=actividad_id, empresa=empresa_actual)
            
            motivo = request.POST.get('motivo')
            
            if not motivo:
                return JsonResponse({'success': False, 'error': 'El motivo de cancelación es obligatorio.'})

            actividad.estado = 'cancelada'
            actividad.motivo_cancelacion = motivo
            actividad.save()
            
            return JsonResponse({'success': True, 'mensaje': 'Actividad cancelada.'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# 7. VISTA: API DATOS ACTIVIDAD
@login_required(login_url='/login/')
@require_sales_permission('actividades', 'ver', json_response=True)
def api_datos_actividad(request, actividad_id):
    empresa_actual = get_empresa_actual(request)
    try:
        # --- SEGURIDAD ---
        actividad = get_object_or_404(Actividad, id=actividad_id, empresa=empresa_actual)
        
        data = model_to_dict(actividad)
        if data.get('fecha'): data['fecha'] = actividad.fecha.strftime('%Y-%m-%d')
        data['cliente'] = actividad.cliente.id if actividad.cliente else None
        data['contacto'] = actividad.contacto.id if actividad.contacto else None
        data['cotizacion'] = actividad.cotizacion.id if actividad.cotizacion else None
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)
    
# 3. API: Obtener Datos del Cliente
@login_required(login_url='/login/')
@csrf_exempt
@require_sales_permission('actividades', 'ver', json_response=True)
def api_cliente_datos(request, cliente_id):
    empresa_actual = get_empresa_actual(request)
    
    if request.method == 'GET':
        try:
            # --- SEGURIDAD: Solo ver clientes de mi empresa ---
            cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa_actual)
            
            # 1. CONSTRUIR LISTA DE CONTACTOS
            contactos_data = []
            for c in ContactoCliente.objects.filter(cliente=cliente):
                contactos_data.append({
                    'id': c.id,
                    'nombre_completo': c.nombre_completo,
                    'correo_1': c.correo_1,
                    'telefono_1': c.telefono_1
                })
            
            # 2. CONSTRUIR LISTA DE COTIZACIONES
            cotizaciones_data = []
            for cot in Cotizacion.objects.filter(cliente=cliente):
                cotizaciones_data.append({
                    'id': cot.id,
                    'folio_completo': cot.folio_completo
                })

            return JsonResponse({
                'cliente': {
                    'email': cliente.email,
                    'calle': cliente.calle,
                    'numero_ext': cliente.numero_ext,
                    'numero_int': cliente.numero_int,
                    'colonia': cliente.colonia,
                    'estado': cliente.estado,
                    'cp': cliente.cp,
                },
                'contactos': contactos_data,
                'cotizaciones': cotizaciones_data
            })
            
        except Exception as e:
            print(f"Error en API: {e}")
            return JsonResponse({'error': str(e)}, status=500)