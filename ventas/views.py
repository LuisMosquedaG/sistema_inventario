from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
from django.db.models import F, Q, Sum
from decimal import Decimal
import openpyxl
from datetime import datetime

from .models import OrdenVenta, DetalleOrdenVenta
from pedidos.models import Pedido, DetallePedido
from panel.models import Empresa
from almacenes.models import Inventario, Almacen, Kardex
from core.models import Producto, Transaccion
from clientes.models import Cliente
from recepciones.models import DetalleRecepcionExtra
from notificaciones.utils import crear_notificacion
from preferencias.permissions import require_sales_permission, user_has_sales_permission

def get_empresa_actual(request):
    username = request.user.username
    if '@' in username:
        subdominio = username.split('@')[1]
        try:
            return Empresa.objects.get(subdominio=subdominio)
        except Empresa.DoesNotExist:
            return None
    return None

@login_required(login_url='/login/')
@require_sales_permission('salidas', 'crear')
def descargar_plantilla_salidas(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Salidas"
    
    # Encabezados: Referencia ayuda a agrupar productos en una misma orden de salida
    headers = [
        'Referencia (Ej: OS-001)', 'Cliente (Nombre o ID)', 'Fecha (YYYY-MM-DD)', 
        'Direccion Envio', 'Producto (Nombre o ID)', 'Cantidad', 'Precio Unitario'
    ]
    
    ws.append(headers)
    
    # Ajustar ancho de columnas
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Salidas.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
@require_sales_permission('salidas', 'crear', json_response=True)
def importar_salidas_ajax(request):
    if request.method == 'POST' and request.FILES.get('archivo_salidas'):
        empresa_actual = get_empresa_actual(request)
        excel_file = request.FILES['archivo_salidas']
        
        try:
            from datetime import datetime
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return JsonResponse({'success': False, 'error': 'El archivo está vacío o no tiene datos.'})
            
            data_rows = rows[1:]
            
            # Agrupar por Referencia
            salidas_data = {}
            for idx, row in enumerate(data_rows, start=2):
                if not any(row): continue
                ref = str(row[0] or f"TEMP-{idx}").strip()
                if ref not in salidas_data:
                    salidas_data[ref] = {
                        'cliente_input': str(row[1] or '').strip(),
                        'fecha_raw': row[2],
                        'direccion': str(row[3] or '').strip(),
                        'items': []
                    }
                salidas_data[ref]['items'].append({
                    'prod_input': str(row[4] or '').strip(),
                    'cantidad': row[5] or 1,
                    'precio': row[6] or 0
                })

            creadas = 0
            errores = []
            
            with transaction.atomic():
                # Obtener sucursal de la sesión
                sucursal_obj = None
                sucursal_id = request.session.get('sucursal_id')
                if sucursal_id:
                    from preferencias.models import Sucursal
                    try:
                        sucursal_obj = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
                    except Sucursal.DoesNotExist:
                        pass

                for ref, data in salidas_data.items():
                    try:
                        # Buscar Cliente
                        cliente = None
                        if data['cliente_input'].isdigit():
                            cliente = Cliente.objects.filter(id=int(data['cliente_input']), empresa=empresa_actual).first()
                        if not cliente:
                            cliente = Cliente.objects.filter(
                                Q(nombre__icontains=data['cliente_input']) | 
                                Q(apellidos__icontains=data['cliente_input']) | 
                                Q(razon_social__icontains=data['cliente_input']),
                                empresa=empresa_actual
                            ).first()
                        
                        if not cliente:
                            errores.append(f"Ref {ref}: No se encontró el cliente '{data['cliente_input']}'.")
                            continue

                        # Crear Orden de Venta (Salida)
                        orden = OrdenVenta.objects.create(
                            empresa=empresa_actual,
                            cliente=cliente,
                            vendedor=request.user,
                            sucursal=sucursal_obj,
                            estado='borrador',
                            direccion_envio=data['direccion']
                        )

                        # Crear Detalles
                        for item in data['items']:
                            producto = None
                            if item['prod_input'].isdigit():
                                producto = Producto.objects.filter(id=int(item['prod_input']), empresa=empresa_actual).first()
                            if not producto:
                                producto = Producto.objects.filter(nombre__icontains=item['prod_input'], empresa=empresa_actual).first()
                            
                            if producto:
                                DetalleOrdenVenta.objects.create(
                                    orden_venta=orden,
                                    producto=producto,
                                    cantidad=item['cantidad'],
                                    precio_unitario=item['precio']
                                )
                        
                        creadas += 1
                    except Exception as e:
                        errores.append(f"Ref {ref}: {str(e)}")
            
            if creadas == 0 and errores:
                return JsonResponse({'success': False, 'error': f"No se pudo importar nada. Errores: {', '.join(errores[:2])}"})

            msg = f'Importación exitosa. {creadas} órdenes de salida creadas.'
            if errores: msg += f' ({len(errores)} errores)'
            return JsonResponse({'success': True, 'message': msg})

        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error al procesar: {str(e)}'})
            
    return JsonResponse({'success': False, 'error': 'Solicitud inválida.'})

@login_required(login_url='/login/')
@require_sales_permission('salidas', 'ver')
def exportar_salidas_excel(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return HttpResponse("No autorizado", status=403)

    # --- LÓGICA DE FILTRADO (Igual que dashboard_ventas) ---
    q = request.GET.get('q', '')
    folio_salida = request.GET.get('folio_salida', '')
    folio_cotizacion = request.GET.get('folio_cotizacion', '')
    folio_pedido = request.GET.get('folio_pedido', '')
    cliente_id = request.GET.get('cliente_id', '')
    estado = request.GET.get('estado', '')
    sucursal_id_filtro = request.GET.get('sucursal', '')

    ordenes = OrdenVenta.objects.filter(empresa=empresa_actual).select_related('pedido_origen', 'cliente', 'sucursal').order_by('-fecha_creacion')

    if q:
        ordenes = ordenes.filter(
            Q(id__icontains=q) | Q(pedido_origen__id__icontains=q) |
            Q(cliente__razon_social__icontains=q) | Q(cliente__nombre__icontains=q) |
            Q(estado__icontains=q)
        ).distinct()
    
    if folio_salida:
        clean_os = folio_salida.upper().replace('OS-', '').replace('OS', '').strip()
        ordenes = ordenes.filter(id__icontains=clean_os)
    if folio_cotizacion:
        clean_cot = folio_cotizacion.upper().replace('COT-', '').replace('COT', '').strip()
        ordenes = ordenes.filter(pedido_origen__cotizacion_origen_id__icontains=clean_cot)
    if folio_pedido:
        clean_ped = folio_pedido.upper().replace('PED-', '').replace('PED', '').strip()
        ordenes = ordenes.filter(pedido_origen__id__icontains=clean_ped)
        
    if cliente_id and cliente_id != 'all':
        ordenes = ordenes.filter(cliente_id=cliente_id)
    if estado:
        ordenes = ordenes.filter(estado=estado)
    if sucursal_id_filtro:
        ordenes = ordenes.filter(sucursal_id=sucursal_id_filtro)

    # --- Generación del Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Salidas"

    headers = [
        'ID', 'Folio', 'Cliente', 'Vendedor', 'Sucursal', 'Pedido Origen', 
        'Fecha Creación', 'Estado', 'Total', 'Dirección Envío'
    ]
    ws.append(headers)

    for o in ordenes:
        cliente_str = o.cliente.razon_social if o.cliente.razon_social else f"{o.cliente.nombre} {o.cliente.apellidos}"
        ws.append([
            o.id, o.folio_display, cliente_str, o.vendedor.get_full_name() or o.vendedor.username,
            o.sucursal.nombre if o.sucursal else "Principal",
            o.pedido_origen_id or "Directa",
            o.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            o.get_estado_display(),
            o.total_orden,
            o.direccion_envio
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Salidas_{empresa_actual.nombre}.xlsx"'
    wb.save(response)
    return response

@login_required
@require_sales_permission('salidas', 'ver')
def dashboard_ventas(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)

    # --- LÓGICA DE FILTRADO ---
    q = request.GET.get('q', '')
    folio_salida = request.GET.get('folio_salida', '')
    folio_cotizacion = request.GET.get('folio_cotizacion', '')
    folio_pedido = request.GET.get('folio_pedido', '')
    fecha_salida = request.GET.get('fecha_salida', '')
    cliente_id = request.GET.get('cliente_id', '')
    estado = request.GET.get('estado', '')
    sucursal_id_filtro = request.GET.get('sucursal', '')

    ordenes = OrdenVenta.objects.filter(empresa=empresa_actual).select_related('pedido_origen', 'cliente', 'sucursal').order_by('-fecha_creacion')

    if q:
        ordenes = ordenes.filter(
            Q(id__icontains=q) |
            Q(pedido_origen__id__icontains=q) |
            Q(pedido_origen__cotizacion_origen_id__icontains=q) |
            Q(cliente__razon_social__icontains=q) |
            Q(cliente__nombre__icontains=q) |
            Q(cliente__apellidos__icontains=q) |
            Q(estado__icontains=q) |
            Q(detalles__producto__nombre__icontains=q)
        ).distinct()
    
    if folio_salida:
        clean_os = folio_salida.upper().replace('OS-', '').replace('OS', '').strip()
        ordenes = ordenes.filter(id__icontains=clean_os)
    if folio_cotizacion:
        clean_cot = folio_cotizacion.upper().replace('COT-', '').replace('COT', '').strip()
        ordenes = ordenes.filter(pedido_origen__cotizacion_origen_id__icontains=clean_cot)
    if folio_pedido:
        clean_ped = folio_pedido.upper().replace('PED-', '').replace('PED', '').strip()
        ordenes = ordenes.filter(pedido_origen__id__icontains=clean_ped)
        
    if cliente_id and cliente_id != 'all':
        ordenes = ordenes.filter(cliente_id=cliente_id)
    if fecha_salida:
        try:
            ordenes = ordenes.filter(fecha_creacion__date=fecha_salida)
        except:
            pass
    if estado:
        ordenes = ordenes.filter(estado=estado)
    if sucursal_id_filtro:
        ordenes = ordenes.filter(sucursal_id=sucursal_id_filtro)

    # Para el buscador visual
    cliente_nombre_display = ""
    if cliente_id and cliente_id != 'all':
        try:
            c_obj = Cliente.objects.get(id=cliente_id, empresa=empresa_actual)
            cliente_nombre_display = c_obj.razon_social if c_obj.razon_social else f"{c_obj.nombre} {c_obj.apellidos}"
        except:
            pass

    filtros = {
        'q': q,
        'folio_salida': folio_salida,
        'folio_cotizacion': folio_cotizacion,
        'folio_pedido': folio_pedido,
        'fecha_salida': fecha_salida,
        'cliente_id': cliente_id,
        'cliente_nombre': cliente_nombre_display,
        'estado': estado,
        'sucursal': sucursal_id_filtro
    }
    # --- FIN LÓGICA DE FILTRADO ---
    from preferencias.models import Sucursal
    sucursales_lista = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')

    paginator = Paginator(ordenes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    cotizaciones_ids = set()
    for orden in page_obj:
        if orden.pedido_origen and orden.pedido_origen.cotizacion_origen_id:
            cotizaciones_ids.add(orden.pedido_origen.cotizacion_origen_id)
    
    fechas_map = {}
    if cotizaciones_ids:
        from cotizaciones.models import Cotizacion 
        datos_cot = Cotizacion.objects.filter(id__in=cotizaciones_ids).values('id', 'creado_en')
        for dato in datos_cot:
            fechas_map[dato['id']] = dato['creado_en']

    for orden in page_obj:
        orden.fecha_cotizacion_display = None
        if orden.pedido_origen and orden.pedido_origen.cotizacion_origen_id:
            orden.fecha_cotizacion_display = fechas_map.get(orden.pedido_origen.cotizacion_origen_id)

    # LISTA DE ALMACENES PARA EL MODAL (Filtrado por sucursal si hay filtro activo)
    almacenes_qs = empresa_actual.almacen_set.all()
    if sucursal_id_filtro:
        almacenes_qs = almacenes_qs.filter(sucursal_id=sucursal_id_filtro)
    
    almacenes = list(almacenes_qs.values('id', 'nombre'))

    contexto = {
        'page_obj': page_obj, 
        'almacenes_json': almacenes, 
        'clientes': Cliente.objects.filter(empresa=empresa_actual), 
        'productos': Producto.objects.filter(empresa=empresa_actual),
        'sucursales': sucursales_lista,
        'filtros': filtros,
        'perm_ventas': {
            'ver': user_has_sales_permission(request, 'salidas', 'ver'),
            'crear': user_has_sales_permission(request, 'salidas', 'crear'),
            'editar': user_has_sales_permission(request, 'salidas', 'editar'),
            'eliminar': user_has_sales_permission(request, 'salidas', 'eliminar'),
            'aprobar': user_has_sales_permission(request, 'salidas', 'aprobar'),
            'imprimir': user_has_sales_permission(request, 'salidas', 'imprimir'),
        }
    }
    return render(request, 'dashboard_ventas.html', contexto)

@login_required
@transaction.atomic
@require_sales_permission('salidas', 'crear')
def crear_salida_directa(request):
    """Crea una Orden de Salida (Venta) directa sin pedido previo"""
    if request.method != 'POST':
        return redirect('dashboard_ventas')
    
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        messages.error(request, "Empresa no detectada")
        return redirect('dashboard_ventas')

    try:
        cliente_id = request.POST.get('cliente')
        cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa_actual)
        
        # Asignar sucursal desde la sesión
        sucursal_obj = None
        sucursal_id = request.session.get('sucursal_id')
        if sucursal_id:
            from preferencias.models import Sucursal
            try:
                sucursal_obj = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
            except Sucursal.DoesNotExist:
                pass

        # 1. Crear la cabecera
        ov = OrdenVenta.objects.create(
            pedido_origen=None,
            cliente=cliente,
            vendedor=request.user,
            empresa=empresa_actual,
            sucursal=sucursal_obj,
            estado='borrador'
        )

        # 2. Crear los detalles
        productos_ids = request.POST.getlist('producto_id[]')
        cantidades = request.POST.getlist('cantidad[]')
        precios = request.POST.getlist('precio_unitario[]')

        if not productos_ids:
            raise ValueError("Debes agregar al menos un artículo a la salida.")

        for i in range(len(productos_ids)):
            prod_id = productos_ids[i]
            producto = get_object_or_404(Producto, id=prod_id, empresa=empresa_actual)
            
            DetalleOrdenVenta.objects.create(
                orden_venta=ov,
                producto=producto,
                cantidad=int(cantidades[i]),
                precio_unitario=Decimal(precios[i])
            )

        crear_notificacion(
            empresa=empresa_actual,
            actor=request.user,
            mensaje=f'creó salida directa {ov.folio_display}',
            propietario=request.user
        )

        messages.success(request, f'Orden de Salida OS-{ov.id:04d} creada correctamente.')
    except Exception as e:
        messages.error(request, f'Error al crear salida: {str(e)}')

    return redirect('dashboard_ventas')

@login_required
@transaction.atomic
@require_sales_permission('salidas', 'crear')
def crear_orden_venta(request, pedido_id):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        messages.error(request, "Empresa no detectada")
        return redirect('dashboard_pedidos')

    pedido = get_object_or_404(Pedido, id=pedido_id, empresa=empresa_actual)
    if pedido.estado not in ['confirmado', 'completo']:
        messages.error(request, 'Solo se pueden generar Órdenes de Venta desde pedidos confirmados.')
        return redirect('dashboard_pedidos')

    # Permitir creación solo si no hay una orden de venta ABIERTA para este pedido
    # (Borrador o Aprobado)
    orden_abierta = OrdenVenta.objects.filter(pedido_origen=pedido, estado__in=['borrador', 'aprobado']).exists()
    if orden_abierta:
        messages.warning(request, 'Este pedido ya tiene una Orden de Salida activa.')
        return redirect('dashboard_ventas')

    lineas_pedido = pedido.detalles.filter(parent_line__isnull=True)
    only_services = lineas_pedido.exists() and all(linea.producto.tipo == 'servicio' for linea in lineas_pedido)

    from django.utils import timezone
    from almacenes.models import Almacen

    if only_services:
        estado_inicial = 'surtido'
        entrega_inicial = 'entregado'
        fecha_surtido_inicial = timezone.now()
        almacen_salida = Almacen.objects.filter(empresa=empresa_actual, sucursal=pedido.sucursal).first()
        if not almacen_salida:
            almacen_salida = Almacen.objects.filter(empresa=empresa_actual).first()
    else:
        estado_inicial = 'borrador'
        entrega_inicial = 'pendiente'
        fecha_surtido_inicial = None
        almacen_salida = None

    ov = OrdenVenta.objects.create(
        pedido_origen=pedido,
        cliente=pedido.cliente,
        vendedor=pedido.vendedor, # Se hereda el vendedor del pedido original
        empresa=empresa_actual,
        sucursal=pedido.sucursal, # Hereda la sucursal del pedido
        estado=estado_inicial,
        estado_entrega=entrega_inicial,
        fecha_surtido=fecha_surtido_inicial,
        almacen=almacen_salida
    )

    for linea in lineas_pedido:
        DetalleOrdenVenta.objects.create(
            orden_venta=ov,
            producto=linea.producto,
            cantidad=linea.cantidad_solicitada,
            precio_unitario=linea.precio_unitario or 0
        )

    # NOTIFICACIÓN
    crear_notificacion(
        empresa=empresa_actual,
        mensaje=f"Se ha generado una Orden de Salida #{ov.id} desde el Pedido #{pedido.id}",
        actor=request.user,
        propietario=pedido.vendedor # El dueño es el vendedor original
    )

    if only_services:
        messages.success(request, f'Orden de Salida #{ov.id} creada y surtida automáticamente al ser únicamente de servicios.')
    else:
        messages.success(request, f'Orden de Salida #{ov.id} creada en estado Borrador.')
    return redirect('dashboard_ventas')

@login_required
@require_sales_permission('salidas', 'aprobar')
def cambiar_estado_ov(request, ov_id, nuevo_estado):
    empresa_actual = get_empresa_actual(request)
    ov = get_object_or_404(OrdenVenta, id=ov_id, empresa=empresa_actual)
    if nuevo_estado == 'aprobado' and ov.estado == 'borrador':
        ov.estado = 'aprobado'
        ov.save()

        # NOTIFICACIÓN
        crear_notificacion(
            empresa=empresa_actual,
            mensaje=f"La Orden de Salida #{ov.id} ha sido Aprobada.",
            actor=request.user,
            propietario=ov.vendedor
        )

        messages.success(request, 'Orden de Salida Aprobada. Lista para surtir.')
    return redirect('dashboard_ventas')

@login_required
@require_sales_permission('salidas', 'ver', json_response=True)
def api_preparar_surtido(request, ov_id):
    empresa_actual = get_empresa_actual(request)
    ov = get_object_or_404(OrdenVenta, id=ov_id, empresa=empresa_actual)

    if ov.estado not in ['borrador', 'aprobado', 'surtido']:
        return JsonResponse({'success': False, 'error': 'Estado inválido para ver detalles.'})

    almacen_id_req = request.GET.get('almacen_id')
    
    # --- RESTRICCIÓN: Solo almacenes de la sucursal de la orden ---
    almacenes_disponibles = Almacen.objects.filter(empresa=empresa_actual, sucursal=ov.sucursal)
    
    # Asegurar que almacen_final_id sea un ID numérico válido
    try:
        if almacen_id_req:
            almacen_final_id = int(almacen_id_req)
        elif ov.almacen_id:
            almacen_final_id = ov.almacen_id
        else:
            almacen_final_id = None
    except (ValueError, TypeError):
        almacen_final_id = None
    
    # Si el almacén pre-seleccionado no es de la sucursal (o no hay), tomar el primero disponible
    if almacen_final_id:
        if not almacenes_disponibles.filter(id=almacen_final_id).exists():
            almacen_final_id = almacenes_disponibles.first().id if almacenes_disponibles.exists() else None
    else:
        almacen_final_id = almacenes_disponibles.first().id if almacenes_disponibles.exists() else None

    cliente = ov.cliente
    pedido_origen = ov.pedido_origen
    
    razon_social = cliente.razon_social if cliente.razon_social else ""
    nombre_completo = f"{cliente.nombre} {cliente.apellidos}" if not razon_social else ""
    correo_cliente = getattr(cliente, 'email', '')
    telefono_cliente = getattr(cliente, 'telefono', '')
    
    dir_parts = []
    if cliente.calle: dir_parts.append(cliente.calle)
    if cliente.numero_ext: dir_parts.append(f"#{cliente.numero_ext}")
    if cliente.numero_int: dir_parts.append(f"Int {cliente.numero_int}")
    if cliente.colonia: dir_parts.append(cliente.colonia)
    if cliente.cp: dir_parts.append(f"CP {cliente.cp}")
    if cliente.estado_dir: dir_parts.append(cliente.estado_dir)
    direccion_completa = ", ".join(dir_parts)

    contacto_nombre, contacto_correo, contacto_telefono = "", "", ""
    if pedido_origen and pedido_origen.cotizacion_origen_id:
        try:
            from cotizaciones.models import Cotizacion
            cot = Cotizacion.objects.get(id=pedido_origen.cotizacion_origen_id, empresa=empresa_actual)
            if cot.contacto:
                contacto_nombre = cot.contacto.nombre_completo
                contacto_correo = cot.contacto.correo_1 or cot.contacto.correo_2 or ""
                contacto_telefono = cot.contacto.telefono_1 or cot.contacto.telefono_2 or ""
        except: pass

    envio_parts = []
    if getattr(cliente, 'envio_calle', ''): envio_parts.append(cliente.envio_calle)
    if getattr(cliente, 'envio_numero_ext', ''): envio_parts.append(f"#{cliente.envio_numero_ext}")
    if getattr(cliente, 'envio_numero_int', ''): envio_parts.append(f"Int {cliente.envio_numero_int}")
    if getattr(cliente, 'envio_colonia', ''): envio_parts.append(cliente.envio_colonia)
    if getattr(cliente, 'envio_cp', ''): envio_parts.append(f"CP {cliente.envio_cp}")
    if getattr(cliente, 'envio_estado', ''): envio_parts.append(cliente.envio_estado)
    direccion_cliente_envio = ", ".join(envio_parts)

    final_quien_recibe = ov.quien_recibe or getattr(cliente, 'envio_quien_recibe', '') or cliente.nombre
    final_telefono = ov.telefono_recibe or getattr(cliente, 'envio_telefono', '') or contacto_telefono or telefono_cliente
    final_correo = ov.contacto_envio or getattr(cliente, 'envio_correo', '') or contacto_correo or correo_cliente
    final_notas = ov.notas_envio or getattr(cliente, 'envio_notas', '')

    detalles_data = []
    detalles_data = []
    for det in ov.detalles.all():
        extras_disponibles = []
        if (det.producto.maneja_lote or det.producto.maneja_serie) and almacen_final_id:
            # Filtro directo por producto_id y por el almacén específico (usando IDs numéricos)
            extras_qs = DetalleRecepcionExtra.objects.filter(
                almacen_id=almacen_final_id, 
                producto_id=det.producto_id
            ).filter(Q(cantidad_lote__gt=0) | Q(tipo='serie'))
            
            for extra in extras_qs:
                # Doble validación de stock para lotes
                if extra.tipo == 'lote' and extra.cantidad_lote <= 0: continue
                extras_disponibles.append({
                    'id': extra.id, 
                    'tipo': extra.tipo, 
                    'lote': extra.lote, 
                    'serie': extra.serie, 
                    'cantidad': extra.cantidad_lote if extra.tipo == 'lote' else 1
                })

        detalles_data.append({
            'id': det.id, 
            'producto_id': det.producto.id, 
            'producto_nombre': det.producto.nombre,
            'cantidad': det.cantidad, 
            'precio': float(det.precio_unitario),
            'iva_porcentaje': float(det.producto.iva or 0),
            'subtotal': float(det.subtotal),
            'iva_monto': float(det.iva_monto),
            'total': float(det.total),
            'maneja_lote': det.producto.maneja_lote, 
            'maneja_serie': det.producto.maneja_serie, 
            'extras': extras_disponibles
        })

    return JsonResponse({
        'success': True, 'id': ov.id, 'almacen_id': almacen_final_id,
        'almacenes_validos': list(almacenes_disponibles.values('id', 'nombre')),
        'folio_display': ov.folio_display,
        'cliente_razon': razon_social, 'cliente_nombre': nombre_completo, 'cliente_correo': correo_cliente, 'cliente_telefono': telefono_cliente, 'cliente_direccion': direccion_completa,
        'contacto_nombre': contacto_nombre, 'contacto_correo': contacto_correo, 'contacto_telefono': contacto_telefono,
        'direccion_envio': ov.direccion_envio or direccion_cliente_envio,
        'quien_recibe': final_quien_recibe, 'telefono_recibe': final_telefono, 'email': final_correo, 'guia': ov.guia or '', 'notas_envio': final_notas,
        'detalles': detalles_data, 
        'subtotal_total': float(ov.calcular_subtotal),
        'iva_total': float(ov.calcular_iva),
        'total': float(ov.calcular_total)
    })

@login_required
@require_sales_permission('salidas', 'imprimir')
def imprimir_salida(request, pk):
    """Genera la vista para impresión de orden de salida (PDF)"""
    empresa_actual = get_empresa_actual(request)
    orden = get_object_or_404(OrdenVenta, id=pk, empresa=empresa_actual)

    # Limpiar nombre del vendedor
    vendedor_nombre = orden.vendedor.get_full_name()
    if not vendedor_nombre:
        vendedor_nombre = orden.vendedor.username.split('@')[0]

    context = {
        'orden': orden,
        'empresa': empresa_actual,
        'vendedor_nombre': vendedor_nombre,
    }
    return render(request, 'ventas/imprimir_salida.html', context)

@login_required
@require_sales_permission('salidas', 'surtir_orden', json_response=True)
def ejecutar_surtido(request, ov_id):
    if request.method != 'POST': return JsonResponse({'success': False, 'error': 'Método no permitido'})
    try:
        with transaction.atomic():
            empresa_actual = get_empresa_actual(request)
            ov = get_object_or_404(OrdenVenta, id=ov_id, empresa=empresa_actual)

            if ov.estado == 'surtido':
                return JsonResponse({'success': False, 'error': 'Esta orden ya ha sido surtida.'})

            almacen_id = request.POST.get('almacen_id')
            if not almacen_id: raise ValueError("Debes seleccionar un almacén de salida.")

            # --- RESTRICCIÓN: Solo almacenes de la sucursal de la orden ---
            almacen = get_object_or_404(Almacen, id=almacen_id, empresa=empresa_actual, sucursal=ov.sucursal)

            ov.almacen = almacen
            ov.direccion_envio, ov.quien_recibe = request.POST.get('direccion'), request.POST.get('quien_recibe')
            ov.telefono_recibe, ov.guia, ov.notas_envio = request.POST.get('telefono_recibe'), request.POST.get('guia'), request.POST.get('notas')
            ov.fecha_surtido = timezone.now()

            # 1. Preparar Orden Hija (Backorder) si es necesario
            orden_hija = None
            detalles_actuales = list(ov.detalles.all())

            for det in detalles_actuales:
                producto = det.producto
                cant_solicitada = det.cantidad
                cant_a_entregar = int(request.POST.get(f'cantidad_entregar_{det.id}', cant_solicitada))

                if cant_a_entregar < 0: raise ValueError(f"Cantidad inválida para {producto.nombre}")
                if cant_a_entregar > cant_solicitada: raise ValueError(f"No puedes entregar más de lo solicitado ({cant_solicitada}) para {producto.nombre}")

                # Lógica de Partición: Si queda pendiente
                if cant_a_entregar < cant_solicitada:
                    if not orden_hija:
                        # Crear la cabecera de la hija una sola vez
                        parent = ov.parent_orden if ov.parent_orden else ov
                        nueva_secuencia = parent.hijas.count() + 1

                        orden_hija = OrdenVenta.objects.create(
                            pedido_origen=ov.pedido_origen,
                            cliente=ov.cliente,
                            vendedor=ov.vendedor,
                            empresa=ov.empresa,
                            sucursal=ov.sucursal, # <--- HEREDAR SUCURSAL
                            estado='aprobado', # Lista para surtir el resto
                            parent_orden=parent,
                            secuencia=nueva_secuencia,
                            direccion_envio=ov.direccion_envio,
                            contacto_envio=ov.contacto_envio,
                            notas_envio=ov.notas_envio,
                            quien_recibe=ov.quien_recibe,
                            telefono_recibe=ov.telefono_recibe
                        )

                    # Crear detalle en la hija con el pendiente
                    DetalleOrdenVenta.objects.create(
                        orden_venta=orden_hija,
                        producto=producto,
                        cantidad=cant_solicitada - cant_a_entregar,
                        precio_unitario=det.precio_unitario or 0
                    )

                    # Ajustar la orden actual con lo que realmente se entrega
                    det.cantidad = cant_a_entregar
                    det.save()

                # 2. Procesar Salida de Almacén (Solo si se entrega algo)
                if cant_a_entregar > 0 and producto.tipo != 'servicio':
                    extra_ids = request.POST.getlist(f'extra_id_{det.id}[]')
                    
                    # Preparar datos de trazabilidad (Lotes/Series)
                    extras_data = []
                    if extra_ids:
                        piezas_pendientes = cant_a_entregar
                        for eid in extra_ids:
                            if piezas_pendientes <= 0: break
                            
                            extra_obj = DetalleRecepcionExtra.objects.get(id=eid, almacen=almacen)
                            qty_descontar = 0
                            
                            if extra_obj.tipo == 'serie':
                                qty_descontar = 1
                            else:
                                qty_manual = request.POST.get(f'extra_qty_{det.id}_{eid}')
                                if qty_manual:
                                    qty_descontar = int(qty_manual)
                                else:
                                    qty_descontar = min(piezas_pendientes, extra_obj.cantidad_lote)
                            
                            if qty_descontar > 0:
                                extras_data.append({'id': eid, 'qty': qty_descontar})
                                piezas_pendientes -= qty_descontar

                    # USAR MÉTODO CENTRALIZADO (Vía Transacción para log centralizado)
                    Transaccion.objects.create(
                        producto=producto,
                        almacen=almacen,
                        tipo='venta',
                        cantidad=cant_a_entregar,
                        total=Decimal(cant_a_entregar) * (det.precio_unitario or 0),
                        empresa=empresa_actual,
                        usuario=request.user,
                        referencia=ov.folio_display,
                        extras_data=extras_data,
                        estado='recibida' # Marcamos como procesada
                    )

                    # --- PUNTO B: Sincronización con Pedido Original ---
                    if ov.pedido_origen:
                        # Buscar la partida correspondiente en el pedido original
                        # (Suponiendo que el producto es la llave, o se podría mejorar con un link directo)
                        detalle_pedido = ov.pedido_origen.detalles.filter(producto=producto).first()
                        if detalle_pedido:
                            detalle_pedido.cantidad_entregada += cant_a_entregar
                            # Actualizar estado de la línea
                            if detalle_pedido.cantidad_entregada >= detalle_pedido.cantidad_solicitada:
                                detalle_pedido.estado_linea = 'completo'
                            else:
                                detalle_pedido.estado_linea = 'parcial'
                            detalle_pedido.save()

                # Si se entregó 0 y no se ha creado hija, hay que forzar creación de hija para no perder la partida
                elif cant_a_entregar == 0 and not orden_hija:
                    # (Mismo bloque de creación de hija que arriba)
                    parent = ov.parent_orden if ov.parent_orden else ov
                    nueva_secuencia = parent.hijas.count() + 1
                    orden_hija = OrdenVenta.objects.create(
                        pedido_origen=ov.pedido_origen, cliente=ov.cliente, vendedor=ov.vendedor,
                        empresa=ov.empresa, sucursal=ov.sucursal, # <--- HEREDAR SUCURSAL
                        estado='aprobado', parent_orden=parent, secuencia=nueva_secuencia,
                        direccion_envio=ov.direccion_envio, contacto_envio=ov.contacto_envio
                    )
                    DetalleOrdenVenta.objects.create(
                        orden_venta=orden_hija, producto=producto, cantidad=cant_solicitada, precio_unitario=det.precio_unitario or 0
                    )
                    det.delete() # Se movió completa a la hija

            # --- FINALIZAR PROCESO DE SURTIDO ---
            ov.estado = 'surtido'
            ov.estado_entrega = 'listo'
            ov.save()

            # Actualizar estado global del pedido si aplica
            if ov.pedido_origen:
                p_madre = ov.pedido_origen
                total_solicitado = p_madre.detalles.aggregate(total=Sum('cantidad_solicitada'))['total'] or 0
                total_entregado = p_madre.detalles.aggregate(total=Sum('cantidad_entregada'))['total'] or 0
                
                if total_entregado >= total_solicitado:
                    p_madre.estado = 'completo'
                else:
                    # Si ya se entregó algo pero no todo, lo mantenemos en confirmado (o podrías usar un estado 'parcial')
                    p_madre.estado = 'confirmado'
                p_madre.save()

            crear_notificacion(
                empresa=empresa_actual,
                actor=request.user,
                mensaje=f'surtió la orden {ov.folio_display}',
                propietario=ov.vendedor
            )

            # Solo actualizar pedido si NO hay más órdenes pendientes para este pedido
            if ov.pedido_origen:
                ordenes_pendientes = OrdenVenta.objects.filter(pedido_origen=ov.pedido_origen).exclude(estado__in=['surtido', 'cancelado']).count()
                if ordenes_pendientes == 0:
                    pedido = ov.pedido_origen
                    pedido.estado = 'completo'
                    pedido.save()

            msg = f'Orden {ov.folio_display} surtida correctamente.'
            if orden_hija:
                msg += f' Se creó la orden pendiente {orden_hija.folio_display}.'

            return JsonResponse({'success': True, 'message': msg})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_sales_permission('salidas', 'actualizar_entrega')
def actualizar_estado_entrega(request, ov_id):
    if request.method != 'POST':
        return redirect('dashboard_ventas')
    
    empresa_actual = get_empresa_actual(request)
    ov = get_object_or_404(OrdenVenta, id=ov_id, empresa=empresa_actual)
    nuevo_estado = request.POST.get('nuevo_estado')
    
    if nuevo_estado in ['transito', 'entregado']:
        ov.estado_entrega = nuevo_estado
        ov.save()
        
        crear_notificacion(
            empresa=empresa_actual,
            actor=request.user,
            mensaje=f'marcó como {ov.get_estado_entrega_display()} la orden {ov.folio_display}',
            propietario=ov.vendedor
        )
        
        messages.success(request, f"Estado de entrega actualizado a {ov.get_estado_entrega_display()}.")
    
    return redirect('dashboard_ventas')

# --- PUNTO DE VENTA (POS) ---

@login_required
def punto_de_venta(request):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'punto_de_venta', 'ver'):
        return render(request, 'error_sin_empresa.html', status=403)

    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)

    sucursal_id = request.session.get('sucursal_id')
    from preferencias.models import Sucursal
    sucursal = None
    if sucursal_id:
        try:
            sucursal = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
        except Sucursal.DoesNotExist:
            pass

    if not sucursal:
        sucursal = Sucursal.objects.filter(empresa=empresa_actual).first()

    if not sucursal:
        messages.warning(request, "Debes configurar al menos una sucursal en Preferencias antes de usar el Punto de Venta.")
        return redirect('dashboard_ventas')

    cliente_defecto = sucursal.cliente_defecto

    from core.models import Producto
    from almacenes.models import Inventario, Almacen
    
    almacen = Almacen.objects.filter(empresa=empresa_actual, sucursal=sucursal).first()
    from categorias.models import ListaPrecioCosto
    listas_precio_globales = ListaPrecioCosto.objects.filter(empresa=empresa_actual, tipo='precio')

    lista_activa_nombre = ""
    for lista in listas_precio_globales:
        if lista.esta_activa_ahora():
            lista_activa_nombre = lista.nombre
            break

    productos_data = []
    productos = Producto.objects.filter(empresa=empresa_actual, estado='activo', mostrar_en_pos=True).order_by('nombre')
    
    for prod in productos:
        stock = 0
        if prod.tipo != 'servicio' and almacen:
            inv = Inventario.objects.filter(almacen=almacen, producto=prod).first()
            physical_stock = (inv.cantidad - inv.reservado) if inv else 0

            from core.models import DetalleReceta
            receta = DetalleReceta.objects.filter(producto_padre=prod)
            if receta.exists():
                import math
                max_posible = None
                for r in receta:
                    if r.componente.tipo != 'servicio':
                        inv_comp = Inventario.objects.filter(almacen=almacen, producto=r.componente).first()
                        stock_disp_comp = (inv_comp.cantidad - inv_comp.reservado) if inv_comp else 0
                        if r.cantidad > 0:
                            posibles_con_este = float(stock_disp_comp) / float(r.cantidad)
                            if max_posible is None or posibles_con_este < max_posible:
                                max_posible = posibles_con_este
                        else:
                            posibles_con_este = 0.0
                            if max_posible is None or posibles_con_este < max_posible:
                                max_posible = posibles_con_este
                
                stock = physical_stock + (math.floor(max_posible) if max_posible is not None else 0)
            else:
                stock = physical_stock
        
        import math
        precio_redondeado = math.ceil(float(prod.precio_venta or 0))
        
        # Obtener y normalizar precios específicos del producto
        precios_articulo = prod.precios_lista or []
        nombres_precios_articulo = set()
        precios_completos = []
        
        for item in precios_articulo:
            if isinstance(item, dict) and 'nombre' in item and 'monto' in item:
                try:
                    precios_completos.append({
                        'nombre': item['nombre'],
                        'monto': float(item['monto'])
                    })
                    nombres_precios_articulo.add(item['nombre'])
                except (ValueError, TypeError):
                    continue

        # Calcular e integrar precios provenientes de listas globales si no están sobreescritos
        base_venta = float(prod.precio_venta or 0)
        for lista in listas_precio_globales:
            if lista.nombre not in nombres_precios_articulo:
                porc = float(lista.porcentaje_extra or 0)
                monto_fijo = float(lista.monto_extra or 0)
                monto_calculado = base_venta + (base_venta * (porc / 100.0)) + monto_fijo
                precios_completos.append({
                    'nombre': lista.nombre,
                    'monto': float(monto_calculado)
                })

        mods_data = []
        if prod.permitir_modificadores:
            for m in prod.modificadores.all().select_related('producto_modificador'):
                mods_data.append({
                    'id': m.id,
                    'producto_modificador_id': m.producto_modificador.id,
                    'nombre': m.producto_modificador.nombre,
                    'permite_extra': m.permite_extra,
                    'permite_sin': m.permite_sin,
                    'precio_extra': float(m.precio_extra),
                    'cantidad_modificadora': float(m.cantidad_modificadora)
                })

        import json
        productos_data.append({
            'id': prod.id,
            'nombre': prod.nombre,
            'sku': prod.clave or '',
            'tipo': prod.tipo,
            'categoria': prod.categoria or '',
            'subcategoria': prod.subcategoria or '',
            'precio_venta': float(precio_redondeado),
            'iva': float(prod.iva or 0),
            'stock': stock,
            'imagen_url': prod.imagen.url if prod.imagen else '',
            'precios_lista_json': json.dumps(precios_completos),
            'modificadores_json': json.dumps(mods_data)
        })

    from tesoreria.models import CajaBanco
    cajas_bancos = CajaBanco.objects.filter(empresa=empresa_actual, activo=True)

    from clientes.models import Cliente
    clientes = Cliente.objects.filter(empresa=empresa_actual, estado='activo').order_by('razon_social', 'nombre')

    categorias_unicas = Producto.objects.filter(empresa=empresa_actual, estado='activo', mostrar_en_pos=True).exclude(categoria__isnull=True).exclude(categoria='').values_list('categoria', flat=True).distinct().order_by('categoria')
    subcategorias_unicas = Producto.objects.filter(empresa=empresa_actual, estado='activo', mostrar_en_pos=True).exclude(subcategoria__isnull=True).exclude(subcategoria='').values_list('subcategoria', flat=True).distinct().order_by('subcategoria')

    from ventas.models import CajaPOS, SesionCajaPOS
    # Verificar sesión activa en base de datos
    sesion_activa = SesionCajaPOS.objects.filter(usuario=request.user, estado='abierta').first()
    if sesion_activa:
        request.session['sesion_caja_id'] = sesion_activa.id
    else:
        if 'sesion_caja_id' in request.session:
            del request.session['sesion_caja_id']

    # Cajas habilitadas (abiertas por administración) y asignadas a este usuario
    cajas_disponibles = CajaPOS.objects.filter(empresa=empresa_actual, sucursal=sucursal, estado='abierta', usuario_asignado=request.user)

    contexto = {
        'empresa': empresa_actual,
        'sucursal': sucursal,
        'almacen': almacen,
        'cliente_defecto': cliente_defecto,
        'productos': productos_data,
        'cajas_bancos': cajas_bancos,
        'clientes': clientes,
        'categorias': list(categorias_unicas),
        'subcategorias': list(subcategorias_unicas),
        'sesion_activa': sesion_activa,
        'cajas_disponibles': cajas_disponibles,
        'lista_activa_nombre': lista_activa_nombre,
        'section': 'pos',
    }
    return render(request, 'ventas/punto_de_venta.html', contexto)


@login_required
@require_POST
def crear_venta_pos_ajax(request):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'punto_de_venta', 'crear'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar ventas.'})

    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'Empresa no encontrada.'})

    from ventas.models import SesionCajaPOS
    sesion_id = request.session.get('sesion_caja_id')
    sesion = None
    if sesion_id:
        sesion = SesionCajaPOS.objects.filter(id=sesion_id, usuario=request.user, estado='abierta').first()
    if not sesion:
        sesion = SesionCajaPOS.objects.filter(usuario=request.user, estado='abierta').first()
    if not sesion:
        return JsonResponse({'success': False, 'error': 'No tienes una sesión de caja abierta. Por favor, abre una sesión primero.'})

    try:
        import json
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos.'})

    cliente_id = data.get('cliente_id')
    items = data.get('items', [])
    pagos = data.get('pagos', [])
    aplica_iva = data.get('aplica_iva', True)

    # Validar permisos granulares de POS en el backend
    # 1. Desactivar IVA
    if not aplica_iva:
        if not user_has_sales_permission(request, 'punto_de_venta', 'desactivar_iva'):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para desactivar el IVA.'})

    # 2. Descuentos
    from decimal import Decimal
    descuento_val = Decimal(str(data.get('descuento', 0)))
    if descuento_val > 0:
        if not user_has_sales_permission(request, 'punto_de_venta', 'descuento'):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para aplicar descuentos.'})

    # 3. Listas de precios manuales
    from categorias.models import ListaPrecioCosto
    listas_precio_globales = ListaPrecioCosto.objects.filter(empresa=empresa_actual, tipo='precio')
    lista_activa_nombre = ""
    for lista in listas_precio_globales:
        if lista.esta_activa_ahora():
            lista_activa_nombre = lista.nombre
            break

    for item in items:
        lista_sel = item.get('lista_seleccionada', '')
        if lista_sel and lista_sel != lista_activa_nombre:
            if not user_has_sales_permission(request, 'punto_de_venta', 'listas'):
                return JsonResponse({'success': False, 'error': f'No tienes permisos para cambiar listas de precios (se detectó selección manual: {lista_sel}).'})

    if not items:
        return JsonResponse({'success': False, 'error': 'El carrito está vacío.'})
    if not pagos:
        return JsonResponse({'success': False, 'error': 'Debes ingresar al menos un método de pago.'})

    from preferencias.models import Sucursal
    sucursal_id = request.session.get('sucursal_id')
    sucursal = None
    if sucursal_id:
        try:
            sucursal = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
        except Sucursal.DoesNotExist:
            pass
    if not sucursal:
        sucursal = Sucursal.objects.filter(empresa=empresa_actual).first()
    if not sucursal:
        return JsonResponse({'success': False, 'error': 'No se detectó sucursal activa.'})

    from clientes.models import Cliente
    cliente = None
    if cliente_id:
        try:
            cliente = Cliente.objects.get(id=cliente_id, empresa=empresa_actual)
        except Cliente.DoesNotExist:
            pass
    if not cliente:
        cliente = sucursal.cliente_defecto
    if not cliente:
        return JsonResponse({'success': False, 'error': 'No hay un cliente configurado por defecto para la sucursal ni seleccionado en el carrito.'})

    from almacenes.models import Almacen, Inventario
    almacen = Almacen.objects.filter(empresa=empresa_actual, sucursal=sucursal).first()
    if not almacen:
        return JsonResponse({'success': False, 'error': 'No se encontró ningún almacén configurado para la sucursal actual para procesar la salida de inventario.'})

    from core.models import Producto, Transaccion
    from pedidos.models import Pedido, DetallePedido
    from ventas.models import OrdenVenta, DetalleOrdenVenta
    from tesoreria.models import PagoPedido, Ingreso, CajaBanco
    from django.db import transaction
    from decimal import Decimal

    try:
        with transaction.atomic():
            # 1. Calcular total_pedido aplicando el factor de descuento proporcional
            descuento = Decimal(str(data.get('descuento', 0)))
            descuento_tipo = data.get('descuento_tipo', 'monto')
            
            subtotal_original = Decimal('0.00')
            for item in items:
                qty = int(item.get('cantidad', 1))
                price = Decimal(str(item.get('precio_unitario', 0)))
                subtotal_original += qty * price

            descuento_monto = Decimal('0.00')
            if descuento > 0:
                if descuento_tipo == 'porcentaje':
                    descuento_monto = subtotal_original * (descuento / 100)
                else:
                    descuento_monto = descuento

            factor = Decimal('1.00')
            if subtotal_original > 0 and descuento_monto > 0:
                factor = (subtotal_original - descuento_monto) / subtotal_original
                if factor < 0:
                    factor = Decimal('0.00')

            total_pedido = Decimal('0.00')
            for item in items:
                prod_id = item.get('producto_id')
                qty = int(item.get('cantidad', 1))
                price = Decimal(str(item.get('precio_unitario', 0))) * factor

                try:
                    producto = Producto.objects.get(id=prod_id, empresa=empresa_actual, estado='activo')
                except Producto.DoesNotExist:
                    raise ValueError(f"El producto con ID {prod_id} no existe o está inactivo.")

                subtotal = qty * price
                porc = (producto.iva or Decimal('0.00')) if aplica_iva else Decimal('0.00')
                iva_monto = subtotal * (porc / 100)
                total_pedido += (subtotal + iva_monto)

            total_pedido = total_pedido.quantize(Decimal('1'), rounding='ROUND_CEILING')

            # 2. Validar y procesar pagos
            total_pagos_ingresados = Decimal('0.00')
            for p in pagos:
                total_pagos_ingresados += Decimal(str(p.get('monto', 0)))

            if total_pagos_ingresados < total_pedido:
                raise ValueError(f"El monto total pagado (${total_pagos_ingresados}) es menor al total de la venta (${total_pedido}).")

            cambio = total_pagos_ingresados - total_pedido

            pagos_ajustados = []
            for p in pagos:
                fp = p.get('forma_pago')
                db_fp = fp
                if fp == 'efectivo':
                    caja_banco = sesion.caja_pos.caja_efectivo
                elif fp in ['tarjeta', 'tarjeta_debito', 'tarjeta_credito']:
                    caja_banco = sesion.caja_pos.banco_tarjeta
                    db_fp = 'tarjeta_debito'
                elif fp == 'transferencia':
                    caja_banco = sesion.caja_pos.banco_transferencia
                else:
                    caja_banco = sesion.caja_pos.caja_efectivo

                if not caja_banco:
                    raise ValueError(f"No se ha configurado la cuenta de destino para el método de pago {fp} en la Caja POS.")

                pagos_ajustados.append({
                    'caja_banco_id': caja_banco.id,
                    'forma_pago': db_fp,
                    'monto': Decimal(str(p.get('monto', 0)))
                })

            if cambio > 0:
                # Restar cambio del pago en efectivo si existe
                efectivo_pago = next((p for p in pagos_ajustados if p['forma_pago'] == 'efectivo'), None)
                if efectivo_pago:
                    if efectivo_pago['monto'] >= cambio:
                        efectivo_pago['monto'] -= cambio
                        cambio = Decimal('0')
                    else:
                        cambio -= efectivo_pago['monto']
                        efectivo_pago['monto'] = Decimal('0')

                # Si aún queda cambio, restar de los demás pagos
                if cambio > 0:
                    for p in pagos_ajustados:
                        if p['monto'] >= cambio:
                            p['monto'] -= cambio
                            cambio = Decimal('0')
                            break
                        else:
                            cambio -= p['monto']
                            p['monto'] = Decimal('0')

            pagos_ajustados = [p for p in pagos_ajustados if p['monto'] > 0]

            # 3. Crear el Pedido y la Orden de Venta (Salida)
            import json
            notas_data = {}
            if descuento > 0:
                notas_data['descuento'] = float(descuento)
                notas_data['descuento_tipo'] = descuento_tipo
                notas_data['descuento_monto'] = float(descuento_monto)

            pedido = Pedido.objects.create(
                cliente=cliente,
                vendedor=request.user,
                empresa=empresa_actual,
                sucursal=sucursal,
                estado='completo',
                aplica_iva=aplica_iva,
                sesion_caja=sesion,
                notas=json.dumps(notas_data) if notas_data else ""
            )

            from django.utils import timezone
            orden_venta = OrdenVenta.objects.create(
                pedido_origen=pedido,
                cliente=cliente,
                vendedor=request.user,
                empresa=empresa_actual,
                sucursal=sucursal,
                estado='surtido',
                estado_entrega='entregado',
                almacen=almacen,
                fecha_surtido=timezone.now()
            )

            # 4. Crear DetallePedido, DetalleOrdenVenta y Transaccion
            for item in items:
                prod_id = item.get('producto_id')
                qty = int(item.get('cantidad', 1))
                price = Decimal(str(item.get('precio_unitario', 0))) * factor
                producto = Producto.objects.get(id=prod_id, empresa=empresa_actual, estado='activo')
                modificadores = item.get('modificadores', [])

                if producto.tipo != 'servicio':
                    from core.models import DetalleReceta, ModificadorProducto
                    receta = DetalleReceta.objects.filter(producto_padre=producto)
                    
                    if receta.exists() or modificadores:
                        componentes_req = {}
                        if receta.exists():
                            for r in receta:
                                componentes_req[r.componente.id] = r.cantidad
                        else:
                            componentes_req[producto.id] = Decimal('1.0000')

                        for m_item in modificadores:
                            m_id = m_item.get('id')
                            m_tipo = m_item.get('tipo')
                            if m_tipo == 'sin':
                                if m_id in componentes_req:
                                    componentes_req[m_id] = Decimal('0.0000')
                            elif m_tipo == 'extra':
                                mod_obj = ModificadorProducto.objects.filter(
                                    producto_padre=producto,
                                    producto_modificador_id=m_id,
                                    empresa=empresa_actual
                                ).first()
                                cant_extra = mod_obj.cantidad_modificadora if mod_obj else Decimal('1.0000')
                                componentes_req[m_id] = componentes_req.get(m_id, Decimal('0.0000')) + cant_extra

                        for comp_id, cant_unit in componentes_req.items():
                            if cant_unit > 0:
                                try:
                                    comp_prod = Producto.objects.get(id=comp_id, empresa=empresa_actual, estado='activo')
                                except Producto.DoesNotExist:
                                    raise ValueError(f"El ingrediente/modificador con ID {comp_id} no existe o está inactivo.")
                                
                                if comp_prod.tipo != 'servicio':
                                    cant_req = cant_unit * qty
                                    inv_comp = Inventario.objects.filter(almacen=almacen, producto=comp_prod).first()
                                    stock_disp_comp = (inv_comp.cantidad - inv_comp.reservado) if inv_comp else 0
                                    if stock_disp_comp < cant_req:
                                        raise ValueError(f"Stock insuficiente de {comp_prod.nombre} para preparar {producto.nombre}. Disponible: {stock_disp_comp}, Requerido: {cant_req}")
                    else:
                        inv = Inventario.objects.filter(almacen=almacen, producto=producto).first()
                        stock_disp = (inv.cantidad - inv.reservado) if inv else 0
                        if stock_disp < qty:
                            raise ValueError(f"Stock insuficiente para {producto.nombre}. Disponible: {stock_disp}, Solicitado: {qty}")

                detalle = DetallePedido.objects.create(
                    pedido=pedido,
                    producto=producto,
                    cantidad_solicitada=qty,
                    cantidad_entregada=qty,
                    precio_unitario=price,
                    estado_linea='completo',
                    modificadores_json=json.dumps(modificadores) if modificadores else None
                )

                DetalleOrdenVenta.objects.create(
                    orden_venta=orden_venta,
                    producto=producto,
                    cantidad=qty,
                    precio_unitario=price,
                    modificadores_json=json.dumps(modificadores) if modificadores else None
                )

                if producto.tipo != 'servicio':
                    from core.models import DetalleReceta
                    receta = DetalleReceta.objects.filter(producto_padre=producto)
                    if receta.exists() or modificadores:
                        if receta.exists():
                            Transaccion.objects.create(
                                producto=producto,
                                almacen=almacen,
                                tipo='produccion',
                                cantidad=qty,
                                total=Decimal(qty) * producto.precio_costo,
                                empresa=empresa_actual,
                                usuario=request.user,
                                referencia=f"Autoproducción POS Pedido #{pedido.id}",
                                estado='recibida'
                            )
                        for comp_id, cant_unit in componentes_req.items():
                            if cant_unit > 0:
                                try:
                                    comp_prod = Producto.objects.get(id=comp_id, empresa=empresa_actual, estado='activo')
                                except Producto.DoesNotExist:
                                    continue
                                cant_req = cant_unit * qty
                                Transaccion.objects.create(
                                    producto=comp_prod,
                                    almacen=almacen,
                                    tipo='produccion',
                                    cantidad=-cant_req,
                                    total=Decimal(cant_req) * comp_prod.precio_costo,
                                    empresa=empresa_actual,
                                    usuario=request.user,
                                    referencia=f"Consumo Receta {producto.nombre} con Modificadores - Pedido #{pedido.id}",
                                    estado='recibida'
                                )

                    if receta.exists() or not modificadores:
                        Transaccion.objects.create(
                            producto=producto,
                            almacen=almacen,
                            tipo='venta',
                            cantidad=-qty,
                            total=Decimal(qty) * price,
                            empresa=empresa_actual,
                            usuario=request.user,
                            referencia=f"Venta POS Pedido #{pedido.id}",
                            estado='recibida'
                        )

            # 5. Crear los registros de pago e ingreso
            for idx, p in enumerate(pagos_ajustados):
                caja_id = p['caja_banco_id']
                try:
                    caja_banco = CajaBanco.objects.get(id=caja_id, empresa=empresa_actual, activo=True)
                except CajaBanco.DoesNotExist:
                    raise ValueError(f"La Caja o Banco con ID {caja_id} no es válido o está inactivo.")

                pago_pedido = PagoPedido.objects.create(
                    fecha_pago=timezone.now().date(),
                    forma_pago=p['forma_pago'],
                    referencia=f"Venta POS Pedido #{pedido.id} (Pago {idx+1})",
                    tipo_cambio=Decimal('1.00'),
                    monto=p['monto'],
                    monto_mxn=p['monto'],
                    caja_banco=caja_banco,
                    empresa=empresa_actual,
                    moneda=caja_banco.moneda,
                    pedido=pedido,
                    estado='aplicado'
                )

                Ingreso.objects.create(
                    fecha=timezone.now().date(),
                    concepto=f"Cobro POS - Pedido #{pedido.id} (Pago {idx+1})",
                    monto=p['monto'],
                    moneda=caja_banco.moneda,
                    tipo_cambio=Decimal('1.00'),
                    monto_mxn=p['monto'],
                    forma_pago=p['forma_pago'],
                    caja_banco=caja_banco,
                    referencia=f"Venta POS Pedido #{pedido.id} (Pago {idx+1})",
                    pago_pedido=pago_pedido,
                    estado='aplicado',
                    empresa=empresa_actual,
                    sucursal=sucursal
                )

            return JsonResponse({'success': True, 'message': 'Venta registrada y pagada correctamente.', 'pedido_id': pedido.id})

    except ValueError as ve:
        return JsonResponse({'success': False, 'error': str(ve)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f"Error interno: {str(e)}"})


# ==========================================
# CORTES DE CAJA VIEWS
# ==========================================
from django.contrib.auth.models import User
from tesoreria.models import CajaBanco, PagoPedido
from django.db.models import Sum
from .models import CajaPOS, SesionCajaPOS, CorteZ

@login_required
def cortes_caja_list(request):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'cortes_de_caja', 'ver'):
        return render(request, 'error_sin_empresa.html', status=403)

    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)

    sucursal_id = request.session.get('sucursal_id')
    from preferencias.models import Sucursal
    sucursal = None
    if sucursal_id:
        try:
            sucursal = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
        except Sucursal.DoesNotExist:
            pass
    if not sucursal:
        sucursal = Sucursal.objects.filter(empresa=empresa_actual).first()

    # Listado de cajas configuradas para esta empresa (todas las sucursales)
    cajas_pos = CajaPOS.objects.filter(empresa=empresa_actual)
    
    # Historial de sesiones (todas las sucursales) y Cortes Z (cierres diarios)
    sesiones_qs = SesionCajaPOS.objects.filter(caja_pos__empresa=empresa_actual).order_by('-fecha_apertura')
    cortes_z_qs = CorteZ.objects.filter(empresa=empresa_actual).order_by('-fecha_creacion')

    # Aplicar filtros
    q = request.GET.get('q', '').strip()
    cajero_id = request.GET.get('cajero', '').strip()
    fecha_str = request.GET.get('fecha', '').strip()
    estado = request.GET.get('estado', '').strip()
    sucursal_filtro_id = request.GET.get('sucursal', '').strip()

    if q:
        from django.db.models import Q
        sesiones_qs = sesiones_qs.filter(
            Q(usuario__username__icontains=q) |
            Q(usuario__first_name__icontains=q) |
            Q(usuario__last_name__icontains=q) |
            Q(caja_pos__nombre__icontains=q) |
            Q(caja_pos__sucursal__nombre__icontains=q)
        )
        cortes_z_qs = cortes_z_qs.filter(
            Q(usuario__username__icontains=q) |
            Q(usuario__first_name__icontains=q) |
            Q(usuario__last_name__icontains=q)
        )

    if cajero_id:
        sesiones_qs = sesiones_qs.filter(usuario_id=cajero_id)
        cortes_z_qs = cortes_z_qs.filter(usuario_id=cajero_id)

    if fecha_str:
        sesiones_qs = sesiones_qs.filter(fecha_apertura__date=fecha_str)
        cortes_z_qs = cortes_z_qs.filter(fecha=fecha_str)

    if estado:
        if estado == 'corte_z':
            sesiones_qs = sesiones_qs.none()
        elif estado in ['abierta', 'cerrada']:
            cortes_z_qs = cortes_z_qs.none()
        else:
            sesiones_qs = sesiones_qs.filter(estado=estado)
            cortes_z_qs = cortes_z_qs.none()

    if sucursal_filtro_id:
        sesiones_qs = sesiones_qs.filter(caja_pos__sucursal_id=sucursal_filtro_id)
        cortes_z_qs = cortes_z_qs.none()

    # Combinar y etiquetar los registros
    sesiones = []
    for s in sesiones_qs:
        s.tipo_corte = 'X'
        s.fecha_orden = s.fecha_apertura
        sesiones.append(s)

    for cz in cortes_z_qs:
        cz.tipo_corte = 'Z'
        cz.fecha_orden = cz.fecha_creacion
        sesiones.append(cz)

    # Ordenar por fecha_orden descendente
    sesiones.sort(key=lambda x: x.fecha_orden, reverse=True)

    # Datos para filtros y modales de creación
    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    usuarios = User.objects.filter(is_active=True, username__contains=f"@{empresa_actual.subdominio}").order_by('username')
    cajas_efectivo = CajaBanco.objects.filter(empresa=empresa_actual, activo=True, tipo='caja')
    if not cajas_efectivo.exists():
        cajas_efectivo = CajaBanco.objects.filter(empresa=empresa_actual, activo=True)
        
    bancos = CajaBanco.objects.filter(empresa=empresa_actual, activo=True, tipo='banco')
    if not bancos.exists():
        bancos = CajaBanco.objects.filter(empresa=empresa_actual, activo=True)

    filtros = {
        'q': q,
        'cajero': cajero_id,
        'fecha': fecha_str,
        'estado': estado,
        'sucursal': sucursal_filtro_id,
    }

    contexto = {
        'cajas_pos': cajas_pos,
        'sesiones': sesiones,
        'usuarios': usuarios,
        'sucursales': sucursales,
        'cajas_efectivo': cajas_efectivo,
        'bancos': bancos,
        'empresa': empresa_actual,
        'sucursal': sucursal,
        'filtros': filtros,
        'section': 'cortes_caja',
    }
    return render(request, 'ventas/cortes_caja.html', contexto)


@login_required
@require_POST
def crear_caja_pos_ajax(request):
    from preferencias.permissions import user_has_sales_permission
    caja_id = request.POST.get('caja_id')
    
    if not user_has_sales_permission(request, 'cortes_de_caja', 'nueva_caja'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para configurar cajas.'})

    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'Empresa no encontrada.'})

    sucursal_id = request.session.get('sucursal_id')
    from preferencias.models import Sucursal
    sucursal = None
    if sucursal_id:
        try:
            sucursal = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
        except Sucursal.DoesNotExist:
            pass
    if not sucursal:
        sucursal = Sucursal.objects.filter(empresa=empresa_actual).first()

    nombre = request.POST.get('nombre')
    usuario_id = request.POST.get('usuario_asignado')
    caja_efectivo_id = request.POST.get('caja_efectivo')
    banco_tarjeta_id = request.POST.get('banco_tarjeta')
    banco_transferencia_id = request.POST.get('banco_transferencia')

    if not (nombre and usuario_id and caja_efectivo_id and banco_tarjeta_id and banco_transferencia_id):
        return JsonResponse({'success': False, 'error': 'Todos los campos son obligatorios.'})

    try:
        usuario = User.objects.get(id=usuario_id, is_active=True)
        caja_efectivo = CajaBanco.objects.get(id=caja_efectivo_id, empresa=empresa_actual)
        banco_tarjeta = CajaBanco.objects.get(id=banco_tarjeta_id, empresa=empresa_actual)
        banco_transferencia = CajaBanco.objects.get(id=banco_transferencia_id, empresa=empresa_actual)

        if caja_id:
            try:
                caja = CajaPOS.objects.get(id=caja_id, empresa=empresa_actual)
                caja.nombre = nombre
                caja.usuario_asignado = usuario
                caja.caja_efectivo = caja_efectivo
                caja.banco_tarjeta = banco_tarjeta
                caja.banco_transferencia = banco_transferencia
                caja.save()
                return JsonResponse({'success': True, 'message': 'Caja de punto de venta actualizada correctamente.'})
            except CajaPOS.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'La caja a editar no existe.'})
        else:
            caja = CajaPOS.objects.create(
                nombre=nombre,
                usuario_asignado=usuario,
                caja_efectivo=caja_efectivo,
                banco_tarjeta=banco_tarjeta,
                banco_transferencia=banco_transferencia,
                estado='cerrada',
                empresa=empresa_actual,
                sucursal=sucursal
            )
            return JsonResponse({'success': True, 'message': 'Caja de punto de venta creada correctamente.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def cambiar_estado_caja_pos_ajax(request, caja_id):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'cortes_de_caja', 'cajas'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para modificar cajas.'})

    empresa_actual = get_empresa_actual(request)
    caja = get_object_or_404(CajaPOS, id=caja_id, empresa=empresa_actual)

    nuevo_estado = request.POST.get('estado')
    if nuevo_estado not in ['abierta', 'cerrada']:
        return JsonResponse({'success': False, 'error': 'Estado inválido.'})

    caja.estado = nuevo_estado
    caja.save()
    return JsonResponse({'success': True, 'message': f'La caja ahora está {caja.get_estado_display().lower()}.'})


@login_required
@require_POST
def apertura_sesion_pos_ajax(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'Empresa no encontrada.'})

    caja_id = request.POST.get('caja_id')
    monto_inicial_str = request.POST.get('monto_inicial', '0.00')

    try:
        monto_inicial = Decimal(monto_inicial_str)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Monto inicial inválido.'})

    if not caja_id:
        return JsonResponse({'success': False, 'error': 'Debes seleccionar una caja registradora.'})

    # Verificar si el usuario ya tiene una sesión abierta
    sesion_abierta = SesionCajaPOS.objects.filter(usuario=request.user, estado='abierta').first()
    if sesion_abierta:
        request.session['sesion_caja_id'] = sesion_abierta.id
        return JsonResponse({'success': True, 'message': 'Ya tenías una sesión activa.', 'sesion_id': sesion_abierta.id})

    try:
        caja_pos = CajaPOS.objects.get(id=caja_id, empresa=empresa_actual, estado='abierta')
        # Crear la sesión
        sesion = SesionCajaPOS.objects.create(
            caja_pos=caja_pos,
            usuario=request.user,
            monto_inicial=monto_inicial,
            estado='abierta'
        )
        request.session['sesion_caja_id'] = sesion.id
        return JsonResponse({'success': True, 'message': 'Sesión de caja abierta correctamente.', 'sesion_id': sesion.id})
    except CajaPOS.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'La caja seleccionada no está habilitada o abierta por administración.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def obtener_totales_sesion_ajax(request):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'punto_de_venta', 'hacer_corte_pos'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para hacer el corte de caja.'}, status=403)

    empresa_actual = get_empresa_actual(request)
    sesion_id = request.session.get('sesion_caja_id')
    sesion = None
    if sesion_id:
        sesion = SesionCajaPOS.objects.filter(id=sesion_id, usuario=request.user, estado='abierta').first()
    if not sesion:
        sesion = SesionCajaPOS.objects.filter(usuario=request.user, estado='abierta').first()

    if not sesion:
        return JsonResponse({'success': False, 'error': 'No hay sesión de caja abierta.'})

    # Calcular montos dinámicamente de los pagos aplicados
    pagos = PagoPedido.objects.filter(pedido__sesion_caja=sesion, estado='aplicado')
    
    ventas_efectivo = pagos.filter(forma_pago='efectivo').aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')
    ventas_tarjeta = pagos.filter(forma_pago__in=['tarjeta_debito', 'tarjeta_credito']).aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')
    ventas_transferencia = pagos.filter(forma_pago='transferencia').aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')

    efectivo_estimado = sesion.monto_inicial + ventas_efectivo

    return JsonResponse({
        'success': True,
        'caja_nombre': sesion.caja_pos.nombre,
        'monto_inicial': float(sesion.monto_inicial),
        'ventas_efectivo': float(ventas_efectivo),
        'ventas_tarjeta': float(ventas_tarjeta),
        'ventas_transferencia': float(ventas_transferencia),
        'efectivo_estimado': float(efectivo_estimado),
        'fecha_apertura': sesion.fecha_apertura.strftime('%Y-%m-%d %H:%M:%S')
    })


@login_required
@require_POST
def cierre_sesion_pos_ajax(request):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'punto_de_venta', 'hacer_corte_pos'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar el corte de caja.'}, status=403)

    empresa_actual = get_empresa_actual(request)
    sesion_id = request.session.get('sesion_caja_id')
    sesion = None
    if sesion_id:
        sesion = SesionCajaPOS.objects.filter(id=sesion_id, usuario=request.user, estado='abierta').first()
    if not sesion:
        sesion = SesionCajaPOS.objects.filter(usuario=request.user, estado='abierta').first()

    if not sesion:
        return JsonResponse({'success': False, 'error': 'No hay sesión de caja activa para cerrar.'})

    monto_final_str = request.POST.get('monto_final_efectivo', '0.00')
    try:
        monto_final = Decimal(monto_final_str)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Monto final de efectivo inválido.'})

    try:
        # Calcular los montos dinámicos del sistema para congelarlos
        pagos = PagoPedido.objects.filter(pedido__sesion_caja=sesion, estado='aplicado')
        
        ventas_efectivo = pagos.filter(forma_pago='efectivo').aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')
        ventas_tarjeta = pagos.filter(forma_pago__in=['tarjeta_debito', 'tarjeta_credito']).aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')
        ventas_transferencia = pagos.filter(forma_pago='transferencia').aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')

        with transaction.atomic():
            # Actualizar la sesión
            sesion.monto_final_efectivo = monto_final
            sesion.total_ventas_efectivo = ventas_efectivo
            sesion.total_ventas_tarjeta = ventas_tarjeta
            sesion.total_ventas_transferencia = ventas_transferencia
            sesion.estado = 'cerrada'
            sesion.fecha_cierre = timezone.now()
            sesion.save()

            # Cerrar también la caja POS asociada (para exigir apertura del administrador la siguiente vez)
            caja = sesion.caja_pos
            caja.estado = 'cerrada'
            caja.save()

            # Limpiar la variable de sesión
            if 'sesion_caja_id' in request.session:
                del request.session['sesion_caja_id']

        return JsonResponse({
            'success': True, 
            'message': 'Corte y cierre de caja procesado correctamente.',
            'sesion_id': sesion.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def cierre_sesion_pos_por_id_ajax(request, sesion_id):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'cortes_de_caja', 'hacer_corte'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar el corte de caja.'}, status=403)

    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'Empresa no encontrada.'})

    from django.shortcuts import get_object_or_404
    from django.utils import timezone
    from decimal import Decimal
    from django.db import transaction
    from django.db.models import Sum
    from tesoreria.models import PagoPedido
    
    sesion = get_object_or_404(SesionCajaPOS, id=sesion_id, caja_pos__empresa=empresa_actual, estado='abierta')

    monto_final_str = request.POST.get('monto_final_efectivo', '0.00')
    try:
        monto_final = Decimal(monto_final_str)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Monto final de efectivo inválido.'})

    try:
        # Calcular los montos dinámicos del sistema para congelarlos
        pagos = PagoPedido.objects.filter(pedido__sesion_caja=sesion, estado='aplicado')
        
        ventas_efectivo = pagos.filter(forma_pago='efectivo').aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')
        ventas_tarjeta = pagos.filter(forma_pago__in=['tarjeta_debito', 'tarjeta_credito']).aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')
        ventas_transferencia = pagos.filter(forma_pago='transferencia').aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')

        with transaction.atomic():
            # Actualizar la sesión
            sesion.monto_final_efectivo = monto_final
            sesion.total_ventas_efectivo = ventas_efectivo
            sesion.total_ventas_tarjeta = ventas_tarjeta
            sesion.total_ventas_transferencia = ventas_transferencia
            sesion.estado = 'cerrada'
            sesion.fecha_cierre = timezone.now()
            sesion.save()

            # Cerrar también la caja POS asociada
            caja = sesion.caja_pos
            caja.estado = 'cerrada'
            caja.save()

        return JsonResponse({
            'success': True, 
            'message': 'Corte y cierre de caja procesado correctamente.',
            'sesion_id': sesion.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def obtener_ventas_sesion_ajax(request, sesion_id):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'cortes_de_caja', 'ver_ventas'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para ver ventas de la sesión.'}, status=403)

    from pedidos.models import Pedido
    from tesoreria.models import PagoPedido
    
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'No se encontró empresa activa.'}, status=403)
        
    try:
        sesion = SesionCajaPOS.objects.get(id=sesion_id, caja_pos__empresa=empresa_actual)
        pedidos = Pedido.objects.filter(sesion_caja=sesion).order_by('-fecha_creacion')
        
        from decimal import Decimal
        total_subtotal = Decimal('0.00')
        total_iva = Decimal('0.00')
        total_final = Decimal('0.00')
        
        ventas_data = []
        for p in pedidos:
            # Obtener desglose de pagos
            pagos = PagoPedido.objects.filter(pedido=p)
            desglose_pagos = ", ".join([f"{pag.get_forma_pago_display()}: ${pag.monto:,.2f}" for pag in pagos])
            if not desglose_pagos:
                desglose_pagos = "Sin pagos registrados"
                
            p_subtotal = Decimal('0.00')
            p_iva = Decimal('0.00')
            for d in p.detalles.all():
                p_subtotal += d.subtotal
                p_iva += d.iva_monto
                
            total_subtotal += p_subtotal
            total_iva += p_iva
            total_final += p.total_pedido
                
            ventas_data.append({
                'id': p.id,
                'fecha': p.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'cliente': str(p.cliente),
                'subtotal': float(p_subtotal),
                'iva': float(p_iva),
                'total': float(p.total_pedido),
                'estado': p.get_estado_display(),
                'pagos': desglose_pagos
            })
            
        return JsonResponse({
            'success': True, 
            'ventas': ventas_data,
            'subtotal': float(total_subtotal),
            'iva': float(total_iva),
            'total': float(total_final)
        })
    except SesionCajaPOS.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'La sesión de caja no existe o no pertenece a la empresa.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def obtener_articulos_sesion_ajax(request, sesion_id):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'cortes_de_caja', 'ver_articulos'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para ver artículos de la sesión.'}, status=403)

    from pedidos.models import DetallePedido
    
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'No se encontró empresa activa.'}, status=403)
        
    try:
        sesion = SesionCajaPOS.objects.get(id=sesion_id, caja_pos__empresa=empresa_actual)
        detalles = DetallePedido.objects.filter(pedido__sesion_caja=sesion).select_related('producto')
        
        from collections import defaultdict
        from decimal import Decimal
        items_map = defaultdict(lambda: {
            'nombre': '',
            'cantidad': 0,
            'subtotal': Decimal('0.00'),
            'iva': Decimal('0.00'),
            'total': Decimal('0.00')
        })
        
        session_subtotal = Decimal('0.00')
        session_iva = Decimal('0.00')
        session_total = Decimal('0.00')
        
        for d in detalles:
            prod_id = d.producto.id
            items_map[prod_id]['nombre'] = d.producto.nombre
            items_map[prod_id]['cantidad'] += d.cantidad_solicitada
            items_map[prod_id]['subtotal'] += d.subtotal
            items_map[prod_id]['iva'] += d.iva_monto
            items_map[prod_id]['total'] += d.total
            
            session_subtotal += d.subtotal
            session_iva += d.iva_monto
            session_total += d.total
            
        articulos_data = []
        for prod_id, info in items_map.items():
            articulos_data.append({
                'id': prod_id,
                'nombre': info['nombre'],
                'cantidad': info['cantidad'],
                'subtotal': float(info['subtotal']),
                'iva': float(info['iva']),
                'total': float(info['total'])
            })
            
        articulos_data.sort(key=lambda x: x['cantidad'], reverse=True)
            
        return JsonResponse({
            'success': True, 
            'articulos': articulos_data,
            'subtotal': float(session_subtotal),
            'iva': float(session_iva),
            'total': float(session_total)
        })
    except SesionCajaPOS.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'La sesión de caja no existe o no pertenece a la empresa.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def imprimir_corte_ticket(request, sesion_id):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'cortes_de_caja', 'imprimir_articulo'):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para imprimir este corte.")

    from django.db.models import Sum
    from tesoreria.models import PagoPedido
    from decimal import Decimal
    from django.http import Http404
    
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Empresa no configurada.")
        
    try:
        sesion = SesionCajaPOS.objects.get(id=sesion_id, caja_pos__empresa=empresa_actual)
    except SesionCajaPOS.DoesNotExist:
        raise Http404("La sesión de caja no existe o no pertenece a la empresa.")
        
    # Calcular montos dinámicamente de los pagos aplicados
    pagos = PagoPedido.objects.filter(pedido__sesion_caja=sesion, estado='aplicado')
    
    ventas_efectivo = pagos.filter(forma_pago='efectivo').aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')
    ventas_tarjeta = pagos.filter(forma_pago__in=['tarjeta_debito', 'tarjeta_credito']).aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')
    ventas_transferencia = pagos.filter(forma_pago='transferencia').aggregate(Sum('monto_mxn'))['monto_mxn__sum'] or Decimal('0.00')
    
    total_ventas = ventas_efectivo + ventas_tarjeta + ventas_transferencia
    efectivo_estimado = sesion.monto_inicial + ventas_efectivo
    
    # Calcular subtotal e IVA de los productos vendidos
    from pedidos.models import DetallePedido, Pedido
    detalles = DetallePedido.objects.filter(pedido__sesion_caja=sesion)
    
    session_subtotal = Decimal('0.00')
    session_iva = Decimal('0.00')
    session_total = Decimal('0.00')
    for d in detalles:
        session_subtotal += d.subtotal
        session_iva += d.iva_monto
        session_total += d.total

    pedidos_sesion = Pedido.objects.filter(sesion_caja=sesion)

    # Calcular descuentos aplicados en la sesión
    total_descuentos = Decimal('0.00')
    for ped in pedidos_sesion:
        if ped.notas:
            try:
                import json
                notas_data = json.loads(ped.notas)
                if isinstance(notas_data, dict):
                    total_descuentos += Decimal(str(notas_data.get('descuento_monto', 0.00)))
            except Exception:
                pass
    original_subtotal = session_subtotal + total_descuentos

    # Calcular desglose de subtotales, IVAs y totales por forma de pago
    subtotal_efectivo = Decimal('0.00')
    iva_efectivo = Decimal('0.00')
    total_efectivo = Decimal('0.00')
    
    subtotal_tarjeta = Decimal('0.00')
    iva_tarjeta = Decimal('0.00')
    total_tarjeta = Decimal('0.00')
    
    subtotal_transferencia = Decimal('0.00')
    iva_transferencia = Decimal('0.00')
    total_transferencia = Decimal('0.00')
    
    for ped in pedidos_sesion:
        ped_subtotal = Decimal('0.00')
        ped_iva = Decimal('0.00')
        ped_total = Decimal('0.00')
        for d in ped.detalles.all():
            ped_subtotal += d.subtotal
            ped_iva += d.iva_monto
            ped_total += d.total
            
        pagos_pedido = PagoPedido.objects.filter(pedido=ped, estado='aplicado')
        ped_pagado = sum(p.monto_mxn for p in pagos_pedido) or Decimal('0.00')
        
        if ped_pagado > 0:
            for p in pagos_pedido:
                proporcion = p.monto_mxn / ped_pagado
                fp = p.forma_pago
                if fp == 'efectivo':
                    subtotal_efectivo += ped_subtotal * proporcion
                    iva_efectivo += ped_iva * proporcion
                    total_efectivo += ped_total * proporcion
                elif fp in ['tarjeta_debito', 'tarjeta_credito']:
                    subtotal_tarjeta += ped_subtotal * proporcion
                    iva_tarjeta += ped_iva * proporcion
                    total_tarjeta += ped_total * proporcion
                elif fp == 'transferencia':
                    subtotal_transferencia += ped_subtotal * proporcion
                    iva_transferencia += ped_iva * proporcion
                    total_transferencia += ped_total * proporcion
    
    if sesion.estado == 'cerrada':
        monto_final = sesion.monto_final_efectivo
        diferencia = monto_final - efectivo_estimado
    else:
        monto_final = None
        diferencia = None

    # Agrupar productos vendidos si se solicita en la URL
    incluir_articulos = request.GET.get('incluir_articulos') == 'true'
    articulos_data = []
    if incluir_articulos:
        from collections import defaultdict
        items_map = defaultdict(lambda: {
            'nombre': '',
            'cantidad': 0,
            'total': Decimal('0.00')
        })
        for d in detalles:
            prod_id = d.producto.id
            items_map[prod_id]['nombre'] = d.producto.nombre
            items_map[prod_id]['cantidad'] += d.cantidad_solicitada
            items_map[prod_id]['total'] += d.total
            
        for prod_id, info in items_map.items():
            articulos_data.append({
                'nombre': info['nombre'],
                'cantidad': info['cantidad'],
                'total': info['total']
            })
        # Ordenar de mayor a menor cantidad
        articulos_data.sort(key=lambda x: x['cantidad'], reverse=True)
        
    context = {
        'sesion': sesion,
        'ventas_efectivo': ventas_efectivo,
        'ventas_tarjeta': ventas_tarjeta,
        'ventas_transferencia': ventas_transferencia,
        'total_ventas': total_ventas,
        'efectivo_estimado': efectivo_estimado,
        'monto_final': monto_final,
        'diferencia': diferencia,
        'empresa': empresa_actual,
        'session_subtotal': session_subtotal,
        'session_iva': session_iva,
        'session_total': session_total,
        'total_descuentos': total_descuentos,
        'original_subtotal': original_subtotal,
        'subtotal_efectivo': subtotal_efectivo,
        'iva_efectivo': iva_efectivo,
        'total_efectivo': total_efectivo,
        'subtotal_tarjeta': subtotal_tarjeta,
        'iva_tarjeta': iva_tarjeta,
        'total_tarjeta': total_tarjeta,
        'subtotal_transferencia': subtotal_transferencia,
        'iva_transferencia': iva_transferencia,
        'total_transferencia': total_transferencia,
        'incluir_articulos': incluir_articulos,
        'articulos': articulos_data
    }
    
    return render(request, 'ventas/ticket_corte.html', context)


# ==============================================================================
# VISTAS PARA EL CORTE Z (CIERRE DIARIO DE CAJAS)
# ==============================================================================

@login_required
@require_POST
def generar_corte_z_ajax(request):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'cortes_de_caja', 'hacer_corte'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para generar un Corte Z.'})

    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'Empresa no configurada.'})

    fecha_str = request.POST.get('fecha')
    if not fecha_str:
        return JsonResponse({'success': False, 'error': 'La fecha es requerida.'})

    try:
        from datetime import datetime
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Formato de fecha inválido.'})

    # Verificar si ya existe un Corte Z para esta fecha y empresa
    from django.db import IntegrityError, transaction
    
    if CorteZ.objects.filter(empresa=empresa_actual, fecha=fecha).exists():
        return JsonResponse({'success': False, 'error': f'Ya se ha generado un Corte Z para la fecha {fecha.strftime("%d/%m/%Y")}.'})

    # Filtrar sesiones cerradas en esa fecha de cierre que no tengan Corte Z asignado
    sesiones_pendientes = SesionCajaPOS.objects.filter(
        caja_pos__empresa=empresa_actual,
        estado='cerrada',
        fecha_cierre__date=fecha,
        corte_z__isnull=True
    )

    if not sesiones_pendientes.exists():
        return JsonResponse({'success': False, 'error': f'No hay turnos cerrados pendientes de Corte Z para la fecha {fecha.strftime("%d/%m/%Y")}.'})

    # Consolidar totales
    from decimal import Decimal
    monto_inicial = Decimal('0.00')
    monto_final_efectivo = Decimal('0.00')
    total_efectivo = Decimal('0.00')
    total_tarjeta = Decimal('0.00')
    total_transferencia = Decimal('0.00')
    total_ventas = Decimal('0.00')

    for s in sesiones_pendientes:
        monto_inicial += s.monto_inicial
        monto_final_efectivo += s.monto_final_efectivo or Decimal('0.00')
        total_efectivo += s.total_ventas_efectivo
        total_tarjeta += s.total_ventas_tarjeta
        total_transferencia += s.total_ventas_transferencia
        total_ventas += s.total_ventas

    try:
        with transaction.atomic():
            corte_z = CorteZ.objects.create(
                empresa=empresa_actual,
                fecha=fecha,
                usuario=request.user,
                monto_inicial=monto_inicial,
                monto_final_efectivo=monto_final_efectivo,
                total_efectivo=total_efectivo,
                total_tarjeta=total_tarjeta,
                total_transferencia=total_transferencia,
                total_ventas=total_ventas
            )
            # Vincular las sesiones al nuevo Corte Z
            sesiones_pendientes.update(corte_z=corte_z)
            
        return JsonResponse({'success': True, 'corte_z_id': corte_z.id})
    except IntegrityError:
        return JsonResponse({'success': False, 'error': 'Error de integridad al guardar el Corte Z. Posiblemente ya existe uno para esta fecha.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al guardar el Corte Z: {str(e)}'})


@login_required
def historial_cortes_z_ajax(request):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'cortes_de_caja', 'ver'):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para ver el historial.'})

    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'Empresa no configurada.'})

    cortes = CorteZ.objects.filter(empresa=empresa_actual).order_by('-fecha')
    data = []
    for c in cortes:
        # Obtener los nombres de las cajas incluidas en este corte z
        cajas_nombres = list(c.sesiones_corte.values_list('caja_pos__nombre', flat=True))
        data.append({
            'id': c.id,
            'fecha': c.fecha.strftime('%Y-%m-%d'),
            'fecha_formateada': c.fecha.strftime('%d/%m/%Y'),
            'fecha_creacion': c.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            'usuario': c.usuario.username.split('@')[0],
            'total_efectivo': float(c.total_efectivo),
            'total_tarjeta': float(c.total_tarjeta),
            'total_transferencia': float(c.total_transferencia),
            'total_ventas': float(c.total_ventas),
            'monto_inicial': float(c.monto_inicial),
            'monto_final_efectivo': float(c.monto_final_efectivo),
            'cajas': ", ".join(cajas_nombres) if cajas_nombres else "Ninguna"
        })

    return JsonResponse({'success': True, 'cortes': data})


@login_required
def imprimir_corte_z_ticket(request, corte_z_id):
    from preferencias.permissions import user_has_sales_permission
    if not user_has_sales_permission(request, 'cortes_de_caja', 'imprimir_articulo'):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("No tienes permisos para imprimir este corte.")

    from django.db.models import Sum
    from tesoreria.models import PagoPedido
    from decimal import Decimal
    from django.http import Http404
    
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Empresa no configurada.")
        
    try:
        corte_z = CorteZ.objects.get(id=corte_z_id, empresa=empresa_actual)
    except CorteZ.DoesNotExist:
        raise Http404("El Corte Z no existe o no pertenece a la empresa.")
        
    # Obtener todas las sesiones vinculadas
    sesiones = corte_z.sesiones_corte.all()
    
    # Consolidar detalles de productos de todas las sesiones de este Corte Z
    from pedidos.models import DetallePedido, Pedido
    detalles = DetallePedido.objects.filter(pedido__sesion_caja__in=sesiones)
    
    session_subtotal = Decimal('0.00')
    session_iva = Decimal('0.00')
    session_total = Decimal('0.00')
    for d in detalles:
        session_subtotal += d.subtotal
        session_iva += d.iva_monto
        session_total += d.total

    pedidos_sesion = Pedido.objects.filter(sesion_caja__in=sesiones)

    # Calcular descuentos aplicados en las sesiones
    total_descuentos = Decimal('0.00')
    for ped in pedidos_sesion:
        if ped.notas:
            try:
                import json
                notas_data = json.loads(ped.notas)
                if isinstance(notas_data, dict):
                    total_descuentos += Decimal(str(notas_data.get('descuento_monto', 0.00)))
            except Exception:
                pass
    original_subtotal = session_subtotal + total_descuentos

    # Calcular desglose de subtotales, IVAs y totales por forma de pago
    subtotal_efectivo = Decimal('0.00')
    iva_efectivo = Decimal('0.00')
    total_efectivo = Decimal('0.00')
    
    subtotal_tarjeta = Decimal('0.00')
    iva_tarjeta = Decimal('0.00')
    total_tarjeta = Decimal('0.00')
    
    subtotal_transferencia = Decimal('0.00')
    iva_transferencia = Decimal('0.00')
    total_transferencia = Decimal('0.00')
    
    for ped in pedidos_sesion:
        ped_subtotal = Decimal('0.00')
        ped_iva = Decimal('0.00')
        ped_total = Decimal('0.00')
        for d in ped.detalles.all():
            ped_subtotal += d.subtotal
            ped_iva += d.iva_monto
            ped_total += d.total
            
        pagos_pedido = PagoPedido.objects.filter(pedido=ped, estado='aplicado')
        ped_pagado = sum(p.monto_mxn for p in pagos_pedido) or Decimal('0.00')
        
        if ped_pagado > 0:
            for p in pagos_pedido:
                proporcion = p.monto_mxn / ped_pagado
                fp = p.forma_pago
                if fp == 'efectivo':
                    subtotal_efectivo += ped_subtotal * proporcion
                    iva_efectivo += ped_iva * proporcion
                    total_efectivo += ped_total * proporcion
                elif fp in ['tarjeta_debito', 'tarjeta_credito']:
                    subtotal_tarjeta += ped_subtotal * proporcion
                    iva_tarjeta += ped_iva * proporcion
                    total_tarjeta += ped_total * proporcion
                elif fp == 'transferencia':
                    subtotal_transferencia += ped_subtotal * proporcion
                    iva_transferencia += ped_iva * proporcion
                    total_transferencia += ped_total * proporcion
    
    efectivo_estimado = corte_z.monto_inicial + corte_z.total_efectivo
    monto_final = corte_z.monto_final_efectivo
    diferencia = monto_final - efectivo_estimado

    # Agrupar productos vendidos
    incluir_articulos = request.GET.get('incluir_articulos') == 'true'
    articulos_data = []
    if incluir_articulos:
        from collections import defaultdict
        items_map = defaultdict(lambda: {
            'nombre': '',
            'cantidad': 0,
            'total': Decimal('0.00')
        })
        for d in detalles:
            prod_id = d.producto.id
            items_map[prod_id]['nombre'] = d.producto.nombre
            items_map[prod_id]['cantidad'] += d.cantidad_solicitada
            items_map[prod_id]['total'] += d.total
            
        for prod_id, info in items_map.items():
            articulos_data.append({
                'nombre': info['nombre'],
                'cantidad': info['cantidad'],
                'total': info['total']
            })
        articulos_data.sort(key=lambda x: x['cantidad'], reverse=True)
        
    context = {
        'corte_z': corte_z,
        'sesiones': sesiones,
        'ventas_efectivo': corte_z.total_efectivo,
        'ventas_tarjeta': corte_z.total_tarjeta,
        'ventas_transferencia': corte_z.total_transferencia,
        'total_ventas': corte_z.total_ventas,
        'efectivo_estimado': efectivo_estimado,
        'monto_final': monto_final,
        'diferencia': diferencia,
        'empresa': empresa_actual,
        'session_subtotal': session_subtotal,
        'session_iva': session_iva,
        'session_total': session_total,
        'total_descuentos': total_descuentos,
        'original_subtotal': original_subtotal,
        'subtotal_efectivo': subtotal_efectivo,
        'iva_efectivo': iva_efectivo,
        'total_efectivo_pagos': total_efectivo,
        'subtotal_tarjeta': subtotal_tarjeta,
        'iva_tarjeta': iva_tarjeta,
        'total_tarjeta_pagos': total_tarjeta,
        'subtotal_transferencia': subtotal_transferencia,
        'iva_transferencia': iva_transferencia,
        'total_transferencia_pagos': total_transferencia,
        'incluir_articulos': incluir_articulos,
        'articulos': articulos_data
    }
    
    return render(request, 'ventas/ticket_corte_z.html', context)


@login_required
def imprimir_pedido_ticket(request, pedido_id):
    """Genera la vista para la impresión del ticket de venta del Pedido (Formato térmico 80mm)"""
    import json
    from decimal import Decimal
    from pedidos.models import Pedido
    
    empresa_actual = get_empresa_actual(request)
    pedido = get_object_or_404(Pedido, id=pedido_id, empresa=empresa_actual)
    
    # Limpiar nombre del vendedor (quitar @empresa si es necesario)
    vendedor_nombre = pedido.vendedor.get_full_name()
    if not vendedor_nombre:
        vendedor_nombre = pedido.vendedor.username.split('@')[0]
        
    items_preparados = []
    for d in pedido.detalles.all():
        mods = []
        if d.modificadores_json:
            try:
                mods_list = json.loads(d.modificadores_json)
                for m in mods_list:
                    tipo_lbl = "Extra" if m.get('tipo') == 'extra' else "Sin"
                    nombre = m.get('nombre', '')
                    mods.append(f"{tipo_lbl} {nombre}")
            except Exception:
                pass
        items_preparados.append({
            'detalle': d,
            'nombre_producto': d.producto.nombre,
            'cantidad': d.cantidad_solicitada,
            'precio_unitario': d.precio_unitario,
            'subtotal': d.subtotal,
            'modificadores': mods
        })
        
    pagos_data = []
    for p in pedido.pagos.filter(estado='aplicado'):
        pagos_data.append({
            'monto': p.monto,
            'metodo': p.caja_banco.nombre if p.caja_banco else 'Efectivo',
            'fecha': p.fecha_registro
        })
        
    context = {
        'pedido': pedido,
        'empresa': empresa_actual,
        'vendedor_nombre': vendedor_nombre,
        'items': items_preparados,
        'pagos': pagos_data
    }
    return render(request, 'ventas/ticket_pedido.html', context)