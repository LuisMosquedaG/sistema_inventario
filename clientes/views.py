from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Cliente, ContactoCliente
from .forms import ClienteForm
from django.http import JsonResponse
from django.forms.models import model_to_dict
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
import json
import openpyxl
from django.http import HttpResponse, JsonResponse
from io import BytesIO
from django.db import transaction
from panel.models import Empresa
from preferencias.permissions import require_sales_permission

# ... existing code ...

@login_required(login_url='/login/')
@require_sales_permission('clientes', 'crear')
def descargar_plantilla_clientes(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Clientes"
    
    # Encabezados basados en los campos del modelo y formulario
    headers = [
        'Nombre', 'Apellidos', 'Razon Social', 'RFC', 'Email', 'Telefono',
        'Calle', 'Num Ext', 'Num Int', 'Colonia', 'Estado Ubicacion', 'CP',
        'Envio Calle', 'Envio Num Ext', 'Envio Num Int', 'Envio Colonia', 
        'Envio Estado', 'Envio CP', 'Envio Quien Recibe', 'Envio Telefono', 
        'Envio Correo', 'Envio Notas', 'Estado (activo/suspendido/inactivo)', 
        'Tipo (prospecto/cliente_nuevo/cliente_activo/cliente_inactivo/vip)', 
        'Relacion (directo/referido/revendedor)'
    ]
    
    ws.append(headers)
    
    # Ajustar ancho de columnas
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Clientes.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
@require_sales_permission('clientes', 'crear', json_response=True)
def importar_clientes_ajax(request):
    if request.method == 'POST' and request.FILES.get('archivo_clientes'):
        empresa_actual = get_empresa_actual(request)
        excel_file = request.FILES['archivo_clientes']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return JsonResponse({'success': False, 'error': 'El archivo está vacío o no tiene datos.'})
            
            data_rows = rows[1:]
            
            creados = 0
            errores = []
            
            with transaction.atomic():
                for idx, row in enumerate(data_rows, start=2):
                    if not any(row): continue # Saltar filas vacías
                    
                    try:
                        # Mapeo por posición (basado en la plantilla)
                        cliente = Cliente(
                            empresa=empresa_actual,
                            nombre=str(row[0] or '').strip(),
                            apellidos=str(row[1] or '').strip(),
                            razon_social=str(row[2] or '').strip(),
                            rfc=str(row[3] or '').strip().upper(),
                            email=str(row[4] or '').strip(),
                            telefono=str(row[5] or '').strip(),
                            calle=str(row[6] or '').strip(),
                            numero_ext=str(row[7] or '').strip(),
                            numero_int=str(row[8] or '').strip(),
                            colonia=str(row[9] or '').strip(),
                            estado_dir=str(row[10] or '').strip(),
                            cp=str(row[11] or '').strip(),
                            envio_calle=str(row[12] or '').strip(),
                            envio_numero_ext=str(row[13] or '').strip(),
                            envio_numero_int=str(row[14] or '').strip(),
                            envio_colonia=str(row[15] or '').strip(),
                            envio_estado=str(row[16] or '').strip(),
                            envio_cp=str(row[17] or '').strip(),
                            envio_quien_recibe=str(row[18] or '').strip(),
                            envio_telefono=str(row[19] or '').strip(),
                            envio_correo=str(row[20] or '').strip(),
                            envio_notas=str(row[21] or '').strip(),
                            estado=(str(row[22] or 'activo')).strip().lower(),
                            tipo=(str(row[23] or 'prospecto')).strip().lower(),
                            relacion=(str(row[24] or 'directo')).strip().lower(),
                        )
                        
                        if not cliente.nombre:
                            errores.append(f"Fila {idx}: El nombre es obligatorio.")
                            continue
                            
                        cliente.save()
                        creados += 1
                    except Exception as row_err:
                        errores.append(f"Fila {idx}: {str(row_err)}")
            
            if creados == 0 and errores:
                return JsonResponse({'success': False, 'error': f"No se pudo importar ningún cliente. Errores: {', '.join(errores[:3])}..."})

            msg = f'Importación exitosa. {creados} clientes creados.'
            if errores:
                msg += f' ({len(errores)} filas con error)'
                
            return JsonResponse({'success': True, 'message': msg})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error al procesar el archivo: {str(e)}'})
            
    return JsonResponse({'success': False, 'error': 'Solicitud inválida.'})


@login_required(login_url='/login/')
@require_sales_permission('clientes', 'ver')
def exportar_clientes_excel(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return HttpResponse("No autorizado", status=403)
        
    # --- Mismo filtrado que dashboard_clientes ---
    q = request.GET.get('q', '')
    cliente_id = request.GET.get('cliente_id', '')
    email = request.GET.get('email', '')
    tipo = request.GET.get('tipo', '')
    relacion = request.GET.get('relacion', '')
    estado = request.GET.get('estado', '')

    clientes = Cliente.objects.filter(empresa=empresa_actual).order_by('-creado_en')

    if q:
        clientes = clientes.filter(
            Q(nombre__icontains=q) | Q(apellidos__icontains=q) |
            Q(razon_social__icontains=q) | Q(email__icontains=q) |
            Q(rfc__icontains=q) | Q(telefono__icontains=q)
        )
    if cliente_id and cliente_id != 'all':
        clientes = clientes.filter(id=cliente_id)
    if email:
        clientes = clientes.filter(email__icontains=email)
    if tipo:
        clientes = clientes.filter(tipo=tipo)
    if relacion:
        clientes = clientes.filter(relacion=relacion)
    if estado:
        clientes = clientes.filter(estado=estado)

    # --- Generación del Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Directorio de Clientes"
    
    headers = [
        'ID', 'Nombre', 'Apellidos', 'Razon Social', 'RFC', 'Email', 'Telefono',
        'Calle', 'Num Ext', 'Num Int', 'Colonia', 'Estado Ubicacion', 'CP',
        'Envio Calle', 'Envio Num Ext', 'Envio Num Int', 'Envio Colonia', 
        'Envio Estado', 'Envio CP', 'Envio Quien Recibe', 'Envio Telefono', 
        'Envio Correo', 'Envio Notas', 'Estado', 'Tipo', 'Relacion', 'Fecha Registro'
    ]
    ws.append(headers)
    
    for c in clientes:
        ws.append([
            c.id, c.nombre, c.apellidos, c.razon_social, c.rfc, c.email, c.telefono,
            c.calle, c.numero_ext, c.numero_int, c.colonia, c.estado_dir, c.cp,
            c.envio_calle, c.envio_numero_ext, c.envio_numero_int, c.envio_colonia,
            c.envio_estado, c.envio_cp, c.envio_quien_recibe, c.envio_telefono,
            c.envio_correo, c.envio_notas, c.get_estado_display(), c.get_tipo_display(),
            c.get_relacion_display(), c.creado_en.strftime('%d/%m/%Y %H:%M') if c.creado_en else ''
        ])
        
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Clientes_{empresa_actual.subdominio}.xlsx"'
    wb.save(response)
    return response

# --- 1. FUNCIÓN AYUDANTE ESTÁNDAR ---
def get_empresa_actual(request):
    username = request.user.username
    if '@' in username:
        subdominio = username.split('@')[1]
        try:
            return Empresa.objects.get(subdominio=subdominio)
        except Empresa.DoesNotExist:
            return None
    return None

from core.models import Producto

from django.db.models import Q

@login_required(login_url='/login/')
@require_sales_permission('clientes', 'ver')
def dashboard_clientes(request):
    empresa_actual = get_empresa_actual(request)
    
    # --- SEGURIDAD: Bloquear si no hay empresa (Eliminar fallback a 'all') ---
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)
    
    # --- LÓGICA DE FILTRADO ---
    q = request.GET.get('q', '')
    cliente_id = request.GET.get('cliente_id', '')
    email = request.GET.get('email', '')
    tipo = request.GET.get('tipo', '')
    relacion = request.GET.get('relacion', '')
    estado = request.GET.get('estado', '')

    lista_clientes = Cliente.objects.filter(empresa=empresa_actual).order_by('-creado_en')

    if q:
        lista_clientes = lista_clientes.filter(
            Q(nombre__icontains=q) |
            Q(apellidos__icontains=q) |
            Q(razon_social__icontains=q) |
            Q(email__icontains=q) |
            Q(rfc__icontains=q) |
            Q(telefono__icontains=q)
        )
    if cliente_id and cliente_id != 'all':
        lista_clientes = lista_clientes.filter(id=cliente_id)
    if email:
        lista_clientes = lista_clientes.filter(email__icontains=email)
    if tipo:
        lista_clientes = lista_clientes.filter(tipo=tipo)
    if relacion:
        lista_clientes = lista_clientes.filter(relacion=relacion)
    if estado:
        lista_clientes = lista_clientes.filter(estado=estado)

    # Para el buscador visual de cliente (se busca a sí mismo)
    cliente_nombre_display = ""
    if cliente_id and cliente_id != 'all':
        try:
            c_obj = Cliente.objects.get(id=cliente_id, empresa=empresa_actual)
            cliente_nombre_display = c_obj.razon_social if c_obj.razon_social else f"{c_obj.nombre} {c_obj.apellidos}"
        except:
            pass

    filtros = {
        'q': q,
        'cliente_id': cliente_id,
        'cliente_nombre': cliente_nombre_display,
        'email': email,
        'tipo': tipo,
        'relacion': relacion,
        'estado': estado
    }
    # --- FIN LÓGICA DE FILTRADO ---
    
    todos_los_productos = Producto.objects.filter(empresa=empresa_actual)
    form = ClienteForm()
    
    contexto = {
        'clientes': lista_clientes,
        'todos_los_clientes': Cliente.objects.filter(empresa=empresa_actual), # Para el selector de filtro
        'productos': todos_los_productos,
        'form': form,
        'filtros': filtros
    }
    return render(request, 'dashboard_clientes.html', contexto)

@login_required(login_url='/login/')
@require_sales_permission('clientes', 'crear', json_response=True)
def crear_cliente(request):
    empresa_actual = get_empresa_actual(request)
    
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'No se pudo detectar tu empresa.'}, status=403)
    
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.empresa = empresa_actual
            cliente.save()
            
            # --- SOPORTE PARA AJAX (MODAL RÁPIDO) ---
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True, 
                    'id': cliente.id, 
                    'nombre_completo': cliente.nombre_completo
                })
            
            messages.success(request, 'Cliente creado correctamente.')
            return redirect('dashboard_clientes')
        else:
            lista_clientes = Cliente.objects.filter(empresa=empresa_actual)
            return render(request, 'dashboard_clientes.html', {'clientes': lista_clientes, 'form': form})
    
    # GET
    lista_clientes = Cliente.objects.filter(empresa=empresa_actual)
    form = ClienteForm()
    return render(request, 'dashboard_clientes.html', {'clientes': lista_clientes, 'form': form})

@login_required(login_url='/login/')
@require_sales_permission('clientes', 'ver', json_response=True)
def obtener_cliente_json(request, cliente_id):
    """Devuelve los datos de un cliente para el modal de edición"""
    try:
        empresa_actual = get_empresa_actual(request)
        
        # --- MEJORA: Filtrar por empresa directo en la query ---
        # Si no existe el cliente con ese ID Y esa empresa, lanza 404 automáticamente.
        cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa_actual)

        data = model_to_dict(cliente)
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)

@login_required(login_url='/login/')
@require_sales_permission('clientes', 'editar')
def actualizar_cliente(request, cliente_id):
    """Guarda los cambios de un cliente existente"""
    empresa_actual = get_empresa_actual(request)
    
    # --- MEJORA: Validación en la query ---
    cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa_actual)

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente actualizado correctamente.')
            return redirect('dashboard_clientes')
        else:
            messages.error(request, 'Por favor corrige los errores.')
            return redirect('dashboard_clientes')
            
    return redirect('dashboard_clientes')

@login_required(login_url='/login/')
@require_sales_permission('clientes', 'agenda_contactos', json_response=True)
def obtener_contactos_cliente(request, cliente_id):
    """Lista los contactos de un cliente"""
    empresa_actual = get_empresa_actual(request)
    
    try:
        # --- MEJORA: Validar cliente en la query ---
        cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa_actual)

        contactos = ContactoCliente.objects.filter(cliente_id=cliente_id)
        data = []
        for c in contactos:
            data.append({
                'id': c.id,
                'nombre': c.nombre_completo,
                't1': c.telefono_1,
                't2': c.telefono_2,
                'e1': c.correo_1,
                'e2': c.correo_2,
                'notas': c.notas
            })
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse([], safe=False)

@login_required(login_url='/login/')
@require_sales_permission('clientes', 'agenda_contactos', json_response=True)
def guardar_contactos_cliente(request, cliente_id):
    """Guarda la lista de contactos"""
    empresa_actual = get_empresa_actual(request)
    
    try:
        # --- MEJORA: Validar cliente en la query ---
        cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa_actual)
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': 'Cliente no encontrado o acceso denegado'})

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            contactos_list = data.get('contactos', [])
            
            ids_recibidos = [int(c['id']) for c in contactos_list if c.get('id')]

            # Eliminar los que no vinieron en la lista (pero solo de este cliente validado)
            if ids_recibidos:
                ContactoCliente.objects.filter(cliente=cliente).exclude(id__in=ids_recibidos).delete()
            else:
                ContactoCliente.objects.filter(cliente=cliente).delete()
            
            count = 0
            for cont_data in contactos_list:
                if cont_data.get('id'):
                    contacto = ContactoCliente.objects.get(id=cont_data['id'], cliente=cliente)
                    contacto.nombre_completo = cont_data['nombre']
                    contacto.telefono_1 = cont_data['t1']
                    contacto.telefono_2 = cont_data['t2']
                    contacto.correo_1 = cont_data['e1']
                    contacto.correo_2 = cont_data['e2']
                    contacto.notas = cont_data['notas']
                    contacto.save()
                else:
                    ContactoCliente.objects.create(
                        cliente=cliente,
                        nombre_completo=cont_data['nombre'],
                        telefono_1=cont_data['t1'],
                        telefono_2=cont_data['t2'],
                        correo_1=cont_data['e1'],
                        correo_2=cont_data['e2'],
                        notas=cont_data['notas']
                    )
                count += 1
            
            return JsonResponse({'success': True, 'count': count})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False})