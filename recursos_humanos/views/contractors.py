from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q, Count
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
import re
from datetime import datetime

from ..models import Empleado, Contrato, Contratista, Beneficiario, ImportacionSUA
from preferencias.models import Sucursal
from preferencias.permissions import require_hr_permission
from .utils import get_empresa_actual

@login_required(login_url='/login/')
@require_hr_permission('contratistas', 'ver')
def lista_contratistas(request):
    empresa_actual = get_empresa_actual(request)
    contratistas = Contratista.objects.filter(empresa=empresa_actual).annotate(total_colaboradores=Count('empleado')).order_by('nombre_razon_social')
    
    q = request.GET.get('q', '')
    f_razon = request.GET.get('razon_social', '')
    f_rfc = request.GET.get('rfc', '')
    f_rp = request.GET.get('reg_patronal', '')
    sucursal_id = request.GET.get('sucursal', '')

    if q:
        contratistas = contratistas.filter(
            Q(nombre_razon_social__icontains=q) |
            Q(rfc__icontains=q) |
            Q(representante_legal__icontains=q) |
            Q(correo__icontains=q) |
            Q(clave__icontains=q)
        )
    
    if f_razon:
        contratistas = contratistas.filter(nombre_razon_social__icontains=f_razon)
    if f_rfc:
        contratistas = contratistas.filter(rfc__icontains=f_rfc)
    if f_rp:
        contratistas = contratistas.filter(registro_patronal__icontains=f_rp)
    if sucursal_id:
        contratistas = contratistas.filter(sucursal_id=sucursal_id)

    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    
    # PAGINACIÓN
    paginator = Paginator(contratistas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'recursos_humanos/lista_contratistas.html', {
        'page_obj': page_obj,
        'sucursales': sucursales,
        'empresa': empresa_actual,
        'filtros': {
            'q': q,
            'razon_social': f_razon,
            'rfc': f_rfc,
            'reg_patronal': f_rp,
            'sucursal': sucursal_id
        }
    })

@login_required(login_url='/login/')
@require_hr_permission('beneficiarios', 'ver', json_response=True)
def obtener_contratista_json(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        cont = Contratista.objects.get(id=id, empresa=empresa_actual)
        data = {
            'id': cont.id, 'clave': cont.clave or '', 'rfc': cont.rfc, 'nombre_razon_social': cont.nombre_razon_social,
            'regimen': cont.regimen or '',
            'correo': cont.correo, 'telefono': cont.telefono, 'registro_patronal': cont.registro_patronal,
            'calle': cont.calle, 'num_ext': cont.num_ext, 'num_int': cont.num_int,
            'entre_calle': cont.entre_calle, 'y_calle': cont.y_calle, 'colonia': cont.colonia,
            'cp': cont.cp, 'municipio_alcaldia': cont.municipio_alcaldia,
            'entidad_federativa': cont.entidad_federativa, 'representante_legal': cont.representante_legal,
            'administrador_unico': cont.administrador_unico, 'num_escritura': cont.num_escritura,
            'nombre_notario_publico': cont.nombre_notario_publico, 'num_notario_publico': cont.num_notario_publico,
            'fecha_escritura_publica': cont.fecha_escritura_publica.isoformat() if cont.fecha_escritura_publica else '',
            'folio_mercantil': cont.folio_mercantil, 'numero_stps': cont.numero_stps,
        }
        return JsonResponse({'success': True, 'data': data})
    except Contratista.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Contratista no encontrado.'})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('contratistas', 'crear', json_response=True)
def crear_contratista_ajax(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual: return JsonResponse({'success': False, 'error': 'No se encontró la empresa.'}, status=403)
    try:
        data = request.POST
        sucursal_id = request.session.get('sucursal_id')
        nuevo = Contratista(
            empresa=empresa_actual, sucursal_id=sucursal_id, 
            clave=data.get('clave', ''),
            rfc=data.get('rfc', '').upper(),
            nombre_razon_social=data.get('nombre_razon_social'), 
            regimen=data.get('regimen'),
            correo=data.get('correo'),
            telefono=data.get('telefono'), registro_patronal=data.get('registro_patronal'),
            calle=data.get('calle'), num_ext=data.get('num_ext'), num_int=data.get('num_int'),
            entre_calle=data.get('entre_calle'), y_calle=data.get('y_calle'), colonia=data.get('colonia'),
            cp=data.get('cp'), municipio_alcaldia=data.get('municipio_alcaldia'),
            entidad_federativa=data.get('entidad_federativa'), representante_legal=data.get('representante_legal'),
            administrador_unico=data.get('administrador_unico'), num_escritura=data.get('num_escritura'),
            nombre_notario_publico=data.get('nombre_notario_publico'), num_notario_publico=data.get('num_notario_publico'),
            fecha_escritura_publica=data.get('fecha_escritura_publica') or None, folio_mercantil=data.get('folio_mercantil'),
            numero_stps=data.get('numero_stps'),
        )
        nuevo.save()
        return JsonResponse({'success': True, 'message': 'Contratista registrado correctamente.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('contratistas', 'editar', json_response=True)
def editar_contratista_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        cont = Contratista.objects.get(id=id, empresa=empresa_actual)
        data = request.POST
        cont.clave = data.get('clave', '')
        cont.rfc = data.get('rfc', '').upper()
        cont.nombre_razon_social = data.get('nombre_razon_social')
        cont.regimen = data.get('regimen')
        cont.correo = data.get('correo'); cont.telefono = data.get('telefono')
        cont.registro_patronal = data.get('registro_patronal'); cont.calle = data.get('calle')
        cont.num_ext = data.get('num_ext'); cont.num_int = data.get('num_int')
        cont.entre_calle = data.get('entre_calle'); cont.y_calle = data.get('y_calle')
        cont.colonia = data.get('colonia'); cont.cp = data.get('cp')
        cont.municipio_alcaldia = data.get('municipio_alcaldia'); cont.entidad_federativa = data.get('entidad_federativa')
        cont.representante_legal = data.get('representante_legal'); cont.administrador_unico = data.get('administrador_unico')
        cont.num_escritura = data.get('num_escritura'); cont.nombre_notario_publico = data.get('nombre_notario_publico')
        cont.num_notario_publico = data.get('num_notario_publico'); cont.fecha_escritura_publica = data.get('fecha_escritura_publica') or None
        cont.folio_mercantil = data.get('folio_mercantil'); cont.numero_stps = data.get('numero_stps')
        cont.save()
        return JsonResponse({'success': True, 'message': 'Contratista actualizado correctamente.'})
    except Contratista.DoesNotExist: return JsonResponse({'success': False, 'error': 'Contratista no encontrado.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('contratistas', 'eliminar', json_response=True)
def eliminar_contratista_ajax(request, id):
    """Eliminar un contratista."""
    empresa_actual = get_empresa_actual(request)
    try:
        cont = Contratista.objects.get(id=id, empresa=empresa_actual)
        cont.delete()
        return JsonResponse({'success': True, 'message': 'Contratista eliminado correctamente.'})
    except Contratista.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Contratista no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@transaction.atomic
def importar_contratistas_ajax(request):
    """Cargador de contratistas desde Excel."""
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'status': 'error', 'message': 'Empresa no encontrada.'})
    
    file = request.FILES.get('archivo')
    if not file:
        return JsonResponse({'status': 'error', 'message': 'No se proporcionó ningún archivo.'})
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        
        def get_val(row, header_name, default=None):
            try:
                idx = headers.index(header_name)
                val = row[idx].value
                return val if val is not None else default
            except (ValueError, IndexError):
                return default

        def to_date(val):
            if isinstance(val, datetime): return val.date()
            if isinstance(val, str):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%dT%H:%M:%S'):
                    try: return datetime.strptime(val.strip(), fmt).date()
                    except: pass
            return None

        sucursal_id = request.session.get('sucursal_id')
        count = 0
        
        for row in sheet.iter_rows(min_row=2):
            rfc = str(get_val(row, 'Registro Federal de Contribuyente', '')).strip().upper()
            if not rfc:
                continue

            Contratista.objects.update_or_create(
                empresa=empresa_actual,
                rfc=rfc,
                defaults={
                    'sucursal_id': sucursal_id,
                    'nombre_razon_social': str(get_val(row, 'Nombre denominacion o razon social', '')).strip(),
                    'correo': str(get_val(row, 'Correo electronico', '')).strip(),
                    'telefono': str(get_val(row, 'Telefono (numero extension)', '')).strip(),
                    'registro_patronal': str(get_val(row, 'Registro patronal', '')).strip(),
                    'calle': str(get_val(row, 'Calle', '')).strip(),
                    'num_ext': str(get_val(row, 'Numero exterior', '')).strip(),
                    'num_int': str(get_val(row, 'Numero interior', '')).strip(),
                    'entre_calle': str(get_val(row, 'Entre calle', '')).strip(),
                    'y_calle': str(get_val(row, 'Y calle', '')).strip(),
                    'colonia': str(get_val(row, 'Colonia', '')).strip(),
                    'cp': str(get_val(row, 'Codigo Postal', '')).strip(),
                    'municipio_alcaldia': str(get_val(row, 'Municipio o Alcaldia', '')).strip(),
                    'entidad_federativa': str(get_val(row, 'Entidad Federativa', '')).strip(),
                    'representante_legal': str(get_val(row, 'Representante legal', '')).strip(),
                    'administrador_unico': str(get_val(row, 'Administrador Unico', '')).strip(),
                    'num_escritura': str(get_val(row, 'Numero de escritura', '')).strip(),
                    'nombre_notario_publico': str(get_val(row, 'Nombre del Notario Publico', '')).strip(),
                    'num_notario_publico': str(get_val(row, 'Numero de Notario Publico', '')).strip(),
                    'fecha_escritura_publica': to_date(get_val(row, 'Fecha de escritura publica')),
                    'folio_mercantil': str(get_val(row, 'Folio mercantil', '')).strip(),
                    'numero_stps': str(get_val(row, 'Numero de registro ante la Secretaria de Trabajo y Prevision Social', '')).strip(),
                }
            )
            count += 1

        return JsonResponse({
            'status': 'success', 
            'message': f'Proceso completado. Se registraron/actualizaron {count} contratistas.'
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error al procesar el archivo: {str(e)}'})

@login_required(login_url='/login/')
def exportar_sisub_contratos(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        contratista = get_object_or_404(Contratista, id=id, empresa=empresa_actual)
        cuatrimestre = request.GET.get('cuatrimestre', '1')
        anio = request.GET.get('anio', '')
        contratos = Contrato.objects.filter(contratista=contratista, empresa=empresa_actual).select_related('beneficiario')

        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SISUB Contratos"

        fill_main = PatternFill(start_color="00b8b9", end_color="00b8b9", fill_type="solid")
        fill_sec = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        fill_head = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(left=Side(style='thin', color="B2B2B2"), right=Side(style='thin', color="B2B2B2"), top=Side(style='thin', color="B2B2B2"), bottom=Side(style='thin', color="B2B2B2"))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=25)
        c1 = ws.cell(row=1, column=1, value="b-Contratos de servicio (cliente)")
        c1.alignment = center_align; c1.font = Font(bold=True, color="FFFFFF", size=12); c1.fill = fill_main; c1.border = border

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2); ws.cell(row=2, column=1, value="periodo")
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=11); ws.cell(row=2, column=3, value="a-Datos generales del contrato")
        ws.merge_cells(start_row=2, start_column=12, end_row=2, end_column=14); ws.cell(row=2, column=12, value="b-Identificacion del beneficiario")
        ws.merge_cells(start_row=2, start_column=15, end_row=2, end_column=25); ws.cell(row=2, column=15, value="c-Domicilio fiscal del beneficiario")
        
        for c in range(1, 26):
            cell = ws.cell(row=2, column=c); cell.fill = fill_sec; cell.font = Font(bold=True); cell.border = border; cell.alignment = center_align

        headers = ['Cuatrimestre', 'Año', 'RFC Sujeto', 'Folio', 'Tipo', 'Objeto', 'Monto', 'Vigencia', 'Inicio', 'Termino', 'Trabajadores', 'RFC Ben', 'Nombre Ben', 'RegPat Ben', 'Calle', 'Ext', 'Int', 'Entre', 'Y', 'Colonia', 'CP', 'Mun', 'Edo', 'Email', 'Tel']
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=h); cell.fill = fill_head; cell.font = Font(bold=True); cell.border = border; cell.alignment = center_align

        for con in contratos:
            ben = con.beneficiario
            row_data = [cuatrimestre, anio, contratista.rfc, con.folio, con.get_tipo_contrato_display(), con.objeto_contrato, con.monto_contrato, str(con.vigencia_contrato or ''), str(con.fecha_inicio or ''), str(con.fecha_fin or ''), con.num_estimado_trabajadores, ben.rfc if ben else '', ben.nombre_razon_social if ben else '', ben.registro_patronal if ben else '', ben.calle if ben else '', ben.num_ext if ben else '', ben.num_int if ben else '', ben.entre_calle if ben else '', ben.y_calle if ben else '', ben.colonia if ben else '', ben.cp if ben else '', ben.municipio_alcaldia if ben else '', ben.entidad_federativa if ben else '', ben.correo if ben else '', ben.telefono if ben else '']
            ws.append(row_data)
            for cell in ws[ws.max_row]: cell.border = border; cell.alignment = Alignment(vertical="center")

        for i in range(1, 26): ws.column_dimensions[get_column_letter(i)].width = 18

        rfc_clean = re.sub(r'[^A-Z0-9]', '', contratista.rfc.upper())
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="SISUB_CONTRATOS_{rfc_clean}.xlsx"'
        wb.save(response)
        return response
    except Exception as e: return HttpResponse(f"Error al generar reporte: {str(e)}", status=500)

@login_required(login_url='/login/')
@require_hr_permission('contratistas', 'ver')
def exportar_icsoe(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        contratista = Contratista.objects.get(id=id, empresa=empresa_actual)
        cuat = request.GET.get('cuatrimestre', '1'); anio = request.GET.get('anio', '')
        if not anio: return HttpResponse("Año requerido", status=400)
        
        periodos_busqueda = {'1': ['FEBRERO', 'ABRIL'], '2': ['JUNIO', 'AGOSTO'], '3': ['OCTUBRE', 'DICIEMBRE']}.get(cuat, [])
        rfc_input_clean = re.sub(r'[^A-Z0-9]', '', contratista.rfc.upper())
        importaciones_qs = ImportacionSUA.objects.filter(empresa=empresa_actual, tipo='bimestral')
        
        importaciones_validas = []
        for imp in importaciones_qs:
            if anio not in imp.periodo: continue
            if not any(mes in imp.periodo.upper() for mes in periodos_busqueda): continue
            rfc_rep_clean = re.sub(r'[^A-Z0-9]', '', (imp.rfc_empresa or '').upper())
            if rfc_input_clean == rfc_rep_clean or rfc_input_clean in rfc_rep_clean or rfc_rep_clean in rfc_input_clean:
                importaciones_validas.append(imp)
        
        # Obtener todos los contratos vigentes de este contratista para consolidar trabajadores
        contratos_vigentes = Contrato.objects.filter(contratista=contratista, empresa=empresa_actual, estado='vigente')
        folios_consolidado = ", ".join([c.folio for c in contratos_vigentes if c.folio]) or contratista.folio_mercantil

        # Consolidar NSS de trabajadores enlazados a CUALQUIERA de los contratos vigentes
        empleados_nss_qs = Empleado.objects.filter(
            empresa=empresa_actual,
            contratos_asignados__in=contratos_vigentes
        ).values_list('nss', flat=True).distinct()
        
        nss_con_contrato = set(re.sub(r'[^0-9]', '', str(n)) for n in empleados_nss_qs if n)

        total_sin_credito = total_con_credito = total_amortizaciones = Decimal('0')
        for imp in importaciones_validas:
            for t in imp.trabajadores.all():
                # Limpiar NSS del SUA para la comparación
                nss_t_clean = re.sub(r'[^0-9]', '', str(t.nss))
                
                # Validación: Solo considerar si el trabajador (vía NSS limpio) está enlazado a un contrato vigente
                if nss_t_clean in nss_con_contrato:
                    val_inf = (t.tipo_valor_infonavit or '').strip()
                    if not val_inf or val_inf == '-': total_sin_credito += t.aportacion_patronal
                    else: total_con_credito += t.aportacion_patronal
                    total_amortizaciones += t.amortizacion

        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "ICSOE Informativo"
        fill_brand = PatternFill(start_color="00b8b9", end_color="00b8b9", fill_type="solid")
        fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        fill_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(left=Side(style='thin', color="B2B2B2"), right=Side(style='thin', color="B2B2B2"), top=Side(style='thin', color="B2B2B2"), bottom=Side(style='thin', color="B2B2B2"))

        ws.merge_cells('A1:AA1'); c1 = ws['A1']; c1.value = "a-Datos Generales"; c1.fill = fill_brand; c1.font = Font(bold=True, color="FFFFFF"); c1.alignment = Alignment(horizontal="center"); c1.border = border

        subgrupos = [("Periodo", 2), ("b-Datos de identificacion", 5), ("c-Domicilio fiscal", 9), ("d-Datos actuales de la escritura publica", 7), ("g-Aportacion y Amortizacion", 3), ("a-Registro en STPS", 1)]
        curr_col = 1
        for nombre, span in subgrupos:
            start_cell = get_column_letter(curr_col) + "2"; end_cell = get_column_letter(curr_col + span - 1) + "2"
            if span > 1: ws.merge_cells(f"{start_cell}:{end_cell}")
            cell = ws[start_cell]; cell.value = nombre; cell.fill = fill_gray; cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center"); cell.border = border
            for c in range(curr_col, curr_col + span): ws.cell(row=2, column=c).border = border
            curr_col += span

        headers = ["cuatrimestre que declara", "anio que se declara", "Registro Federal de Contribuyente", "Nombre denominacion o razon social", "Correo electronico", "Telefono (numero extension)", "Registro patronal", "Calle", "Numero exterior", "Numero interior", "Entre calle", "Y calle", "Colonia", "Codigo Postal", "Municipio o Alcaldia", "Entidad Federativa", "Representante legal", "Administrador Unico", "Numero de escritura", "Nombre del Notario Publico", "Numero de Notario Publico", "Fecha de escritura publica", "Folio mercantil", "Aportacion sin credito de los trabajadores del contrato", "Aportacion con credito de los trabajadores del contrato", "Amortizacion de los trabajadores del contrato", "Numero de registro ante la Secretaria de Trabajo y Prevision Social"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=h); cell.fill = fill_light; cell.font = Font(bold=True); cell.border = border; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); ws.column_dimensions[get_column_letter(i)].width = 20

        # Única fila consolidada con redondeo matemático (.50 sube, .49 baja)
        total_sin_credito_red = total_sin_credito.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        total_con_credito_red = total_con_credito.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        total_amortizaciones_red = total_amortizaciones.quantize(Decimal('1'), rounding=ROUND_HALF_UP)

        data_row = [
            cuat, anio, contratista.rfc, contratista.nombre_razon_social, 
            contratista.correo, contratista.telefono, contratista.registro_patronal, 
            contratista.calle, contratista.num_ext, contratista.num_int, contratista.entre_calle, 
            contratista.y_calle, contratista.colonia, contratista.cp, contratista.municipio_alcaldia, 
            contratista.entidad_federativa, contratista.representante_legal, contratista.administrador_unico, 
            contratista.num_escritura, contratista.nombre_notario_publico, contratista.num_notario_publico, 
            str(contratista.fecha_escritura_publica) if contratista.fecha_escritura_publica else '', 
            contratista.folio_mercantil, 
            total_sin_credito_red, total_con_credito_red, total_amortizaciones_red, 
            contratista.numero_stps
        ]
        ws.append(data_row)
        
        for col_idx in [24, 25, 26]: ws.cell(row=4, column=col_idx).number_format = '"$"#,##0.00'
        for col_idx in range(1, 28): ws.cell(row=4, column=col_idx).border = border

        # Ajustar el merge inicial de la fila 1
        ws.merge_cells('A1:AA1')
        c1 = ws['A1']; c1.alignment = Alignment(horizontal="center")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="ICSOE_INFO_{rfc_input_clean}_{anio}_C{cuat}.xlsx"'; wb.save(response)
        return response
    except Exception as e: return HttpResponse(str(e), status=500)
