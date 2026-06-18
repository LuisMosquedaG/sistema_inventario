from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
import math
import re
import csv
from datetime import datetime, date

from ..models import Empleado, Contrato, Contratista, Nomina, Beneficiario, ImportacionSUA, TrabajadorSUA
from preferencias.models import Sucursal
from preferencias.permissions import require_hr_permission
from .utils import get_empresa_actual

@login_required(login_url='/login/')
@require_hr_permission('nomina', 'ver')
def lista_nomina(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html')

    nominas = Nomina.objects.filter(empresa=empresa_actual).select_related('empleado', 'sucursal').order_by('-fecha_pago', 'nombre')

    # Filtros
    q = request.GET.get('q', '').strip()
    f_folio = request.GET.get('folio', '').strip()
    f_uuid = request.GET.get('uuid', '').strip()
    f_colaborador = request.GET.get('colaborador', '').strip()
    f_rfc_contratista = request.GET.get('rfc_contratista', '').strip()
    f_fecha_pago = request.GET.get('fecha_pago', '').strip()
    f_sucursal = request.GET.get('sucursal', '').strip()

    if q:
        nominas = nominas.filter(
            Q(nombre__icontains=q) |
            Q(rfc__icontains=q) |
            Q(curp__icontains=q) |
            Q(nss__icontains=q) |
            Q(folio__icontains=q) |
            Q(uuid__icontains=q) |
            Q(rfc_contratista__icontains=q)
        ).distinct()
    
    if f_folio: nominas = nominas.filter(folio__icontains=f_folio)
    if f_uuid: nominas = nominas.filter(uuid__icontains=f_uuid)
    if f_colaborador: nominas = nominas.filter(nombre__icontains=f_colaborador)
    if f_rfc_contratista: nominas = nominas.filter(rfc_contratista__icontains=f_rfc_contratista)
    if f_fecha_pago: nominas = nominas.filter(fecha_pago=f_fecha_pago)
    if f_sucursal: nominas = nominas.filter(sucursal_id=f_sucursal)

    nominas = nominas.order_by('-fecha_pago', 'nombre')
    paginator = Paginator(nominas, 20); page_number = request.GET.get('page'); page_obj = paginator.get_page(page_number)
    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    empleados = Empleado.objects.filter(empresa=empresa_actual).order_by('apellido_paterno')
    contratistas = Contratista.objects.filter(empresa=empresa_actual).order_by('nombre_razon_social')

    return render(request, 'recursos_humanos/lista_nomina.html', {
        'page_obj': page_obj, 'sucursales': sucursales, 'empleados': empleados, 'contratistas': contratistas, 'empresa': empresa_actual,
        'filtros': {'q': q, 'folio': f_folio, 'uuid': f_uuid, 'colaborador': f_colaborador, 'rfc_contratista': f_rfc_contratista, 'fecha_pago': f_fecha_pago, 'sucursal': f_sucursal}
    })

@login_required(login_url='/login/')
@require_hr_permission('nomina', 'ver', json_response=True)
def obtener_nomina_json(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        nom = Nomina.objects.get(id=id, empresa=empresa_actual)
        def safe_decimal(val): return str(val) if val is not None else '0.00'
        data = {
            'id': nom.id, 'empleado': nom.empleado_id or '', 'periodo': nom.periodo, 'uso_cfdi': nom.uso_cfdi, 'uuid': nom.uuid or '', 'tipo_nomina': nom.tipo_nomina, 'serie': nom.serie or '', 'folio': nom.folio or '', 'fecha_emision': nom.fecha_emision.strftime('%Y-%m-%dT%H:%M') if nom.fecha_emision else '', 'fecha_certificacion': nom.fecha_certificacion.strftime('%Y-%m-%dT%H:%M') if nom.fecha_certificacion else '', 'fecha_pago': nom.fecha_pago.isoformat() if nom.fecha_pago else '', 'fecha_inicial_pago': nom.fecha_inicial_pago.isoformat() if nom.fecha_inicial_pago else '', 'fecha_final_pago': nom.fecha_final_pago.isoformat() if nom.fecha_final_pago else '', 'dias_pagados': safe_decimal(nom.dias_pagados), 'rfc': nom.rfc, 'curp': nom.curp, 'nss': nom.nss, 'nombre': nom.nombre, 'rfc_contratista': nom.rfc_contratista or '', 'sdi': safe_decimal(nom.sdi), 'sbc': safe_decimal(nom.sbc), 'vacaciones_exento': safe_decimal(nom.vacaciones_exento), 'vacaciones_dignas_exento': safe_decimal(nom.vacaciones_dignas_exento), 'aguinaldo_exento': safe_decimal(nom.aguinaldo_exento), 'sueldo_gravado': safe_decimal(nom.sueldo_gravado), 'vacaciones_gravado': safe_decimal(nom.vacaciones_gravado), 'vacaciones_dignas_gravado': safe_decimal(nom.vacaciones_dignas_gravado), 'aguinaldo_gravado': safe_decimal(nom.aguinaldo_gravado),
        }
        return JsonResponse({'success': True, 'data': data})
    except Nomina.DoesNotExist: return JsonResponse({'success': False, 'error': 'Registro de nómina no encontrado.'})
    except Exception as e: return JsonResponse({'success': False, 'error': f'Error en el servidor: {str(e)}'})

@login_required(login_url='/login/')
@require_POST
def crear_nomina_ajax(request):
    empresa_actual = get_empresa_actual(request)
    try:
        data = request.POST; sucursal_id = request.session.get('sucursal_id'); empleado_id = data.get('empleado')
        empleado = Empleado.objects.get(id=empleado_id, empresa=empresa_actual) if empleado_id else None
        nueva_nom = Nomina(
            empresa=empresa_actual, sucursal_id=sucursal_id, empleado=empleado, periodo=data.get('periodo'), uso_cfdi=data.get('uso_cfdi', 'CN01'), uuid=data.get('uuid'), tipo_nomina=data.get('tipo_nomina', 'O'), serie=data.get('serie'), folio=data.get('folio'), fecha_emision=data.get('fecha_emision') or None, fecha_certificacion=data.get('fecha_certificacion') or None, fecha_pago=data.get('fecha_pago') or None, fecha_inicial_pago=data.get('fecha_inicial_pago') or None, fecha_final_pago=data.get('fecha_final_pago') or None, dias_pagados=Decimal(data.get('dias_pagados', '0')), rfc=data.get('rfc', '').upper(), curp=data.get('curp', '').upper(), nss=data.get('nss', ''), nombre=data.get('nombre', ''), rfc_contratista=data.get('rfc_contratista', '').upper(), sdi=Decimal(data.get('sdi', '0')), sbc=Decimal(data.get('sbc', '0')), vacaciones_exento=Decimal(data.get('vacaciones_exento', '0')), vacaciones_dignas_exento=Decimal(data.get('vacaciones_dignas_exento', '0')), aguinaldo_exento=Decimal(data.get('aguinaldo_exento', '0')), sueldo_gravado=Decimal(data.get('sueldo_gravado', '0')), vacaciones_gravado=Decimal(data.get('vacaciones_gravado', '0')), vacaciones_dignas_gravado=Decimal(data.get('vacaciones_dignas_gravado', '0')), aguinaldo_gravado=Decimal(data.get('aguinaldo_gravado', '0')),
        )
        nueva_nom.save(); return JsonResponse({'success': True, 'message': 'Nómina registrada correctamente.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
def editar_nomina_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        nom = Nomina.objects.get(id=id, empresa=empresa_actual); data = request.POST
        nom.empleado_id = data.get('empleado') or None; nom.periodo = data.get('periodo'); nom.uso_cfdi = data.get('uso_cfdi', 'CN01'); nom.uuid = data.get('uuid'); nom.tipo_nomina = data.get('tipo_nomina'); nom.serie = data.get('serie'); nom.folio = data.get('folio'); nom.fecha_emision = data.get('fecha_emision') or None; nom.fecha_certificacion = data.get('fecha_certificacion') or None; nom.fecha_pago = data.get('fecha_pago') or None; nom.fecha_inicial_pago = data.get('fecha_inicial_pago') or None; nom.fecha_final_pago = data.get('fecha_final_pago') or None; nom.dias_pagados = Decimal(data.get('dias_pagados', '0')); nom.rfc = data.get('rfc', '').upper(); nom.curp = data.get('curp', '').upper(); nom.nss = data.get('nss', ''); nom.nombre = data.get('nombre', ''); nom.rfc_contratista = data.get('rfc_contratista', '').upper(); nom.sdi = Decimal(data.get('sdi', '0')); nom.sbc = Decimal(data.get('sbc', '0')); nom.vacaciones_exento = Decimal(data.get('vacaciones_exento', '0')); nom.vacaciones_dignas_exento = Decimal(data.get('vacaciones_dignas_exento', '0')); nom.aguinaldo_exento = Decimal(data.get('aguinaldo_exento', '0')); nom.sueldo_gravado = Decimal(data.get('sueldo_gravado', '0')); nom.vacaciones_gravado = Decimal(data.get('vacaciones_gravado', '0')); nom.vacaciones_dignas_gravado = Decimal(data.get('vacaciones_dignas_gravado', '0')); nom.aguinaldo_gravado = Decimal(data.get('aguinaldo_gravado', '0'))
        nom.save(); return JsonResponse({'success': True, 'message': 'Nómina actualizada correctamente.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
def eliminar_nomina_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        nom = Nomina.objects.get(id=id, empresa=empresa_actual); nom.delete()
        return JsonResponse({'success': True, 'message': 'Registro de nómina eliminado correctamente.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
def importar_nomina_ajax(request):
    empresa_actual = get_empresa_actual(request); archivo = request.FILES.get('archivo')
    if not archivo: return JsonResponse({'success': False, 'error': 'No se proporcionó ningún archivo.'})
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True); sheet = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]; col_map = {h: i for i, h in enumerate(headers) if h}
        def get_val(row, h, d=None): i = col_map.get(h); return row[i].value if i is not None and row[i].value is not None else d
        def to_dec(v):
            try: x = str(v or 0).replace(',', '').replace('$', '').strip(); return Decimal(x) if x else Decimal('0')
            except: return Decimal('0')
        def to_d(v):
            if isinstance(v, datetime): return v.date()
            if isinstance(v, str):
                for f in ('%Y-%m-%d','%d/%m/%Y'):
                    try: return datetime.strptime(v,f).date()
                    except: pass
            return None
        sucursal_id = request.session.get('sucursal_id'); count = 0
        for row in sheet.iter_rows(min_row=2):
            curp_f = str(get_val(row, 'CURP', '')).strip().upper(); nss_f = str(get_val(row, 'No Seguro social', '')).strip()
            if not curp_f or not nss_f: continue
            empleado = Empleado.objects.filter(empresa=empresa_actual, curp=curp_f, nss=nss_f).first()
            tipo_nom = 'E' if 'EXTRAORDINARIA' in str(get_val(row,'Tipo nomina','O')).upper() else 'O'
            nueva_nom = Nomina(
                empresa=empresa_actual, sucursal_id=sucursal_id, empleado=empleado, periodo=str(get_val(row,'Periodo','')), uso_cfdi=str(get_val(row,'Uso CFDI','CN01')), uuid=str(get_val(row,'UUID','')), tipo_nomina=tipo_nom, serie=str(get_val(row,'Serie','')), folio=str(get_val(row,'Folio','')), fecha_pago=to_d(get_val(row,'Fecha pago')), rfc=str(get_val(row,'RFC receptor','')), curp=curp_f, nss=nss_f, nombre=str(get_val(row,'Razon receptor','')), rfc_contratista=str(get_val(row,'RFC emisor','')), sdi=to_dec(get_val(row,'Salario diario integrado')), sbc=to_dec(get_val(row,'Salario base cot apor')), sueldo_gravado=to_dec(get_val(row,'001/P001/Gravado/SUELDO'))
            )
            nueva_nom.save(); count += 1
        return JsonResponse({'success': True, 'message': f'Se importaron {count} registros.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
def actualizar_datos_trabajadores_nomina_ajax(request):
    empresa_actual = get_empresa_actual(request)
    try:
        empleados_map = {re.sub(r'[^A-Z0-9]', '', (e.rfc or '').upper()): e for e in Empleado.objects.filter(empresa=empresa_actual) if e.rfc}
        nominas = Nomina.objects.filter(empresa=empresa_actual); actualizados = 0
        for nom in nominas:
            rfc_nom = re.sub(r'[^A-Z0-9]', '', (nom.rfc or '').upper()); empleado = empleados_map.get(rfc_nom)
            if empleado:
                c = False
                if not nom.nss and empleado.nss: nom.nss = empleado.nss; c = True
                if not nom.curp and empleado.curp: nom.curp = empleado.curp; c = True
                if (not nom.sdi or nom.sdi == 0) and empleado.sdi: nom.sdi = empleado.sdi; c = True
                if not nom.empleado: nom.empleado = empleado; c = True
                if c: nom.save(); actualizados += 1
        return JsonResponse({'success': True, 'message': f'Actualizados {actualizados} recibos.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
def exportar_nominas_excel(request):
    empresa_actual = get_empresa_actual(request)
    nominas = Nomina.objects.filter(empresa=empresa_actual).select_related('empleado', 'sucursal')
    q = request.GET.get('q','').strip(); f_folio = request.GET.get('folio','').strip(); f_uuid = request.GET.get('uuid','').strip(); f_col = request.GET.get('colaborador','').strip(); f_rfc = request.GET.get('rfc_contratista','').strip(); f_pago = request.GET.get('fecha_pago','').strip(); f_suc = request.GET.get('sucursal','').strip()
    if q: nominas = nominas.filter(Q(nombre__icontains=q)|Q(rfc__icontains=q)|Q(curp__icontains=q)|Q(nss__icontains=q)|Q(folio__icontains=q)|Q(uuid__icontains=q)).distinct()
    if f_folio: nominas = nominas.filter(folio__icontains=f_folio)
    if f_uuid: nominas = nominas.filter(uuid__icontains=f_uuid)
    if f_col: nominas = nominas.filter(nombre__icontains=f_col)
    if f_rfc: nominas = nominas.filter(rfc_contratista__icontains=f_rfc)
    if f_pago: nominas = nominas.filter(fecha_pago=f_pago)
    if f_suc: nominas = nominas.filter(sucursal_id=f_suc)
    nominas = nominas.order_by('-fecha_pago', 'nombre')
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Nominas"
    from openpyxl.styles import Font, Alignment
    h_font = Font(bold=True)
    headers = [
        "Folio", "Serie", "UUID", "Uso CFDI", "Tipo Nomina", "Periodo", 
        "Fecha Emision", "Fecha Certificacion", "Fecha Pago", "Fecha Inicial", "Fecha Final", 
        "Dias Pagados", "Colaborador", "RFC", "CURP", "NSS", "RFC Contratista", "SDI", "SBC",
        "Sueldo (Gravado)", "Vacaciones (Gravado)", "Vacaciones (Exento)", 
        "Vacaciones Dignas (Gravado)", "Vacaciones Dignas (Exento)", 
        "Aguinaldo (Gravado)", "Aguinaldo (Exento)", "Sucursal"
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.font = h_font; c.alignment = Alignment(horizontal="center")
    for r_idx, nom in enumerate(nominas, 2):
        data = [
            nom.folio, nom.serie, nom.uuid, nom.uso_cfdi, "O" if nom.tipo_nomina=='O' else "E", nom.periodo, 
            nom.fecha_emision, nom.fecha_certificacion, nom.fecha_pago, nom.fecha_inicial_pago, nom.fecha_final_pago, 
            nom.dias_pagados, nom.nombre, nom.rfc, nom.curp, nom.nss, nom.rfc_contratista, nom.sdi, nom.sbc,
            nom.sueldo_gravado, nom.vacaciones_gravado, nom.vacaciones_exento, 
            nom.vacaciones_dignas_gravado, nom.vacaciones_dignas_exento, 
            nom.aguinaldo_gravado, nom.aguinaldo_exento,
            nom.sucursal.nombre if nom.sucursal else "General"
        ]
        for c_idx, val in enumerate(data, 1):
            cl = ws.cell(row=r_idx, column=c_idx, value=val)
            if isinstance(val, (datetime, date)):
                cl.number_format = 'yyyy-mm-dd'
            elif isinstance(val, Decimal):
                cl.number_format = '#,##0.00'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'); response['Content-Disposition'] = 'attachment; filename="Listado_Nominas.xlsx"'; wb.save(response); return response

@login_required(login_url='/login/')
def exportar_sisub_trabajadores(request, id):
    empresa_actual = get_empresa_actual(request); contratista = get_object_or_404(Contratista, id=id, empresa=empresa_actual)
    cuat = int(request.GET.get('cuatrimestre', 1)); anio_val = request.GET.get('anio', ''); formato = request.GET.get('formato', 'excel')
    if not anio_val: return HttpResponse("Año requerido", status=400)
    anio = int(anio_val); meses_filtro = {1:[1,2,3,4], 2:[5,6,7,8], 3:[9,10,11,12]}.get(cuat)
    contratos = Contrato.objects.filter(contratista=contratista, empresa=empresa_actual, estado='vigente').prefetch_related('empleados', 'beneficiario')
    emp_map = {}
    for c in contratos:
        for e in c.empleados.all():
            n = re.sub(r'[^0-9]', '', e.nss) if e.nss else ""; curp = re.sub(r'[^A-Z0-9]', '', (e.curp or '').upper()); name = f"{e.nombre} {e.apellido_paterno} {e.apellido_materno}".strip().upper()
            if n: emp_map[n] = c
            if curp: emp_map[curp] = c
            if name: emp_map[name] = c
    recibos = [r for r in Nomina.objects.filter(empresa=empresa_actual, fecha_pago__year=anio, fecha_pago__month__in=meses_filtro) if emp_map.get(re.sub(r'[^0-9]','',r.nss or '')) or emp_map.get(re.sub(r'[^A-Z0-9]','',(r.curp or '').upper())) or emp_map.get((r.nombre or '').strip().upper())]
    
    rfc_clean_input = re.sub(r'[^A-Z0-9]', '', contratista.rfc.upper())
    sua_data = {}
    ultimo_sdi_trabajador = {}
    nombres = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    for m in meses_filtro:
        imps = ImportacionSUA.objects.filter(
            empresa=empresa_actual,
            periodo__icontains=nombres[m].upper()
        ).filter(
            periodo__icontains=str(anio)
        )
        for imp in imps:
            rfc_imp_clean = re.sub(r'[^A-Z0-9]', '', (imp.rfc_empresa or '').upper())
            if rfc_clean_input == rfc_imp_clean or rfc_clean_input in rfc_imp_clean or rfc_imp_clean in rfc_clean_input:
                bim = (m+1)//2
                for ts in imp.trabajadores.all():
                    tn = re.sub(r'[^0-9]','',ts.nss) if ts.nss else ""; tc = re.sub(r'[^A-Z0-9]','',(ts.rfc_curp or '').upper()); tm = (ts.nombre or "").strip().upper()
                    
                    sdi_val = ts.sdi
                    last_known_sdi = None
                    if tn and tn in ultimo_sdi_trabajador:
                        last_known_sdi = ultimo_sdi_trabajador[tn]
                    elif tc and tc in ultimo_sdi_trabajador:
                        last_known_sdi = ultimo_sdi_trabajador[tc]
                    elif tm and tm in ultimo_sdi_trabajador:
                        last_known_sdi = ultimo_sdi_trabajador[tm]
                    
                    if (not sdi_val or sdi_val == 0) and last_known_sdi:
                        sdi_val = last_known_sdi
                    elif sdi_val and sdi_val > 0:
                        if tn: ultimo_sdi_trabajador[tn] = sdi_val
                        if tc: ultimo_sdi_trabajador[tc] = sdi_val
                        if tm: ultimo_sdi_trabajador[tm] = sdi_val
                    
                    v = {'sdi': sdi_val, 'inc': 0}
                    try: v['inc'] = int(float(ts.incapacidades or 0))
                    except: pass
                    if tn: sua_data[(tn, bim)] = v
                    if tc: sua_data[(tc, bim)] = v
                    if tm: sua_data[(tm, bim)] = v



    headers = ["cuatrimestre que declara", "anio que se declara", "bimestre", "Registro Federal de Contribuyente del sujeto obligado", "Numero de contrato", "Registro Patronal ante el IMSS", "Numero de Seguro Social del trabajador", "Calle (centro del trabajo)", "Numero exterior (centro del trabajo)", "Numero interior (centro de trabajo)", "Colonia (centro de trabajo)", "Codigo Postal (centro de trabajo)", "Municipio o Alcaldia (centro de trabajo)", "Entidad federativa (centro de trabajo)", "Monto Percepciones variables", "Monto Percepciones fijas", "Dias de Incapacidad", "Percepciones no integrables al SBA", "salario no excedente (VSM)"]
    
    grouped_data = {}
    for r in recibos:
        n = re.sub(r'[^0-9]','',r.nss or '')
        curp = re.sub(r'[^A-Z0-9]','',(r.curp or '').upper())
        name = (r.nombre or '').strip().upper()
        if not r.fecha_pago:
            continue
        bim = (r.fecha_pago.month + 1) // 2
        
        # Clave única del trabajador en este bimestre
        worker_key = n if n else (curp if curp else name)
        if not worker_key:
            continue
            
        key = (worker_key, bim)
        per = (r.vacaciones_exento or 0) + (r.vacaciones_dignas_exento or 0) + (r.aguinaldo_exento or 0) + (r.sueldo_gravado or 0) + (r.vacaciones_gravado or 0) + (r.vacaciones_dignas_gravado or 0) + (r.aguinaldo_gravado or 0)
        
        if key not in grouped_data:
            con = emp_map.get(n) or emp_map.get(curp) or emp_map.get(name)
            ben = con.beneficiario if con else None
            s = sua_data.get((n, bim)) or sua_data.get((curp, bim)) or sua_data.get((name, bim)) or {'sdi': Decimal('0.00'), 'inc': 0}
            
            sdi_val = s['sdi']
            if not sdi_val or sdi_val == 0:
                # 1. Fallback to last known non-zero SDI in this cuatrimestre
                last_known = None
                if n and n in ultimo_sdi_trabajador:
                    last_known = ultimo_sdi_trabajador[n]
                elif curp and curp in ultimo_sdi_trabajador:
                    last_known = ultimo_sdi_trabajador[curp]
                elif name and name in ultimo_sdi_trabajador:
                    last_known = ultimo_sdi_trabajador[name]
                
                if last_known and last_known > 0:
                    sdi_val = last_known
                else:
                    # 2. Fallback to employee profile SDI
                    empleado = None
                    if n: empleado = Empleado.objects.filter(empresa=empresa_actual, nss=n).first()
                    if not empleado and curp: empleado = Empleado.objects.filter(empresa=empresa_actual, curp=curp).first()
                    
                    if empleado and empleado.sdi and empleado.sdi > 0:
                        sdi_val = empleado.sdi
            
            grouped_data[key] = {
                'cuat': cuat,
                'anio': anio,
                'bim': bim,
                'contratista_rfc': contratista.rfc,
                'contrato_folio': con.folio if con and con.folio else "S/F",
                'registro_patronal': contratista.registro_patronal,
                'nss': n or r.nss,
                'calle': ben.calle if ben else '',
                'num_ext': ben.num_ext if ben else '',
                'num_int': ben.num_int if ben else '',
                'colonia': ben.colonia if ben else '',
                'cp': ben.cp if ben else '',
                'municipio': ben.municipio_alcaldia if ben else '',
                'entidad': ben.entidad_federativa if ben else '',
                'percepciones': Decimal('0.00'),
                'incapacidades': s['inc'],
                'sdi': sdi_val
            }
            
        grouped_data[key]['percepciones'] += per

    data_rows = []
    for key in sorted(grouped_data.keys(), key=lambda x: (x[1], x[0])):
        d = grouped_data[key]
        per_rounded = int(Decimal(d['percepciones']).quantize(Decimal('1'), ROUND_HALF_UP))
        data_rows.append([
            d['cuat'], d['anio'], d['bim'], d['contratista_rfc'], d['contrato_folio'],
            d['registro_patronal'], d['nss'], d['calle'], d['num_ext'], d['num_int'],
            d['colonia'], d['cp'], d['municipio'], d['entidad'], 0,
            per_rounded, d['incapacidades'], 0, d['sdi']
        ])

    if formato == 'csv':
        resp = HttpResponse(content_type='text/csv'); resp['Content-Disposition'] = f'attachment; filename="SISUB_TRABAJADORES_{contratista.rfc}.csv"'; resp.write(u'\ufeff'.encode('utf8')); w = csv.writer(resp); w.writerow(headers); w.writerows(data_rows); return resp
    else:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "SISUB"
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        fb = PatternFill(start_color="00b8b9", end_color="00b8b9", fill_type="solid"); fg = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"); fl = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"); br = Border(left=Side(style='thin', color="B2B2B2"), right=Side(style='thin', color="B2B2B2"), top=Side(style='thin', color="B2B2B2"), bottom=Side(style='thin', color="B2B2B2"))
        ws.merge_cells('A1:N1'); ws['A1'] = "d-Informacion de los trabajadores"; ws.merge_cells('O1:S1'); ws['O1'] = "e-Determinacion del salario base de aportacion"
        for c in range(1, 20): cl = ws.cell(row=1, column=c); cl.fill = fb; cl.font = Font(bold=True, color="FFFFFF"); cl.alignment = Alignment(horizontal="center"); cl.border = br
        ws.merge_cells('A2:C2'); ws['A2'] = "Periodo"; ws.merge_cells('D2:N2'); ws['D2'] = "a-Identificacion"; ws.merge_cells('O2:S2'); ws['O2'] = "a-Percepciones por bimestre"
        for c in range(1, 20): cl = ws.cell(row=2, column=c); cl.fill = fg; cl.font = Font(bold=True); cl.alignment = Alignment(horizontal="center"); cl.border = br
        for i, h in enumerate(headers, 1): cl = ws.cell(row=3, column=i); cl.value = h; cl.font = Font(bold=True); cl.fill = fl; cl.alignment = Alignment(horizontal="center"); cl.border = br
        for r_i, rd in enumerate(data_rows, 4):
            for c_i, v in enumerate(rd, 1):
                cl = ws.cell(row=r_i, column=c_i, value=v)
                if c_i == 19: cl.number_format = '#,##0.00'
                cl.border = br
        for i in range(1, 20): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22
        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'); resp['Content-Disposition'] = f'attachment; filename="SISUB_TRABAJADORES_{contratista.rfc}.xlsx"'; wb.save(resp); return resp
