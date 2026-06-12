from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q
from decimal import Decimal
import re
import csv
import openpyxl
import math
import uuid
from datetime import datetime
from django.http import JsonResponse, HttpResponse
try:
    import pdfplumber
except ImportError:
    pdfplumber = None

from panel.models import Empresa
from .models import (
    Empleado, Contrato, Contratista, Beneficiario, 
    ImportacionSUA, TrabajadorSUA, Nomina,
    FielContratista, SolicitudDescargaSAT
)
from .security_utils import (
    get_master_key, generate_data_key, encrypt_data, decrypt_data,
    cifrar_archivos_fiel, descifrar_archivo
)
from preferencias.models import Sucursal
from preferencias.permissions import require_hr_permission

def limpiar_basura_header(texto):
    """Elimina textos innecesarios del encabezado del SUA como convenios y versiones."""
    if not texto: return ""
    # Patrones que suelen "pegarse" al final de los campos reales
    patrones = [
        r'Convenio\s+de\s+Re?mbolso:.*',
        r'Aportación\s+Patronal:.*',
        r'V\s?\d\.\d\.\d.*',
        r'Página:.*',
        r'Hoja:.*'
    ]
    for p in patrones:
        texto = re.sub(p, '', texto, flags=re.I)
    return texto.strip()

def get_empresa_actual(request):
    username = request.user.username
    if '@' in username:
        subdominio = username.split('@')[1]
        try:
            return Empresa.objects.get(subdominio=subdominio)
        except Empresa.DoesNotExist:
            return None
    return None

def get_sucursal_actual_id(request):
    """Auxiliar para obtener la sucursal de la sesión o el filtro."""
    suc_id = request.POST.get('sucursal') or request.GET.get('sucursal') or request.session.get('sucursal_id')
    return suc_id

@login_required(login_url='/login/')
@require_hr_permission('empleados', 'ver')
def lista_empleados(request):
    empresa_actual = get_empresa_actual(request)
    empleados = Empleado.objects.filter(empresa=empresa_actual).select_related('sucursal', 'contratista', 'beneficiario').order_by('apellido_paterno')
    
    # Obtener parámetros de filtros
    q = request.GET.get('q', '')
    f_nss = request.GET.get('nss', '')
    f_empleado = request.GET.get('empleado', '')
    f_contratista = request.GET.get('contratista', '')
    f_beneficiario = request.GET.get('beneficiario', '')
    f_sucursal = request.GET.get('sucursal', '')
    f_estado = request.GET.get('estado', '')

    # Aplicar Filtros
    if q:
        empleados = empleados.filter(
            Q(nombre__icontains=q) |
            Q(apellido_paterno__icontains=q) |
            Q(apellido_materno__icontains=q) |
            Q(nss__icontains=q) |
            Q(curp__icontains=q) |
            Q(rfc__icontains=q) |
            Q(puesto__icontains=q)
        )
    
    if f_nss:
        empleados = empleados.filter(nss__icontains=f_nss)
    
    if f_empleado:
        empleados = empleados.filter(
            Q(nombre__icontains=f_empleado) |
            Q(apellido_paterno__icontains=f_empleado) |
            Q(apellido_materno__icontains=f_empleado)
        )

    if f_contratista:
        empleados = empleados.filter(contratista_id=f_contratista)
    
    if f_beneficiario:
        empleados = empleados.filter(beneficiario_id=f_beneficiario)
            
    if f_sucursal:
        empleados = empleados.filter(sucursal_id=f_sucursal)

    if f_estado:
        if f_estado == 'baja':
            empleados = empleados.filter(estado__icontains='baja')
        else:
            empleados = empleados.filter(estado=f_estado)
            
    from preferencias.models import Sucursal
    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    contratistas = Contratista.objects.filter(empresa=empresa_actual).order_by('nombre_razon_social')
    beneficiarios = Beneficiario.objects.filter(empresa=empresa_actual).order_by('nombre_razon_social')
    contratos_disponibles = Contrato.objects.filter(empresa=empresa_actual).select_related('contratista', 'beneficiario')

    # PAGINACIÓN
    paginator = Paginator(empleados, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'recursos_humanos/lista_empleados.html', {
        'page_obj': page_obj,
        'sucursales': sucursales,
        'contratistas': contratistas,
        'beneficiarios': beneficiarios,
        'contratos_disponibles': contratos_disponibles,
        'empresa': empresa_actual,
        'filtros': {
            'q': q,
            'nss': f_nss,
            'empleado': f_empleado,
            'contratista': f_contratista,
            'beneficiario': f_beneficiario,
            'sucursal': f_sucursal,
            'estado': f_estado,
        }
    })

@login_required(login_url='/login/')
@require_hr_permission('empleados', 'ver', json_response=True)
def obtener_empleado_json(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        emp = Empleado.objects.get(id=id, empresa=empresa_actual)
        contrato_vinculado = emp.contratos_asignados.order_by('-fecha_inicio').first()

        data = {
            'id': emp.id,
            'contrato_id': contrato_vinculado.id if contrato_vinculado else '',
            'curp': emp.curp,
            'rfc': emp.rfc,
            'nss': emp.nss,
            'id_oficial': emp.id_oficial,
            'fecha_nacimiento': emp.fecha_nacimiento.isoformat() if emp.fecha_nacimiento else '',
            'genero': emp.genero,
            'nacionalidad': emp.nacionalidad,
            'nombre': emp.nombre,
            'apellido_paterno': emp.apellido_paterno,
            'apellido_materno': emp.apellido_materno,
            'estado_civil': emp.estado_civil,
            'correo_personal': emp.correo_personal,
            'telefono_movil': emp.telefono_movil,
            'telefono_fijo': emp.telefono_fijo,
            'calle': emp.calle,
            'num_ext': emp.num_ext,
            'num_int': emp.num_int,
            'colonia': emp.colonia,
            'cp': emp.cp,
            'ciudad': emp.ciudad,
            'estado_dir': emp.estado_dir,
            'num_empleado': emp.num_empleado,
            'estado': emp.estado,
            'fecha_ingreso': emp.fecha_ingreso.isoformat() if emp.fecha_ingreso else '',
            'fecha_antiguedad': emp.fecha_antiguedad.isoformat() if emp.fecha_antiguedad else '',
            'fecha_expiracion': emp.fecha_expiracion.isoformat() if emp.fecha_expiracion else '',
            'tipo_contrato': emp.tipo_contrato,
            'jornada': emp.jornada,
            'puesto': emp.puesto,
            'departamento': emp.departamento,
            'supervisor': emp.supervisor,
            'clave_ubicacion': emp.clave_ubicacion or '',
            'notas': emp.notas or '',
            'riesgo_trabajo': emp.riesgo_trabajo,
            'tipo_trabajador': emp.tipo_trabajador,
            'salario_diario_ordinario': str(emp.salario_diario_ordinario),
            'sbc': str(emp.sbc),
            'sdi': str(emp.sdi),
            'forma_pago': emp.forma_pago,
            'clave_percepcion_sat': emp.clave_percepcion_sat,
            'tipo_salario': emp.tipo_salario,
            'registro_patronal': emp.registro_patronal,
            'num_infonavit': emp.num_infonavit,
            'num_fonacot': emp.num_fonacot,
            'fondo_ahorro': emp.fondo_ahorro,
            'porcentaje_fondo': str(emp.porcentaje_fondo),
            'banco_nombre': emp.banco_nombre,
            'clabe': emp.clabe,
            'num_cuenta': emp.num_cuenta,
            'tipo_cuenta': emp.tipo_cuenta,
            'sucursal': emp.sucursal_id or '',
            'contratista_id': emp.contratista_id or '',
            'beneficiario_id': emp.beneficiario_id or '',
        }
        return JsonResponse({'success': True, 'data': data})
    except Empleado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Empleado no encontrado.'})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('empleados', 'editar', json_response=True)
def editar_empleado_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        emp = Empleado.objects.get(id=id, empresa=empresa_actual)
        data = request.POST

        num_empleado = data.get('num_empleado')
        if num_empleado != emp.num_empleado and Empleado.objects.filter(empresa=empresa_actual, num_empleado=num_empleado).exists():
            return JsonResponse({'success': False, 'error': f'El número de empleado {num_empleado} ya existe.'})

        emp.curp = data.get('curp', '').upper()
        emp.rfc = data.get('rfc', '').upper()
        emp.nss = data.get('nss', '')
        emp.id_oficial = data.get('id_oficial')
        emp.fecha_nacimiento = data.get('fecha_nacimiento') or None
        emp.genero = data.get('genero')
        emp.nacionalidad = data.get('nacionalidad', 'Mexicana')
        emp.nombre = data.get('nombre')
        emp.apellido_paterno = data.get('apellido_paterno')
        emp.apellido_materno = data.get('apellido_materno')
        emp.estado_civil = data.get('estado_civil')
        emp.correo_personal = data.get('correo_personal')
        emp.telefono_movil = data.get('telefono_movil')
        emp.telefono_fijo = data.get('telefono_fijo')
        emp.calle = data.get('calle')
        emp.num_ext = data.get('num_ext')
        emp.num_int = data.get('num_int')
        emp.colonia = data.get('colonia')
        emp.cp = data.get('cp')
        emp.ciudad = data.get('ciudad')
        emp.estado_dir = data.get('estado_dir')
        emp.num_empleado = num_empleado
        emp.estado = data.get('estado')
        emp.fecha_ingreso = data.get('fecha_ingreso') or None
        emp.fecha_antiguedad = data.get('fecha_antiguedad') or None
        emp.tipo_contrato = data.get('tipo_contrato')
        emp.jornada = data.get('jornada')
        emp.puesto = data.get('puesto')
        emp.departamento = data.get('departamento')
        emp.supervisor = data.get('supervisor')
        emp.riesgo_trabajo = data.get('riesgo_trabajo')
        emp.tipo_trabajador = data.get('tipo_trabajador')
        
        # Relaciones directas
        emp.contratista_id = data.get('contratista') or None
        emp.beneficiario_id = data.get('beneficiario') or None
        
        emp.salario_diario_ordinario = Decimal(data.get('salario_diario_ordinario', '0'))
        emp.sbc = Decimal(data.get('sbc', '0'))
        emp.sdi = Decimal(data.get('sdi', '0'))
        emp.forma_pago = data.get('forma_pago')
        emp.clave_percepcion_sat = data.get('clave_percepcion_sat', '001')
        emp.tipo_salario = data.get('tipo_salario')
        emp.registro_patronal = data.get('registro_patronal')
        emp.num_infonavit = data.get('num_infonavit')
        emp.num_fonacot = data.get('num_fonacot')
        emp.fondo_ahorro = (data.get('fondo_ahorro') == 'on')
        emp.porcentaje_fondo = Decimal(data.get('porcentaje_fondo', '0'))
        emp.banco_nombre = data.get('banco_nombre')
        emp.clabe = data.get('clabe')
        emp.num_cuenta = data.get('num_cuenta')
        emp.tipo_cuenta = data.get('tipo_cuenta')
        
        emp.save()

        contrato_id = data.get('contrato_id')
        if contrato_id:
            try:
                contrato = Contrato.objects.get(id=contrato_id, empresa=empresa_actual)
                emp.contratos_asignados.clear()
                contrato.empleados.add(emp)
            except Contrato.DoesNotExist:
                pass
        else:
            emp.contratos_asignados.clear()

        return JsonResponse({'success': True, 'message': 'Empleado actualizado correctamente.'})
    except Empleado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Empleado no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('empleados', 'eliminar', json_response=True)
def eliminar_empleado_ajax(request, id):
    """Eliminar un empleado."""
    empresa_actual = get_empresa_actual(request)
    try:
        emp = Empleado.objects.get(id=id, empresa=empresa_actual)
        emp.delete()
        return JsonResponse({'success': True, 'message': 'Empleado eliminado correctamente.'})
    except Empleado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Empleado no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('empleados', 'crear', json_response=True)
def crear_empleado_ajax(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'No se encontró la empresa.'}, status=403)

    try:
        data = request.POST
        num_empleado = data.get('num_empleado')
        if Empleado.objects.filter(empresa=empresa_actual, num_empleado=num_empleado).exists():
            return JsonResponse({'success': False, 'error': f'El número de empleado {num_empleado} ya existe.'})

        sucursal_id = request.session.get('sucursal_id')

        nuevo_empleado = Empleado(
            empresa=empresa_actual,
            sucursal_id=sucursal_id,
            curp=data.get('curp', '').upper(),
            rfc=data.get('rfc', '').upper(),
            nss=data.get('nss', ''),
            id_oficial=data.get('id_oficial'),
            fecha_nacimiento=data.get('fecha_nacimiento') or None,
            genero=data.get('genero'),
            nacionalidad=data.get('nacionalidad', 'Mexicana'),
            nombre=data.get('nombre'),
            apellido_paterno=data.get('apellido_paterno'),
            apellido_materno=data.get('apellido_materno'),
            estado_civil=data.get('estado_civil'),
            correo_personal=data.get('correo_personal'),
            telefono_movil=data.get('telefono_movil'),
            telefono_fijo=data.get('telefono_fijo'),
            calle=data.get('calle'),
            num_ext=data.get('num_ext'),
            num_int=data.get('num_int'),
            colonia=data.get('colonia'),
            cp=data.get('cp'),
            ciudad=data.get('ciudad'),
            estado_dir=data.get('estado_dir'),
            num_empleado=num_empleado,
            estado=data.get('estado'),
            fecha_ingreso=data.get('fecha_ingreso') or None,
            fecha_antiguedad=data.get('fecha_antiguedad') or data.get('fecha_ingreso') or None,
            fecha_expiracion=data.get('fecha_expiracion') or None,
            tipo_contrato=data.get('tipo_contrato'),
            jornada=data.get('jornada'),
            puesto=data.get('puesto'),
            departamento=data.get('departamento'),
            supervisor=data.get('supervisor'),
            clave_ubicacion=data.get('clave_ubicacion'),
            notas=data.get('notas'),
            riesgo_trabajo=data.get('riesgo_trabajo'),
            tipo_trabajador=data.get('tipo_trabajador'),
            contratista_id=data.get('contratista') or None,
            beneficiario_id=data.get('beneficiario') or None,
            salario_diario_ordinario=Decimal(data.get('salario_diario_ordinario', '0')),
            sbc=Decimal(data.get('sbc', '0')),
            sdi=Decimal(data.get('sdi', '0')),
            forma_pago=data.get('forma_pago'),
            clave_percepcion_sat=data.get('clave_percepcion_sat', '001'),
            tipo_salario=data.get('tipo_salario'),
            registro_patronal=data.get('registro_patronal'),
            num_infonavit=data.get('num_infonavit'),
            num_fonacot=data.get('num_fonacot'),
            fondo_ahorro=(data.get('fondo_ahorro') == 'on'),
            porcentaje_fondo=Decimal(data.get('porcentaje_fondo', '0')),
            caja_ahorro=(data.get('caja_ahorro') == 'on'),
            banco_nombre=data.get('banco_nombre'),
            clabe=data.get('clabe'),
            num_cuenta=data.get('num_cuenta'),
            tipo_cuenta=data.get('tipo_cuenta'),
            tarjeta_nomina=(data.get('tarjeta_nomina') == 'on'),
            num_tarjeta=data.get('num_tarjeta'),
        )
        nuevo_empleado.save()

        contrato_id = data.get('contrato_id')
        if contrato_id:
            try:
                contrato = Contrato.objects.get(id=contrato_id, empresa=empresa_actual)
                contrato.empleados.add(nuevo_empleado)
            except Contrato.DoesNotExist:
                pass

        return JsonResponse({'success': True, 'message': 'Empleado registrado correctamente.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_hr_permission('contratos', 'ver')
def lista_contratos(request):
    empresa_actual = get_empresa_actual(request)
    contratos = Contrato.objects.filter(empresa=empresa_actual).prefetch_related('empleados', 'empleados__sucursal').order_by('-fecha_inicio')
    
    q = request.GET.get('q', '')
    empleado_id = request.GET.get('empleado_id', '')
    estado = request.GET.get('estado', '')
    sucursal_id = request.GET.get('sucursal', '')

    if q:
        contratos = contratos.filter(
            Q(folio__icontains=q) |
            Q(beneficiario__nombre_razon_social__icontains=q) |
            Q(contratista__nombre_razon_social__icontains=q) |
            Q(empleados__nombre__icontains=q) |
            Q(empleados__apellido_paterno__icontains=q) |
            Q(notas__icontains=q)
        ).distinct()
    
    if empleado_id:
        contratos = contratos.filter(empleados__id=empleado_id)
    
    if estado:
        contratos = contratos.filter(estado=estado)
            
    if sucursal_id:
        contratos = contratos.filter(sucursal_id=sucursal_id)

    empleados = Empleado.objects.filter(empresa=empresa_actual).order_by('apellido_paterno')
    contratistas = Contratista.objects.filter(empresa=empresa_actual).order_by('nombre_razon_social')
    beneficiarios = Beneficiario.objects.filter(empresa=empresa_actual).order_by('nombre_razon_social')
    
    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')

    # PAGINACIÓN
    paginator = Paginator(contratos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Re-obtener parámetros para el dict de filtros
    folio = request.GET.get('folio', '')
    beneficiario_id = request.GET.get('beneficiario_id', '')
    contratista_id = request.GET.get('contratista_id', '')

    return render(request, 'recursos_humanos/lista_contratos.html', {
        'page_obj': page_obj,
        'empleados': empleados,
        'contratistas': contratistas,
        'beneficiarios': beneficiarios,
        'sucursales': sucursales,
        'empresa': empresa_actual,
        'filtros': {
            'q': q,
            'folio': folio,
            'beneficiario_id': beneficiario_id,
            'contratista_id': contratista_id,
            'empleado_id': empleado_id,
            'estado': estado,
            'sucursal': sucursal_id
        }
    })

@login_required(login_url='/login/')
@require_hr_permission('contratos', 'ver', json_response=True)
def obtener_contrato_json(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        con = Contrato.objects.get(id=id, empresa=empresa_actual)
        data = {
            'id': con.id,
            'empleados_ids': list(con.empleados.values_list('id', flat=True)),
            'contratista': con.contratista_id,
            'beneficiario': con.beneficiario_id,
            'folio': con.folio,
            'tipo_contrato': con.tipo_contrato,
            'objeto_contrato': con.objeto_contrato,
            'monto_contrato': str(con.monto_contrato),
            'fecha_inicio': con.fecha_inicio.isoformat() if con.fecha_inicio else '',
            'fecha_fin': con.fecha_fin.isoformat() if con.fecha_fin else '',
            'vigencia_contrato': con.vigencia_contrato.isoformat() if con.vigencia_contrato else '',
            'num_estimado_trabajadores': con.num_estimado_trabajadores,
            'estado': con.estado,
            'notas': con.notas,
        }
        return JsonResponse({'success': True, 'data': data})
    except Contrato.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Contrato no encontrado.'})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('contratos', 'crear', json_response=True)
def crear_contrato_ajax(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'success': False, 'error': 'No se encontró la empresa.'}, status=403)

    try:
        data = request.POST
        sucursal_id = request.session.get('sucursal_id')
        with transaction.atomic():
            nuevo_contrato = Contrato(
                empresa=empresa_actual,
                sucursal_id=sucursal_id,
                contratista_id=data.get('contratista') or None,
                beneficiario_id=data.get('beneficiario') or None,
                folio=data.get('folio'),
                tipo_contrato=data.get('tipo_contrato'),
                objeto_contrato=data.get('objeto_contrato'),
                monto_contrato=Decimal(data.get('monto_contrato') or '0'),
                fecha_inicio=data.get('fecha_inicio') or None,
                fecha_fin=data.get('fecha_fin') or None,
                vigencia_contrato=data.get('vigencia_contrato') or None,
                num_estimado_trabajadores=int(data.get('num_estimado_trabajadores') or 0),
                estado=data.get('estado', 'vigente'),
                notas=data.get('notas', '')
            )
            nuevo_contrato.save()
            
            empleados_ids = request.POST.getlist('empleados[]') or request.POST.getlist('empleados')
            if empleados_ids:
                nuevo_contrato.empleados.set(empleados_ids)
                
            return JsonResponse({'success': True, 'message': 'Contrato registrado correctamente.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('contratos', 'editar', json_response=True)
def editar_contrato_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        con = Contrato.objects.get(id=id, empresa=empresa_actual)
        data = request.POST

        with transaction.atomic():
            con.contratista_id = data.get('contratista') or None
            con.beneficiario_id = data.get('beneficiario') or None
            con.folio = data.get('folio')
            con.tipo_contrato = data.get('tipo_contrato')
            con.objeto_contrato = data.get('objeto_contrato')
            con.monto_contrato = Decimal(data.get('monto_contrato') or '0')
            con.fecha_inicio = data.get('fecha_inicio') or None
            con.fecha_fin = data.get('fecha_fin') or None
            con.vigencia_contrato = data.get('vigencia_contrato') or None
            con.num_estimado_trabajadores = int(data.get('num_estimado_trabajadores') or 0)
            con.estado = data.get('estado')
            con.notas = data.get('notas', '')
            con.save()
            
            empleados_ids = request.POST.getlist('empleados[]') or request.POST.getlist('empleados')
            con.empleados.set(empleados_ids)
            
            return JsonResponse({'success': True, 'message': 'Contrato actualizado correctamente.'})
    except Contrato.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Contrato no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('contratos', 'eliminar', json_response=True)
def eliminar_contrato_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        con = Contrato.objects.get(id=id, empresa=empresa_actual)
        con.delete()
        return JsonResponse({'success': True, 'message': 'Contrato eliminado correctamente.'})
    except Contrato.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Contrato no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_hr_permission('contratistas', 'ver')
def lista_contratistas(request):
    empresa_actual = get_empresa_actual(request)
    from django.db.models import Count
    contratistas = Contratista.objects.filter(empresa=empresa_actual).annotate(total_colaboradores=Count('empleado')).order_by('nombre_razon_social')
    
    q = request.GET.get('q', '')
    f_razon = request.GET.get('razon_social', '')
    f_rfc = request.GET.get('rfc', '')
    f_rp = request.GET.get('reg_patronal', '')
    sucursal_id = request.GET.get('sucursal', '')

    if q:
        contratistas = contratistas.filter(
            Q(nombre_razon_social__icontains=q) |
            Q(rfc__icontains=q) |
            Q(representante_legal__icontains=q) |
            Q(correo__icontains=q) |
            Q(clave__icontains=q)
        )
    
    if f_razon:
        contratistas = contratistas.filter(nombre_razon_social__icontains=f_razon)
    if f_rfc:
        contratistas = contratistas.filter(rfc__icontains=f_rfc)
    if f_rp:
        contratistas = contratistas.filter(registro_patronal__icontains=f_rp)
    if sucursal_id:
        contratistas = contratistas.filter(sucursal_id=sucursal_id)

    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    
    # PAGINACIÓN
    paginator = Paginator(contratistas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'recursos_humanos/lista_contratistas.html', {
        'page_obj': page_obj,
        'sucursales': sucursales,
        'empresa': empresa_actual,
        'filtros': {
            'q': q,
            'razon_social': f_razon,
            'rfc': f_rfc,
            'reg_patronal': f_rp,
            'sucursal': sucursal_id
        }
    })

@login_required(login_url='/login/')
@require_hr_permission('beneficiarios', 'ver', json_response=True)
def obtener_contratista_json(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        cont = Contratista.objects.get(id=id, empresa=empresa_actual)
        data = {
            'id': cont.id, 'clave': cont.clave or '', 'rfc': cont.rfc, 'nombre_razon_social': cont.nombre_razon_social,
            'regimen': cont.regimen or '',
            'correo': cont.correo, 'telefono': cont.telefono, 'registro_patronal': cont.registro_patronal,
            'calle': cont.calle, 'num_ext': cont.num_ext, 'num_int': cont.num_int,
            'entre_calle': cont.entre_calle, 'y_calle': cont.y_calle, 'colonia': cont.colonia,
            'cp': cont.cp, 'municipio_alcaldia': cont.municipio_alcaldia,
            'entidad_federativa': cont.entidad_federativa, 'representante_legal': cont.representante_legal,
            'administrador_unico': cont.administrador_unico, 'num_escritura': cont.num_escritura,
            'nombre_notario_publico': cont.nombre_notario_publico, 'num_notario_publico': cont.num_notario_publico,
            'fecha_escritura_publica': cont.fecha_escritura_publica.isoformat() if cont.fecha_escritura_publica else '',
            'folio_mercantil': cont.folio_mercantil, 'numero_stps': cont.numero_stps,
        }
        return JsonResponse({'success': True, 'data': data})
    except Contratista.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Contratista no encontrado.'})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('contratistas', 'crear', json_response=True)
def crear_contratista_ajax(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual: return JsonResponse({'success': False, 'error': 'No se encontró la empresa.'}, status=403)
    try:
        data = request.POST
        sucursal_id = request.session.get('sucursal_id')
        nuevo = Contratista(
            empresa=empresa_actual, sucursal_id=sucursal_id, 
            clave=data.get('clave', ''),
            rfc=data.get('rfc', '').upper(),
            nombre_razon_social=data.get('nombre_razon_social'), 
            regimen=data.get('regimen'),
            correo=data.get('correo'),
            telefono=data.get('telefono'), registro_patronal=data.get('registro_patronal'),
            calle=data.get('calle'), num_ext=data.get('num_ext'), num_int=data.get('num_int'),
            entre_calle=data.get('entre_calle'), y_calle=data.get('y_calle'), colonia=data.get('colonia'),
            cp=data.get('cp'), municipio_alcaldia=data.get('municipio_alcaldia'),
            entidad_federativa=data.get('entidad_federativa'), representante_legal=data.get('representante_legal'),
            administrador_unico=data.get('administrador_unico'), num_escritura=data.get('num_escritura'),
            nombre_notario_publico=data.get('nombre_notario_publico'), num_notario_publico=data.get('num_notario_publico'),
            fecha_escritura_publica=data.get('fecha_escritura_publica') or None, folio_mercantil=data.get('folio_mercantil'),
            numero_stps=data.get('numero_stps'),
        )
        nuevo.save()
        return JsonResponse({'success': True, 'message': 'Contratista registrado correctamente.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('contratistas', 'editar', json_response=True)
def editar_contratista_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        cont = Contratista.objects.get(id=id, empresa=empresa_actual)
        data = request.POST
        cont.clave = data.get('clave', '')
        cont.rfc = data.get('rfc', '').upper()
        cont.nombre_razon_social = data.get('nombre_razon_social')
        cont.regimen = data.get('regimen')
        cont.correo = data.get('correo'); cont.telefono = data.get('telefono')
        cont.registro_patronal = data.get('registro_patronal'); cont.calle = data.get('calle')
        cont.num_ext = data.get('num_ext'); cont.num_int = data.get('num_int')
        cont.entre_calle = data.get('entre_calle'); cont.y_calle = data.get('y_calle')
        cont.colonia = data.get('colonia'); cont.cp = data.get('cp')
        cont.municipio_alcaldia = data.get('municipio_alcaldia'); cont.entidad_federativa = data.get('entidad_federativa')
        cont.representante_legal = data.get('representante_legal'); cont.administrador_unico = data.get('administrador_unico')
        cont.num_escritura = data.get('num_escritura'); cont.nombre_notario_publico = data.get('nombre_notario_publico')
        cont.num_notario_publico = data.get('num_notario_publico'); cont.fecha_escritura_publica = data.get('fecha_escritura_publica') or None
        cont.folio_mercantil = data.get('folio_mercantil'); cont.numero_stps = data.get('numero_stps')
        cont.save()
        return JsonResponse({'success': True, 'message': 'Contratista actualizado correctamente.'})
    except Contratista.DoesNotExist: return JsonResponse({'success': False, 'error': 'Contratista no encontrado.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('contratistas', 'eliminar', json_response=True)
def eliminar_contratista_ajax(request, id):
    """Eliminar un contratista."""
    empresa_actual = get_empresa_actual(request)
    try:
        cont = Contratista.objects.get(id=id, empresa=empresa_actual)
        cont.delete()
        return JsonResponse({'success': True, 'message': 'Contratista eliminado correctamente.'})
    except Contratista.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Contratista no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_hr_permission('beneficiarios', 'ver')
def lista_beneficiarios(request):
    empresa_actual = get_empresa_actual(request)
    from django.db.models import Count
    beneficiarios = Beneficiario.objects.filter(empresa=empresa_actual).annotate(total_empleados=Count('empleado')).order_by('nombre_razon_social')
    
    q = request.GET.get('q', '')
    f_razon = request.GET.get('razon_social', '')
    f_rfc = request.GET.get('rfc', '')
    f_clave = request.GET.get('clave', '')
    f_cont = request.GET.get('contacto', '')
    f_rp = request.GET.get('reg_patronal', '')
    sucursal_id = request.GET.get('sucursal', '')

    if q:
        beneficiarios = beneficiarios.filter(
            Q(nombre_razon_social__icontains=q) | 
            Q(rfc__icontains=q) | 
            Q(correo__icontains=q) |
            Q(clave__icontains=q)
        )
    
    if f_razon:
        beneficiarios = beneficiarios.filter(nombre_razon_social__icontains=f_razon)
    if f_rfc:
        beneficiarios = beneficiarios.filter(rfc__icontains=f_rfc)
    if f_clave:
        beneficiarios = beneficiarios.filter(clave__icontains=f_clave)
    if f_cont:
        beneficiarios = beneficiarios.filter(Q(correo__icontains=f_cont) | Q(telefono__icontains=f_cont))
    if f_rp:
        beneficiarios = beneficiarios.filter(registro_patronal__icontains=f_rp)
    if sucursal_id:
        beneficiarios = beneficiarios.filter(sucursal_id=sucursal_id)

    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    
    # PAGINACIÓN
    paginator = Paginator(beneficiarios, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'recursos_humanos/lista_beneficiarios.html', {
        'page_obj': page_obj, 
        'sucursales': sucursales, 
        'empresa': empresa_actual, 
        'filtros': {
            'q': q, 
            'razon_social': f_razon, 
            'rfc': f_rfc, 
            'clave': f_clave, 
            'contacto': f_cont, 
            'reg_patronal': f_rp, 
            'sucursal': sucursal_id
        }
    })

@login_required(login_url='/login/')
@require_hr_permission('beneficiarios', 'ver', json_response=True)
def obtener_beneficiario_json(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        ben = Beneficiario.objects.get(id=id, empresa=empresa_actual)
        data = {
            'id': ben.id, 'clave': ben.clave or '', 'rfc': ben.rfc, 'nombre_razon_social': ben.nombre_razon_social,
            'registro_patronal': ben.registro_patronal, 'calle': ben.calle, 'num_ext': ben.num_ext,
            'num_int': ben.num_int, 'entre_calle': ben.entre_calle, 'y_calle': ben.y_calle,
            'colonia': ben.colonia, 'cp': ben.cp, 'municipio_alcaldia': ben.municipio_alcaldia,
            'entidad_federativa': ben.entidad_federativa, 'correo': ben.correo, 'telefono': ben.telefono,
        }
        return JsonResponse({'success': True, 'data': data})
    except Beneficiario.DoesNotExist: return JsonResponse({'success': False, 'error': 'Beneficiario no encontrado.'})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('beneficiarios', 'crear', json_response=True)
def crear_beneficiario_ajax(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual: return JsonResponse({'success': False, 'error': 'No se encontró la empresa.'}, status=403)
    try:
        data = request.POST; sucursal_id = request.session.get('sucursal_id')
        nuevo = Beneficiario(
            empresa=empresa_actual, sucursal_id=sucursal_id, 
            clave=data.get('clave', ''),
            rfc=data.get('rfc', '').upper(),
            nombre_razon_social=data.get('nombre_razon_social'), registro_patronal=data.get('registro_patronal'),
            calle=data.get('calle'), num_ext=data.get('num_ext'), num_int=data.get('num_int'),
            entre_calle=data.get('entre_calle'), y_calle=data.get('y_calle'), colonia=data.get('colonia'),
            cp=data.get('cp'), municipio_alcaldia=data.get('municipio_alcaldia'),
            entidad_federativa=data.get('entidad_federativa'), correo=data.get('correo'), telefono=data.get('telefono'),
        )
        nuevo.save()
        return JsonResponse({'success': True, 'message': 'Beneficiario registrado correctamente.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('beneficiarios', 'editar', json_response=True)
def editar_beneficiario_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        ben = Beneficiario.objects.get(id=id, empresa=empresa_actual)
        data = request.POST
        ben.clave = data.get('clave', '')
        ben.rfc = data.get('rfc', '').upper(); ben.nombre_razon_social = data.get('nombre_razon_social')
        ben.registro_patronal = data.get('registro_patronal'); ben.calle = data.get('calle')
        ben.num_ext = data.get('num_ext'); ben.num_int = data.get('num_int')
        ben.entre_calle = data.get('entre_calle'); ben.y_calle = data.get('y_calle')
        ben.colonia = data.get('colonia'); ben.cp = data.get('cp')
        ben.municipio_alcaldia = data.get('municipio_alcaldia'); ben.entidad_federativa = data.get('entidad_federativa')
        ben.correo = data.get('correo'); ben.telefono = data.get('telefono')
        ben.save()
        return JsonResponse({'success': True, 'message': 'Beneficiario actualizado correctamente.'})
    except Beneficiario.DoesNotExist: return JsonResponse({'success': False, 'error': 'Beneficiario no encontrado.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('beneficiarios', 'eliminar', json_response=True)
def eliminar_beneficiario_ajax(request, id):
    """Eliminar un beneficiario."""
    empresa_actual = get_empresa_actual(request)
    try:
        ben = Beneficiario.objects.get(id=id, empresa=empresa_actual)
        ben.delete()
        return JsonResponse({'success': True, 'message': 'Beneficiario eliminado correctamente.'})
    except Beneficiario.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Beneficiario no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_hr_permission('sua', 'ver')
def lista_sua(request):
    empresa_actual = get_empresa_actual(request)
    importaciones = ImportacionSUA.objects.filter(empresa=empresa_actual).select_related('sucursal').order_by('-fecha_importacion')
    
    q_reg_pat = request.GET.get('reg_patronal', '')
    q_razon = request.GET.get('razon_social', '')
    q_periodo = request.GET.get('periodo', '')
    sucursal_id = request.GET.get('sucursal', '')

    if q_reg_pat:
        importaciones = importaciones.filter(registro_patronal__icontains=q_reg_pat)
    if q_razon:
        importaciones = importaciones.filter(nombre_razon_social__icontains=q_razon)
    if q_periodo:
        importaciones = importaciones.filter(periodo__icontains=q_periodo)
    if sucursal_id:
        importaciones = importaciones.filter(sucursal_id=sucursal_id)

    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    
    # PAGINACIÓN
    paginator = Paginator(importaciones, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'recursos_humanos/lista_sua.html', {
        'page_obj': page_obj, 
        'sucursales': sucursales, 
        'empresa': empresa_actual, 
        'filtros': {
            'reg_patronal': q_reg_pat, 
            'razon_social': q_razon, 
            'periodo': q_periodo, 
            'sucursal': sucursal_id
        }
    })

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('sua', 'importar', json_response=True)
def importar_sua_ajax(request):
    empresa_actual = get_empresa_actual(request)
    pdf_file = request.FILES.get('archivo_sua')
    tipo_importacion = request.POST.get('tipo', 'bimestral')
    if not pdf_file: return JsonResponse({'success': False, 'error': 'No se proporcionó ningún archivo.'})
    try:
        if not pdfplumber: return JsonResponse({'success': False, 'error': 'Librería pdfplumber no instalada.'})
        
        text_lines = []
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                text = page.extract_text(layout=True)
                if text:
                    text_lines.extend(text.split('\n'))
        
        full_text = "\n".join(text_lines)
        
        reg_pat_val, rfc_emp_val, area_val = "", "", ""
        nom_razon_val, deleg_val = "", ""
        actividad_val, subdeleg_val = "", ""
        domicilio_val, mun_alc_val = "", ""
        cp_val, entidad_val, prima_val = "", "", ""
        periodo_val = "Desconocido"

        for line in text_lines:
            if "Registro Patronal:" in line:
                m_rp = re.search(r'Registro Patronal:\s*([\w-]+)', line, re.I)
                m_rfc = re.search(r'RFC:\s*([\w\d-]+)', line, re.I)
                m_area = re.search(r'Area Geográfica:\s*(.*?)(?=\s{2,}|Delegación|$)', line, re.I)
                if m_rp: reg_pat_val = m_rp.group(1).strip()
                if m_rfc: rfc_emp_val = m_rfc.group(1).strip()
                if m_area: area_val = m_area.group(1).strip()
            
            if "Nombre o Razón Social:" in line:
                m_nom = re.search(r'Nombre o Razón Social:\s*(.*?)(?=\s{2,}|Delegación|Convenio|Aportación|V \d|$)', line, re.I)
                m_del = re.search(r'Delegación IMSS:\s*(.*?)(?=\s{2,}|Fecha|Convenio|Aportación|V \d|$)', line, re.I)
                if m_nom: nom_razon_val = m_nom.group(1).strip()
                if m_del: deleg_val = m_del.group(1).strip()

            if "Actividad:" in line:
                m_act = re.search(r'Actividad:\s*(.*?)(?=\s{2,}|Subdelegación|Convenio|Aportación|V \d|$)', line, re.I)
                m_sub = re.search(r'SubDelegación IMSS:\s*(.*?)(?=\s{2,}|Area|Convenio|Aportación|V \d|$)', line, re.I)
                if m_act: actividad_val = m_act.group(1).strip()
                if m_sub: subdeleg_val = m_sub.group(1).strip()

            if "Domicilio:" in line:
                m_dom = re.search(r'Domicilio:\s*(.*?)(?=\s{2,}|Pob|Convenio|Aportación|V \d|$)', line, re.I)
                m_mun = re.search(r'Pob\., Mun\. / Alcaldía:\s*(.*?)(?=\s{2,}|Prima|Convenio|Aportación|V \d|$)', line, re.I)
                if m_dom: domicilio_val = m_dom.group(1).strip()
                if m_mun: mun_alc_val = m_mun.group(1).strip()

            if "Código Postal:" in line:
                m_cp = re.search(r'Código Postal:\s*(\d+)', line, re.I)
                m_ent = re.search(r'Entidad:\s*(.*?)(?=\s{2,}|Prima|Convenio|Aportación|V \d|$)', line, re.I)
                m_pri = re.search(r'Prima de R\.T\.\s*([\d\.,%]+)', line, re.I)
                if m_cp: cp_val = m_cp.group(1).strip()
                if m_ent: entidad_val = m_ent.group(1).strip()
                if m_pri: prima_val = m_pri.group(1).strip()
            
            if "Proceso:" in line:
                m_per = re.search(r'(?:Período|Bimestre)\s+de\s+Proceso:\s*([\w\d-]+)', line, re.I)
                if m_per: periodo_val = m_per.group(1).strip()

        # Validación de Duplicados
        periodo_final = limpiar_basura_header(periodo_val).strip()
        razon_final = nom_razon_val.strip()
        reg_pat_final = reg_pat_val.strip()
        
        if ImportacionSUA.objects.filter(
            empresa=empresa_actual,
            periodo=periodo_final,
            nombre_razon_social=razon_final,
            registro_patronal=reg_pat_final,
            tipo=tipo_importacion
        ).exists():
            return JsonResponse({
                'success': False, 
                'error': f'Error: Esta cédula ya fue registrada (Periodo: {periodo_final}, Empresa: {razon_final}, Tipo: {tipo_importacion.upper()}).'
            })

        total_reporte = 0
        m_total_rep = re.search(r'Total\s+de\s+cotizaciones:\s*(\d+)', full_text, re.I)
        if m_total_rep:
            total_reporte = int(m_total_rep.group(1))

        sucursal_id = request.session.get('sucursal_id')
        created_count = 0
        nss_encontrados = set()

        with transaction.atomic():
            importacion = ImportacionSUA.objects.create(
                empresa=empresa_actual, sucursal_id=sucursal_id,
                registro_patronal=reg_pat_val,
                rfc_empresa=rfc_emp_val,
                nombre_razon_social=nom_razon_val,
                actividad=limpiar_basura_header(actividad_val),
                domicilio=limpiar_basura_header(domicilio_val),
                cp=limpiar_basura_header(cp_val),
                entidad=limpiar_basura_header(entidad_val),
                area_geografica=limpiar_basura_header(area_val),
                delegacion_imss=limpiar_basura_header(deleg_val),
                subdelegacion_imss=limpiar_basura_header(subdeleg_val),
                municipio_alcaldia=limpiar_basura_header(mun_alc_val),
                prima_rt=prima_val,
                periodo=periodo_final,
                tipo=tipo_importacion
            )

            current_worker_info = None
            stop_workers = False
            
            KEYWORDS_EXCLUDE = [
                'REGISTRO PATRONAL', 'RFC', 'NOMBRE O RAZÓN SOCIAL', 'DELEGACIÓN', 
                'ACTIVIDAD', 'DOMICILIO', 'BIMESTRE', 'PERIODO', 'POB.', 
                'CÓDIGO POSTAL', 'ENTIDAD', 'PRIMA', 'PAGINA', 'HOJA'
            ]

            for line in text_lines:
                l_clean = line.strip()
                if not l_clean: continue

                # Detector de Paro
                if "TOTAL DE COTIZACIONES" in l_clean.upper() or \
                   "TOTALES" in l_clean.upper() or \
                   re.search(r'([_-]\s?){15,}', l_clean):
                    current_worker_info = None
                    stop_workers = True 
                    continue
                if stop_workers: continue

                # Identificar Asegurado (L1: NSS -> NOMBRE -> RFC -> UBICACION)
                # El NSS puede venir con guiones (12-34...) o directo (1234...)
                nss_match = re.search(r'(\d{2}-?\d{2}-?\d{2}-?\d{4}-?\d)', l_clean)
                if nss_match:
                    nss_val = nss_match.group(1)
                    current_worker_info = None 
                    remainder = l_clean.split(nss_val)[-1].strip()
                    
                    # Buscamos el RFC/CURP como ancla pivotante (bloque alfanumérico de 10-18 chars)
                    # Usamos una búsqueda más amplia para no omitir variantes
                    m_rfc = re.search(r'([A-Z]{3,4}[0-9]{6}[A-Z0-9]{0,9})', remainder)
                    
                    if m_rfc:
                        rfc = m_rfc.group(1)
                        nombre = remainder[:m_rfc.start()].strip()
                        post_rfc = remainder[m_rfc.end():].strip()
                        
                        # Ubicación: lo que queda antes del primer bloque numérico (Días SDI)
                        # Los datos numéricos empiezan con un bloque de 1-2 dígitos (Días) + espacio + SDI
                        m_data_start = re.search(r'(\d{1,2})\s+([\d\.,]+)', post_rfc)
                        if m_data_start:
                            ubic = post_rfc[:m_data_start.start()].strip() or "-"
                            l_clean = post_rfc[m_data_start.start():].strip()
                        else:
                            ubic = post_rfc or "-"
                            l_clean = ""
                        
                        if nombre and rfc:
                            current_worker_info = {
                                'nss': nss_val, 'nombre': nombre, 'rfc': rfc, 'clave_u': ubic
                            }
                            nss_encontrados.add(nss_val)
                    else:
                        # Fallback agresivo para líneas sin RFC claro o con espacios simples
                        parts = [p.strip() for p in re.split(r'\s+', remainder) if p.strip()]
                        if len(parts) >= 2:
                            # Asumimos Nombre y RFC por posición si no hay ancla
                            # Pero intentamos validar que el segundo bloque parezca un ID
                            nombre = parts[0]
                            rfc = parts[1]
                            ubic = " ".join(parts[2:]) if len(parts) > 2 else "-"
                            
                            current_worker_info = {
                                'nss': nss_val, 'nombre': nombre, 'rfc': rfc, 'clave_u': ubic
                            }
                            nss_encontrados.add(nss_val)
                            l_clean = ""
                    
                    # NO usar continue. En bimestrales, los datos suelen estar en la misma línea.
                    if not l_clean: continue

                if current_worker_info:
                    m_header = re.match(r'^([^0-9\s,]{2,})?\s*(\d{2}/\d{2}/\d{4})?\s*(.*)', l_clean, re.I)
                    clave_mov = m_header.group(1).strip().rstrip(',') if m_header and m_header.group(1) else '-'
                    fecha_mov = m_header.group(2) or '' if m_header else ''
                    resto_linea = m_header.group(3).strip() if m_header else l_clean
                    
                    # 1. Extraer tokens base (Priorizando fechas y códigos antes que números simples)
                    raw_tokens = re.findall(r'\d{2}/\d{2}/\d{4}|(?:FD|\$|%)\s*[\d\.,]+%?|[A-Z]{3}|[\d\.,]+%?|FD|%|\$', resto_linea)
                    raw_tokens = [t.strip() for t in raw_tokens if t.strip()]
                    
                    # 2. Función interna para unir prefijos huérfanos con su número
                    def merge_tokens_sua(items):
                        res = []
                        skip = False
                        for i in range(len(items)):
                            if skip:
                                skip = False
                                continue
                            token = items[i]
                            if token in ['FD', '$', '%'] and (i + 1) < len(items):
                                next_t = items[i+1]
                                if re.match(r'^[\d\.,]+', next_t):
                                    res.append(f"{token} {next_t}")
                                    skip = True
                                    continue
                            res.append(token)
                        return res

                    tokens = merge_tokens_sua(raw_tokens)
                    tokens_clean = [t.replace(',', '') for t in tokens]
                    
                    def clean_dec(val):
                        if not val or val == '-': return Decimal('0')
                        c = re.sub(r'[^\d.]', '', str(val))
                        return Decimal(c) if c else Decimal('0')

                    try:
                        num_idx = 0
                        if tokens_clean[num_idx] == 'FD': num_idx += 1
                        
                        dias_val = int(float(clean_dec(tokens_clean[num_idx])))
                        tiene_sdi = len(tokens_clean) > (num_idx + 1) and re.match(r'^[^\d]*[\d\.]+[^\d]*$', tokens_clean[num_idx+1]) and float(clean_dec(tokens_clean[num_idx+1])) > 0
                        es_movimiento_datos = (dias_val <= 99 and tiene_sdi)
                        if es_movimiento_datos:
                            # Re-ajustar tokens para que el índice 0 sea 'dias'
                            tokens = tokens[num_idx:]
                            tokens_clean = tokens_clean[num_idx:]
                    except: es_movimiento_datos = False

                    if (clave_mov.lower() in ['baja', 'reingreso', 'modificación', 'alta'] and fecha_mov != "") or es_movimiento_datos:
                        trabajador_data = {
                            'importacion': importacion, 'nss': current_worker_info['nss'], 'nombre': current_worker_info['nombre'],
                            'rfc_curp': current_worker_info['rfc'], 'clave_ubicacion': current_worker_info['clave_u'],
                            'clave_mov': clave_mov, 'fecha_mov': fecha_mov, 'dias': 0, 'sdi': 0, 'licencias': 0, 'incapacidades': 0, 'ausentismos': 0, 'total_general': 0
                        }
                        if es_movimiento_datos:
                            try:
                                # Mapeo base RCV (Siempre presente)
                                trabajador_data.update({
                                    'dias': int(float(clean_dec(tokens_clean[0]))), 
                                    'sdi': clean_dec(tokens_clean[1]), 
                                    'licencias': int(float(clean_dec(tokens_clean[2]))), 
                                    'incapacidades': int(float(clean_dec(tokens_clean[3]))), 
                                    'ausentismos': int(float(clean_dec(tokens_clean[4])))
                                })
                                
                                if tipo_importacion == 'mensual':
                                    if len(tokens_clean) >= 19:
                                        trabajador_data.update({
                                            'cuota_fija': clean_dec(tokens_clean[5]), 
                                            'excedente_patronal': clean_dec(tokens_clean[6]), 
                                            'excedente_obrera': clean_dec(tokens_clean[7]), 
                                            'prestaciones_dinero_patronal': clean_dec(tokens_clean[8]), 
                                            'prestaciones_dinero_obrera': clean_dec(tokens_clean[9]), 
                                            'gastos_medicos_patronal': clean_dec(tokens_clean[10]), 
                                            'gastos_medicos_obrera': clean_dec(tokens_clean[11]), 
                                            'riesgo_trabajo_cuota': clean_dec(tokens_clean[12]), 
                                            'invalidez_vida_patronal': clean_dec(tokens_clean[13]), 
                                            'invalidez_vida_obrera': clean_dec(tokens_clean[14]), 
                                            'guarderias_ps': clean_dec(tokens_clean[15]), 
                                            'imss_patronal': clean_dec(tokens_clean[16]), 
                                            'imss_obrera': clean_dec(tokens_clean[17]), 
                                            'imss_subtotal': clean_dec(tokens_clean[18]), 
                                            'total_general': clean_dec(tokens_clean[18])
                                        })
                                else:
                                    # --- LÓGICA DE PIVOT UNIFICADO PARA BIMESTRAL ---
                                    
                                    # 1. Encontrar el pivot de Infonavit (Factor: FD, $ o %)
                                    pivot_idx = -1
                                    for idx, t in enumerate(tokens):
                                        if any(pref in t for pref in ['FD', '$', '%']) and re.search(r'\d', t):
                                            pivot_idx = idx
                                            break
                                    
                                    if pivot_idx != -1:
                                        # Mapeo relativo al pivot (mucho más robusto)
                                        # Estructura: [..., Retiro, PatRCV, ObrRCV, SumaRCV, ApPatInf, FACTOR, Amortiz, SumaInf, ...]
                                        try:
                                            # RCV (hacia atrás desde el pivot)
                                            trabajador_data.update({
                                                'retiro': clean_dec(tokens_clean[pivot_idx - 5]),
                                                'patronal': clean_dec(tokens_clean[pivot_idx - 4]),
                                                'obrera': clean_dec(tokens_clean[pivot_idx - 3]),
                                                'subtotal': clean_dec(tokens_clean[pivot_idx - 2])
                                            })
                                            
                                            # Infonavit
                                            ap_pat = clean_dec(tokens_clean[pivot_idx - 1])
                                            amort = clean_dec(tokens_clean[pivot_idx + 1])
                                            # Calculamos la suma manualmente para asegurar exactitud matemática
                                            suma_inf = ap_pat + amort
                                            
                                            trabajador_data.update({
                                                'aportacion_patronal': ap_pat,
                                                'tipo_valor_infonavit': tokens[pivot_idx],
                                                'amortizacion': amort,
                                                'suma_infonavit': suma_inf
                                            })

                                            # Columnas adicionales (Crédito, Movimientos)
                                            if len(tokens) > pivot_idx + 2:
                                                for extra_idx in range(pivot_idx + 3, len(tokens)):
                                                    t_extra = tokens[extra_idx]
                                                    if re.match(r'^\d{8,11}$', t_extra): # Número de crédito
                                                        trabajador_data['cred_vivienda'] = t_extra
                                                    elif re.match(r'^[A-Z]{3}$', t_extra): # Código MVD, FSD...
                                                        trabajador_data['tipo_mov_credito'] = t_extra
                                                    elif re.match(r'^\d{2}/\d{2}/\d{4}$', t_extra): # Fecha mov.
                                                        trabajador_data['fecha_mov_credito'] = t_extra
                                        except: pass
                                    else:
                                        # Fallback sin crédito (Aportación Patronal 5% suele ser token 9)
                                        try:
                                            if len(tokens_clean) >= 9:
                                                trabajador_data.update({
                                                    'retiro': clean_dec(tokens_clean[5]), 
                                                    'patronal': clean_dec(tokens_clean[6]), 
                                                    'obrera': clean_dec(tokens_clean[7]), 
                                                    'subtotal': clean_dec(tokens_clean[8])
                                                })
                                            if len(tokens_clean) >= 10:
                                                ap_pat = clean_dec(tokens_clean[9])
                                                trabajador_data.update({
                                                    'aportacion_patronal': ap_pat,
                                                    'suma_infonavit': ap_pat, # Sin amortización, la suma es solo la aportación
                                                    'amortizacion': Decimal('0'),
                                                    'tipo_valor_infonavit': '-'
                                                })
                                        except: pass
                                    
                                    # Total final de la fila (Suma RCV + Suma INF)
                                    s_rcv = clean_dec(trabajador_data.get('subtotal', 0))
                                    s_inf = clean_dec(trabajador_data.get('suma_infonavit', 0))
                                    trabajador_data['total_general'] = s_rcv + s_inf
                            except: pass
                        TrabajadorSUA.objects.create(**trabajador_data)
                        created_count += 1
                    else:
                        if not nss_match: current_worker_info = None

            if created_count == 0: raise Exception("No se detectaron trabajadores válidos.")
            unique_count = len(nss_encontrados)
            msg_validacion = f" Advertencia: Se detectaron {unique_count} trabajadores únicos pero el reporte indica un total de {total_reporte}." if total_reporte > 0 and unique_count != total_reporte else ""

        return JsonResponse({'success': True, 'message': f'Importación exitosa: {created_count} registros procesados.{msg_validacion}'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_hr_permission('sua', 'ver', json_response=True)
def obtener_registro_sua_json(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        imp = ImportacionSUA.objects.get(id=id, empresa=empresa_actual)
        trabajadores = []
        totales = {
            'dias': 0, 'total_general': 0, 'retiro': 0, 'patronal_rcv': 0, 'obrera_rcv': 0, 'total_rcv': 0, 'ap_pat_inf': 0, 'tipo_val_inf': 0, 'amortiz': 0, 'total_inf': 0,
            'cuota_fija': 0, 'exc_pat': 0, 'exc_obr': 0, 'pd_pat': 0, 'pd_obr': 0, 'gm_pat': 0, 'gm_obr': 0, 'rt': 0, 'iv_pat': 0, 'iv_obr': 0, 'gps': 0, 'imss_pat': 0, 'imss_obr': 0, 'imss_sub': 0
        }
        for t in imp.trabajadores.all().order_by('id'):
            t_dict = {'nss': t.nss, 'nombre': t.nombre, 'rfc': t.rfc_curp, 'clave_u': t.clave_ubicacion, 'clave_mov': t.clave_mov, 'fecha_mov': t.fecha_mov, 'dias': t.dias, 'sdi': str(t.sdi), 'lic': t.licencias, 'inc': t.incapacidades, 'aus': t.ausentismos, 'total_general': str(t.total_general)}
            if imp.tipo == 'mensual':
                t_dict.update({'cf': str(t.cuota_fija), 'exc_pat': str(t.excedente_patronal), 'exc_obr': str(t.excedente_obrera), 'pd_pat': str(t.prestaciones_dinero_patronal), 'pd_obr': str(t.prestaciones_dinero_obrera), 'gm_pat': str(t.gastos_medicos_patronal), 'gm_obr': str(t.gastos_medicos_obrera), 'rt': str(t.riesgo_trabajo_cuota), 'iv_pat': str(t.invalidez_vida_patronal), 'iv_obr': str(t.invalidez_vida_obrera), 'gps': str(t.guarderias_ps), 'imss_pat': str(t.imss_patronal), 'imss_obr': str(t.imss_obrera), 'imss_sub': str(t.imss_subtotal)})
                totales['cuota_fija'] += float(t.cuota_fija); totales['exc_pat'] += float(t.excedente_patronal); totales['exc_obr'] += float(t.excedente_obrera); totales['pd_pat'] += float(t.prestaciones_dinero_patronal); totales['pd_obr'] += float(t.prestaciones_dinero_obrera); totales['gm_pat'] += float(t.gastos_medicos_patronal); totales['gm_obr'] += float(t.gastos_medicos_obrera); totales['rt'] += float(t.riesgo_trabajo_cuota); totales['iv_pat'] += float(t.invalidez_vida_patronal); totales['iv_obr'] += float(t.invalidez_vida_obrera); totales['gps'] += float(t.guarderias_ps); totales['imss_pat'] += float(t.imss_patronal); totales['imss_obr'] += float(t.imss_obrera); totales['imss_sub'] += float(t.imss_subtotal)
            else:
                t_dict.update({'retiro': str(t.retiro), 'patronal_rcv': str(t.patronal), 'obrera_rcv': str(t.obrera), 'total_rcv': str(t.subtotal), 'ap_pat_inf': str(t.aportacion_patronal), 'tipo_val_inf': t.tipo_valor_infonavit or '-', 'amortiz': str(t.amortizacion), 'total_inf': str(t.suma_infonavit), 'cred_viv': t.cred_vivienda or '', 'tipo_mov_cred': t.tipo_mov_credito or '', 'fecha_mov_cred': t.fecha_mov_credito or ''})
                totales['retiro'] += float(t.retiro); totales['patronal_rcv'] += float(t.patronal); totales['obrera_rcv'] += float(t.obrera); totales['total_rcv'] += float(t.subtotal); totales['ap_pat_inf'] += float(t.aportacion_patronal); totales['amortiz'] += float(t.amortizacion); totales['total_inf'] += float(t.suma_infonavit)
                if t.tipo_valor_infonavit and t.tipo_valor_infonavit != '-':
                    v_l = re.sub(r'[^\d.]', '', t.tipo_valor_infonavit)
                    if v_l: totales['tipo_val_inf'] += float(v_l)
            totales['dias'] += t.dias; totales['total_general'] += float(t.total_general); trabajadores.append(t_dict)
        for k in totales:
            if k == 'dias': totales[k] = int(totales[k])
            else: totales[k] = "{:,.2f}".format(totales[k])
        data = {'empresa': {'razon_social': imp.nombre_razon_social, 'rfc': imp.rfc_empresa, 'reg_patronal': imp.registro_patronal, 'actividad': limpiar_basura_header(imp.actividad), 'domicilio': limpiar_basura_header(imp.domicilio), 'cp': limpiar_basura_header(imp.cp), 'entidad': limpiar_basura_header(imp.entidad), 'periodo': limpiar_basura_header(imp.periodo), 'tipo': imp.get_tipo_display(), 'tipo_raw': imp.tipo}, 'trabajadores': trabajadores, 'totales': totales}
        return JsonResponse({'success': True, 'data': data})
    except ImportacionSUA.DoesNotExist: return JsonResponse({'success': False, 'error': 'No se encontró la importación.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('sua', 'eliminar', json_response=True)
def eliminar_sua_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        imp = ImportacionSUA.objects.get(id=id, empresa=empresa_actual)
        imp.delete(); return JsonResponse({'success': True, 'message': 'Registro eliminado correctamente.'})
    except ImportacionSUA.DoesNotExist: return JsonResponse({'success': False, 'error': 'No se encontró la importación.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_hr_permission('sua', 'ver')
def exportar_sua_excel(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        imp = ImportacionSUA.objects.get(id=id, empresa=empresa_actual)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="SUA_{imp.periodo}_{imp.registro_patronal}.csv"'
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.writer(response)
        writer.writerow(['REPORTE DE INTEGRACIÓN SUA'])
        writer.writerow(['Empresa', imp.nombre_razon_social]); writer.writerow(['Registro Patronal', imp.registro_patronal]); writer.writerow(['Periodo', imp.periodo]); writer.writerow(['Tipo', imp.get_tipo_display()]); writer.writerow([])
        if imp.tipo == 'mensual':
            headers = ['NSS', 'Nombre', 'RFC/CURP', 'Ubicación', 'Movimiento', 'Fecha Mov.', 'Días', 'SDI', 'Lic.', 'Inc.', 'Aus.', 'C.F.', 'Exc. Pat.', 'Exc. Obr.', 'P.D. Pat.', 'P.D. Obr.', 'G.M.P. Pat.', 'G.M.P. Obr.', 'R.T.', 'I.V. Pat.', 'I.V. Obr.', 'G.P.S.', 'Patronal', 'Obrera', 'Subtotal']
        else:
            headers = ['NSS', 'Nombre', 'RFC/CURP', 'Ubicación', 'Movimiento', 'Fecha Mov.', 'Días', 'SDI', 'Lic.', 'Inc.', 'Aus.', 'Retiro', 'Patronal RCV', 'Obrera RCV', 'Suma RCV', 'Ap. Pat. Infonavit', '%/$ /FD', 'Amortización', 'Suma Infonavit', 'Total General', 'Créd. Vivienda', 'Tipo Mov. Crédito', 'Fecha Mov. Crédito']
        writer.writerow(headers)
        for t in imp.trabajadores.all().order_by('id'):
            if imp.tipo == 'mensual':
                writer.writerow([t.nss, t.nombre, t.rfc_curp, t.clave_ubicacion, t.clave_mov, t.fecha_mov, t.dias, t.sdi, t.licencias, t.incapacidades, t.ausentismos, t.cuota_fija, t.excedente_patronal, t.excedente_obrera, t.prestaciones_dinero_patronal, t.prestaciones_dinero_obrera, t.gastos_medicos_patronal, t.gastos_medicos_obrera, t.riesgo_trabajo_cuota, t.invalidez_vida_patronal, t.invalidez_vida_obrera, t.guarderias_ps, t.imss_patronal, t.imss_obrera, t.imss_subtotal])
            else:
                writer.writerow([t.nss, t.nombre, t.rfc_curp, t.clave_ubicacion, t.clave_mov, t.fecha_mov, t.dias, t.sdi, t.licencias, t.incapacidades, t.ausentismos, t.retiro, t.patronal, t.obrera, t.subtotal, t.aportacion_patronal, t.tipo_valor_infonavit, t.amortizacion, t.suma_infonavit, t.total_general, t.cred_vivienda, t.tipo_mov_credito, t.fecha_mov_credito])
        return response
    except ImportacionSUA.DoesNotExist: return HttpResponse("No se encontró la importación", status=404)

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('empleados', 'crear', json_response=True)
def alta_empleados_sua_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        importacion = ImportacionSUA.objects.get(id=id, empresa=empresa_actual)
        trabajadores = importacion.trabajadores.all()
        sucursal_id = request.session.get('sucursal_id')
        creados = 0; actualizados = 0; vinculados_a_contrato = 0
        beneficiarios_map = {b.clave.strip().upper(): b for b in Beneficiario.objects.filter(empresa=empresa_actual) if b.clave}
        
        # Limpieza de datos del reporte para búsqueda unificada de contratista
        rfc_reporte = re.sub(r'[^A-Z0-9]', '', (importacion.rfc_empresa or '').upper()).strip()[:13]
        rp_reporte = re.sub(r'[^A-Z0-9]', '', (importacion.registro_patronal or '').upper()).strip()
        nombre_reporte = (importacion.nombre_razon_social or '').strip()

        filtros_or = Q()
        if rfc_reporte and rfc_reporte != "POR_DEFINIR":
            filtros_or |= Q(rfc__iexact=rfc_reporte)
        if rp_reporte:
            filtros_or |= Q(registro_patronal__iexact=rp_reporte)
        if nombre_reporte:
            filtros_or |= Q(nombre_razon_social__icontains=nombre_reporte)
            
        contratista_obj = Contratista.objects.filter(Q(empresa=empresa_actual) & filtros_or).first()

        status_cont = "Existente"
        if not contratista_obj:
            contratista_obj = Contratista.objects.create(
                empresa=empresa_actual, sucursal_id=sucursal_id,
                registro_patronal=rp_reporte, nombre_razon_social=nombre_reporte,
                rfc=rfc_reporte or "POR_DEFINIR", calle=importacion.domicilio,
                cp=importacion.cp, entidad_federativa=importacion.entidad,
                correo=f"contacto@{rp_reporte or 'empresa'}.com"
            )
            status_cont = "NUEVO REGISTRO"
        
        # Pre-cache de contratos vigentes para este contratista
        contratos_activos = list(Contrato.objects.filter(
            empresa=empresa_actual, 
            contratista=contratista_obj,
            estado='vigente'
        ))

        info_contratista = f"{contratista_obj.nombre_razon_social} (ID: {contratista_obj.id}, {status_cont})"

        with transaction.atomic():
            for t in trabajadores:
                nss_clean = re.sub(r'[^0-9]', '', t.nss).strip()[:11]
                curp_clean = re.sub(r'[^A-Z0-9]', '', (t.rfc_curp or '').upper()).strip()[:18]
                empleado = Empleado.objects.filter(Q(nss=nss_clean) | Q(curp=curp_clean), empresa=empresa_actual).first()
                beneficiario_obj = beneficiarios_map.get((t.clave_ubicacion or "").strip().upper())
                
                nombre_partes = t.nombre.strip().split(' ')
                paterno = ""; materno = ""; nombres = t.nombre
                if len(nombre_partes) >= 3:
                    paterno = nombre_partes[0]; materno = nombre_partes[1]; nombres = " ".join(nombre_partes[2:])
                elif len(nombre_partes) == 2:
                    paterno = nombre_partes[0]; nombres = nombre_partes[1]

                if not empleado:
                    fecha_imp = importacion.fecha_importacion.strftime('%d/%m/%Y')
                    audit_nota = f"Importado el día {fecha_imp} de la cédula {importacion.periodo} del contratista {importacion.nombre_razon_social}"
                    empleado = Empleado(
                        empresa=empresa_actual, sucursal_id=sucursal_id, nss=nss_clean, curp=curp_clean,
                        nombre=nombres, apellido_paterno=paterno, apellido_materno=materno,
                        sdi=t.sdi, contratista=contratista_obj, beneficiario=beneficiario_obj,
                        puesto="", departamento="General", clave_ubicacion=t.clave_ubicacion,
                        notas=audit_nota, estado='activo'
                    )
                    empleado.save(); creados += 1
                else:
                    empleado.sdi = t.sdi
                    if beneficiario_obj: empleado.beneficiario = beneficiario_obj
                    if contratista_obj: empleado.contratista = contratista_obj
                    if t.clave_ubicacion: empleado.clave_ubicacion = t.clave_ubicacion
                    empleado.save(); actualizados += 1
                
                # --- VINCULACIÓN AUTOMÁTICA CON CONTRATO ---
                if beneficiario_obj:
                    # Intentamos encontrar el contrato que una a este beneficiario con el contratista del reporte
                    # Buscamos por Beneficiario (que es el ancla fuerte en el SUA) 
                    # y validamos que el contratista del contrato coincida con el del reporte (vía RFC o Registro)
                    contrato_final = Contrato.objects.filter(
                        empresa=empresa_actual,
                        beneficiario=beneficiario_obj,
                        estado='vigente'
                    ).filter(
                        Q(contratista__rfc__iexact=rfc_reporte) | 
                        Q(contratista__registro_patronal__iexact=rp_reporte) |
                        Q(contratista=contratista_obj)
                    ).first()

                    if contrato_final:
                        contrato_final.empleados.add(empleado)
                        vinculados_a_contrato += 1

        return JsonResponse({
            'success': True, 
            'message': f'Proceso completado: {creados} nuevos, {actualizados} actualizados. {vinculados_a_contrato} vinculaciones a contratos realizadas. Contratista identificado: {info_contratista}'
        })
    except ImportacionSUA.DoesNotExist: return JsonResponse({'success': False, 'error': 'No se encontró el registro SUA.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
def exportar_sisub_contratos(request, id):
    import re
    empresa_actual = get_empresa_actual(request)
    try:
        contratista = get_object_or_404(Contratista, id=id, empresa=empresa_actual)
        
        cuatrimestre = request.GET.get('cuatrimestre', '1')
        anio = request.GET.get('anio', '')
        
        contratos = Contrato.objects.filter(contratista=contratista, empresa=empresa_actual).select_related('beneficiario')

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SISUB Contratos"

        fill_main = PatternFill(start_color="00b8b9", end_color="00b8b9", fill_type="solid")
        fill_sec = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        fill_head = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(
        left=Side(style='thin', color="B2B2B2"), 
        right=Side(style='thin', color="B2B2B2"), 
        top=Side(style='thin', color="B2B2B2"), 
        bottom=Side(style='thin', color="B2B2B2")
    )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # NIVEL 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=25)
        c1 = ws.cell(row=1, column=1, value="b-Contratos de servicio (cliente)")
        c1.alignment = center_align
        c1.font = Font(bold=True, color="FFFFFF", size=12)
        c1.fill = fill_main
        c1.border = border

        # NIVEL 2
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
        ws.cell(row=2, column=1, value="periodo")
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=11)
        ws.cell(row=2, column=3, value="a-Datos generales del contrato")
        ws.merge_cells(start_row=2, start_column=12, end_row=2, end_column=14)
        ws.cell(row=2, column=12, value="b-Identificacion del beneficiario")
        ws.merge_cells(start_row=2, start_column=15, end_row=2, end_column=25)
        ws.cell(row=2, column=15, value="c-Domicilio fiscal del beneficiario")
        
        for c in range(1, 26):
            cell = ws.cell(row=2, column=c)
            cell.fill = fill_sec
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = center_align

        # NIVEL 3
        headers = [
            'Cuatrimestre', 'Año', 'RFC Sujeto', 'Folio', 'Tipo', 'Objeto', 'Monto', 'Vigencia', 'Inicio', 'Termino', 
            'Trabajadores', 'RFC Ben', 'Nombre Ben', 'RegPat Ben', 'Calle', 'Ext', 'Int', 'Entre', 'Y', 'Colonia', 
            'CP', 'Mun', 'Edo', 'Email', 'Tel'
        ]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.fill = fill_head
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = center_align

        # Datos
        for con in contratos:
            ben = con.beneficiario
            row_data = [
                cuatrimestre, anio, contratista.rfc, con.folio, con.get_tipo_contrato_display(), 
                con.objeto_contrato, con.monto_contrato, str(con.vigencia_contrato or ''), 
                str(con.fecha_inicio or ''), str(con.fecha_fin or ''), con.num_estimado_trabajadores, 
                ben.rfc if ben else '', ben.nombre_razon_social if ben else '', 
                ben.registro_patronal if ben else '', ben.calle if ben else '', 
                ben.num_ext if ben else '', ben.num_int if ben else '', 
                ben.entre_calle if ben else '', ben.y_calle if ben else '', 
                ben.colonia if ben else '', ben.cp if ben else '', 
                ben.municipio_alcaldia if ben else '', ben.entidad_federativa if ben else '', 
                ben.correo if ben else '', ben.telefono if ben else ''
            ]
            ws.append(row_data)
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        for i in range(1, 26):
            ws.column_dimensions[get_column_letter(i)].width = 18

        rfc_clean = re.sub(r'[^A-Z0-9]', '', contratista.rfc.upper())
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="SISUB_CONTRATOS_{rfc_clean}.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        return HttpResponse(f"Error al generar reporte: {str(e)}", status=500)


@login_required(login_url='/login/')
@require_hr_permission('contratistas', 'ver')
def exportar_icsoe(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        contratista = Contratista.objects.get(id=id, empresa=empresa_actual)
        cuat = request.GET.get('cuatrimestre', '1')
        anio = request.GET.get('anio', '')
        
        if not anio: return HttpResponse("Año requerido", status=400)
        
        # Mapeo de periodos bimestrales por cuatrimestre (meses pares que cierran el bimestre)
        periodos_busqueda = []
        if cuat == '1':
            periodos_busqueda = ['FEBRERO', 'ABRIL']
        elif cuat == '2':
            periodos_busqueda = ['JUNIO', 'AGOSTO']
        elif cuat == '3':
            periodos_busqueda = ['OCTUBRE', 'DICIEMBRE']

        # Limpiamos el RFC del contratista (quitamos guiones y espacios)
        rfc_input_clean = re.sub(r'[^A-Z0-9]', '', contratista.rfc.upper())
        
        # Buscamos todas las importaciones bimestrales de la empresa actual
        importaciones_qs = ImportacionSUA.objects.filter(
            empresa=empresa_actual,
            tipo='bimestral'
        )
        
        # Filtrado manual para mayor precisión con RFCs formateados distinto
        importaciones_validas = []
        for imp in importaciones_qs:
            # Validar Año en el periodo
            if anio not in imp.periodo:
                continue
            
            # Validar Mes (alguno de los dos del cuatrimestre)
            mes_valido = False
            for mes in periodos_busqueda:
                if mes.upper() in imp.periodo.upper():
                    mes_valido = True
                    break
            if not mes_valido:
                continue
                
            # Validar RFC (limpiamos el del reporte para comparar)
            rfc_rep_clean = re.sub(r'[^A-Z0-9]', '', (imp.rfc_empresa or '').upper())
            # Match si el RFC limpio coincide exactamente o si uno contiene al otro
            if rfc_input_clean == rfc_rep_clean or rfc_input_clean in rfc_rep_clean or rfc_rep_clean in rfc_input_clean:
                importaciones_validas.append(imp)
        
        total_sin_credito = Decimal('0')
        total_con_credito = Decimal('0')
        total_amortizaciones = Decimal('0')
        
        # Obtener NSS de empleados vinculados a contratos de este contratista
        # Solo tomamos en cuenta los que pertenecen a la empresa actual y limpiamos el NSS
        empleados_nss_qs = Empleado.objects.filter(
            empresa=empresa_actual, 
            contratos_asignados__contratista=contratista
        ).values_list('nss', flat=True).distinct()
        
        # Guardamos los NSS limpios (solo números) para una comparación robusta
        nss_con_beneficiario = set(re.sub(r'[^0-9]', '', str(n)) for n in empleados_nss_qs if n)
        
        registros_encontrados = []
        for imp in importaciones_validas:
            registros_encontrados.append(imp.periodo)
            for t in imp.trabajadores.all():
                # Limpiar NSS del SUA para la comparación
                nss_t_clean = re.sub(r'[^0-9]', '', str(t.nss))
                
                # Nueva validación: Solo considerar si el trabajador (vía NSS limpio) está enlazado a un beneficiario en el sistema
                if nss_t_clean not in nss_con_beneficiario:
                    continue

                val_inf = (t.tipo_valor_infonavit or '').strip()
                if not val_inf or val_inf == '-':
                    total_sin_credito += t.aportacion_patronal
                else:
                    total_con_credito += t.aportacion_patronal
                
                total_amortizaciones += t.amortizacion
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ICSOE Informativo"
        
        # Estilos
        fill_brand = PatternFill(start_color="00b8b9", end_color="00b8b9", fill_type="solid")
        fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        fill_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        font_white = Font(bold=True, color="FFFFFF")
        font_bold = Font(bold=True)
        border = Border(
        left=Side(style='thin', color="B2B2B2"), 
        right=Side(style='thin', color="B2B2B2"), 
        top=Side(style='thin', color="B2B2B2"), 
        bottom=Side(style='thin', color="B2B2B2")
    )

        # FILA 1: a-Datos Generales (A1:AA1)
        ws.merge_cells('A1:AA1')
        c1 = ws['A1']
        c1.value = "a-Datos Generales"
        c1.fill = fill_brand
        c1.font = font_white
        c1.alignment = Alignment(horizontal="center")
        c1.border = border

        # FILA 2: Subgrupos
        subgrupos = [
            ("Periodo", 2),
            ("b-Datos de identificacion", 5),
            ("c-Domicilio fiscal", 9),
            ("d-Datos actuales de la escritura publica", 7),
            ("g-Aportacion y Amortización", 3),
            ("a-Registro en STPS", 1)
        ]
        
        curr_col = 1
        for nombre, span in subgrupos:
            start_cell = get_column_letter(curr_col) + "2"
            end_cell = get_column_letter(curr_col + span - 1) + "2"
            if span > 1:
                ws.merge_cells(f"{start_cell}:{end_cell}")
            
            cell = ws[start_cell]
            cell.value = nombre
            cell.fill = fill_gray
            cell.font = font_bold
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            
            # Aplicar bordes a todas las celdas del merge
            for c in range(curr_col, curr_col + span):
                ws.cell(row=2, column=c).border = border
                
            curr_col += span

        # FILA 3: Campos específicos
        headers = [
            # Periodo
            "cuatrimestre que declara", "año que se declara",
            # Identificación
            "Registro Federal de Contribuyente", "Nombre denominacion o razon social", "Correo electronico", "Telefono (numero extension)", "Registro patronal",
            # Domicilio
            "Calle", "Numero exterior", "Numero interior", "Entre calle", "Y calle", "Colonia", "Codigo Postal", "Municipio o Alcaldia", "Entidad Federativa",
            # Escritura
            "Representante legal", "Administrador Unico", "Numero de escritura", "Nombre del Notario Publico", "Numero de Notario Publico", "Fecha de escritura publica", "Folio mercantil",
            # Aportación y Amortización
            "Aportacion sin credito de los trabajadores del contrato", "Aportacion con credito de los trabajadores del contrato", "Amortizacion de los trabajadores del contrato",
            # STPS
            "Numero de registro ante la Secretaria de Trabajo y Prevision Social"
        ]

        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.fill = fill_light
            cell.font = font_bold
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = 20

        # FILA 4: Datos
        data_row = [
            # Periodo
            f"Cuatrimestre {cuat}", anio,
            # Identificación
            contratista.rfc, contratista.nombre_razon_social, contratista.correo, contratista.telefono, contratista.registro_patronal,
            # Domicilio
            contratista.calle, contratista.num_ext, contratista.num_int, contratista.entre_calle, contratista.y_calle, contratista.colonia, contratista.cp, contratista.municipio_alcaldia, contratista.entidad_federativa,
            # Escritura
            contratista.representante_legal, contratista.administrador_unico, contratista.num_escritura, contratista.nombre_notario_publico, contratista.num_notario_publico, str(contratista.fecha_escritura_publica) if contratista.fecha_escritura_publica else '', contratista.folio_mercantil,
            # Aportación y Amortización (Calculados Fase 1)
            total_sin_credito, total_con_credito, total_amortizaciones,
            # STPS
            contratista.numero_stps
        ]
        
        ws.append(data_row)
        
        # Formato de moneda para las sumatorias
        for col_idx in [24, 25, 26]:
            ws.cell(row=4, column=col_idx).number_format = '"$"#,##0.00'
        
        # Bordes a la fila de datos
        for col_idx in range(1, 28):
            ws.cell(row=4, column=col_idx).border = border

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="ICSOE_INFO_{rfc_input_clean}_{anio}_C{cuat}.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(str(e), status=500)

# ==============================================================================
# SUBMÓDULO: NÓMINA
# ==============================================================================

@login_required(login_url='/login/')
@require_hr_permission('nomina', 'ver')
def lista_nomina(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html')

    nominas = Nomina.objects.filter(empresa=empresa_actual).select_related('empleado', 'sucursal').order_by('-fecha_pago', 'nombre')

    # Filtros
    q = request.GET.get('q', '')
    f_folio = request.GET.get('folio', '')
    f_uuid = request.GET.get('uuid', '')
    f_colaborador = request.GET.get('colaborador', '')
    f_rfc_contratista = request.GET.get('rfc_contratista', '')
    f_fecha_pago = request.GET.get('fecha_pago', '')
    f_sucursal = request.GET.get('sucursal', '')

    if q:
        nominas = nominas.filter(
            Q(nombre__icontains=q) |
            Q(rfc__icontains=q) |
            Q(curp__icontains=q) |
            Q(folio__icontains=q) |
            Q(uuid__icontains=q) |
            Q(rfc_contratista__icontains=q)
        )
    
    if f_folio:
        nominas = nominas.filter(folio__icontains=f_folio)
    
    if f_uuid:
        nominas = nominas.filter(uuid__icontains=f_uuid)
    
    if f_colaborador:
        nominas = nominas.filter(nombre__icontains=f_colaborador)

    if f_rfc_contratista:
        nominas = nominas.filter(rfc_contratista__icontains=f_rfc_contratista)

    if f_fecha_pago:
        nominas = nominas.filter(fecha_pago=f_fecha_pago)
        
    if f_sucursal:
        nominas = nominas.filter(sucursal_id=f_sucursal)

    # Paginación
    paginator = Paginator(nominas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    empleados = Empleado.objects.filter(empresa=empresa_actual).order_by('apellido_paterno')
    contratistas = Contratista.objects.filter(empresa=empresa_actual).order_by('nombre_razon_social')

    return render(request, 'recursos_humanos/lista_nomina.html', {
        'page_obj': page_obj,
        'sucursales': sucursales,
        'empleados': empleados,
        'contratistas': contratistas,
        'empresa': empresa_actual,
        'filtros': {
            'q': q,
            'folio': f_folio,
            'uuid': f_uuid,
            'colaborador': f_colaborador,
            'rfc_contratista': f_rfc_contratista,
            'fecha_pago': f_fecha_pago,
            'sucursal': f_sucursal
        }
    })

@login_required(login_url='/login/')
@require_hr_permission('nomina', 'ver', json_response=True)
def obtener_nomina_json(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        nom = Nomina.objects.get(id=id, empresa=empresa_actual)
        
        def safe_decimal(val):
            return str(val) if val is not None else '0.00'

        data = {
            'id': nom.id,
            'empleado': nom.empleado_id or '',
            'periodo': nom.periodo,
            'uso_cfdi': nom.uso_cfdi,
            'uuid': nom.uuid or '',
            'tipo_nomina': nom.tipo_nomina,
            'serie': nom.serie or '',
            'folio': nom.folio or '',
            'fecha_emision': nom.fecha_emision.strftime('%Y-%m-%dT%H:%M') if nom.fecha_emision else '',
            'fecha_certificacion': nom.fecha_certificacion.strftime('%Y-%m-%dT%H:%M') if nom.fecha_certificacion else '',
            'fecha_pago': nom.fecha_pago.isoformat() if nom.fecha_pago else '',
            'fecha_inicial_pago': nom.fecha_inicial_pago.isoformat() if nom.fecha_inicial_pago else '',
            'fecha_final_pago': nom.fecha_final_pago.isoformat() if nom.fecha_final_pago else '',
            'dias_pagados': safe_decimal(nom.dias_pagados),
            'rfc': nom.rfc,
            'curp': nom.curp,
            'nss': nom.nss,
            'nombre': nom.nombre,
            'rfc_contratista': nom.rfc_contratista or '',
            'sdi': safe_decimal(nom.sdi),
            'sbc': safe_decimal(nom.sbc),
            'vacaciones_exento': safe_decimal(nom.vacaciones_exento),
            'vacaciones_dignas_exento': safe_decimal(nom.vacaciones_dignas_exento),
            'aguinaldo_exento': safe_decimal(nom.aguinaldo_exento),
            'sueldo_gravado': safe_decimal(nom.sueldo_gravado),
            'vacaciones_gravado': safe_decimal(nom.vacaciones_gravado),
            'vacaciones_dignas_gravado': safe_decimal(nom.vacaciones_dignas_gravado),
            'aguinaldo_gravado': safe_decimal(nom.aguinaldo_gravado),
        }
        return JsonResponse({'success': True, 'data': data})
    except Nomina.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Registro de nómina no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error en el servidor: {str(e)}'})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('nomina', 'crear', json_response=True)
def crear_nomina_ajax(request):
    empresa_actual = get_empresa_actual(request)
    try:
        data = request.POST
        sucursal_id = request.session.get('sucursal_id')
        empleado_id = data.get('empleado')
        empleado = Empleado.objects.get(id=empleado_id, empresa=empresa_actual) if empleado_id else None

        nueva_nom = Nomina(
            empresa=empresa_actual, sucursal_id=sucursal_id, empleado=empleado,
            periodo=data.get('periodo'), uso_cfdi=data.get('uso_cfdi', 'CN01'), uuid=data.get('uuid'),
            tipo_nomina=data.get('tipo_nomina', 'O'), serie=data.get('serie'), folio=data.get('folio'),
            fecha_emision=data.get('fecha_emision') or None, fecha_certificacion=data.get('fecha_certificacion') or None,
            fecha_pago=data.get('fecha_pago') or None, fecha_inicial_pago=data.get('fecha_inicial_pago') or None, fecha_final_pago=data.get('fecha_final_pago') or None,
            dias_pagados=Decimal(data.get('dias_pagados', '0')),
            rfc=data.get('rfc', '').upper(), curp=data.get('curp', '').upper(), nss=data.get('nss', ''), nombre=data.get('nombre', ''),
            rfc_contratista=data.get('rfc_contratista', '').upper(), sdi=Decimal(data.get('sdi', '0')), sbc=Decimal(data.get('sbc', '0')),
            vacaciones_exento=Decimal(data.get('vacaciones_exento', '0')), vacaciones_dignas_exento=Decimal(data.get('vacaciones_dignas_exento', '0')), aguinaldo_exento=Decimal(data.get('aguinaldo_exento', '0')),
            sueldo_gravado=Decimal(data.get('sueldo_gravado', '0')), vacaciones_gravado=Decimal(data.get('vacaciones_gravado', '0')), vacaciones_dignas_gravado=Decimal(data.get('vacaciones_dignas_gravado', '0')), aguinaldo_gravado=Decimal(data.get('aguinaldo_gravado', '0')),
        )
        nueva_nom.save()
        return JsonResponse({'success': True, 'message': 'Nómina registrada correctamente.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('nomina', 'editar', json_response=True)
def editar_nomina_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        nom = Nomina.objects.get(id=id, empresa=empresa_actual)
        data = request.POST
        nom.empleado_id = data.get('empleado') or None
        nom.periodo = data.get('periodo'); nom.uso_cfdi = data.get('uso_cfdi', 'CN01'); nom.uuid = data.get('uuid')
        nom.tipo_nomina = data.get('tipo_nomina'); nom.serie = data.get('serie'); nom.folio = data.get('folio')
        nom.fecha_emision = data.get('fecha_emision') or None; nom.fecha_certificacion = data.get('fecha_certificacion') or None
        nom.fecha_pago = data.get('fecha_pago') or None; nom.fecha_inicial_pago = data.get('fecha_inicial_pago') or None; nom.fecha_final_pago = data.get('fecha_final_pago') or None
        nom.dias_pagados = Decimal(data.get('dias_pagados', '0'))
        nom.rfc = data.get('rfc', '').upper(); nom.curp = data.get('curp', '').upper(); nom.nss = data.get('nss', ''); nom.nombre = data.get('nombre', '')
        nom.rfc_contratista = data.get('rfc_contratista', '').upper(); nom.sdi = Decimal(data.get('sdi', '0')); nom.sbc = Decimal(data.get('sbc', '0'))
        nom.vacaciones_exento = Decimal(data.get('vacaciones_exento', '0')); nom.vacaciones_dignas_exento = Decimal(data.get('vacaciones_dignas_exento', '0')); nom.aguinaldo_exento = Decimal(data.get('aguinaldo_exento', '0'))
        nom.sueldo_gravado = Decimal(data.get('sueldo_gravado', '0')); nom.vacaciones_gravado = Decimal(data.get('vacaciones_gravado', '0')); nom.vacaciones_dignas_gravado = Decimal(data.get('vacaciones_dignas_gravado', '0')); nom.aguinaldo_gravado = Decimal(data.get('aguinaldo_gravado', '0'))
        nom.save()
        return JsonResponse({'success': True, 'message': 'Nómina actualizada correctamente.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('nomina', 'eliminar', json_response=True)
def eliminar_nomina_ajax(request, id):
    empresa_actual = get_empresa_actual(request)
    try:
        nom = Nomina.objects.get(id=id, empresa=empresa_actual)
        nom.delete()
        return JsonResponse({'success': True, 'message': 'Registro de nómina eliminado correctamente.'})
    except Exception as e: return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/login/')
@require_POST
@require_hr_permission('nomina', 'crear', json_response=True)
def importar_nomina_ajax(request):
    empresa_actual = get_empresa_actual(request)
    archivo = request.FILES.get('archivo')
    if not archivo: return JsonResponse({'success': False, 'error': 'No se proporcionó ningún archivo.'})
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        col_map = {h: i for i, h in enumerate(headers) if h}

        def get_val(row, header, default=None):
            idx = col_map.get(header)
            return row[idx].value if idx is not None and row[idx].value is not None else default

        def to_decimal(val):
            try:
                v = str(val or 0).replace(',', '').replace('$', '').strip()
                return Decimal(v) if v else Decimal('0')
            except: return Decimal('0')

        def to_date(val):
            if isinstance(val, datetime): return val.date()
            if isinstance(val, str):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
                    try: return datetime.strptime(val, fmt).date()
                    except: pass
            return None

        def to_datetime(val):
            if isinstance(val, datetime): return val
            if isinstance(val, str):
                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%d/%m/%Y %H:%M'):
                    try: return datetime.strptime(val, fmt)
                    except: pass
            return None

        sucursal_id = request.session.get('sucursal_id')
        count = 0
        for row in sheet.iter_rows(min_row=2):
            curp_file = str(get_val(row, 'CURP', '')).strip().upper()
            nss_file = str(get_val(row, 'No Seguro social', '')).strip()
            if not curp_file or not nss_file: continue

            # 2. Buscar Empleado por CURP y NSS
            empleado = Empleado.objects.filter(empresa=empresa_actual, curp=curp_file, nss=nss_file).first()
            if empleado:
                # Actualizar datos del empleado
                empleado.rfc = str(get_val(row, 'RFC receptor', empleado.rfc)).strip().upper()
                empleado.cp = str(get_val(row, 'Domicilio receptor', empleado.cp)).strip()
                fecha_ing = to_date(get_val(row, 'Fecha inicio relacion laboral'))
                if fecha_ing: empleado.fecha_ingreso = fecha_ing
                
                empleado.jornada_sat = str(get_val(row, 'Tipo jornada', empleado.jornada_sat or '')).strip()
                empleado.tipo_regimen_sat = str(get_val(row, 'Tipo regimen', empleado.tipo_regimen_sat or '')).strip()
                empleado.num_empleado = str(get_val(row, 'Num empleado', empleado.num_empleado)).strip()
                empleado.antiguedad_sat = str(get_val(row, 'Antiguedad', empleado.antiguedad_sat or '')).strip()
                empleado.puesto = str(get_val(row, 'Puesto', empleado.puesto)).strip()
                empleado.periodicidad_pago_sat = str(get_val(row, 'Periodicidad pago', empleado.periodicidad_pago_sat or '')).strip()
                
                sbc_val = to_decimal(get_val(row, 'Salario base cot apor'))
                if sbc_val > 0: empleado.sbc = sbc_val
                empleado.save()

            tipo_nom_raw = str(get_val(row, 'Tipo nomina', 'O')).strip().upper()
            if 'EXTRAORDINARIA' in tipo_nom_raw or tipo_nom_raw == 'E':
                tipo_nom = 'E'
            else:
                tipo_nom = 'O'
            
            nueva_nom = Nomina(
                empresa=empresa_actual, sucursal_id=sucursal_id, empleado=empleado,
                periodo=str(get_val(row, 'Periodo', '')), uso_cfdi=str(get_val(row, 'Uso CFDI', 'CN01')), uuid=str(get_val(row, 'UUID', '')),
                tipo_nomina=tipo_nom, serie=str(get_val(row, 'Serie', '')), folio=str(get_val(row, 'Folio', '')),
                fecha_emision=to_datetime(get_val(row, 'Fecha emision')),
                fecha_certificacion=to_datetime(get_val(row, 'Fecha certificacion')),
                fecha_pago=to_date(get_val(row, 'Fecha pago')), 
                fecha_inicial_pago=to_date(get_val(row, 'Fecha inicial pago')), 
                fecha_final_pago=to_date(get_val(row, 'Fecha final pago')),
                dias_pagados=to_decimal(get_val(row, 'Dias pagados')),
                
                nombre=(empleado.nombre + " " + empleado.apellido_paterno + " " + empleado.apellido_materno) if empleado else str(get_val(row, 'Razon receptor', '')),
                rfc=empleado.rfc if empleado else str(get_val(row, 'RFC receptor', '')),
                curp=empleado.curp if empleado else curp_file, nss=empleado.nss if empleado else nss_file,
                rfc_contratista=str(get_val(row, 'RFC emisor', '')),
                sdi=empleado.sdi if empleado else to_decimal(get_val(row, 'Salario diario integrado')),
                sbc=empleado.sbc if empleado else to_decimal(get_val(row, 'Salario base cot apor')),

                vacaciones_exento=to_decimal(get_val(row, '001/P009/Exento/VACACIONES')),
                vacaciones_dignas_exento=to_decimal(get_val(row, '001/P009/Exento/VACACIONES DIGNAS')),
                aguinaldo_exento=to_decimal(get_val(row, '002/P004/Exento/AGUINALDO')),
                
                sueldo_gravado=to_decimal(get_val(row, '001/P001/Gravado/SUELDO')),
                vacaciones_gravado=to_decimal(get_val(row, '001/P009/Gravado/VACACIONES')),
                vacaciones_dignas_gravado=to_decimal(get_val(row, '001/P009/Gravado/VACACIONES DIGNAS')),
                aguinaldo_gravado=to_decimal(get_val(row, '002/P004/Gravado/AGUINALDO')),
            )
            nueva_nom.save(); count += 1
            
        return JsonResponse({'success': True, 'message': f'Se importaron {count} registros correctamente.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al procesar el archivo: {str(e)}'})

@login_required(login_url='/login/')
def exportar_sisub_trabajadores(request, id):
    empresa_actual = get_empresa_actual(request)
    contratista = get_object_or_404(Contratista, id=id, empresa=empresa_actual)
    
    cuat = int(request.GET.get('cuatrimestre', 1))
    anio = int(request.GET.get('anio', datetime.now().year))

    # Definición de meses del cuatrimestre
    cuat_meses_map = {
        1: [1, 2, 3, 4],
        2: [5, 6, 7, 8],
        3: [9, 10, 11, 12],
    }
    meses_filtro = cuat_meses_map.get(cuat, [1, 2, 3, 4])

    # 1. Obtener Nóminas de todo el cuatrimestre filtradas por contratista
    nominas_qs = Nomina.objects.filter(
        empresa=empresa_actual,
        fecha_pago__year=anio,
        fecha_pago__month__in=meses_filtro
    ).filter(
        Q(empleado__contratista=contratista) | Q(rfc_contratista__iexact=contratista.rfc.strip())
    ).select_related('empleado', 'empleado__contratista')

    # 2. Buscar datos de Incapacidad en SUA (Buscamos en todos los meses del cuatrimestre)
    incapacidades = {} # Llave: (nss, bimestre)
    for mes in meses_filtro:
        nombre_mes = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"][mes]
        sua_periodo = f"{nombre_mes} {anio}"
        sua_reg = ImportacionSUA.objects.filter(empresa=empresa_actual, periodo__icontains=sua_periodo).first()
        if sua_reg:
            bimestre_actual = (mes + 1) // 2
            trabajadores_sua = TrabajadorSUA.objects.filter(importacion=sua_reg)
            for ts in trabajadores_sua:
                nss_key = ts.nss.strip() if ts.nss else ""
                if nss_key:
                    inc_val = 0
                    try: inc_val = int(float(ts.inc or 0))
                    except: pass
                    # Acumulamos por NSS y Bimestre
                    k = (nss_key, bimestre_actual)
                    incapacidades[k] = incapacidades.get(k, 0) + inc_val

    # 3. Preparar Excel
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trabajadores SISUB"

    # Definición de Estilos
    fill_brand = PatternFill(start_color="00b8b9", end_color="00b8b9", fill_type="solid")
    fill_gray_dark = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_gray_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    font_white = Font(bold=True, color="FFFFFF")
    font_bold = Font(bold=True)
    border = Border(
        left=Side(style='thin', color="B2B2B2"), 
        right=Side(style='thin', color="B2B2B2"), 
        top=Side(style='thin', color="B2B2B2"), 
        bottom=Side(style='thin', color="B2B2B2")
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # NIVEL 1: Azul de marca
    ws.merge_cells('A1:N1')
    ws['A1'] = "d-Informacion de los trabajadores"
    ws.merge_cells('O1:S1')
    ws['O1'] = "e-Determinacion del salario base de aportacion"

    for c in range(1, 20):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill_brand
        cell.font = font_white
        cell.alignment = center_align
        cell.border = border

    # NIVEL 2: Gris obscuro
    ws.merge_cells('A2:C2')
    ws['A2'] = "Periodo"
    ws.merge_cells('D2:N2')
    ws['D2'] = "a-Identificacion"
    ws.merge_cells('O2:S2')
    ws['O2'] = "a-Percepciones por bimestre 1"

    for c in range(1, 20):
        cell = ws.cell(row=2, column=c)
        cell.fill = fill_gray_dark
        cell.font = font_bold
        cell.alignment = center_align
        cell.border = border

    # NIVEL 3: Gris claro
    headers = [
        "cuatrimestre que declara", "año que se declara", "bimestre",
        "Registro Federal de Contribuyente del sujeto obligado", "Numero de contrato",
        "Registro Patronal ante el IMSS", "Numero de Seguro Social del trabajador",
        "Calle (centro del trabajo)", "Numero exterior (centro del trabajo)",
        "Numero interior (centro de trabajo)", "Colonia (centro de trabajo)",
        "Codigo Postal (centro de trabajo)", "Municipio o Alcaldia (centro de trabajo)",
        "Entidad federativa (centro de trabajo)",
        "Monto Percepciones variables", "Monto Percepciones fijas",
        "Dias de Incapacidad", "Percepciones no integrables al SBA",
        "salario no excedente (VSM)"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = font_bold
        cell.fill = fill_gray_light
        cell.alignment = center_align
        cell.border = border

    # 4. Llenar Datos
    row_num = 4
    # Agrupamos por (NSS, Bimestre) para consolidar percepciones
    data_procesada = {}

    for nom in nominas_qs:
        bimestre_nom = (nom.fecha_pago.month + 1) // 2
        nss_val = nom.nss.strip() if nom.nss else f"REF-{nom.id}"
        key = (nss_val, bimestre_nom)
        
        if key not in data_procesada:
            contrato = None
            if nom.empleado:
                contrato = Contrato.objects.filter(empleados=nom.empleado, contratista=contratista).first()
            
            beneficiario = contrato.beneficiario if contrato else (nom.empleado.beneficiario if nom.empleado else None)

            data_procesada[key] = {
                'nss': nom.nss,
                'sdi': nom.sdi,
                'bimestre': bimestre_nom,
                'percepciones_fijas': 0,
                'contrato_folio': contrato.folio if (contrato and contrato.folio) else "1",
                'beneficiario': beneficiario
            }
        
        total_p = (
            nom.vacaciones_exento + nom.vacaciones_dignas_exento + nom.aguinaldo_exento +
            nom.sueldo_gravado + nom.vacaciones_gravado + nom.vacaciones_dignas_gravado + nom.aguinaldo_gravado
        )
        data_procesada[key]['percepciones_fijas'] += total_p

    # Ordenar por bimestre y luego por nss
    for k in sorted(data_procesada.keys(), key=lambda x: (x[1], x[0])):
        data = data_procesada[k]
        b = data['beneficiario']
        ws.cell(row=row_num, column=1).value = cuat
        ws.cell(row=row_num, column=2).value = anio
        ws.cell(row=row_num, column=3).value = data['bimestre']
        ws.cell(row=row_num, column=4).value = contratista.rfc
        ws.cell(row=row_num, column=5).value = data['contrato_folio']
        ws.cell(row=row_num, column=6).value = contratista.registro_patronal
        ws.cell(row=row_num, column=7).value = data['nss']
        
        if b:
            ws.cell(row=row_num, column=8).value = b.calle
            ws.cell(row=row_num, column=9).value = b.num_ext
            ws.cell(row=row_num, column=10).value = b.num_int
            ws.cell(row=row_num, column=11).value = b.colonia
            ws.cell(row=row_num, column=12).value = b.cp
            ws.cell(row=row_num, column=13).value = b.municipio_alcaldia
            ws.cell(row=row_num, column=14).value = b.entidad_federativa
        
        ws.cell(row=row_num, column=15).value = 0 # Monto Percepciones variables (Ceil if used)
        ws.cell(row=row_num, column=16).value = math.ceil(float(data['percepciones_fijas']))
        ws.cell(row=row_num, column=17).value = incapacidades.get((data['nss'].strip(), data['bimestre']), 0)
        ws.cell(row=row_num, column=18).value = 0 # Percepciones no integrables (Keep decimals)
        ws.cell(row=row_num, column=19).value = data['sdi'] # Salario no excedente (Keep decimals)
        
        # Formato de número para columnas con decimales (19)
        ws.cell(row=row_num, column=19).number_format = '#,##0.00'
        
        for c in range(1, 20):
            ws.cell(row=row_num, column=c).border = border
            
        row_num += 1

    from openpyxl.utils import get_column_letter
    for i in range(1, 20):
        ws.column_dimensions[get_column_letter(i)].width = 22

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="SISUB_TRABAJADORES_{contratista.rfc}_{anio}_C{cuat}.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
@require_POST
@transaction.atomic
def importar_contratos_ajax(request):
    """Cargador de contratos y beneficiarios desde Excel."""
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'status': 'error', 'message': 'Empresa no encontrada.'})
    
    file = request.FILES.get('archivo')
    if not file:
        return JsonResponse({'status': 'error', 'message': 'No se proporcionó ningún archivo.'})
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active
        # Obtener encabezados y limpiar espacios
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        
        def get_val(row, header_name, default=None):
            try:
                idx = headers.index(header_name)
                val = row[idx].value
                return val if val is not None else default
            except (ValueError, IndexError):
                return default

        def to_decimal(val):
            if val is None: return Decimal('0')
            try:
                v = str(val).replace(',', '').replace('$', '').strip()
                return Decimal(v) if v else Decimal('0')
            except: return Decimal('0')

        def to_date(val):
            if isinstance(val, datetime): return val.date()
            if isinstance(val, str):
                # Intentar varios formatos comunes
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%dT%H:%M:%S'):
                    try: return datetime.strptime(val.strip(), fmt).date()
                    except: pass
            return None

        TIPO_CONTRATO_MAP = {
            '01': '01', 'tiempo indeterminado': '01', 'indeterminado': '01',
            '02': '02', 'obra determinada': '02',
            '03': '03', 'tiempo determinado': '03', 'determinado': '03',
            '05': '05', 'prueba': '05',
            '06': '06', 'capacitacion': '06', 'capacitación': '06',
        }

        sucursal_id = request.session.get('sucursal_id')
        count = 0
        
        for row in sheet.iter_rows(min_row=2):
            rfc_beneficiario = str(get_val(row, 'Registro Federal de Contribuyentes', '')).strip().upper()
            if not rfc_beneficiario:
                continue

            # 1. Buscar contratista por RFC (Sujeto Obligado)
            rfc_sujeto_obligado = str(get_val(row, 'Registro Federal de Contribuyente del sujeto obligado', '')).strip().upper()
            contratista = None
            if rfc_sujeto_obligado:
                contratista = Contratista.objects.filter(empresa=empresa_actual, rfc=rfc_sujeto_obligado).first()

            # 2. Buscar o crear Beneficiario
            beneficiario, created = Beneficiario.objects.get_or_create(
                empresa=empresa_actual,
                rfc=rfc_beneficiario,
                defaults={
                    'sucursal_id': sucursal_id,
                    'nombre_razon_social': str(get_val(row, 'Nombre denominacion o razon social', '')).strip(),
                    'registro_patronal': str(get_val(row, 'Registro Patronal ante el IMSS', '')).strip(),
                    'calle': str(get_val(row, 'Calle', '')).strip(),
                    'num_ext': str(get_val(row, 'Numero exterior', '')).strip(),
                    'num_int': str(get_val(row, 'Numero interior', '')).strip(),
                    'entre_calle': str(get_val(row, 'Entre calle', '')).strip(),
                    'y_calle': str(get_val(row, 'Y calle', '')).strip(),
                    'colonia': str(get_val(row, 'Colonia', '')).strip(),
                    'cp': str(get_val(row, 'Codigo Postal', '')).strip(),
                    'municipio_alcaldia': str(get_val(row, 'Municipio o Alcaldia', '')).strip(),
                    'entidad_federativa': str(get_val(row, 'Entidad Federativa', '')).strip(),
                    'correo': str(get_val(row, 'Correo electronico', '')).strip(),
                    'telefono': str(get_val(row, 'telefono (numero extension)', '')).strip(),
                }
            )

            # 3. Extraer datos de Contrato
            folio = str(get_val(row, 'Numero de contrato', '')).strip()
            tipo_raw = str(get_val(row, 'Tipo de contrato', '01')).strip().lower()
            tipo_contrato = '01'
            for key, code in TIPO_CONTRATO_MAP.items():
                if key in tipo_raw:
                    tipo_contrato = code
                    break
            
            objeto = str(get_val(row, 'Objeto del contrato', '')).strip()
            monto = to_decimal(get_val(row, 'Monto del contrato'))
            vigencia = to_date(get_val(row, 'Vigencia (del contrato)'))
            f_inicio = to_date(get_val(row, 'Fecha de inicio (del contrato)'))
            f_fin = to_date(get_val(row, 'Fecha de termino (del contrato)'))
            num_trabajadores_raw = get_val(row, 'Numero estimado mensual de trabajadores que se pondran a disposicion (del contrato)', 0)
            try:
                num_trabajadores = int(num_trabajadores_raw) if num_trabajadores_raw else 0
            except:
                num_trabajadores = 0

            # 4. Crear el Contrato
            Contrato.objects.create(
                empresa=empresa_actual,
                sucursal_id=sucursal_id,
                beneficiario=beneficiario,
                contratista=contratista,
                folio=folio,
                tipo_contrato=tipo_contrato,
                objeto_contrato=objeto,
                monto_contrato=monto,
                vigencia_contrato=vigencia,
                fecha_inicio=f_inicio or datetime.now().date(),
                fecha_fin=f_fin,
                num_estimado_trabajadores=num_trabajadores,
                estado='vigente'
            )
            count += 1

        return JsonResponse({
            'status': 'success', 
            'message': f'Proceso completado. Se registraron {count} contratos.'
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error al procesar el archivo: {str(e)}'})

@login_required(login_url='/login/')
@require_POST
@transaction.atomic
def importar_contratistas_ajax(request):
    """Cargador de contratistas desde Excel."""
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return JsonResponse({'status': 'error', 'message': 'Empresa no encontrada.'})
    
    file = request.FILES.get('archivo')
    if not file:
        return JsonResponse({'status': 'error', 'message': 'No se proporcionó ningún archivo.'})
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        
        def get_val(row, header_name, default=None):
            try:
                idx = headers.index(header_name)
                val = row[idx].value
                return val if val is not None else default
            except (ValueError, IndexError):
                return default

        def to_date(val):
            if isinstance(val, datetime): return val.date()
            if isinstance(val, str):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%dT%H:%M:%S'):
                    try: return datetime.strptime(val.strip(), fmt).date()
                    except: pass
            return None

        sucursal_id = request.session.get('sucursal_id')
        count = 0
        
        for row in sheet.iter_rows(min_row=2):
            rfc = str(get_val(row, 'Registro Federal de Contribuyente', '')).strip().upper()
            if not rfc:
                continue

            # Buscar o crear Contratista
            contratista, created = Contratista.objects.update_or_create(
                empresa=empresa_actual,
                rfc=rfc,
                defaults={
                    'sucursal_id': sucursal_id,
                    'nombre_razon_social': str(get_val(row, 'Nombre denominacion o razon social', '')).strip(),
                    'correo': str(get_val(row, 'Correo electronico', '')).strip(),
                    'telefono': str(get_val(row, 'Telefono (numero extension)', '')).strip(),
                    'registro_patronal': str(get_val(row, 'Registro patronal', '')).strip(),
                    'calle': str(get_val(row, 'Calle', '')).strip(),
                    'num_ext': str(get_val(row, 'Numero exterior', '')).strip(),
                    'num_int': str(get_val(row, 'Numero interior', '')).strip(),
                    'entre_calle': str(get_val(row, 'Entre calle', '')).strip(),
                    'y_calle': str(get_val(row, 'Y calle', '')).strip(),
                    'colonia': str(get_val(row, 'Colonia', '')).strip(),
                    'cp': str(get_val(row, 'Codigo Postal', '')).strip(),
                    'municipio_alcaldia': str(get_val(row, 'Municipio o Alcaldia', '')).strip(),
                    'entidad_federativa': str(get_val(row, 'Entidad Federativa', '')).strip(),
                    'representante_legal': str(get_val(row, 'Representante legal', '')).strip(),
                    'administrador_unico': str(get_val(row, 'Administrador Unico', '')).strip(),
                    'num_escritura': str(get_val(row, 'Numero de escritura', '')).strip(),
                    'nombre_notario_publico': str(get_val(row, 'Nombre del Notario Publico', '')).strip(),
                    'num_notario_publico': str(get_val(row, 'Numero de Notario Publico', '')).strip(),
                    'fecha_escritura_publica': to_date(get_val(row, 'Fecha de escritura publica')),
                    'folio_mercantil': str(get_val(row, 'Folio mercantil', '')).strip(),
                    'numero_stps': str(get_val(row, 'Numero de registro ante la Secretaria de Trabajo y Prevision Social', '')).strip(),
                }
            )
            count += 1

        return JsonResponse({
            'status': 'success', 
            'message': f'Proceso completado. Se registraron/actualizaron {count} contratistas.'
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error al procesar el archivo: {str(e)}'})

from .sat_service import SATService

# ==============================================================================
# INTEGRACIÓN SAT XML (Descarga Masiva Real)
# ==============================================================================

@login_required(login_url='/login/')
@require_POST
def diagnostico_fiel_ajax(request):
    """Extrae información de un certificado .cer para diagnóstico."""
    archivo_cer = request.FILES.get('archivo_cer')
    if not archivo_cer:
        return JsonResponse({'success': False, 'error': 'No se proporcionó el archivo .cer'})
    
    res = SATService.obtener_info_certificado(archivo_cer.read())
    return JsonResponse(res)

@login_required(login_url='/login/')
@require_POST
@transaction.atomic
def solicitar_descarga_sat_ajax(request):
    """Envía una solicitud real al Web Service del SAT."""
    empresa_actual = get_empresa_actual(request)
    data = request.POST
    contratista_id = data.get('contratista_id')
    password_fiel = data.get('password_fiel')
    fecha_inicio_str = data.get('fecha_inicio')
    fecha_fin_str = data.get('fecha_fin')
    estatus = data.get('estatus', 'vigente')

    if not contratista_id or not password_fiel:
        return JsonResponse({'status': 'error', 'message': 'Faltan datos obligatorios (Contratista o Contraseña).'})

    contratista = get_object_or_404(Contratista, id=contratista_id, empresa=empresa_actual)
    
    # Manejo de Archivos FIEL (Cifrado si se proporcionan nuevos)
    archivo_cer = request.FILES.get('archivo_cer')
    archivo_key = request.FILES.get('archivo_key')
    master_key = get_master_key()
    
    if archivo_cer and archivo_key:
        try:
            cer_content = archivo_cer.read()
            # Validación Local previa
            info_cert = SATService.obtener_info_certificado(cer_content)
            if not info_cert['success']:
                return JsonResponse({'status': 'error', 'message': f"El archivo .cer no es válido: {info_cert['error']}"})
            
            if not info_cert['es_fiel']:
                return JsonResponse({'status': 'error', 'message': "El certificado subido es de SELLOS (CSD). Para este trámite se requiere la e.firma (FIEL)."})

            # VALIDACIÓN DE RFC CRUZADA
            rfc_cert = info_cert['rfc'].upper().strip()
            rfc_cont = contratista.rfc.upper().strip()
            if rfc_cert != rfc_cont:
                return JsonResponse({
                    'status': 'error', 
                    'message': f"El RFC del certificado ({rfc_cert}) no coincide con el RFC del contratista seleccionado ({rfc_cont})."
                })

            cer_cifrado, key_cifrado, data_key_cifrada = cifrar_archivos_fiel(
                cer_content, 
                archivo_key.read(), 
                master_key
            )
            FielContratista.objects.update_or_create(
                contratista=contratista,
                defaults={
                    'certificado_cifrado': cer_cifrado,
                    'llave_privada_cifrada': key_cifrado,
                    'data_key_cifrada': data_key_cifrada,
                    'rfc_fiel': info_cert['rfc'] # Usamos el RFC real del certificado
                }
            )
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Error al procesar archivos FIEL: {str(e)}'})

    try:
        sat_service = SATService(contratista)
        # El SAT suele preferir datetime (con hora) para el rango
        f_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
        f_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        
        id_solicitud = sat_service.solicitar_descarga(password_fiel, f_inicio, f_fin, estatus)

        if not id_solicitud:
            return JsonResponse({'status': 'error', 'message': 'El SAT no devolvió un ID de solicitud.'})

        SolicitudDescargaSAT.objects.create(
            empresa=empresa_actual,
            contratista=contratista,
            id_solicitud=id_solicitud,
            fecha_inicio=f_inicio.date(),
            fecha_fin=f_fin.date(),
            estado='solicitada'
        )

        return JsonResponse({
            'status': 'success',
            'message': f'Solicitud enviada al SAT exitosamente. ID: {id_solicitud}'
        })
    except Exception as e:
        import traceback
        err_detail = str(e)
        if not err_detail:
            err_detail = f"Excepción de tipo {type(e).__name__} sin mensaje descriptivo."
        
        print("--- ERROR EN SOLICITUD SAT ---")
        print(traceback.format_exc())
        return JsonResponse({'status': 'error', 'message': f'Error del SAT o Servidor: {err_detail}'})

@login_required(login_url='/login/')
def verificar_fiel_contratista_ajax(request, contratista_id):
    """Verifica si un contratista ya tiene archivos FIEL cargados."""
    empresa_actual = get_empresa_actual(request)
    contratista = get_object_or_404(Contratista, id=contratista_id, empresa=empresa_actual)
    fiel = FielContratista.objects.filter(contratista=contratista).first()
    
    return JsonResponse({
        'tiene_fiel': fiel is not None,
        'rfc': fiel.rfc_fiel if fiel else contratista.rfc,
        'nombre': contratista.nombre_razon_social
    })

@login_required(login_url='/login/')
def listar_solicitudes_sat_ajax(request):
    """Retorna el historial de solicitudes de la empresa."""
    empresa_actual = get_empresa_actual(request)
    solicitudes = SolicitudDescargaSAT.objects.filter(empresa=empresa_actual).order_by('-fecha_creacion')[:10]
    
    data = []
    for s in solicitudes:
        data.append({
            'id': s.id,
            'id_solicitud': s.id_solicitud,
            'periodo': f"{s.fecha_inicio} al {s.fecha_fin}",
            'estado': s.estado,
            'fecha': s.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        })
    
    return JsonResponse({'solicitudes': data})

@login_required(login_url='/login/')
@require_POST
def verificar_estatus_sat_ajax(request, solicitud_id):
    """Consulta al SAT si la solicitud ya está terminada."""
    empresa_actual = get_empresa_actual(request)
    solicitud = get_object_or_404(SolicitudDescargaSAT, id=solicitud_id, empresa=empresa_actual)
    password = request.POST.get('password_fiel')
    
    if not password:
        return JsonResponse({'status': 'error', 'message': 'Se requiere la contraseña de la FIEL.'})
    
    try:
        sat_service = SATService(solicitud.contratista)
        res = sat_service.verificar_estatus(solicitud.id_solicitud, password)
        
        # Mapeo de estados del SAT: 1=Aceptada, 2=En Proceso, 3=Terminada, 4=Error...
        estado_map = {'1': 'solicitada', '2': 'en_proceso', '3': 'terminada', '4': 'error'}
        solicitud.estado = estado_map.get(str(res['estado']), 'en_proceso')
        solicitud.save()
        
        return JsonResponse({'status': 'success', 'nuevo_estado': solicitud.estado})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@login_required(login_url='/login/')
@require_POST
def integrar_xml_sat_ajax(request, solicitud_id):
    """Baja los paquetes reales del SAT e integra las nóminas."""
    empresa_actual = get_empresa_actual(request)
    solicitud = get_object_or_404(SolicitudDescargaSAT, id=solicitud_id, empresa=empresa_actual)
    password = request.POST.get('password_fiel')
    
    if solicitud.estado != 'terminada':
        return JsonResponse({'status': 'error', 'message': 'Solicitud no terminada en el SAT.'})
    
    try:
        sat_service = SATService(solicitud.contratista)
        res = sat_service.verificar_estatus(solicitud.id_solicitud, password)
        paquetes = res.get('paquetes', [])
        
        if not paquetes:
            return JsonResponse({'status': 'error', 'message': 'No se encontraron paquetes para descargar.'})
            
        count, archivos_encontrados = sat_service.descargar_e_integrar(
            solicitud.id_solicitud, paquetes, password, 
            empresa_actual, request.session.get('sucursal_id')
        )
        
        if count > 0:
            solicitud.estado = 'procesada'
            solicitud.save()
            return JsonResponse({
                'status': 'success',
                'message': f'Integración exitosa. Se procesaron {count} XMLs de nómina reales.'
            })
        else:
            lista_archivos = ", ".join(archivos_encontrados[:5]) + ("..." if len(archivos_encontrados) > 5 else "")
            return JsonResponse({
                'status': 'success',
                'message': f'Se procesaron 0 XMLs de nómina. Archivos encontrados en el paquete: {lista_archivos or "Ninguno"}. Verifique que el periodo solicitado contenga CFDI de Nómina.'
            })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
