from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import F, Q, Sum, Avg
from decimal import Decimal
import openpyxl
from datetime import datetime
from .models import OrdenProduccion, DetalleOrdenProduccion
from core.models import Producto, DetalleReceta
from almacenes.models import Inventario, Almacen
from panel.models import Empresa
from notificaciones.utils import crear_notificacion
from django.contrib.auth.models import User
from preferencias.permissions import require_production_permission, user_has_production_permission
import json

# --- HELPER MULTI-TENANCY ---
def get_empresa_actual(request):
    username = request.user.username
    if '@' in username:
        subdominio = username.split('@')[1]
        try:
            return Empresa.objects.get(subdominio=subdominio)
        except Empresa.DoesNotExist:
            return None
    return None

# --- 1. IMPORTADOR Y EXPORTADOR ---

@login_required(login_url='/login/')
@require_production_permission('tablero_control', 'crear')
def descargar_plantilla_produccion(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Produccion"
    
    # Encabezados: Referencia ayuda a agrupar componentes en una misma OP
    headers = [
        'Referencia (Ej: OP-001)', 'Producto Final (Nombre o ID)', 'Cantidad Producir', 
        'Almacen PT (ID o Nombre)', 'Almacen MP (ID o Nombre)', 'Responsable (User)',
        'Componente (Nombre o ID)', 'Cantidad Componente', 'Notas'
    ]
    
    ws.append(headers)
    
    # Ajustar ancho de columnas
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Produccion.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
@require_production_permission('tablero_control', 'crear', json_response=True)
def importar_produccion_ajax(request):
    if request.method == 'POST' and request.FILES.get('archivo_produccion'):
        empresa_actual = get_empresa_actual(request)
        excel_file = request.FILES['archivo_produccion']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return JsonResponse({'success': False, 'error': 'El archivo está vacío.'})
            
            data_rows = rows[1:]
            
            # Agrupar por Referencia
            ordenes_data = {}
            for idx, row in enumerate(data_rows, start=2):
                if not any(row): continue
                ref = str(row[0] or f"TEMP-{idx}").strip()
                if ref not in ordenes_data:
                    ordenes_data[ref] = {
                        'prod_final_input': str(row[1] or '').strip(),
                        'cant_producir': row[2] or 1,
                        'alm_pt_input': str(row[3] or '').strip(),
                        'alm_mp_input': str(row[4] or '').strip(),
                        'resp_input': str(row[5] or '').strip(),
                        'notas': str(row[8] or '').strip(),
                        'items': []
                    }
                # Si hay componente en esta fila
                if row[6]:
                    ordenes_data[ref]['items'].append({
                        'comp_input': str(row[6]).strip(),
                        'cant_comp': row[7] or 1
                    })

            creadas = 0
            errores = []
            
            # Obtener sucursal de la sesión
            sucursal_obj = None
            sucursal_id = request.session.get('sucursal_id')
            if sucursal_id:
                from preferencias.models import Sucursal
                try:
                    sucursal_obj = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
                except Sucursal.DoesNotExist:
                    pass

            for ref, data in ordenes_data.items():
                try:
                    with transaction.atomic():
                        # Buscar Producto Final
                        producto_final = None
                        if data['prod_final_input'].isdigit():
                            producto_final = Producto.objects.filter(id=int(data['prod_final_input']), empresa=empresa_actual, tipo_abastecimiento='produccion').first()
                        if not producto_final:
                            producto_final = Producto.objects.filter(nombre__icontains=data['prod_final_input'], empresa=empresa_actual, tipo_abastecimiento='produccion').first()
                        
                        if not producto_final:
                            errores.append(f"Ref {ref}: No se encontró el producto final '{data['prod_final_input']}' o no es de producción.")
                            continue

                        # Buscar Almacenes
                        alm_pt = None
                        if data['alm_pt_input']:
                            if data['alm_pt_input'].isdigit(): alm_pt = Almacen.objects.filter(id=int(data['alm_pt_input']), empresa=empresa_actual).first()
                            else: alm_pt = Almacen.objects.filter(nombre__icontains=data['alm_pt_input'], empresa=empresa_actual).first()

                        alm_mp = None
                        if data['alm_mp_input']:
                            if data['alm_mp_input'].isdigit(): alm_mp = Almacen.objects.filter(id=int(data['alm_mp_input']), empresa=empresa_actual).first()
                            else: alm_mp = Almacen.objects.filter(nombre__icontains=data['alm_mp_input'], empresa=empresa_actual).first()

                        # Responsable
                        responsable_user = None
                        if data['resp_input']:
                            responsable_user = User.objects.filter(username=data['resp_input']).first()

                        # Crear Orden de Producción
                        orden = OrdenProduccion.objects.create(
                            empresa=empresa_actual,
                            producto=producto_final,
                            cantidad=data['cant_producir'],
                            almacen=alm_pt if alm_pt else Almacen.objects.filter(empresa=empresa_actual).first(),
                            almacen_materia_prima=alm_mp,
                            responsable=responsable_user,
                            solicitante=request.user,
                            sucursal=sucursal_obj,
                            estado='borrador',
                            notas=data['notas']
                        )

                        # Si traía items personalizados en el Excel
                        if data['items']:
                            for item in data['items']:
                                comp_prod = None
                                if item['comp_input'].isdigit():
                                    comp_prod = Producto.objects.filter(id=int(item['comp_input']), empresa=empresa_actual).first()
                                else:
                                    comp_prod = Producto.objects.filter(nombre__icontains=item['comp_input'], empresa=empresa_actual).first()
                                
                                if comp_prod:
                                    DetalleOrdenProduccion.objects.create(
                                        orden_produccion=orden,
                                        producto=comp_prod,
                                        cantidad=item['cant_comp']
                                    )
                        else:
                            # Cargar receta por defecto si no hay items en Excel
                            receta = DetalleReceta.objects.filter(producto_padre=producto_final)
                            for r in receta:
                                DetalleOrdenProduccion.objects.create(
                                    orden_produccion=orden,
                                    producto=r.componente,
                                    cantidad=r.cantidad * orden.cantidad
                                )
                        
                        creadas += 1
                except Exception as e:
                    errores.append(f"Ref {ref}: {str(e)}")
            
            if creadas == 0 and errores:
                return JsonResponse({'success': False, 'error': f"No se pudo importar ninguna orden. Primer error: {errores[0]}"})

            msg = f'Importación finalizada.\nÓrdenes de producción creadas: {creadas}\n'
            if errores:
                msg += f'\nFilas con error: {len(errores)}\n'
                msg += '\nDetalle de primeros errores:\n'
                msg += "\n".join(errores[:10])
                if len(errores) > 10:
                    msg += "\n... (revisa el resto de tu archivo Excel)"
                
            return JsonResponse({'success': True, 'message': msg})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
            
    return JsonResponse({'success': False, 'error': 'Solicitud inválida.'})

@login_required(login_url='/login/')
@require_production_permission('tablero_control', 'ver')
def exportar_produccion_excel(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return HttpResponse("No autorizado", status=403)

    # --- Filtrado (Sincronizado con dashboard) ---
    q = request.GET.get('q', '')
    folio_op = request.GET.get('folio_op', '')
    producto_id = request.GET.get('producto_id', '')
    estado = request.GET.get('estado', '')
    sucursal_id_filtro = request.GET.get('sucursal', '')

    ordenes = OrdenProduccion.objects.filter(empresa=empresa_actual).select_related(
        'producto', 'almacen', 'responsable', 'solicitante', 'sucursal'
    ).order_by('-fecha_creacion')

    if q:
        ordenes = ordenes.filter(
            Q(id__icontains=q) | Q(producto__nombre__icontains=q) |
            Q(responsable__username__icontains=q) | Q(estado__icontains=q)
        )
    if folio_op:
        ordenes = ordenes.filter(id__icontains=folio_op.replace('OP-', ''))
    if producto_id and producto_id != 'all':
        ordenes = ordenes.filter(producto_id=producto_id)
    if estado:
        ordenes = ordenes.filter(estado=estado)
    if sucursal_id_filtro:
        ordenes = ordenes.filter(sucursal_id=sucursal_id_filtro)

    # --- Generar Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Produccion"

    headers = [
        'ID', 'Folio', 'Producto Final', 'Cantidad', 'Almacén PT', 
        'Responsable', 'Estado', 'Fecha Creación', 'Fecha Inicio', 'Fecha Terminado', 'Sucursal'
    ]
    ws.append(headers)

    for o in ordenes:
        ws.append([
            o.id, o.folio, o.producto.nombre, o.cantidad, o.almacen.nombre,
            o.responsable.username if o.responsable else "N/A",
            o.get_estado_display(),
            o.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            o.fecha_inicio.strftime('%Y-%m-%d %H:%M') if o.fecha_inicio else "N/A",
            o.fecha_terminado.strftime('%Y-%m-%d %H:%M') if o.fecha_terminado else "N/A",
            o.sucursal.nombre if o.sucursal else "Principal"
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Produccion_{empresa_actual.nombre}.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
@require_production_permission('tablero_control', 'ver')
def dashboard_produccion(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)

    # --- LÓGICA DE FILTRADO ---
    # Obtener sucursal de sesión para filtrar por defecto
    sucursal_sesion_id = request.session.get('sucursal_id')
    
    q = request.GET.get('q', '')
    folio_op = request.GET.get('folio_op', '')
    producto_id = request.GET.get('producto_id', '')
    estado = request.GET.get('estado', '')
    
    # Si no viene sucursal en el GET ni almacén, usamos la de la sesión por defecto
    sucursal_id = request.GET.get('sucursal')
    if sucursal_id is None and sucursal_sesion_id:
        sucursal_id = str(sucursal_sesion_id)

    ordenes_qs = OrdenProduccion.objects.filter(empresa=empresa_actual).select_related(
        'producto', 'pedido_origen', 'almacen', 'responsable', 'solicitante', 'producto__test_calidad', 'sucursal'
    ).prefetch_related('detalles', 'resultados_test').order_by('-fecha_creacion')

    if q:
        ordenes_qs = ordenes_qs.filter(
            Q(id__icontains=q) |
            Q(producto__nombre__icontains=q) |
            Q(cliente__nombre__icontains=q) |
            Q(cliente__razon_social__icontains=q) |
            Q(notas__icontains=q)
        ).distinct()

    if folio_op:
        clean_op = folio_op.upper().replace('OP-', '').replace('OP', '').strip()
        ordenes_qs = ordenes_qs.filter(id__icontains=clean_op)

    if producto_id and producto_id != 'all':
        ordenes_qs = ordenes_qs.filter(producto_id=producto_id)

    if estado:
        ordenes_qs = ordenes_qs.filter(estado=estado)
        
    if sucursal_id:
        ordenes_qs = ordenes_qs.filter(sucursal_id=sucursal_id)

    # Para el buscador visual de producto
    producto_nombre_display = ""
    if producto_id and producto_id != 'all':
        try:
            prod_obj = Producto.objects.get(id=producto_id, empresa=empresa_actual)
            producto_nombre_display = prod_obj.nombre
        except:
            pass
            
    from preferencias.models import Sucursal
    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')

    filtros = {
        'q': q,
        'folio_op': folio_op,
        'producto_id': producto_id,
        'producto_nombre': producto_nombre_display,
        'estado': estado,
        'sucursal': sucursal_id or ''
    }
    # --- FIN LÓGICA DE FILTRADO ---

    # PAGINACIÓN
    paginator = Paginator(ordenes_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # AGREGAMOS LOS CÁLCULOS DE AVANCE A CADA ORDEN (EN EL PAGE_OBJ)
    mapa_estados = {
        'borrador': 0,
        'en_proceso': 33,
        'testeo': 66,
        'terminado': 100,
        'cancelada': 0
    }

    for o in page_obj:
        # 1. Avance por Estado
        o.porcentaje_estado = mapa_estados.get(o.estado, 0)
        
        # 2. Avance por Calidad (Tareas)
        o.porcentaje_calidad = 0
        if o.producto.test_calidad:
            total_tareas = o.producto.test_calidad.items.count()
            if total_tareas > 0:
                completadas = o.resultados_test.filter(completado=True).count()
                o.porcentaje_calidad = int((completadas / total_tareas) * 100)
        else:
            o.porcentaje_calidad = 100 if o.estado == 'terminado' else 0
            
        o.faltantes_info = []
        if o.estado == 'borrador':
            # --- LÓGICA FLEXIBLE DE VALIDACIÓN DE STOCK ---
            # Si el usuario eligió un almacén de MP, validamos SOLO ese.
            # Si no, validamos en toda la sucursal de la orden.
            target_almacen = o.almacen_materia_prima
            target_sucursal = o.sucursal
            
            for det in o.detalles.all():
                inv_qs = Inventario.objects.filter(producto=det.producto)
                
                if target_almacen:
                    inv_qs = inv_qs.filter(almacen=target_almacen)
                elif target_sucursal:
                    inv_qs = inv_qs.filter(sucursal=target_sucursal)
                else:
                    inv_qs = inv_qs.filter(empresa=empresa_actual)
                
                # Sumar existencias disponibles
                from django.db.models.functions import Coalesce
                stock_data = inv_qs.aggregate(
                    total_disp=Coalesce(Sum(F('cantidad') - Coalesce(F('reservado'), 0)), 0)
                )
                disponible = stock_data['total_disp']
                
                if disponible < det.cantidad:
                    o.faltantes_info.append({
                        'nombre': det.producto.nombre,
                        'cantidad': float(det.cantidad - disponible)
                    })

    # PRODUCTOS QUE SE PUEDEN PRODUCIR
    productos_finales = Producto.objects.filter(empresa=empresa_actual, tipo_abastecimiento='produccion')
    
    # --- FILTRADO DE ALMACENES POR SUCURSAL ---
    almacenes = Almacen.objects.filter(empresa=empresa_actual)
    if sucursal_id:
        almacenes = almacenes.filter(sucursal_id=sucursal_id)
    
    todos_productos_qs = Producto.objects.filter(empresa=empresa_actual)
    todos_productos_json = list(todos_productos_qs.values('id', 'nombre'))
    
    from clientes.models import Cliente
    clientes = Cliente.objects.filter(empresa=empresa_actual)
    
    contexto = {
        'page_obj': page_obj,
        'productos_finales': productos_finales,
        'almacenes': almacenes,
        'sucursales': sucursales,
        'clientes': clientes,
        'productos_json': todos_productos_json,
        'todos_productos_qs': todos_productos_qs,
        'section': 'produccion',
        'filtros': filtros
    }
    return render(request, 'produccion/dashboard_produccion.html', contexto)

# ==========================================
# VISTAS PARA CATÁLOGO DE TESTS
# ==========================================

from .models import Test, ItemTest

@login_required(login_url='/login/')
@require_production_permission('catalogos_test', 'ver')
def lista_tests(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)
    
    tests = Test.objects.filter(empresa=empresa_actual).prefetch_related('items')
    
    contexto = {
        'tests': tests,
        'section': 'produccion_tests'
    }
    return render(request, 'produccion/dashboard_tests.html', contexto)

@login_required
def api_obtener_test_orden(request, orden_id):
    """Jala las tareas del test y los resultados que ya se hayan guardado"""
    empresa_actual = get_empresa_actual(request)
    orden = get_object_or_404(OrdenProduccion, id=orden_id, empresa=empresa_actual)
    producto = orden.producto
    
    if not producto.test_calidad:
        return JsonResponse({'has_test': False})
    
    test = producto.test_calidad
    
    # Jalar resultados ya guardados para esta orden
    from .models import ResultadoTestOP
    resultados_dict = {r.item_test_id: {
        'completado': r.completado,
        'usuario': r.usuario_verifico.username.split('@')[0] if r.usuario_verifico else '?',
        'fecha': r.fecha_chequeo.strftime('%d/%m %H:%M') if r.fecha_chequeo else ''
    } for r in ResultadoTestOP.objects.filter(orden_produccion=orden)}

    tareas = []
    for item in test.items.all():
        res = resultados_dict.get(item.id, {'completado': False, 'usuario': '', 'fecha': ''})
        tareas.append({
            'id': item.id,
            'tarea': item.tarea,
            'completado': res['completado'],
            'usuario': res['usuario'],
            'fecha': res['fecha']
        })
        
    return JsonResponse({
        'has_test': True,
        'test_nombre': test.nombre,
        'tareas': tareas
    })

@login_required
@transaction.atomic
@require_production_permission('tablero_control', 'guardar_avance', json_response=True)
def guardar_avance_test_ajax(request, orden_id):
    """Guarda los checks sin finalizar la orden"""
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            orden = get_object_or_404(OrdenProduccion, id=orden_id, empresa=empresa_actual)
            data = json.loads(request.body)
            checks_enviados = [int(id) for id in data.get('checks', [])]

            from .models import ResultadoTestOP, ItemTest
            
            # 1. Marcar como completados los que vienen en la lista
            for item_id in checks_enviados:
                ResultadoTestOP.objects.update_or_create(
                    orden_produccion=orden,
                    item_test_id=item_id,
                    defaults={
                        'completado': True,
                        'fecha_chequeo': timezone.now(),
                        'usuario_verifico': request.user
                    }
                )
            
            # 2. Desmarcar los que NO vienen (por si alguien quita un check)
            test_id = orden.producto.test_calidad_id
            todas_tareas_ids = ItemTest.objects.filter(test_id=test_id).values_list('id', flat=True)
            ids_a_desmarcar = [id for id in todas_tareas_ids if id not in checks_enviados]
            
            ResultadoTestOP.objects.filter(orden_produccion=orden, item_test_id__in=ids_a_desmarcar).delete()

            return JsonResponse({'success': True, 'message': 'Avance guardado mi chingon.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@transaction.atomic
def finalizar_con_test_ajax(request, orden_id):
    """Guarda checks y finaliza orden vía AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        empresa_actual = get_empresa_actual(request)
        orden = get_object_or_404(OrdenProduccion, id=orden_id, empresa=empresa_actual)
        
        data = json.loads(request.body)
        checks = data.get('checks', [])

        # 1. Calidad
        from .models import ResultadoTestOP
        for item_id in checks:
            ResultadoTestOP.objects.update_or_create(
                orden_produccion=orden,
                item_test_id=item_id,
                defaults={'completado': True, 'fecha_chequeo': timezone.now(), 'usuario_verifico': request.user}
            )

        # 2. Finalizar
        res = finalizar_produccion_logica(request, orden)
        return JsonResponse({'success': res.get('success', False), 'message': res.get('message', ''), 'error': res.get('error', '')})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@transaction.atomic
@require_production_permission('catalogos_test', 'crear', json_response=True)
def crear_test_ajax(request):
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            data = json.loads(request.body)
            
            nombre = data.get('nombre')
            descripcion = data.get('descripcion', '')
            tareas = data.get('tareas', [])
            
            if not nombre:
                return JsonResponse({'success': False, 'error': 'El nombre del test es obligatorio.'})
            
            # 1. Crear Cabecera
            nuevo_test = Test.objects.create(
                empresa=empresa_actual,
                nombre=nombre,
                descripcion=descripcion
            )
            
            # 2. Crear Tareas
            for i, t in enumerate(tareas):
                if t.strip():
                    ItemTest.objects.create(
                        test=nuevo_test,
                        tarea=t.strip(),
                        orden=i
                    )
            
            return JsonResponse({'success': True, 'message': 'Test guardado correctamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
def api_detalle_test(request, test_id):
    """Devuelve los datos de un test y sus tareas para edición"""
    try:
        empresa_actual = get_empresa_actual(request)
        test = get_object_or_404(Test, id=test_id, empresa=empresa_actual)
        
        tareas = list(test.items.all().values('tarea', 'orden'))
        
        return JsonResponse({
            'success': True,
            'id': test.id,
            'nombre': test.nombre,
            'descripcion': test.descripcion,
            'tareas': tareas
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@transaction.atomic
@require_production_permission('catalogos_test', 'editar', json_response=True)
def actualizar_test_ajax(request, test_id):
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            test = get_object_or_404(Test, id=test_id, empresa=empresa_actual)
            
            data = json.loads(request.body)
            test.nombre = data.get('nombre')
            test.descripcion = data.get('descripcion', '')
            test.save()
            
            # Re-crear tareas (borramos las anteriores y ponemos las nuevas)
            test.items.all().delete()
            for i, t in enumerate(data.get('tareas', [])):
                if t.strip():
                    ItemTest.objects.create(test=test, tarea=t.strip(), orden=i)
            
            return JsonResponse({'success': True, 'message': 'Test actualizado correctamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@transaction.atomic
@require_production_permission('catalogos_test', 'eliminar', json_response=True)
def eliminar_test_ajax(request, test_id):
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            test = get_object_or_404(Test, id=test_id, empresa=empresa_actual)
            
            # Verificamos si algún producto lo está usando antes de borrar
            from core.models import Producto
            if Producto.objects.filter(test_calidad=test).exists():
                return JsonResponse({'success': False, 'error': 'No se puede eliminar porque hay productos que usan este test.'})
            
            test.delete()
            return JsonResponse({'success': True, 'message': 'Test eliminado.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
def api_obtener_receta(request, producto_id):
    """Devuelve los componentes de un producto final"""
    empresa_actual = get_empresa_actual(request)
    producto = get_object_or_404(Producto, id=producto_id, empresa=empresa_actual)
    
    receta = DetalleReceta.objects.filter(producto_padre=producto).select_related('componente')
    items = []
    for r in receta:
        items.append({
            'id': r.componente.id,
            'nombre': r.componente.nombre,
            'cantidad': float(r.cantidad)
        })
    
    return JsonResponse({'success': True, 'producto': producto.nombre, 'items': items})

@login_required
@transaction.atomic
@require_production_permission('tablero_control', 'crear')
def crear_orden_produccion(request):
    """Crea órdenes de producción desde el nuevo modal multicartas"""
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            cliente_id = request.POST.get('cliente')
            almacen_mp_id = request.POST.get('almacen_salida')
            almacen_pt_id = request.POST.get('almacen_entrada')
            
            # Listas de artículos
            productos_ids = request.POST.getlist('producto_id[]')
            cantidades = request.POST.getlist('cantidad[]')
            
            if not productos_ids:
                raise Exception("Debes agregar al menos un artículo a producir.")

            count = 0
            for i in range(len(productos_ids)):
                p_id = productos_ids[i]
                qty = int(cantidades[i])
                
                if not p_id or qty <= 0: continue
                
                producto = get_object_or_404(Producto, id=p_id, empresa=empresa_actual)
                
                # --- OBTENER SUCURSAL DE SESIÓN ---
                from preferencias.models import Sucursal
                sucursal_id = request.session.get('sucursal_id')
                sucursal_obj = None
                if sucursal_id:
                    try:
                        sucursal_obj = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
                    except Sucursal.DoesNotExist:
                        pass

                # 1. Crear OP
                op = OrdenProduccion.objects.create(
                    empresa=empresa_actual,
                    cliente_id=cliente_id if cliente_id else None,
                    producto=producto,
                    cantidad=qty,
                    almacen_id=almacen_pt_id,
                    almacen_materia_prima_id=almacen_mp_id,
                    solicitante=request.user,
                    estado='borrador',
                    sucursal=sucursal_obj
                )
                
                # 2. Copiar Componentes (Receta)
                receta_base = DetalleReceta.objects.filter(producto_padre=producto)
                for r in receta_base:
                    DetalleOrdenProduccion.objects.create(
                        orden_produccion=op,
                        producto=r.componente,
                        cantidad=r.cantidad * qty
                    )
                count += 1
                
            messages.success(request, f'Se han generado {count} órdenes de producción exitosamente.')
            return redirect('dashboard_produccion')
        except Exception as e:
            messages.error(request, f'Error al crear: {str(e)}')
            return redirect('dashboard_produccion')
    return redirect('dashboard_produccion')

@login_required
@require_production_permission('tablero_control', 'ver', json_response=True)
def api_detalle_orden(request, orden_id):
    """Retorna datos de la orden y sus componentes para el modal de visualización"""
    empresa_actual = get_empresa_actual(request)
    orden = get_object_or_404(OrdenProduccion, id=orden_id, empresa=empresa_actual)
    
    detalles = []
    for d in orden.detalles.all():
        detalles.append({
            'id': d.id,
            'producto_id': d.producto.id,
            'producto_nombre': d.producto.nombre,
            'cantidad': float(d.cantidad)
        })
    
    # Cálculos de avance (Misma lógica que en dashboard_produccion)
    mapa_estados = {
        'borrador': 0,
        'en_proceso': 33,
        'testeo': 66,
        'terminado': 100,
        'cancelada': 0
    }
    porcentaje_estado = mapa_estados.get(orden.estado, 0)
    
    porcentaje_calidad = 0
    if orden.producto.test_calidad:
        total_tareas = orden.producto.test_calidad.items.count()
        if total_tareas > 0:
            completadas = orden.resultados_test.filter(completado=True).count()
            porcentaje_calidad = int((completadas / total_tareas) * 100)
    else:
        porcentaje_calidad = 100 if orden.estado == 'terminado' else 0
        
    data = {
        'id': orden.id,
        'folio': orden.folio,
        'cliente_nombre': orden.cliente.nombre_completo if orden.cliente else 'Sin cliente',
        'fecha_op': orden.fecha_creacion.strftime('%d/%m/%Y'),
        'almacen_entrada_id': orden.almacen.id,
        'almacen_entrada': orden.almacen.name if hasattr(orden.almacen, 'name') else orden.almacen.nombre,
        'almacen_salida_id': orden.almacen_materia_prima.id if orden.almacen_materia_prima else '',
        'almacen_salida': orden.almacen_materia_prima.nombre if orden.almacen_materia_prima else 'No asignado',
        'pedido_id': orden.pedido_origen.id if orden.pedido_origen else 'Manual',
        'pedido_folio': f"PED-{orden.pedido_origen.id:04d}" if orden.pedido_origen else 'Manual',
        'pedido_fecha': orden.pedido_origen.fecha_creacion.strftime('%d/%m/%Y') if orden.pedido_origen else '--',
        'producto_id': orden.producto.id,
        'producto_nombre': orden.producto.nombre,
        'maneja_lote': orden.producto.maneja_lote,
        'maneja_serie': orden.producto.maneja_serie,
        'cantidad_producir': orden.cantidad,
        'solicitante': orden.solicitante.username.split('@')[0] if orden.solicitante else 'Sistema',
        'estado': orden.estado,
        'estado_display': orden.get_estado_display(),
        'porcentaje_estado': porcentaje_estado,
        'porcentaje_calidad': porcentaje_calidad,
        'notas': orden.notas,
        'detalles': detalles
    }
    return JsonResponse(data)

@login_required
@transaction.atomic
@require_production_permission('tablero_control', 'finalizar_trabajo', json_response=True)
def finalizar_produccion_completo(request):
    """Procesa el ingreso a inventario con series/lotes y finaliza la OP"""
    if request.method == 'POST':
        try:
            orden_id = request.POST.get('orden_id')
            almacen_id = request.POST.get('almacen_id')
            empresa_actual = get_empresa_actual(request)
            orden = get_object_or_404(OrdenProduccion, id=orden_id, empresa=empresa_actual)
            almacen_destino = get_object_or_404(Almacen, id=almacen_id, empresa=empresa_actual)
            
            # 1. Validar que no esté ya terminada
            if orden.estado == 'terminado':
                return JsonResponse({'success': False, 'error': 'Esta orden ya fue finalizada.'})

            # Actualizamos el almacén si el usuario lo cambió en el modal
            if int(almacen_id) != orden.almacen.id:
                orden.almacen = almacen_destino
                orden.save()

            # 2. Lógica de trazabilidad (Series / Lotes)
            from recepciones.models import DetalleRecepcionExtra
            
            lote_global = request.POST.get('lote_global')
            series = request.POST.getlist('serie[]')
            
            referencia_extra = ""
            if lote_global:
                DetalleRecepcionExtra.objects.create(
                    producto=orden.producto,
                    tipo='lote',
                    lote=lote_global,
                    cantidad_lote=orden.cantidad,
                    almacen=almacen_destino
                )
                referencia_extra = f" | Lote: {lote_global}"
            elif any(series):
                for s in series:
                    if s.strip():
                        DetalleRecepcionExtra.objects.create(
                            producto=orden.producto,
                            tipo='serie',
                            serie=s.strip(),
                            cantidad_lote=1,
                            almacen=almacen_destino
                        )
                series_limpias = [s for s in series if s.strip()]
                if series_limpias:
                    referencia_extra = f" | Series: {', '.join(series_limpias[:3])}..."

            # 3. Ejecutar la lógica de finalización estándar (Descontar materia prima y sumar PT)
            res = finalizar_produccion_logica(request, orden)
            
            if res.get('success'):
                messages.success(request, f'Orden {orden.folio} terminada y mercancía ingresada.{referencia_extra}', extra_tags='produccion')
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': res.get('error')})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@transaction.atomic
@require_production_permission('tablero_control', 'editar', json_response=True)
def actualizar_orden_produccion(request, orden_id):
    """Guarda los cambios realizados en el modal de edición"""
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            orden = get_object_or_404(OrdenProduccion, id=orden_id, empresa=empresa_actual)
            
            if orden.estado != 'borrador':
                return JsonResponse({'success': False, 'error': 'Solo se pueden editar órdenes en Borrador.'})

            # 1. Actualizar Notas, Cantidad, Producto Principal y Almacenes
            data = json.loads(request.body)
            orden.notas = data.get('notas', '')
            if 'cantidad_padre' in data:
                orden.cantidad = int(data['cantidad_padre'])
            if 'producto_id' in data:
                orden.producto_id = int(data['producto_id'])
            
            # NUEVO: Actualizar Almacenes
            if 'almacen_entrada' in data:
                orden.almacen_id = data['almacen_entrada']
            if 'almacen_salida' in data:
                orden.almacen_materia_prima_id = data['almacen_salida'] or None
            
            orden.save()

            # 2. Actualizar Componentes
            # Borramos los actuales y recreamos según lo editado
            orden.detalles.all().delete()
            
            for item in data.get('componentes', []):
                DetalleOrdenProduccion.objects.create(
                    orden_produccion=orden,
                    producto_id=item['producto_id'],
                    cantidad=item['cantidad']
                )

            return JsonResponse({'success': True, 'message': 'Orden actualizada correctamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@transaction.atomic
def avanzar_estado_produccion(request, orden_id):
    """Mueve la orden al siguiente paso de la cadena con validaciones"""
    empresa_actual = get_empresa_actual(request)
    orden = get_object_or_404(OrdenProduccion, id=orden_id, empresa=empresa_actual)
    
    nuevo_estado = request.GET.get('estado')
    
    # 1. BORRADOR -> EN PROCESO (CON RESERVA DE STOCK)
    if nuevo_estado == 'en_proceso' and orden.estado == 'borrador':
        if not user_has_production_permission(request, 'tablero_control', 'iniciar_trabajo'):
             messages.error(request, "No cuentas con permiso para iniciar trabajos.")
             return redirect('dashboard_produccion')
        try:
            detalles = orden.detalles.all()
            if not detalles.exists():
                raise Exception("La orden no tiene componentes asignados.")

            # Bloqueamos y validamos stock
            # Si no hay almacén de MP, validamos en toda la sucursal
            almacen_especifico = orden.almacen_materia_prima
            sucursal_orden = orden.sucursal
            
            faltantes = []
            with transaction.atomic():
                for d in detalles:
                    # 1. Definir QuerySet de búsqueda
                    inv_qs = Inventario.objects.select_for_update().filter(producto=d.producto)
                    if almacen_especifico:
                        inv_qs = inv_qs.filter(almacen=almacen_especifico)
                    elif sucursal_orden:
                        inv_qs = inv_qs.filter(sucursal=sucursal_orden)
                    else:
                        inv_qs = inv_qs.filter(empresa=empresa_actual)

                    # 2. Calcular disponibilidad
                    from django.db.models.functions import Coalesce
                    stock_data = inv_qs.aggregate(
                        total_disp=Coalesce(Sum(F('cantidad') - Coalesce(F('reservado'), 0)), 0)
                    )
                    disponible = stock_data['total_disp']
                    
                    if disponible < d.cantidad:
                        faltante_cant = d.cantidad - disponible
                        faltantes.append(f"{d.producto.nombre} ({int(faltante_cant)} pz)")
                    
                if faltantes:
                    ubicacion_nombre = almacen_especifico.nombre if almacen_especifico else (sucursal_orden.nombre if sucursal_orden else "la empresa")
                    msg = f"Faltan piezas en {ubicacion_nombre} para la producción {orden.folio}: " + ", ".join(faltantes)
                    messages.error(request, msg)
                    return redirect('dashboard_produccion')

                # Si no hay faltantes, procedemos a reservar
                for d in detalles:
                    # Si se especificó almacén, reservamos ahí. 
                    # Si no, buscamos el primer almacén con stock en la sucursal para reservar
                    if almacen_especifico:
                        target_inv = Inventario.objects.filter(producto=d.producto, almacen=almacen_especifico).first()
                        if not target_inv:
                            target_inv = Inventario.objects.create(
                                producto=d.producto, almacen=almacen_especifico, 
                                cantidad=0, empresa=empresa_actual, sucursal=almacen_especifico.sucursal
                            )
                    else:
                        # Buscamos en qué almacén de la sucursal hay stock para reservar
                        # Priorizamos el almacén de PT si tiene stock, si no el primero que encontremos
                        target_inv = Inventario.objects.filter(
                            producto=d.producto, sucursal=sucursal_orden, cantidad__gt=0
                        ).first() or Inventario.objects.filter(
                            producto=d.producto, almacen=orden.almacen
                        ).first()
                        
                        if not target_inv:
                             target_inv = Inventario.objects.create(
                                producto=d.producto, almacen=orden.almacen, 
                                cantidad=0, empresa=empresa_actual, sucursal=orden.almacen.sucursal
                            )
                    
                    target_inv.reservado = F('reservado') + d.cantidad
                    target_inv.save()

            # Si todo bien, iniciamos
            orden.estado = 'en_proceso'
            orden.fecha_inicio = timezone.now()
            orden.responsable = request.user
            orden.save()
            
            crear_notificacion(
                empresa=empresa_actual,
                actor=request.user,
                mensaje=f'inició la producción de {orden.folio} ({orden.producto.nombre})',
                propietario=orden.solicitante
            )
            
            messages.success(request, f'Orden {orden.folio} iniciada. Los materiales han sido reservados en bodega.')
        except Exception as e:
            messages.error(request, str(e))
    
    # 2. EN PROCESO -> TESTEO
    elif nuevo_estado == 'testeo' and orden.estado == 'en_proceso':
        if not user_has_production_permission(request, 'tablero_control', 'enviar_testeo'):
             messages.error(request, "No cuentas con permiso para enviar a testeo.")
             return redirect('dashboard_produccion')
        orden.estado = 'testeo'
        orden.save()
        
        crear_notificacion(
            empresa=empresa_actual,
            actor=request.user,
            mensaje=f'envió {orden.folio} a Testeo / Calidad',
            propietario=orden.solicitante
        )
        
        messages.info(request, f'Orden {orden.folio} enviada a Testeo / Calidad.')
        
    # 3. TESTEO -> TERMINADO
    elif nuevo_estado == 'terminado' and orden.estado == 'testeo':
        # --- CANDADO DE SEGURIDAD ---
        if orden.producto.test_calidad:
            return JsonResponse({'success': False, 'error': 'Esta orden requiere validación de checklist.'}, status=400)
        else:
            res = finalizar_produccion_logica(request, orden)
            if res.get('success'):
                crear_notificacion(
                    empresa=empresa_actual,
                    actor=request.user,
                    mensaje=f'finalizó el trabajo {orden.folio}',
                    propietario=orden.solicitante
                )
                messages.success(request, res['message'])
            else:
                messages.error(request, res.get('error'))

    return redirect('dashboard_produccion')

def finalizar_produccion_logica(request, orden):
    """Función interna para cerrar la orden, descontar físico y reserva, y sumar producto terminado"""
    try:
        producto = orden.producto
        almacen_destino = orden.almacen  # PT
        almacen_origen = orden.almacen_materia_prima or orden.almacen # MP
        cantidad_producir = orden.cantidad

        detalles_orden = orden.detalles.all()

        # 1. Ejecutar Movimientos de Salida (Limpiar Reserva y Quitar Físico)
        for det in detalles_orden:
            # Usamos el método centralizado para que se registre en el Kardex
            Inventario.registrar_salida(
                almacen=almacen_origen,
                producto=det.producto,
                cantidad_salida=det.cantidad,
                referencia=f"OP-{orden.id:04d} (Consumo)"
            )
            # Limpiamos la reserva manualmente ya que registrar_salida no toca 'reservado'
            Inventario.objects.filter(producto=det.producto, almacen=almacen_origen).update(
                reservado=F('reservado') - det.cantidad
            )

        # 2. Sumar Producto Terminado
        Inventario.registrar_ingreso(
            almacen=almacen_destino,
            producto=producto,
            cantidad_ingreso=cantidad_producir,
            costo_unitario=producto.precio_costo, # O el costo calculado de la receta
            referencia=f"OP-{orden.id:04d} (Ensamble)"
        )

        # 5. Finalizar Orden
        orden.estado = 'terminado'
        orden.fecha_terminado = timezone.now()
        orden.save()

        # 6. Sincronizar Pedido
        if orden.pedido_origen:
            from pedidos.models import DetallePedido
            detalles_pedido = DetallePedido.objects.filter(pedido=orden.pedido_origen, producto=producto)
            for dp in detalles_pedido:
                if dp.estado_linea != 'completo':
                    dp.estado_linea = 'pendiente'
                    dp.save()

        return {'success': True, 'message': f'Orden {orden.folio} terminada exitosamente.'}
    except Exception as e:
        return {'success': False, 'error': str(e)}

@login_required
@require_production_permission('tablero_control', 'cancelar_orden')
def cancelar_produccion(request, orden_id):
    empresa_actual = get_empresa_actual(request)
    orden = get_object_or_404(OrdenProduccion, id=orden_id, empresa=empresa_actual)
    
    if orden.estado != 'terminado' and orden.estado != 'cancelada':
        # Si la orden ya estaba en proceso o testeo, tenía materiales reservados
        if orden.estado in ['en_proceso', 'testeo']:
            almacen_origen = orden.almacen_materia_prima or orden.almacen
            for det in orden.detalles.all():
                Inventario.objects.filter(
                    producto=det.producto, 
                    almacen=almacen_origen
                ).update(reservado=F('reservado') - det.cantidad)

        orden.estado = 'cancelada'
        orden.save()
        messages.success(request, f'Orden {orden.folio} cancelada y reservas liberadas.')
    return redirect('dashboard_produccion')

@login_required
@require_production_permission('tablero_control', 'imprimir')
def imprimir_orden_produccion(request, pk):
    """Genera la vista para impresión de orden de producción (PDF)"""
    empresa_actual = get_empresa_actual(request)
    orden = get_object_or_404(OrdenProduccion, id=pk, empresa=empresa_actual)
    
    # Limpiar nombre del solicitante
    solicitante_nombre = orden.solicitante.get_full_name() if orden.solicitante else 'Sistema'
    if not solicitante_nombre and orden.solicitante:
        solicitante_nombre = orden.solicitante.username.split('@')[0]
        
    # Limpiar nombre del responsable
    responsable_nombre = orden.responsable.get_full_name() if orden.responsable else '--'
    if not responsable_nombre and orden.responsable:
        responsable_nombre = orden.responsable.username.split('@')[0]

    context = {
        'orden': orden,
        'empresa': empresa_actual,
        'solicitante_nombre': solicitante_nombre,
        'responsable_nombre': responsable_nombre,
    }
    return render(request, 'produccion/imprimir_op.html', context)
@login_required
def api_stock_almacen(request, almacen_id):
    """Retorna el stock disponible de todos los productos en un almacén específico"""
    empresa_actual = get_empresa_actual(request)
    almacen = get_object_or_404(Almacen, id=almacen_id, empresa=empresa_actual)
    
    # Obtenemos todos los productos de la empresa
    productos = Producto.objects.filter(empresa=empresa_actual)
    
    # Obtenemos los inventarios para este almacén
    inventarios = Inventario.objects.filter(almacen=almacen).values('producto_id', 'cantidad', 'reservado')
    stock_dict = {i['producto_id']: i['cantidad'] - (i['reservado'] or 0) for i in inventarios}
    
    data = []
    for p in productos:
        data.append({
            'id': p.id,
            'nombre': p.nombre,
            'stock': stock_dict.get(p.id, 0)
        })
        
    return JsonResponse(data, safe=False)
