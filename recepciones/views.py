from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum, F
import openpyxl
from datetime import datetime
from .models import Recepcion, DetalleRecepcion, DetalleRecepcionExtra
from compras.models import OrdenCompra, DetalleCompra
from almacenes.models import Almacen, Inventario
from .services import procesar_recepcion_servicio
from panel.models import Empresa
from notificaciones.utils import crear_notificacion

@login_required(login_url='/login/')
def descargar_plantilla_recepciones(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Recepciones"
    
    # Encabezados: Referencia ayuda a agrupar productos en una misma recepción
    headers = [
        'Referencia (Ej: REC-001)', 'Orden Compra (ID)', 'Almacen (Nombre o ID)',
        'Fecha (YYYY-MM-DD)', 'Factura', 'Producto (Nombre o ID)', 'Cantidad Recibida'
    ]
    
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Recepciones.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
def importar_recepciones_ajax(request):
    if request.method == 'POST' and request.FILES.get('archivo_recepciones'):
        empresa_actual = get_empresa_actual(request)
        excel_file = request.FILES['archivo_recepciones']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return JsonResponse({'success': False, 'error': 'El archivo está vacío.'})
            
            data_rows = rows[1:]
            
            # Agrupar por Referencia
            recepciones_data = {}
            for idx, row in enumerate(data_rows, start=2):
                if not any(row): continue
                ref = str(row[0] or f"TEMP-{idx}").strip()
                if ref not in recepciones_data:
                    recepciones_data[ref] = {
                        'oc_input': str(row[1] or '').strip(),
                        'alm_input': str(row[2] or '').strip(),
                        'fecha_raw': row[3],
                        'factura': str(row[4] or '').strip(),
                        'items': []
                    }
                recepciones_data[ref]['items'].append({
                    'prod_input': str(row[5] or '').strip(),
                    'cantidad': row[6] or 0
                })

            creadas = 0
            errores = []
            
            with transaction.atomic():
                # Obtener sucursal de la sesión
                sucursal_obj = None
                sucursal_id = request.session.get('sucursal_id')
                if sucursal_id:
                    from preferencias.models import Sucursal
                    try:
                        sucursal_obj = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
                    except Sucursal.DoesNotExist:
                        pass

                for ref, data in recepciones_data.items():
                    try:
                        # Buscar Orden de Compra
                        orden = None
                        if data['oc_input'].isdigit():
                            orden = OrdenCompra.objects.filter(id=int(data['oc_input']), empresa=empresa_actual).first()
                        if not orden:
                            errores.append(f"Ref {ref}: No se encontró la OC '{data['oc_input']}'.")
                            continue

                        # Buscar Almacén
                        almacen = None
                        if data['alm_input'].isdigit():
                            almacen = Almacen.objects.filter(id=int(data['alm_input']), empresa=empresa_actual).first()
                        if not almacen:
                            almacen = Almacen.objects.filter(nombre__icontains=data['alm_input'], empresa=empresa_actual).first()
                        
                        if not almacen:
                            errores.append(f"Ref {ref}: No se encontró el almacén '{data['alm_input']}'.")
                            continue

                        # Procesar Fecha
                        fecha_rec = timezone.now().date()
                        if data['fecha_raw']:
                            if isinstance(data['fecha_raw'], datetime):
                                fecha_rec = data['fecha_raw'].date()
                            elif isinstance(data['fecha_raw'], str):
                                try: fecha_rec = datetime.strptime(data['fecha_raw'], '%Y-%m-%d').date()
                                except: pass

                        # Crear Recepción
                        recepcion = Recepcion.objects.create(
                            empresa=empresa_actual,
                            orden_compra=orden,
                            almacen=almacen,
                            sucursal=sucursal_obj,
                            fecha=fecha_rec,
                            factura=data['factura'],
                            estado='completada'
                        )

                        # Crear Detalles
                        for item in data['items']:
                            from core.models import Producto
                            producto = None
                            if item['prod_input'].isdigit():
                                producto = Producto.objects.filter(id=int(item['prod_input']), empresa=empresa_actual).first()
                            if not producto:
                                producto = Producto.objects.filter(nombre__icontains=item['prod_input'], empresa=empresa_actual).first()
                            
                            if producto:
                                # Buscar el detalle de compra correspondiente para vincular
                                det_compra = DetalleCompra.objects.filter(orden_compra=orden, producto=producto).first()
                                
                                DetalleRecepcion.objects.create(
                                    recepcion=recepcion,
                                    producto=producto,
                                    detalle_compra=det_compra,
                                    cantidad_recibida=item['cantidad'],
                                    costo_unitario=det_compra.precio_costo if det_compra else (producto.precio_costo or 0)
                                )
                        
                        creadas += 1
                    except Exception as e:
                        errores.append(f"Ref {ref}: {str(e)}")
            
            if creadas == 0 and errores:
                return JsonResponse({'success': False, 'error': f"Error: {errores[0]}"})

            msg = f'Importación exitosa. {creadas} recepciones creadas.'
            if errores: msg += f' ({len(errores)} errores)'
            return JsonResponse({'success': True, 'message': msg})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
            
    return JsonResponse({'success': False, 'error': 'Solicitud inválida.'})

@login_required(login_url='/login/')
def exportar_recepciones_excel(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return HttpResponse("No autorizado", status=403)

    # --- Filtrado ---
    from django.db.models import Q
    q = request.GET.get('q', '')
    folio_recepcion = request.GET.get('folio_recepcion', '')
    folio_oc = request.GET.get('folio_oc', '')
    proveedor_id = request.GET.get('proveedor_id', '')
    almacen_id = request.GET.get('almacen_id', '')
    fecha = request.GET.get('fecha', '')
    estado = request.GET.get('estado', '')
    sucursal_id_filtro = request.GET.get('sucursal', '')

    recepciones = Recepcion.objects.filter(empresa=empresa_actual).select_related('orden_compra', 'almacen', 'orden_compra__proveedor', 'sucursal').order_by('-fecha', '-id')

    if q:
        recepciones = recepciones.filter(
            Q(id__icontains=q) | Q(orden_compra__id__icontains=q) |
            Q(orden_compra__proveedor__razon_social__icontains=q) | Q(factura__icontains=q)
        )
    if folio_recepcion:
        recepciones = recepciones.filter(id__icontains=folio_recepcion.replace('REC-', ''))
    if folio_oc:
        recepciones = recepciones.filter(orden_compra__id__icontains=folio_oc.replace('OC-', ''))
    if proveedor_id and proveedor_id != 'all':
        recepciones = recepciones.filter(orden_compra__proveedor_id=proveedor_id)
    if almacen_id:
        recepciones = recepciones.filter(almacen_id=almacen_id)
    if fecha:
        try: recepciones = recepciones.filter(fecha=fecha)
        except: pass
    if estado:
        recepciones = recepciones.filter(estado=estado)
    if sucursal_id_filtro:
        recepciones = recepciones.filter(sucursal_id=sucursal_id_filtro)

    # --- Generar Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Recepciones"

    headers = [
        'ID', 'Folio', 'OC Origen', 'Proveedor', 'Almacén', 'Fecha', 'Factura', 'Estado', 'Total'
    ]
    ws.append(headers)

    for r in recepciones:
        ws.append([
            r.id, f"REC-{r.id:04d}", f"OC-{r.orden_compra.id:04d}",
            r.orden_compra.proveedor.razon_social, r.almacen.nombre,
            r.fecha.strftime('%Y-%m-%d'), r.factura, r.get_estado_display(), r.total
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Recepciones_{empresa_actual.nombre}.xlsx"'
    wb.save(response)
    return response

# --- 1. FUNCIÓN AYUDANTE ESTÁNDAR ---
def get_empresa_actual(request):
    username = request.user.username
    if '@' in username:
        subdominio = username.split('@')[1]
        try:
            return Empresa.objects.get(subdominio=subdominio)
        except Empresa.DoesNotExist:
            return None
    return None

from django.db.models import Q
from proveedores.models import Proveedor

@login_required
def dashboard_recepciones(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)

    # --- LÓGICA DE FILTRADO ---
    q = request.GET.get('q', '')
    folio_recepcion = request.GET.get('folio_recepcion', '')
    folio_oc = request.GET.get('folio_oc', '')
    proveedor_id = request.GET.get('proveedor_id', '')
    almacen_id = request.GET.get('almacen_id', '')
    fecha = request.GET.get('fecha', '')
    estado = request.GET.get('estado', '')
    sucursal_id_filtro = request.GET.get('sucursal', '')

    recepciones_list = Recepcion.objects.filter(empresa=empresa_actual).select_related('orden_compra', 'almacen', 'orden_compra__proveedor', 'sucursal').order_by('-fecha', '-id')

    if q:
        recepciones_list = recepciones_list.filter(
            Q(id__icontains=q) |
            Q(orden_compra__id__icontains=q) |
            Q(orden_compra__proveedor__razon_social__icontains=q) |
            Q(factura__icontains=q) |
            Q(pedimento__icontains=q) |
            Q(aduana__icontains=q)
        )
    if folio_recepcion:
        clean_rec = folio_recepcion.upper().replace('REC-', '').replace('REC', '').strip()
        recepciones_list = recepciones_list.filter(id__icontains=clean_folio)
    if folio_oc:
        clean_oc = folio_oc.upper().replace('OC-', '').replace('OC', '').strip()
        recepciones_list = recepciones_list.filter(orden_compra__id__icontains=clean_oc)
    if proveedor_id and proveedor_id != 'all':
        recepciones_list = recepciones_list.filter(orden_compra__proveedor_id=proveedor_id)
    if almacen_id:
        recepciones_list = recepciones_list.filter(almacen_id=almacen_id)
    if fecha:
        try:
            recepciones_list = recepciones_list.filter(fecha=fecha)
        except:
            pass
    if estado:
        recepciones_list = recepciones_list.filter(estado=estado)
    if sucursal_id_filtro:
        recepciones_list = recepciones_list.filter(sucursal_id=sucursal_id_filtro)

    # Para el buscador visual de proveedor
    proveedor_nombre_display = ""
    if proveedor_id and proveedor_id != 'all':
        try:
            p_obj = Proveedor.objects.get(id=proveedor_id, empresa=empresa_actual)
            proveedor_nombre_display = p_obj.razon_social
        except:
            pass

    filtros = {
        'q': q,
        'folio_recepcion': folio_recepcion,
        'folio_oc': folio_oc,
        'proveedor_id': proveedor_id,
        'proveedor_nombre': proveedor_nombre_display,
        'almacen_id': almacen_id,
        'fecha': fecha,
        'estado': estado,
        'sucursal': sucursal_id_filtro
    }
    # --- FIN LÓGICA DE FILTRADO ---
    from preferencias.models import Sucursal
    sucursales_lista = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')

    paginator = Paginator(recepciones_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    ordenes_aprobadas = OrdenCompra.objects.filter(
        empresa=empresa_actual,
        estado__in=['aprobada', 'parcial', 'borrador']
    )

    almacenes = Almacen.objects.filter(empresa=empresa_actual)
    todos_los_proveedores = Proveedor.objects.filter(empresa=empresa_actual)
    oc_preseleccionada = request.GET.get('oc_id')

    context = {
        'page_obj': page_obj,
        'ordenes_compra_aprobadas': ordenes_aprobadas,
        'almacenes': almacenes,
        'todos_los_proveedores': todos_los_proveedores,
        'oc_preseleccionada': oc_preseleccionada,
        'sucursales': sucursales_lista,
        'section': 'recepciones',
        'filtros': filtros
    }
    return render(request, 'dashboard_recepciones.html', context)


@login_required
def obtener_items_orden_compra(request, oc_id):
    empresa_actual = get_empresa_actual(request)

    try:
        # Buscamos la orden de forma segura
        orden = OrdenCompra.objects.filter(id=oc_id, empresa=empresa_actual).first()

        if not orden:
            return JsonResponse({'error': 'La Orden de Compra no existe o no pertenece a tu empresa.'}, status=404)

        detalles = DetalleCompra.objects.filter(orden_compra=orden).order_by('id')

        # Preparamos el paquete de datos con moneda y tipo de cambio
        data = {
            'id': orden.id,
            'moneda': orden.moneda.siglas if (orden.moneda and orden.moneda.siglas) else 'MXN',
            'tipo_cambio': str(orden.tipo_cambio) if orden.tipo_cambio else '1.0000',
            'items': []
        }

        for det in detalles:
            # Sumamos lo recibido anteriormente para saber cuánto falta
            total_recibido = DetalleRecepcion.objects.filter(
                detalle_compra=det
            ).exclude(recepcion__estado='cancelada').aggregate(total=Sum('cantidad_recibida'))['total'] or 0

            data['items'].append({
                'id': det.id,
                'nombre': det.producto.nombre,
                'cant_ordenada': det.cantidad,
                'cant_recibida_anterior': total_recibido,
                'costo': float(det.precio_costo),
                'maneja_lote': det.producto.maneja_lote,
                'maneja_serie': det.producto.maneja_serie,
            })

        return JsonResponse(data)

    except Exception as e:
        print(f"Error en obtener_items_orden_compra: {str(e)}")
        return JsonResponse({'error': 'Error interno al cargar artículos.'}, status=500)

@login_required
@transaction.atomic
def crear_recepcion(request):
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            if not empresa_actual:
                return JsonResponse({'success': False, 'error': 'Empresa no detectada.'})

            recepcion = procesar_recepcion_servicio(
                data_post=request.POST,
                empresa_actual=empresa_actual,
                usuario=request.user
            )

            crear_notificacion(
                empresa=empresa_actual,
                actor=request.user,
                mensaje=f'procesó recepción REC-{recepcion.id:04d} para OC-{recepcion.orden_compra.id:04d}',
                propietario=recepcion.orden_compra.usuario
            )

            return JsonResponse({'success': True, 'message': 'Recepción procesada exitosamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
def cambiar_estado_recepcion(request, recepcion_id):
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        empresa_actual = get_empresa_actual(request)
        recepcion = get_object_or_404(Recepcion, id=recepcion_id, empresa=empresa_actual)

        if nuevo_estado in ['completada', 'cancelada']:
            recepcion.estado = nuevo_estado
            recepcion.save()
            return JsonResponse({'success': True, 'message': 'Estado actualizado'})
    return JsonResponse({'success': False, 'error': 'Error'})

@login_required
def api_detalle_recepcion(request, recepcion_id):
    try:
        empresa_actual = get_empresa_actual(request)
        rec = get_object_or_404(Recepcion, id=recepcion_id, empresa=empresa_actual)
        oc = rec.orden_compra

        data = {
            'titulo': 'Detalle Recepción',
            'folio': f"REC-{rec.id:04d}",
            'fecha': rec.fecha.strftime('%d/%m/%Y') if rec.fecha else '-',
            'fecha_comprobante': rec.fecha_comprobante.strftime('%d/%m/%Y') if rec.fecha_comprobante else '-',
            'estado': rec.estado.upper(),
            'proveedor': str(oc.proveedor) if oc else "Directo / Sin OC",
            'sucursal': oc.sucursal.nombre if (oc and oc.sucursal) else 'Matriz / Principal',
            'almacen': str(rec.almacen),
            'factura': rec.factura or '-',
            'total': float(rec.total),
            'pedimento': rec.pedimento or '-',
            'aduana': rec.aduana or '-',
            'fecha_pedimento': rec.fecha_pedimento.strftime('%d/%m/%Y') if rec.fecha_pedimento else '-',
            'oc_id': oc.id if oc else None,
            'oc_folio': f"OC-{oc.id:04d}" if oc else '-',
            'oc_fecha': oc.fecha.strftime('%d/%m/%Y') if (oc and oc.fecha) else '-',
            'detalles': [
                {
                    'producto': d.producto.nombre,
                    'cant': d.cantidad_recibida,
                    'precio': float(d.costo_unitario),
                    'subtotal': float(d.subtotal)
                } for d in rec.detalles.all()
            ]
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

@login_required
@transaction.atomic
def cancelar_recepcion(request, recepcion_id):
    try:
        empresa_actual = get_empresa_actual(request)
        recepcion = get_object_or_404(Recepcion, id=recepcion_id, empresa=empresa_actual)

        if recepcion.estado == 'cancelada':
            return JsonResponse({'success': False, 'error': 'La recepción ya está cancelada.'})

        orden_compra = recepcion.orden_compra
        almacen = recepcion.almacen

        for det in recepcion.detalles.all():
            producto = det.producto
            if producto.tipo != 'servicio':
                try:
                    inventario = Inventario.objects.select_for_update().get(producto=producto, almacen=almacen)
                    inventario.cantidad = F('cantidad') - det.cantidad_recibida
                    inventario.save()
                except Inventario.DoesNotExist: pass

        # Calcular nuevo estado de OC
        nuevo_estado_oc = 'aprobada'
        total_pedido = orden_compra.detalles.aggregate(total=Sum('cantidad'))['total'] or 0
        if total_pedido > 0:
            recibido_valido = DetalleRecepcion.objects.filter(recepcion__orden_compra=orden_compra).exclude(recepcion__estado='cancelada').exclude(recepcion__id=recepcion.id).aggregate(total=Sum('cantidad_recibida'))['total'] or 0
            if recibido_valido >= total_pedido: nuevo_estado_oc = 'recibida'
            elif recibido_valido > 0: nuevo_estado_oc = 'parcial'
            else: nuevo_estado_oc = 'aprobada'

        orden_compra.estado = nuevo_estado_oc
        orden_compra.save()
        recepcion.estado = 'cancelada'
        recepcion.save()

        return JsonResponse({'success': True, 'message': f'Recepción cancelada. La OC regresó a estado: {nuevo_estado_oc.upper()}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def imprimir_recepcion(request, pk):
    """Genera la vista para impresión de recepción de mercancía (PDF)"""
    empresa_actual = get_empresa_actual(request)
    recepcion = get_object_or_404(Recepcion, id=pk, empresa=empresa_actual)
    
    context = {
        'recepcion': recepcion,
        'empresa': empresa_actual,
    }
    return render(request, 'recepciones/imprimir_recepcion.html', context)
