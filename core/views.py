from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from .models import Producto, Transaccion, DetalleReceta
from django.contrib.auth.decorators import login_required
from panel.models import Empresa
import json
import openpyxl
from datetime import datetime
from django.db import transaction
from django.db.models import Sum, Q, Subquery, OuterRef, DecimalField, Avg, ExpressionWrapper, F, FloatField, IntegerField
from django.db.models.functions import Coalesce
from collections import defaultdict
from categorias.models import Categoria, Subcategoria
from almacenes.models import Almacen, Inventario
from proveedores.models import Proveedor
from preferencias.models import Moneda
from django.core.paginator import Paginator
from recepciones.models import Recepcion, DetalleRecepcion, DetalleRecepcionExtra
from compras.models import OrdenCompra
from ventas.models import DetalleOrdenVenta

from django.views.decorators.http import require_POST
from preferencias.permissions import require_inventory_permission

@login_required(login_url='/login/')
@require_inventory_permission('inventario', 'crear')
def descargar_plantilla_articulos(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Articulos"

    headers = [
        'Clave', 'Nombre*', 'Tipo (producto/servicio)', 'Abastecimiento (stock/produccion/compra)',
        'Categoria', 'Subcategoria', 'Marca', 'Modelo', 'Unidad Medida (H87/E48)', 'Costo*', 'Precio Venta*',
        'IVA (%)', 'Stock Minimo', 'Stock Maximo', 'Maneja Lote (si/no)', 'Maneja Serie (si/no)',
        'Estado (activo/inactivo/descontinuado/revision)'
    ]
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Articulos.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
@require_inventory_permission('inventario', 'crear', json_response=True)
def importar_articulos_ajax(request):
    if request.method == 'POST' and request.FILES.get('archivo_articulos'):
        empresa_actual = get_empresa_actual(request)
        excel_file = request.FILES['archivo_articulos']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return JsonResponse({'success': False, 'error': 'El archivo está vacío.'})
            
            data_rows = rows[1:]
            creados = 0
            actualizados = 0
            errores = []
            
            for idx, row in enumerate(data_rows, start=2):
                if not any(row): continue
                
                try:
                    with transaction.atomic():
                        clave = str(row[0] or '').strip()
                        nombre = str(row[1] or '').strip()

                        if not nombre:
                            errores.append(f"Fila {idx}: El nombre es obligatorio.")
                            continue

                        # --- LÓGICA DE IDENTIFICACIÓN: CLAVE Y NOMBRE ---
                        if not clave:
                            # Escenario 1: Clave vacía en Excel -> Buscar mismo nombre con clave vacía/nula en DB
                            producto = Producto.objects.filter(
                                Q(clave__isnull=True) | Q(clave=''),
                                nombre=nombre,
                                empresa=empresa_actual
                            ).first()
                        else:
                            # Escenario 2: Clave con datos en Excel -> Coincidencia exacta de Clave y Nombre
                            producto = Producto.objects.filter(
                                clave=clave,
                                nombre=nombre,
                                empresa=empresa_actual
                            ).first()

                        if producto:
                            # --- ACTUALIZAR EXISTENTE (Solo campos no vacíos en Excel) ---
                            # Tipo
                            val_tipo = str(row[2] or '').strip().lower()
                            if val_tipo in ['producto', 'servicio']: 
                                producto.tipo = val_tipo
                            
                            # Abastecimiento
                            val_abast = str(row[3] or '').strip().lower()
                            if val_abast in ['stock', 'produccion', 'compra']:
                                producto.tipo_abastecimiento = val_abast

                            # Categoría y Subcategoría (con auto-creación)
                            val_cat = str(row[4] or '').strip()
                            val_sub = str(row[5] or '').strip()
                            if val_cat:
                                producto.categoria = val_cat
                                cat_obj, _ = Categoria.objects.get_or_create(nombre=val_cat, empresa=empresa_actual)
                                if val_sub:
                                    producto.subcategoria = val_sub
                                    Subcategoria.objects.get_or_create(nombre=val_sub, categoria=cat_obj, empresa=empresa_actual)

                            if str(row[6] or '').strip(): producto.marca = str(row[6]).strip()
                            if str(row[7] or '').strip(): producto.modelo = str(row[7]).strip()
                            if str(row[8] or '').strip(): producto.unidad_medida = str(row[8]).strip().upper()
                            
                            if row[9] is not None: producto.precio_costo = float(row[9])
                            if row[10] is not None: producto.precio_venta = float(row[10])
                            if row[11] is not None: producto.iva = float(row[11])
                            if row[12] is not None: producto.stock_minimo = int(row[12])
                            if row[13] is not None: producto.stock_maximo = int(row[13])
                            
                            if row[14] is not None: producto.maneja_lote = str(row[14]).strip().lower() == 'si'
                            if row[15] is not None: producto.maneja_serie = str(row[15]).strip().lower() == 'si'
                            
                            # Estado
                            if len(row) > 16 and row[16] is not None:
                                val_estado = str(row[16]).strip().lower()
                                if val_estado in ['activo', 'inactivo', 'descontinuado', 'revision']:
                                    producto.estado = val_estado
                            
                            producto.save()
                            actualizados += 1
                        else:
                            # --- CREAR NUEVO ---
                            cat_name = str(row[4] or '').strip()
                            subcat_name = str(row[5] or '').strip()
                            
                            if cat_name:
                                cat_obj, _ = Categoria.objects.get_or_create(nombre=cat_name, empresa=empresa_actual)
                                if subcat_name:
                                    Subcategoria.objects.get_or_create(nombre=subcat_name, categoria=cat_obj, empresa=empresa_actual)

                            # Determinar estado inicial
                            nuevo_estado = 'activo'
                            if len(row) > 16 and row[16] is not None:
                                val_estado = str(row[16]).strip().lower()
                                if val_estado in ['activo', 'inactivo', 'descontinuado', 'revision']:
                                    nuevo_estado = val_estado

                            Producto.objects.create(
                                empresa=empresa_actual,
                                clave=clave,
                                nombre=nombre,
                                tipo=str(row[2] or 'producto').strip().lower() if str(row[2] or '').strip().lower() in ['producto', 'servicio'] else 'producto',
                                tipo_abastecimiento=str(row[3] or 'compra').strip().lower() if str(row[3] or '').strip().lower() in ['stock', 'produccion', 'compra'] else 'compra',
                                categoria=cat_name,
                                subcategoria=subcat_name,
                                marca=str(row[6] or '').strip(),
                                modelo=str(row[7] or '').strip(),
                                unidad_medida=str(row[8] or 'H87').strip().upper(),
                                precio_costo=float(row[9] or 0),
                                precio_venta=float(row[10] or 0),
                                iva=float(row[11] or 0),
                                stock_minimo=int(row[12] or 0),
                                stock_maximo=int(row[13] or 1000),
                                maneja_lote=str(row[14] or 'no').strip().lower() == 'si',
                                maneja_serie=str(row[15] or 'no').strip().lower() == 'si',
                                estado=nuevo_estado
                            )
                            creados += 1
                except Exception as row_err:
                    errores.append(f"Fila {idx}: {str(row_err)}")
            
            if creados == 0 and actualizados == 0 and errores:
                return JsonResponse({'success': False, 'error': f"No se pudo procesar ningún artículo. Primer error: {errores[0]}"})

            msg = f'Importación finalizada.\nArtículos creados: {creados}\nArtículos actualizados: {actualizados}\n'
            if errores:
                msg += f'\nFilas con error: {len(errores)}\n'
                msg += '\nDetalle de primeros errores:\n'
                msg += "\n".join(errores[:10])
                if len(errores) > 10:
                    msg += "\n... (revisa el resto de tu archivo Excel)"
                
            return JsonResponse({'success': True, 'message': msg})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
            
    return JsonResponse({'success': False, 'error': 'Solicitud inválida.'})

@login_required(login_url='/login/')
@require_inventory_permission('inventario', 'ver')
def exportar_existencias_excel(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return HttpResponse("No autorizado", status=403)

    # --- Filtrado (Misma lógica que el dashboard) ---
    q = request.GET.get('q', '')
    almacen_id = request.GET.get('almacen', '')
    sucursal_id = request.GET.get('sucursal', '')
    categoria = request.GET.get('categoria', '')
    estado = request.GET.get('estado', '')
    stock_status = request.GET.get('existencias', '')

    productos_qs = Producto.objects.filter(empresa=empresa_actual)

    if q:
        productos_qs = productos_qs.filter(
            Q(nombre__icontains=q) | Q(clave__icontains=q) | Q(marca__icontains=q) | Q(modelo__icontains=q) |
            Q(categoria__icontains=q) | Q(subcategoria__icontains=q)
        )
    if estado:
        productos_qs = productos_qs.filter(estado=estado)
    if categoria:
        productos_qs = productos_qs.filter(categoria=categoria)

    # --- Anotaciones de Stock (Misma lógica que el dashboard) ---
    def get_stock_sub(field):
        qs = Inventario.objects.filter(producto=OuterRef('pk'))
        if sucursal_id and sucursal_id != 'all': qs = qs.filter(sucursal_id=sucursal_id)
        if almacen_id and almacen_id != 'all': qs = qs.filter(almacen_id=almacen_id)
        return qs.values('producto').annotate(total=Sum(field)).values('total')[:1]

    # Para el costo promedio, si hay filtro de almacén/sucursal usamos el promedio de esa zona
    # Si no, el costo_promedio_global del modelo (o anotación similar)
    costo_prom_qs = Inventario.objects.filter(producto=OuterRef('pk'))
    if sucursal_id and sucursal_id != 'all': costo_prom_qs = costo_prom_qs.filter(sucursal_id=sucursal_id)
    if almacen_id and almacen_id != 'all': costo_prom_qs = costo_prom_qs.filter(almacen_id=almacen_id)
    
    costo_prom_subquery = costo_prom_qs.values('producto').annotate(avg=Avg('costo_promedio')).values('avg')[:1]

    productos = productos_qs.annotate(
        stock_fisico_anotado=Coalesce(Subquery(get_stock_sub('cantidad')), 0),
        stock_res_anotado=Coalesce(Subquery(get_stock_sub('reservado')), 0),
        costo_prom_anotado=Coalesce(Subquery(costo_prom_subquery), F('precio_costo'))
    ).annotate(
        stock_disponible_anotado=F('stock_fisico_anotado') - F('stock_res_anotado')
    )

    if stock_status == 'con':
        productos = productos.filter(stock_fisico_anotado__gt=0)
    elif stock_status == 'sin':
        productos = productos.filter(stock_fisico_anotado__lte=0)

    # --- Generar Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Existencias Inventario"

    # Determinar nombre del almacén para la columna
    nombre_almacen_col = "GLOBAL"
    if almacen_id and almacen_id != 'all':
        alm_obj = Almacen.objects.filter(id=almacen_id).first()
        if alm_obj: nombre_almacen_col = alm_obj.nombre
    elif sucursal_id and sucursal_id != 'all':
        from preferencias.models import Sucursal
        suc_obj = Sucursal.objects.filter(id=sucursal_id).first()
        if suc_obj: nombre_almacen_col = f"SUC: {suc_obj.nombre}"

    headers = [
        'Clave', 'Producto', 'Zona/Almacén', 'Stock Físico', 'Reservado', 'Disponible', 
        'Unidad de medida', 'Costo Unit.', 'Costo Promedio', 'Valor Inventario'
    ]
    ws.append(headers)

    for p in productos:
        valor = float(p.stock_fisico_anotado) * float(p.costo_prom_anotado)
        # Obtener el nombre legible de la unidad de medida
        um_display = dict(p.UNIDAD_OPCIONES).get(p.unidad_medida, p.unidad_medida)
        
        ws.append([
            p.clave or '', p.nombre, nombre_almacen_col, p.stock_fisico_anotado, p.stock_res_anotado, p.stock_disponible_anotado,
            um_display, p.precio_costo, p.costo_prom_anotado, valor
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Existencias_{empresa_actual.nombre}.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
@require_inventory_permission('inventario', 'crear')
def descargar_plantilla_recetas(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Recetas"
    
    # Encabezados
    headers = ['Clave Producto Final (Producción)', 'Clave Componente (Stock/Compra)', 'Cantidad Requerida']
    ws.append(headers)
    
    # Ajustar ancho
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 35

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Recetas_Masivas.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
@require_inventory_permission('inventario', 'crear', json_response=True)
def importar_recetas_ajax(request):
    if request.method == 'POST' and request.FILES.get('archivo_articulos'):
        empresa_actual = get_empresa_actual(request)
        excel_file = request.FILES['archivo_articulos']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return JsonResponse({'success': False, 'error': 'El archivo está vacío.'})
            
            data_rows = rows[1:]
            
            # Agrupar por Producto Final
            recetas_data = defaultdict(list)
            for idx, row in enumerate(data_rows, start=2):
                if not any(row): continue
                clave_padre = str(row[0] or '').strip()
                clave_comp = str(row[1] or '').strip()
                cant = row[2] or 0
                
                if not clave_padre or not clave_comp: continue
                
                recetas_data[clave_padre].append({
                    'clave_comp': clave_comp,
                    'cantidad': cant,
                    'fila': idx
                })

            creadas = 0
            errores = []
            
            for clave_padre, items in recetas_data.items():
                try:
                    with transaction.atomic():
                        # 1. Validar Padre
                        padre = Producto.objects.filter(clave=clave_padre, empresa=empresa_actual).first()
                        if not padre:
                            errores.append(f"Padre '{clave_padre}': No existe en el catálogo.")
                            continue
                        
                        if padre.tipo_abastecimiento != 'produccion':
                            errores.append(f"Padre '{clave_padre}': No es de tipo producción.")
                            continue

                        # Limpiar receta anterior para evitar duplicados
                        DetalleReceta.objects.filter(producto_padre=padre).delete()

                        # 2. Validar y crear componentes
                        for item in items:
                            comp = Producto.objects.filter(clave=item['clave_comp'], empresa=empresa_actual).first()
                            if not comp:
                                errores.append(f"Padre '{clave_padre}', Fila {item['fila']}: El componente '{item['clave_comp']}' no existe.")
                                continue
                            
                            if comp.tipo_abastecimiento not in ['stock', 'compra']:
                                errores.append(f"Padre '{clave_padre}', Fila {item['fila']}: El componente '{item['clave_comp']}' debe ser Stock o Compra.")
                                continue
                            
                            DetalleReceta.objects.create(
                                producto_padre=padre,
                                componente=comp,
                                cantidad=item['cantidad']
                            )
                        
                        creadas += 1
                except Exception as e:
                    errores.append(f"Padre '{clave_padre}': {str(e)}")
            
            if creadas == 0 and errores:
                return JsonResponse({'success': False, 'error': f"No se pudo procesar ninguna receta. Primer error: {errores[0]}"})

            msg = f'Importación de recetas finalizada.\nRecetas configuradas: {creadas}\n'
            if errores:
                msg += f'\nFilas/Padres con error: {len(errores)}\n'
                msg += '\nDetalle de primeros errores:\n'
                msg += "\n".join(errores[:10])
                if len(errores) > 10:
                    msg += "\n... (revisa el resto de tu archivo Excel)"
                
            return JsonResponse({'success': True, 'message': msg})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
            
    return JsonResponse({'success': False, 'error': 'Solicitud inválida.'})

# --- 1. FUNCIÓN AYUDANTE ESTÁNDAR ---
def get_empresa_actual(request):
    username = request.user.username
    if '@' in username:
        subdominio = username.split('@')[1]
        try:
            return Empresa.objects.get(subdominio=subdominio)
        except Empresa.DoesNotExist:
            return None
    return None

@require_POST
@login_required(login_url='/login/')
@require_inventory_permission('inventario', 'crear', json_response=True)
def crear_producto_rapido(request):
    empresa = get_empresa_actual(request)
    if not empresa:
        return JsonResponse({'success': False, 'error': 'No se encontró la empresa'}, status=403)
    
    try:
        nombre = request.POST.get('nombre')
        tipo = request.POST.get('tipo', 'producto')
        tipo_abastecimiento = request.POST.get('tipo_abastecimiento', 'compra')
        categoria_id = request.POST.get('categoria', '')
        subcategoria = request.POST.get('subcategoria', '')
        
        categoria_nombre = ""
        if categoria_id:
            from categorias.models import Categoria as CategoriaCatalogo
            cat_obj = CategoriaCatalogo.objects.filter(id=categoria_id, empresa=empresa).first()
            if cat_obj:
                categoria_nombre = cat_obj.nombre
        
        if not nombre:
            return JsonResponse({'success': False, 'error': 'El nombre es obligatorio'}, status=400)
            
        estado = request.POST.get('estado', 'activo')

        producto = Producto.objects.create(
            nombre=nombre,
            tipo=tipo,
            tipo_abastecimiento=tipo_abastecimiento,
            categoria=categoria_nombre,
            subcategoria=subcategoria,
            estado=estado,
            precio_costo=0.00,
            precio_venta=0.00,
            iva=16.00,
            tiene_iva=True,
            empresa=empresa
        )
        
        return JsonResponse({
            'success': True,
            'id': producto.id,
            'nombre': producto.nombre,
            'precio_venta': str(producto.precio_venta),
            'iva': str(producto.iva)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required(login_url='/login/')
def punto_de_venta(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)
    productos = Producto.objects.filter(empresa=empresa_actual)
    if request.method == 'POST':
        producto_id = request.POST.get('producto_id')
        cantidad = int(request.POST.get('cantidad', 1))
        try:
            producto = get_object_or_404(Producto, id=producto_id, empresa=empresa_actual)
            Transaccion.objects.create(producto=producto, tipo='venta', cantidad=cantidad, empresa=empresa_actual)
            messages.success(request, f'Vendido: {cantidad} x {producto.nombre}')
            return redirect('punto_de_venta')
        except Exception as e:
            messages.error(request, str(e))
    return render(request, 'core/vender.html', {'productos': productos})

@login_required(login_url='/login/')
@require_inventory_permission('inventario', 'crear', json_response=True)
def crear_producto_ajax(request):
    if request.method == 'POST':
        try:
            from .forms import ProductoForm
            empresa_actual = get_empresa_actual(request)
            if not empresa_actual:
                return JsonResponse({'success': False, 'error': 'Empresa no detectada'})
            form = ProductoForm(request.POST)
            if form.is_valid():
                producto = form.save(commit=False)
                producto.empresa = empresa_actual
                
                # Manejo manual de switches
                producto.maneja_lote = request.POST.get('maneja_lote') == 'on'
                producto.maneja_serie = request.POST.get('maneja_serie') == 'on'
                producto.tiene_iva = request.POST.get('tiene_iva') == 'on'

                if not producto.tiene_iva:
                    producto.iva = 0.00

                test_id = request.POST.get('test_calidad')
                if test_id:
                    from produccion.models import Test
                    producto.test_calidad = Test.objects.filter(id=test_id, empresa=empresa_actual).first()
                producto.save()
                return JsonResponse({'success': True, 'message': 'Artículo creado correctamente.'})
            else:
                return JsonResponse({'success': False, 'error': form.errors})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required(login_url='/login/')
@require_inventory_permission('inventario', 'ver')
def dashboard_inventario(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)
    
    # --- FILTROS ---
    sucursal_sesion_id = request.session.get('sucursal_id')
    sucursal_id = request.GET.get('sucursal')
    almacen_id = request.GET.get('almacen')
    
    # Solo aplicamos la sucursal de sesión si NO hay filtros explícitos de sucursal o almacén
    # Y si el usuario no ha pedido explícitamente "ver todo" (sucursal='')
    if sucursal_id is None and almacen_id is None and sucursal_sesion_id:
        sucursal_id = str(sucursal_sesion_id)
        
    q = request.GET.get('q')
    estado = request.GET.get('estado')
    vista = request.GET.get('vista', 'existencias')
    stock_status = request.GET.get('existencias')
    categoria_id = request.GET.get('categoria')

    productos_qs = Producto.objects.filter(empresa=empresa_actual)

    # NOTA: El filtrado por sucursal/almacén NO debe reducir el catálogo de productos.
    # Solo debe afectar a las cantidades que se muestran en las columnas de Stock.
    # El filtrado de la lista (QuerySet) solo se aplica para búsqueda, estado o categoría.

    if q:
        productos_qs = productos_qs.filter(
            Q(nombre__icontains=q) | Q(marca__icontains=q) | Q(modelo__icontains=q) |
            Q(categoria__icontains=q) | Q(subcategoria__icontains=q)
        )
    if estado:
        productos_qs = productos_qs.filter(estado=estado)

    if categoria_id:
        productos_qs = productos_qs.filter(categoria=categoria_id)

    # 1. PRECIO PROMEDIO VENTA (GLOBAL)
    avg_venta_subquery = DetalleOrdenVenta.objects.filter(producto=OuterRef('pk')).values('producto').annotate(promedio=Avg('precio_unitario')).values('promedio')[:1]
    # 2. PRECIO MÁXIMO VENTA (GLOBAL)
    max_precio_venta = DetalleOrdenVenta.objects.filter(producto=OuterRef('pk')).order_by('-precio_unitario').values('precio_unitario')[:1]
    # 3. COSTO MÁXIMO COMPRA (GLOBAL, EN PESOS)
    max_costo_compra = DetalleRecepcion.objects.filter(producto=OuterRef('pk')).annotate(costo_pesos=ExpressionWrapper(F('costo_unitario') * F('recepcion__tipo_cambio'), output_field=DecimalField(max_digits=10, decimal_places=2))).order_by('-costo_pesos').values('costo_pesos')[:1]
    # 4. COSTO PROMEDIO COMPRAS (GLOBAL, EN PESOS)
    costo_promedio_subquery = DetalleRecepcion.objects.filter(producto=OuterRef('pk')).annotate(costo_pesos=ExpressionWrapper(F('costo_unitario') * F('recepcion__tipo_cambio'), output_field=FloatField())).values('producto').annotate(promedio_historico=Avg('costo_pesos')).values('promedio_historico')[:1]

    # 5. VALOR TOTAL INVENTARIO (POR ALMACÉN SI SE FILTRA)
    inv_val_filter = Q(producto=OuterRef('pk'), cantidad__gt=0)
    if sucursal_id:
        inv_val_filter &= Q(sucursal_id=sucursal_id)
    if almacen_id:
        inv_val_filter &= Q(almacen_id=almacen_id)
    costo_inventario_subquery = Inventario.objects.filter(inv_val_filter).values('producto').annotate(valor_total=Sum(ExpressionWrapper(F('cantidad') * F('costo_promedio'), output_field=FloatField()))).values('valor_total')[:1]

    # 6. STOCK FISICO Y RESERVADO (DINÁMICO)
    def get_stock_sub(field):
        qs = Inventario.objects.filter(producto=OuterRef('pk'))
        if sucursal_id: qs = qs.filter(sucursal_id=sucursal_id)
        if almacen_id: qs = qs.filter(almacen_id=almacen_id)
        return qs.values('producto').annotate(total=Sum(field)).values('total')[:1]

    productos = productos_qs.annotate(
        precio_promedio_venta=Subquery(avg_venta_subquery),
        precio_max_venta=Subquery(max_precio_venta),
        costo_max_compra=Subquery(max_costo_compra),
        costo_promedio_anotado=Subquery(costo_promedio_subquery),
        costo_inventario_anotado=Subquery(costo_inventario_subquery),
        stock_fisico=Subquery(get_stock_sub('cantidad')),
        stock_res=Subquery(get_stock_sub('reservado'))
    ).annotate(
        stock_disponible_anotado=ExpressionWrapper(
            Coalesce(F('stock_fisico'), 0) - Coalesce(F('stock_res'), 0),
            output_field=IntegerField()
        )
    )

    if stock_status == 'con':
        productos = productos.filter(stock_fisico__gt=0)
    elif stock_status == 'sin':
        productos = productos.filter(Q(stock_fisico__lte=0) | Q(stock_fisico__isnull=True))

    productos = productos.order_by('nombre')
    
    # --- PAGINACIÓN ---
    paginator = Paginator(productos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    from preferencias.models import Sucursal
    sucursales = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')

    almacenes_todos = Almacen.objects.filter(empresa=empresa_actual).order_by('nombre')
    almacenes = almacenes_todos
    if sucursal_id:
        almacenes = almacenes.filter(sucursal_id=sucursal_id)

    todas_categorias = Categoria.objects.filter(empresa=empresa_actual).order_by('nombre')
    from produccion.models import Test
    tests_calidad = Test.objects.filter(empresa=empresa_actual).order_by('nombre')
    productos_padre_receta = Producto.objects.filter(
        empresa=empresa_actual,
        tipo_abastecimiento='produccion'
    ).order_by('nombre')
    productos_componentes_receta = Producto.objects.filter(
        empresa=empresa_actual
    ).order_by('nombre')

    # Monedas para conversión en modal
    monedas = Moneda.objects.filter(empresa=empresa_actual)

    contexto = {
        'page_obj': page_obj,
        'almacenes': almacenes,
        'almacenes_todos': almacenes_todos,
        'sucursales': sucursales,
        'categorias': todas_categorias,
        'tests_calidad': tests_calidad,
        'productos_padre_receta': productos_padre_receta,
        'productos_componentes_receta': productos_componentes_receta,
        'monedas': monedas,
        'filtros': {
            'sucursal': sucursal_id or '',
            'almacen': almacen_id or '', 
            'q': q or '',
            'estado': estado or '',
            'vista': vista,
            'existencias': stock_status or '',
            'categoria': categoria_id or ''
        },
        'section': 'inventario'
    }
    return render(request, 'dashboard_inventario.html', contexto)

@login_required
@require_inventory_permission('inventario', 'existencias', json_response=True)
def api_detalle_producto_inventario(request, producto_id):
    try:
        empresa_actual = get_empresa_actual(request)
        producto = get_object_or_404(Producto, id=producto_id, empresa=empresa_actual)
        
        # 1. HISTORIAL DE RECEPCIONES (EXISTENCIAS)
        detalles = DetalleRecepcion.objects.filter(producto_id=producto_id).select_related('recepcion', 'detalle_compra__orden_compra').order_by('-recepcion__fecha')
        historial_data = []
        for d in detalles:
            folio_oc, proveedor = "Directa", "Directo / Sin OC"
            if d.detalle_compra and d.detalle_compra.orden_compra:
                oc = d.detalle_compra.orden_compra
                folio_oc, proveedor = f"OC-{oc.id:04d}", oc.proveedor
            historial_data.append({
                'folio_oc': folio_oc, 'oc_id': d.detalle_compra.orden_compra.id if d.detalle_compra and d.detalle_compra.orden_compra else None,
                'folio_rec': f"REC-{d.recepcion.id:04d}", 'rec_id': d.recepcion.id, 'proveedor': str(proveedor),
                'fecha': d.recepcion.fecha.strftime('%d/%m/%Y'), 'cantidad': d.cantidad_recibida,
                'costo': float(d.costo_unitario), 'total': float(d.cantidad_recibida * d.costo_unitario)
            })

        # 2. LOTES Y SERIES (DESDE EXTRAS)
        extras = DetalleRecepcionExtra.objects.filter(
            Q(detalle_recepcion__producto_id=producto_id) | Q(producto_id=producto_id)
        ).select_related('almacen', 'detalle_recepcion__recepcion')
        
        lotes_data = []
        series_data = []
        
        for e in extras:
            # Determinar fecha: prioridad a recepción, luego fecha creación propia
            fecha_disp = "--"
            if e.detalle_recepcion and e.detalle_recepcion.recepcion:
                fecha_disp = e.detalle_recepcion.recepcion.fecha.strftime('%d/%m/%Y')
            elif e.fecha_creacion:
                fecha_disp = e.fecha_creacion.strftime('%d/%m/%Y')

            if e.tipo == 'lote':
                lotes_data.append({
                    'lote': e.lote,
                    'cantidad': e.cantidad_lote,
                    'almacen': e.almacen.nombre if e.almacen else 'N/A',
                    'fecha': fecha_disp
                })
            elif e.tipo == 'serie':
                series_data.append({
                    'serie': e.serie,
                    'almacen': e.almacen.nombre if e.almacen else 'N/A',
                    'fecha': fecha_disp
                })

        # 3. PEDIMENTOS (DESDE RECEPCIÓN)
        # Obtenemos las recepciones únicas que tengan pedimento para este producto
        recepciones_con_pedimento = Recepcion.objects.filter(detalles__producto_id=producto_id, pedimento__isnull=False).exclude(pedimento='').distinct()
        pedimentos_data = []
        for r in recepciones_con_pedimento:
            pedimentos_data.append({
                'pedimento': r.pedimento,
                'aduana': r.aduana or 'N/A',
                'fecha': r.fecha_pedimento.strftime('%d/%m/%Y') if r.fecha_pedimento else 'N/A',
                'folio_rec': f"REC-{r.id:04d}",
                'rec_id': r.id
            })

        return JsonResponse({
            'historial': historial_data,
            'lotes': lotes_data,
            'series': series_data,
            'pedimentos': pedimentos_data
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@login_required
@require_inventory_permission('inventario', 'ver', json_response=True)
def obtener_producto_json(request, producto_id):
    try:
        empresa_actual = get_empresa_actual(request)
        producto = get_object_or_404(Producto, id=producto_id, empresa=empresa_actual)
        
        # Obtener plantillas globales de listas
        from categorias.models import ListaPrecioCosto
        listas_maestras = ListaPrecioCosto.objects.filter(empresa=empresa_actual).values(
            'id', 'nombre', 'porcentaje_extra', 'monto_extra', 'tipo'
        )
        
        return JsonResponse({
            'id': producto.id, 'clave': producto.clave or '', 'nombre': producto.nombre, 'descripcion': producto.descripcion,
            'tipo': producto.tipo, 'tipo_abastecimiento': producto.tipo_abastecimiento, 'estado': producto.estado,
            'categoria': producto.categoria or '', 'subcategoria': producto.subcategoria or '',
            'marca': producto.marca or '', 'modelo': producto.modelo or '', 'linea': producto.linea or '',
            'unidad_medida': producto.unidad_medida, 'iva': str(producto.iva), 'ieps': str(producto.ieps),
            'precio_costo': str(producto.precio_costo), 'precio_venta': str(producto.precio_venta),
            'precios_lista': producto.precios_lista, 'costos_lista': producto.costos_lista,
            'listas_maestras': list(listas_maestras),
            'stock_minimo': producto.stock_minimo, 'stock_maximo': producto.stock_maximo,
            'maneja_lote': producto.maneja_lote, 'maneja_serie': producto.maneja_serie,
            'tiene_iva': producto.tiene_iva,
            'test_calidad_id': producto.test_calidad.id if producto.test_calidad else "",
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

@login_required(login_url='/login/')
@require_inventory_permission('inventario', 'editar', json_response=True)
def actualizar_producto_ajax(request, producto_id):
    if request.method == 'POST':
        try:
            from .forms import ProductoForm
            empresa_actual = get_empresa_actual(request)
            producto = get_object_or_404(Producto, id=producto_id, empresa=empresa_actual)
            
            form = ProductoForm(request.POST, instance=producto)
            if form.is_valid():
                producto = form.save(commit=False)
                # Maneja el switch manual para lote/serie ya que vienen como 'on'
                producto.maneja_lote = request.POST.get('maneja_lote') == 'on'
                producto.maneja_serie = request.POST.get('maneja_serie') == 'on'
                producto.tiene_iva = request.POST.get('tiene_iva') == 'on'
                
                if not producto.tiene_iva:
                    producto.iva = 0.00
                
                test_id = request.POST.get('test_calidad')
                if test_id:
                    from produccion.models import Test
                    producto.test_calidad = Test.objects.filter(id=test_id, empresa=empresa_actual).first()
                else:
                    producto.test_calidad = None
                
                producto.save()
                return JsonResponse({'success': True, 'message': 'Artículo actualizado correctamente.'})
            else:
                return JsonResponse({'success': False, 'error': form.errors})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@require_inventory_permission('inventario', 'ver', json_response=True)
def api_detalle_documento(request):
    doc_tipo, doc_id = request.GET.get('tipo'), request.GET.get('id')
    empresa_actual = get_empresa_actual(request)
    try:
        if doc_tipo == 'oc':
            obj = get_object_or_404(OrdenCompra, id=doc_id, empresa=empresa_actual)
            data = {'titulo': 'Detalle Orden de Compra', 'folio': f"OC-{obj.id:04d}", 'proveedor': str(obj.proveedor), 'fecha': obj.fecha.strftime('%d/%m/%Y'), 'estado': obj.estado.upper(), 'total': float(obj.total), 'detalles': [{'producto': d.producto.nombre, 'cant': d.cantidad, 'precio': float(d.precio_costo), 'subtotal': float(d.cantidad * d.precio_costo)} for d in obj.detalles.all()]}
        elif doc_tipo == 'rec':
            obj = get_object_or_404(Recepcion, id=doc_id, empresa=empresa_actual)
            data = {'titulo': 'Detalle Recepción', 'folio': f"REC-{obj.id:04d}", 'oc_folio': f"OC-{obj.orden_compra.id:04d}" if obj.orden_compra else '-', 'proveedor': str(obj.orden_compra.proveedor) if obj.orden_compra else "Directo / Sin OC", 'almacen': str(obj.almacen), 'fecha': obj.fecha.strftime('%d/%m/%Y'), 'estado': obj.estado.upper(), 'factura': obj.factura or '-', 'total': float(obj.total), 'pedimento': obj.pedimento or '-', 'aduana': obj.aduana or '-', 'fecha_pedimento': obj.fecha_pedimento.strftime('%d/%m/%Y') if obj.fecha_pedimento else '-', 'detalles': [{'producto': d.producto.nombre, 'cant': d.cantidad_recibida, 'precio': float(d.costo_unitario), 'subtotal': float(d.cantidad_recibida * d.costo_unitario)} for d in obj.detalles.all()]}
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

@login_required
@transaction.atomic
@require_inventory_permission('inventario', 'receta', json_response=True)
def guardar_receta(request):
    if request.method == 'POST':
        try:
            import json
            empresa_actual, data = get_empresa_actual(request), json.loads(request.body.decode('utf-8'))
            producto = get_object_or_404(Producto, id=data.get('producto_id'), empresa=empresa_actual)
            if producto.tipo_abastecimiento != 'produccion': return JsonResponse({'success': False, 'error': 'Solo para productos de PRODUCCIÓN.'})
            DetalleReceta.objects.filter(producto_padre=producto).delete()
            if data.get('componentes'):
                for item in json.loads(data.get('componentes')):
                    if int(item.get('id')) == producto.id: continue
                    DetalleReceta.objects.create(producto_padre=producto, componente=get_object_or_404(Producto, id=item.get('id'), empresa=empresa_actual), cantidad=item.get('cant', 1))
            return JsonResponse({'success': True, 'message': 'Receta guardada correctamente.'})
        except Exception as e: return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@transaction.atomic
@require_inventory_permission('inventario', 'recetas', json_response=True)
def ejecutar_produccion(request):
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            producto, almacen = get_object_or_404(Producto, id=request.POST.get('producto_id'), empresa=empresa_actual), get_object_or_404(Almacen, id=request.POST.get('almacen_id'), empresa=empresa_actual)
            cant_prod, receta = int(request.POST.get('cantidad')), DetalleReceta.objects.filter(producto_padre=producto)
            if not receta.exists(): return JsonResponse({'success': False, 'error': 'Sin receta configurada.'})
            invs = {inv.producto_id: inv for inv in Inventario.objects.select_for_update().filter(producto_id__in=[r.componente_id for r in receta], almacen=almacen)}
            errs = [f"Falta {r.componente.nombre}" for r in receta if invs.get(r.componente_id, Inventario(cantidad=0)).cantidad < (r.cantidad * cant_prod)]
            if errs: return JsonResponse({'success': False, 'error': 'Stock insuficiente: ' + ', '.join(errs)})
            for r in receta:
                inv = invs[r.componente_id]
                inv.cantidad -= (r.cantidad * cant_prod)
                inv.save()
            inv_f, _ = Inventario.objects.get_or_create(producto=producto, almacen=almacen, defaults={'cantidad': 0, 'costo_promedio': Decimal('0.00')})
            inv_f.cantidad += cant_prod
            inv_f.save()
            return JsonResponse({'success': True, 'message': f'Producción exitosa: {cant_prod} {producto.nombre}.'})
        except Exception as e: return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@require_inventory_permission('inventario', 'recetas', json_response=True)
def obtener_receta(request, producto_id):
    try:
        empresa_actual = get_empresa_actual(request)
        get_object_or_404(Producto, id=producto_id, empresa=empresa_actual)
        receta_data = []
        for r in DetalleReceta.objects.filter(producto_padre_id=producto_id):
            receta_data.append({
                'id': r.componente.id, 
                'nombre': r.componente.nombre, 
                'cant': r.cantidad, 
                'costo': float(r.componente.precio_costo),
                'iva': float(r.componente.iva),
                'tiene_iva': r.componente.tiene_iva
            })
        return JsonResponse(receta_data, safe=False)
    except Exception as e: return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_inventory_permission('inventario', 'precios', json_response=True)
def actualizar_precio_producto(request, producto_id):
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            producto = get_object_or_404(Producto, id=producto_id, empresa=empresa_actual)
            
            nuevo_costo = request.POST.get('precio_costo')
            nueva_venta = request.POST.get('precio_venta')
            
            if nuevo_costo: producto.precio_costo = nuevo_costo
            if nueva_venta: producto.precio_venta = nueva_venta
            
            if request.POST.get('precios_extra'): 
                producto.precios_lista = json.loads(request.POST.get('precios_extra'))
            if request.POST.get('costos_extra'): 
                producto.costos_lista = json.loads(request.POST.get('costos_extra'))
                
            producto.save()
            return JsonResponse({'success': True, 'message': 'Datos actualizados.'})
        except Exception as e: return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})
