from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.forms.models import model_to_dict
from django.db import transaction
import openpyxl
from datetime import datetime
from collections import defaultdict
from panel.models import Empresa
from pedidos.models import Pedido, DetallePedido
from .models import SolicitudCompra, DetalleSolicitudCompra
from core.models import Producto, DetalleReceta
from categorias.models import ListaPrecioCosto, Categoria as CategoriaCatalogo
from produccion.models import OrdenProduccion
from proveedores.models import Proveedor
from almacenes.models import Almacen
from preferencias.models import Moneda
from notificaciones.utils import crear_notificacion
from preferencias.permissions import require_sales_permission

@login_required(login_url='/login/')
def descargar_plantilla_solicitudes(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Solicitudes"
    
    # Encabezados: Referencia ayuda a agrupar productos en una misma solicitud
    headers = [
        'Referencia (Ej: SOL-001)', 'Producto (Nombre o ID)', 'Cantidad', 
        'Proveedor (Nombre o ID)', 'Notas'
    ]
    
    ws.append(headers)
    
    # Ajustar ancho de columnas
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Solicitudes.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='/login/')
def importar_solicitudes_ajax(request):
    if request.method == 'POST' and request.FILES.get('archivo_solicitudes'):
        empresa_actual = get_empresa_actual(request)
        excel_file = request.FILES['archivo_solicitudes']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return JsonResponse({'success': False, 'error': 'El archivo está vacío.'})
            
            data_rows = rows[1:]
            
            # Agrupar por Referencia
            solicitudes_data = {}
            for idx, row in enumerate(data_rows, start=2):
                if not any(row): continue
                ref = str(row[0] or f"TEMP-{idx}").strip()
                if ref not in solicitudes_data:
                    solicitudes_data[ref] = {
                        'notas': str(row[4] or '').strip(),
                        'items': []
                    }
                solicitudes_data[ref]['items'].append({
                    'prod_input': str(row[1] or '').strip(),
                    'cantidad': row[2] or 1,
                    'prov_input': str(row[3] or '').strip()
                })

            creadas = 0
            errores = []
            
            with transaction.atomic():
                # Obtener sucursal de la sesión
                sucursal_obj = None
                sucursal_id = request.session.get('sucursal_id')
                if sucursal_id:
                    from preferencias.models import Sucursal
                    try:
                        sucursal_obj = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
                    except Sucursal.DoesNotExist:
                        pass

                for ref, data in solicitudes_data.items():
                    try:
                        # Crear Solicitud
                        solicitud = SolicitudCompra.objects.create(
                            empresa=empresa_actual,
                            solicitante=request.user,
                            sucursal=sucursal_obj,
                            estado='borrador',
                            notas=data['notas']
                        )

                        # Crear Detalles
                        for item in data['items']:
                            # Buscar Producto
                            producto = None
                            if item['prod_input'].isdigit():
                                producto = Producto.objects.filter(id=int(item['prod_input']), empresa=empresa_actual).first()
                            if not producto:
                                producto = Producto.objects.filter(nombre__icontains=item['prod_input'], empresa=empresa_actual).first()
                            
                            if not producto:
                                errores.append(f"Ref {ref}: No se encontró el producto '{item['prod_input']}'.")
                                continue

                            # Buscar Proveedor (opcional)
                            proveedor = None
                            if item['prov_input']:
                                if item['prov_input'].isdigit():
                                    proveedor = Proveedor.objects.filter(id=int(item['prov_input']), empresa=empresa_actual).first()
                                if not proveedor:
                                    proveedor = Proveedor.objects.filter(razon_social__icontains=item['prov_input'], empresa=empresa_actual).first()

                            DetalleSolicitudCompra.objects.create(
                                solicitud=solicitud,
                                producto=producto,
                                cantidad_solicitada=item['cantidad'],
                                proveedor=proveedor,
                                costo_unitario=producto.precio_costo or 0
                            )
                        
                        creadas += 1
                    except Exception as e:
                        errores.append(f"Ref {ref}: {str(e)}")
            
            if creadas == 0 and errores:
                return JsonResponse({'success': False, 'error': f"Error: {errores[0]}"})

            msg = f'Importación exitosa. {creadas} solicitudes creadas.'
            if errores: msg += f' ({len(errores)} errores)'
            return JsonResponse({'success': True, 'message': msg})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
            
    return JsonResponse({'success': False, 'error': 'Solicitud inválida.'})

@login_required(login_url='/login/')
def exportar_solicitudes_excel(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return HttpResponse("No autorizado", status=403)

    # --- Filtrado ---
    q = request.GET.get('q', '')
    folio_solicitud = request.GET.get('folio_solicitud', '')
    folio_cotizacion = request.GET.get('folio_cotizacion', '')
    folio_pedido = request.GET.get('folio_pedido', '')
    proveedor_id = request.GET.get('proveedor_id', '')
    fecha_solicitud = request.GET.get('fecha_solicitud', '')
    estado = request.GET.get('estado', '')
    sucursal_id_filtro = request.GET.get('sucursal', '')

    solicitudes = SolicitudCompra.objects.filter(empresa=empresa_actual).select_related('pedido_origen', 'solicitante', 'sucursal').order_by('-fecha_creacion')

    from django.db.models import Q
    if q:
        solicitudes = solicitudes.filter(
            Q(id__icontains=q) | Q(pedido_origen__id__icontains=q) |
            Q(solicitante__username__icontains=q) | Q(estado__icontains=q)
        ).distinct()
    
    if folio_solicitud:
        clean_sol = folio_solicitud.upper().replace('SOL-', '').replace('SOL', '').strip()
        solicitudes = solicitudes.filter(id__icontains=clean_sol)
    if folio_cotizacion:
        clean_cot = folio_cotizacion.upper().replace('COT-', '').replace('COT', '').strip()
        solicitudes = solicitudes.filter(pedido_origen__cotizacion_origen_id__icontains=clean_cot)
    if folio_pedido:
        clean_ped = folio_pedido.upper().replace('PED-', '').replace('PED', '').strip()
        solicitudes = solicitudes.filter(pedido_origen__id__icontains=clean_ped)
    if proveedor_id and proveedor_id != 'all':
        solicitudes = solicitudes.filter(detalles__proveedor_id=proveedor_id).distinct()
    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    if sucursal_id_filtro:
        solicitudes = solicitudes.filter(sucursal_id=sucursal_id_filtro)

    # --- Generar Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Solicitudes"

    headers = [
        'ID', 'Solicitante', 'Sucursal', 'Pedido Origen', 
        'Fecha Creación', 'Estado', 'Partidas', 'Notas'
    ]
    ws.append(headers)

    for s in solicitudes:
        ws.append([
            s.id, s.solicitante.get_full_name() or s.solicitante.username,
            s.sucursal.nombre if s.sucursal else "Principal",
            s.pedido_origen_id or "N/A",
            s.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            s.get_estado_display(),
            s.total_items,
            s.notas
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Solicitudes_{empresa_actual.nombre}.xlsx"'
    wb.save(response)
    return response

def get_empresa_actual(request):
    username = request.user.username
    if '@' in username:
        subdominio = username.split('@')[1]
        try:
            return Empresa.objects.get(subdominio=subdominio)
        except Empresa.DoesNotExist:
            return None
    return None

from django.db.models import Q
from proveedores.models import Proveedor

@login_required(login_url='/login/')
def dashboard_solicitudcompras(request):
    empresa_actual = get_empresa_actual(request)
    if not empresa_actual:
        return render(request, 'error_sin_empresa.html', status=403)

    # --- LÓGICA DE FILTRADO ---
    q = request.GET.get('q', '')
    folio_solicitud = request.GET.get('folio_solicitud', '')
    folio_cotizacion = request.GET.get('folio_cotizacion', '')
    folio_pedido = request.GET.get('folio_pedido', '')
    proveedor_id = request.GET.get('proveedor_id', '')
    fecha_solicitud = request.GET.get('fecha_solicitud', '')
    estado = request.GET.get('estado', '')
    sucursal_id_filtro = request.GET.get('sucursal', '')

    solicitudes = SolicitudCompra.objects.filter(empresa=empresa_actual).select_related('pedido_origen', 'solicitante', 'sucursal').order_by('-fecha_creacion')

    if q:
        solicitudes = solicitudes.filter(
            Q(id__icontains=q) |
            Q(pedido_origen__id__icontains=q) |
            Q(pedido_origen__cotizacion_origen_id__icontains=q) |
            Q(solicitante__username__icontains=q) |
            Q(estado__icontains=q) |
            Q(notas__icontains=q) |
            Q(detalles__producto__nombre__icontains=q)
        ).distinct()
    
    if folio_solicitud:
        clean_sol = folio_solicitud.upper().replace('SOL-', '').replace('SOL', '').strip()
        solicitudes = solicitudes.filter(id__icontains=clean_sol)
    if folio_cotizacion:
        clean_cot = folio_cotizacion.upper().replace('COT-', '').replace('COT', '').strip()
        solicitudes = solicitudes.filter(pedido_origen__cotizacion_origen_id__icontains=clean_cot)
    if folio_pedido:
        clean_ped = folio_pedido.upper().replace('PED-', '').replace('PED', '').strip()
        solicitudes = solicitudes.filter(pedido_origen__id__icontains=clean_ped)
    if proveedor_id and proveedor_id != 'all':
        solicitudes = solicitudes.filter(detalles__proveedor_id=proveedor_id).distinct()
    if fecha_solicitud:
        try:
            solicitudes = solicitudes.filter(fecha_creacion__date=fecha_solicitud)
        except:
            pass
    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    if sucursal_id_filtro:
        solicitudes = solicitudes.filter(sucursal_id=sucursal_id_filtro)

    # Para el buscador visual de proveedor
    proveedor_nombre_display = ""
    if proveedor_id and proveedor_id != 'all':
        try:
            p_obj = Proveedor.objects.get(id=proveedor_id, empresa=empresa_actual)
            proveedor_nombre_display = p_obj.razon_social
        except:
            pass

    filtros = {
        'q': q,
        'folio_solicitud': folio_solicitud,
        'folio_cotizacion': folio_cotizacion,
        'folio_pedido': folio_pedido,
        'proveedor_id': proveedor_id,
        'proveedor_nombre': proveedor_nombre_display,
        'fecha_solicitud': fecha_solicitud,
        'estado': estado,
        'sucursal': sucursal_id_filtro
    }
    
    # PAGINACIÓN
    paginator = Paginator(solicitudes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # --- FIN LÓGICA DE FILTRADO ---
    from preferencias.models import Sucursal
    sucursales_lista = Sucursal.objects.filter(empresa=empresa_actual).order_by('nombre')
    
    # --- NUEVO: Obtener catálogos para los selects del modal ---
    proveedores_activos = Proveedor.objects.filter(empresa=empresa_actual, estado='activo')
    todos_los_proveedores = Proveedor.objects.filter(empresa=empresa_actual)
    lista_almacenes = Almacen.objects.filter(empresa=empresa_actual).values('id', 'nombre')
    lista_monedas = Moneda.objects.filter(empresa=empresa_actual).values('id', 'siglas', 'simbolo')
    lista_productos = Producto.objects.filter(empresa=empresa_actual).values('id', 'nombre', 'precio_costo')
    listas_costos = ListaPrecioCosto.objects.filter(empresa=empresa_actual, tipo='costo')
    todas_categorias = CategoriaCatalogo.objects.filter(empresa=empresa_actual)
    
    contexto = {
        'page_obj': page_obj,
        'proveedores': proveedores_activos,
        'todos_los_proveedores': todos_los_proveedores,
        'almacenes': list(lista_almacenes),
        'monedas': list(lista_monedas),
        'productos': list(lista_productos),
        'categorias_catalogo': todas_categorias,
        'listas_costos': listas_costos,
        'sucursales': sucursales_lista,
        'filtros': filtros,
        'section': 'solicitudcompras'
    }
    return render(request, 'dashboard_solicitudcompras.html', contexto)

from preferencias.permissions import require_sales_permission

@login_required(login_url='/login/')
@require_sales_permission('pedidos', 'solicitar')
def crear_solicitud_desde_pedido(request, detalle_id):
    empresa_actual = get_empresa_actual(request)
    detalle = get_object_or_404(DetallePedido, id=detalle_id)
    
    # Seguridad: Verificar que el pedido pertenezca a la empresa
    pedido = get_object_or_404(Pedido, id=detalle.pedido.id, empresa=empresa_actual)
    producto = detalle.producto

    # Verificamos si YA existe una solicitud abierta para este pedido para agrupar
    solicitud_existente = SolicitudCompra.objects.filter(
        pedido_origen=pedido, 
        estado='borrador', 
        empresa=empresa_actual
    ).first()

    if solicitud_existente:
        target_solicitud = solicitud_existente
    else:
        target_solicitud = SolicitudCompra.objects.create(
            pedido_origen=pedido, 
            solicitante=request.user, 
            empresa=empresa_actual, 
            sucursal=pedido.sucursal, # Hereda la sucursal del pedido
            estado='borrador'
        )

    # --- INTELIGENCIA: ¿ES PRODUCCIÓN? ---
    if producto.tipo_abastecimiento == 'produccion':
        # Buscar un almacén válido (Prioridad: Sucursal Pedido -> Cualquier Almacén de Empresa)
        almacen_destino = Almacen.objects.filter(empresa=empresa_actual, sucursal=pedido.sucursal).first()
        if not almacen_destino:
            almacen_destino = Almacen.objects.filter(empresa=empresa_actual).first()

        if not almacen_destino:
            error_msg = f'No se pudo generar la orden de producción para "{producto.nombre}" porque no hay almacenes registrados. Por favor, cree un almacén primero.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('detalle_pedido', pedido_id=pedido.id)

        # 1. Crear la Orden de Producción vinculada al pedido (En Borrador)
        op = OrdenProduccion.objects.create(
            empresa=empresa_actual,
            producto=producto,
            cantidad=detalle.cantidad_solicitada,
            pedido_origen=pedido,
            solicitante=request.user,
            almacen=almacen_destino,
            estado='borrador',
            notas=f"Generada desde Pedido #{pedido.id} (Partida Individual)",
            sucursal=pedido.sucursal
        )

        # 2. Obtener Receta y copiar a DetalleOrdenProduccion
        receta = DetalleReceta.objects.filter(producto_padre=producto)
        
        if not receta.exists():
            error_msg = f'El producto "{producto.nombre}" es de producción pero NO TIENE RECETA. No se puede generar solicitud.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('dashboard_pedidos')

        # 3. Calcular necesidades y crear detalles
        items_a_solicitar = []
        for item in receta:
            # Copiar a la tabla de la orden específica
            from produccion.models import DetalleOrdenProduccion
            DetalleOrdenProduccion.objects.create(
                orden_produccion=op,
                producto=item.componente,
                cantidad=item.cantidad * detalle.cantidad_solicitada
            )

            componente = item.componente
            cantidad_necesaria = item.cantidad * detalle.cantidad_solicitada
            
            stock_componente = componente.stock_disponible
            
            if stock_componente < cantidad_necesaria:
                faltante = cantidad_necesaria - stock_componente
                items_a_solicitar.append({
                    'producto': componente,
                    'cantidad': faltante,
                    'costo': componente.precio_costo
                })

        if not items_a_solicitar:
            msg = f'Para producir {producto.nombre} ya tienes todos los componentes necesarios en stock. No se generó solicitud.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': msg})
            messages.info(request, msg)
            return redirect('dashboard_pedidos')

        # 3. Crear Detalles en la Solicitud
        for item in items_a_solicitar:
            # Verificar si ya existe ese componente en la solicitud actual para no duplicar
            det_existente = DetalleSolicitudCompra.objects.filter(
                solicitud=target_solicitud,
                producto=item['producto']
            ).first()
            
            if det_existente:
                det_existente.cantidad_solicitada += item['cantidad']
                det_existente.save()
            else:
                DetalleSolicitudCompra.objects.create(
                    solicitud=target_solicitud,
                    producto=item['producto'],
                    cantidad_solicitada=item['cantidad'],
                    costo_unitario=item['costo'],
                    detalle_pedido_origen=detalle # Vinculamos al pedido original
                )
        
        detalle.estado_linea = 'en_proceso' # Marcamos que ya se inició el trámite
        detalle.save()
        
        msg = f'Solicitud de componentes para "{producto.nombre}" generada correctamente.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': msg})
        messages.success(request, msg)

    else:
        # --- CASO NORMAL: COMPRAR EL PRODUCTO DIRECTO ---
        DetalleSolicitudCompra.objects.create(
            solicitud=target_solicitud,
            producto=producto,
            cantidad_solicitada=detalle.cantidad_solicitada,
            costo_unitario=producto.precio_costo,
            detalle_pedido_origen=detalle
        )
        
        detalle.estado_linea = 'en_proceso'
        detalle.save()
        
        msg = f'Producto "{producto.nombre}" agregado a la Solicitud #{target_solicitud.id}.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': msg})
        messages.success(request, msg)

    return redirect('dashboard_solicitudcompras')

def obtener_solicitud_json(request, solicitud_id):
    """Devuelve los datos de una solicitud específica para el Modal"""
    try:
        empresa_actual = get_empresa_actual(request) 
        solicitud = get_object_or_404(SolicitudCompra, id=solicitud_id, empresa=empresa_actual)
        
        # Convertimos el modelo principal a diccionario
        data = model_to_dict(solicitud)
        
        # --- Construir el objeto anidado del Pedido ---
        if solicitud.pedido_origen:
            data['pedido_origen'] = {
                'id': solicitud.pedido_origen.id,
                'cotizacion_origen_id': solicitud.pedido_origen.cotizacion_origen_id,
                'cliente': {
                    'nombre': solicitud.pedido_origen.cliente.nombre,
                    'apellidos': solicitud.pedido_origen.cliente.apellidos,
                    'razon_social': solicitud.pedido_origen.cliente.razon_social,
                    'nombre_completo': solicitud.pedido_origen.cliente.nombre_completo
                }
            }
        else:
            data['pedido_origen'] = None

        # --- CORRECCIÓN: Mantener lista plana para Edición y agrupada para Ver ---
        detalles_list = []
        detalles_agrupados = {}
        
        for det in solicitud.detalles.all().select_related('producto', 'proveedor', 'almacen', 'moneda', 'detalle_pedido_origen__producto'):
            # Info básica para ambos usos
            item_data = {
                'id': det.id,
                'producto_id': det.producto.id,
                'producto_nombre': det.producto.nombre,
                'cantidad': det.cantidad_solicitada,
                'proveedor_id': det.proveedor.id if det.proveedor else None,
                'proveedor_nombre': det.proveedor.razon_social if det.proveedor else 'No asignado',
                'sucursal_id': det.sucursal.id if det.sucursal else None,
                'sucursal_nombre': det.sucursal.nombre if det.sucursal else '',
                'almacen_id': det.almacen.id if det.almacen else None,
                'almacen_nombre': det.almacen.nombre if det.almacen else 'No asignado',
                'costo_unitario': str(det.costo_unitario),
                'subtotal': str(det.subtotal),
                'iva_porcentaje': str(det.producto.iva or 0),
                'iva_monto': str(det.iva_monto),
                'total': str(det.total),
                'moneda_id': det.moneda.id if det.moneda else None,
                'moneda_siglas': det.moneda.siglas if det.moneda else 'MXN',
                'tipo_cambio': str(det.moneda.factor) if det.moneda else '1.0000',
                'lista_id': det.lista.id if det.lista else None,
                'detalle_pedido_origen_id': det.detalle_pedido_origen.id if det.detalle_pedido_origen else None
            }
            
            # Agregar a lista plana (Edición)
            detalles_list.append(item_data)

            # Identificar el grupo (Visualización Árbol)
            if det.detalle_pedido_origen and det.detalle_pedido_origen.producto.tipo_abastecimiento == 'produccion':
                grupo_nombre = f"Producción de: {det.detalle_pedido_origen.producto.nombre}"
            else:
                grupo_nombre = "Compras Directas / Reabastecimiento"
            
            if grupo_nombre not in detalles_agrupados:
                detalles_agrupados[grupo_nombre] = []
                
            detalles_agrupados[grupo_nombre].append(item_data)
            
        data['detalles'] = detalles_list
        data['grupos'] = []
        # Asegurar que 'Compras Directas' aparezca al final si existe
        nombres_grupos = sorted([g for g in detalles_agrupados.keys() if g != "Compras Directas / Reabastecimiento"])
        if "Compras Directas / Reabastecimiento" in detalles_agrupados:
            nombres_grupos.append("Compras Directas / Reabastecimiento")
            
        for g_nom in nombres_grupos:
            data['grupos'].append({
                'nombre': g_nom,
                'items': detalles_agrupados[g_nom]
            })

        data['subtotal_total'] = str(solicitud.calcular_subtotal)
        data['iva_total'] = str(solicitud.calcular_iva)
        data['gran_total'] = str(solicitud.calcular_total)
        data['solicitante_nombre'] = solicitud.solicitante.username.split('@')[0] if solicitud.solicitante else 'Sistema'
        data['fecha_creacion'] = solicitud.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        data['estado_display'] = solicitud.get_estado_display()

        return JsonResponse(data)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required(login_url='/login/')
def actualizar_solicitud(request, solicitud_id):
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            solicitud = get_object_or_404(SolicitudCompra, id=solicitud_id, empresa=empresa_actual)

            # SEGURIDAD: Solo editar si está en borrador
            if solicitud.estado != 'borrador':
                return JsonResponse({'success': False, 'error': 'Solo se pueden editar solicitudes en Borrador.'})

            solicitud.notas = request.POST.get('notas', '')
            solicitud.save()

            # Borrar y recrear detalles (simplificación para este ejemplo)
            solicitud.detalles.all().delete()

            productos_ids = request.POST.getlist('producto_id[]')
            cantidades = request.POST.getlist('cantidad[]')
            proveedores_ids = request.POST.getlist('proveedor_id[]')
            sucursales_ids = request.POST.getlist('sucursal_id[]') # <--- NUEVO
            costos = request.POST.getlist('costo_unitario[]')
            almacenes_ids = request.POST.getlist('almacen_id[]')
            monedas_ids = request.POST.getlist('moneda_id[]') 
            listas_ids = request.POST.getlist('lista_id[]')
            pedidos_det_ids = request.POST.getlist('pedido_det_id[]') 

            # Zip de todas las listas
            for p_id, cant, prov_id, suc_id, cost, alm_id, mon_id, p_det_id, l_id in zip(productos_ids, cantidades, proveedores_ids, sucursales_ids, costos, almacenes_ids, monedas_ids, pedidos_det_ids, listas_ids):
                if p_id and p_id != '':
                    DetalleSolicitudCompra.objects.create(
                        solicitud=solicitud,
                        producto_id=p_id,
                        cantidad_solicitada=cant,
                        proveedor_id=prov_id if prov_id else None,
                        sucursal_id=suc_id if (suc_id and suc_id != '') else None,
                        costo_unitario=cost,
                        almacen_id=alm_id if alm_id else None,
                        moneda_id=mon_id if mon_id else None,
                        lista_id=l_id if l_id else None,
                        detalle_pedido_origen_id=p_det_id if p_det_id else None 
                    )
            
            messages.success(request, f'Solicitud #{solicitud.id} actualizada.')
            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required(login_url='/login/')
@transaction.atomic
def crear_solicitud_manual(request):
    if request.method == 'POST':
        try:
            empresa_actual = get_empresa_actual(request)
            if not empresa_actual:
                return JsonResponse({'success': False, 'error': 'Empresa no detectada'})

            # Asignar sucursal desde la sesión
            sucursal_obj = None
            sucursal_id = request.session.get('sucursal_id')
            if sucursal_id:
                from preferencias.models import Sucursal
                try:
                    sucursal_obj = Sucursal.objects.get(id=sucursal_id, empresa=empresa_actual)
                except Sucursal.DoesNotExist:
                    pass

            notas = request.POST.get('notas', '')
            solicitud = SolicitudCompra.objects.create(
                solicitante=request.user,
                empresa=empresa_actual,
                sucursal=sucursal_obj,
                estado='borrador',
                notas=notas
            )

            productos_ids = request.POST.getlist('producto_id[]')
            cantidades = request.POST.getlist('cantidad[]')
            proveedores_ids = request.POST.getlist('proveedor_id[]')
            sucursales_ids = request.POST.getlist('sucursal_id[]')
            costos = request.POST.getlist('costo_unitario[]')
            almacenes_ids = request.POST.getlist('almacen_id[]')
            monedas_ids = request.POST.getlist('moneda_id[]')
            listas_ids = request.POST.getlist('lista_id[]')

            count = 0
            for p_id, cant, prov_id, suc_id, cost, alm_id, mon_id, l_id in zip(productos_ids, cantidades, proveedores_ids, sucursales_ids, costos, almacenes_ids, monedas_ids, listas_ids):
                if p_id and p_id != '':
                    DetalleSolicitudCompra.objects.create(
                        solicitud=solicitud,
                        producto_id=p_id,
                        cantidad_solicitada=cant,
                        proveedor_id=prov_id if prov_id else None,
                        sucursal_id=suc_id if (suc_id and suc_id != '') else None,
                        costo_unitario=cost,
                        almacen_id=alm_id if alm_id else None,
                        moneda_id=mon_id if mon_id else None,
                        lista_id=l_id if l_id else None
                    )
                    count += 1
            
            if count == 0:
                raise ValueError("Debes agregar al menos un producto a la solicitud.")

            messages.success(request, f'Solicitud #{solicitud.id} creada correctamente.')
            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required(login_url='/login/')
@transaction.atomic 
def autorizar_solicitud(request, solicitud_id):
    empresa_actual = get_empresa_actual(request)
    solicitud = get_object_or_404(SolicitudCompra, id=solicitud_id, empresa=empresa_actual)

    if solicitud.estado != 'borrador':
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Solo se pueden autorizar solicitudes en estado Borrador.'})
        messages.warning(request, 'Solo se pueden autorizar solicitudes en estado Borrador.')
        return redirect('dashboard_solicitudcompras')

    # 1. Validar que todas las partidas tengan proveedor, almacén y moneda
    for det in solicitud.detalles.all():
        if not det.proveedor or not det.almacen or not det.moneda:
            error_msg = f'Faltan datos (Proveedor, Almacén o Moneda) en el producto {det.producto.nombre}.'
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('dashboard_solicitudcompras')

    # 2. Agrupar ítems por (Proveedor, Sucursal, Moneda, Almacén)
    items_agrupados = defaultdict(list)
    for det in solicitud.detalles.all():
        llave = (det.proveedor, det.sucursal, det.moneda, det.almacen)
        items_agrupados[llave].append(det)

    # 3. Si es una previsualización (AJAX)
    if request.GET.get('preview') == '1':
        desglose = []
        # Agrupamos por proveedor para el mensaje de alerta
        conteo_por_proveedor = defaultdict(int)
        for (prov, suc, mon, alm) in items_agrupados.keys():
            conteo_por_proveedor[prov.razon_social] += 1

        for prov_nombre, total_oc in conteo_por_proveedor.items():
            desglose.append({
                'proveedor': prov_nombre,
                'total_oc': total_oc
            })

        return JsonResponse({
            'success': True,
            'total_ordenes': len(items_agrupados),
            'desglose': desglose
        })

    # 4. Procesar creación física (POST)
    if request.method == 'POST':
        from compras.models import OrdenCompra, DetalleCompra

        ordenes_creadas_count = 0
        for (proveedor, sucursal, moneda, almacen), lista_detalles in items_agrupados.items():
            nueva_oc = OrdenCompra.objects.create(
                proveedor=proveedor,
                sucursal=sucursal, # Esta es la sucursal del proveedor
                usuario=request.user,
                empresa=empresa_actual,
                sucursal_empresa=solicitud.sucursal, # Hereda la sucursal de la empresa de la solicitud
                estado='borrador',
                almacen_destino=almacen,
                moneda=moneda,
                tipo_cambio=moneda.factor if moneda else 1.0000,
                notas=f"Generada desde Solicitud #{solicitud.id}.",
                solicitud_origen=solicitud 
            )

            for det_solicitud in lista_detalles:
                DetalleCompra.objects.create(
                    orden_compra=nueva_oc,
                    producto=det_solicitud.producto,
                    cantidad=det_solicitud.cantidad_solicitada,
                    precio_costo=det_solicitud.costo_unitario,
                    detalle_pedido_origen=det_solicitud.detalle_pedido_origen
                )

                if det_solicitud.detalle_pedido_origen:
                    det_solicitud.detalle_pedido_origen.estado_linea = 'comprado'
                    det_solicitud.detalle_pedido_origen.save()

            ordenes_creadas_count += 1

        solicitud.estado = 'atendida'
        solicitud.save()

        # NOTIFICACIÓN
        crear_notificacion(
            empresa=empresa_actual,
            mensaje=f"La Solicitud #{solicitud.id} ha sido autorizada y se generaron {ordenes_creadas_count} órdenes de compra.",
            actor=request.user,
            propietario=solicitud.solicitante
        )

        msg = f'Solicitud autorizada. Se generaron {ordenes_creadas_count} órdenes de compra.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': msg})

        messages.success(request, msg)
        return redirect('dashboard_solicitudcompras')

    return redirect('dashboard_solicitudcompras')

@login_required(login_url='/login/')
@transaction.atomic
def cancelar_solicitud(request, solicitud_id):
    """
    Cancela una solicitud de compra y libera las partidas del pedido origen
    para que puedan ser solicitadas nuevamente.
    """
    empresa_actual = get_empresa_actual(request)
    solicitud = get_object_or_404(SolicitudCompra, id=solicitud_id, empresa=empresa_actual)

    if solicitud.estado == 'atendida':
        msg = "No se puede cancelar una solicitud que ya ha sido atendida (OC generada)."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': msg})
        messages.error(request, msg)
        return redirect('dashboard_solicitudcompras')

    if solicitud.estado == 'cancelada':
        msg = "Esta solicitud ya está cancelada."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': msg})
        messages.warning(request, msg)
        return redirect('dashboard_solicitudcompras')

    # 1. Liberar las partidas del pedido
    from almacenes.models import Inventario
    detalles_solicitud = solicitud.detalles.all()
    for detalle in detalles_solicitud:
        if detalle.detalle_pedido_origen:
            detalle_pedido = detalle.detalle_pedido_origen
            
            # Si estaba reservado, liberar el stock físico en el inventario
            if detalle_pedido.estado_linea == 'reservado':
                # Buscamos el inventario en el almacén que corresponda (o el primero con reserva)
                # Intentamos ser específicos si el detalle_pedido tuviera almacén, si no, buscamos por producto
                inv_reserva = Inventario.objects.filter(
                    producto=detalle_pedido.producto,
                    reservado__gt=0
                ).first()
                if inv_reserva:
                    inv_reserva.reservado = max(0, inv_reserva.reservado - detalle_pedido.cantidad_solicitada)
                    inv_reserva.save()

            # Revertir al estado original de abastecimiento según el tipo de producto
            # Ampliamos para que cualquier estado que no sea el final (completo) se libere
            if detalle_pedido.estado_linea in ['en_proceso', 'comprado', 'reservado', 'pendiente']:
                if detalle_pedido.producto.tipo_abastecimiento == 'produccion':
                    detalle_pedido.estado_linea = 'produccion'
                else:
                    detalle_pedido.estado_linea = 'compra'
                detalle_pedido.save()
    
    # 2. Si la solicitud tiene un pedido_origen, y se generaron órdenes de producción
    # en borrador vinculadas a este pedido, también las cancelamos para no duplicar.
    if solicitud.pedido_origen:
        from produccion.models import OrdenProduccion
        # También verificamos si el pedido estaba en 'completo' y lo regresamos a 'revision'
        if solicitud.pedido_origen.estado == 'completo':
            solicitud.pedido_origen.estado = 'revision'
            solicitud.pedido_origen.save()

        ordenes_produccion = OrdenProduccion.objects.filter(
            pedido_origen=solicitud.pedido_origen,
            estado='borrador',
            empresa=empresa_actual
        )
        for op in ordenes_produccion:
            op.estado = 'cancelado'
            op.save()

    # 3. Marcar solicitud como cancelada
    solicitud.estado = 'cancelada'
    solicitud.save()

    # NOTIFICACIÓN
    crear_notificacion(
        empresa=empresa_actual,
        mensaje=f"La Solicitud #{solicitud.id} ha sido CANCELADA y las partidas del pedido liberadas.",
        actor=request.user,
        propietario=solicitud.solicitante
    )
    
    # Notificar también al vendedor del pedido si es una persona distinta
    if solicitud.pedido_origen and solicitud.pedido_origen.vendedor != solicitud.solicitante:
        crear_notificacion(
            empresa=empresa_actual,
            mensaje=f"La solicitud vinculada al Pedido #{solicitud.pedido_origen.id} ha sido CANCELADA.",
            actor=request.user,
            propietario=solicitud.pedido_origen.vendedor
        )

    msg = f'Solicitud SOL-{solicitud.id:04d} cancelada correctamente.'
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': msg})

    messages.success(request, msg)
    return redirect('dashboard_solicitudcompras')

@login_required
def imprimir_solicitud(request, pk):
    """Genera la vista para impresión de solicitud de compra (PDF)"""
    empresa_actual = get_empresa_actual(request)
    solicitud = get_object_or_404(SolicitudCompra, id=pk, empresa=empresa_actual)

    # Limpiar nombre del solicitante
    solicitante_nombre = solicitud.solicitante.get_full_name()
    if not solicitante_nombre:
        solicitante_nombre = solicitud.solicitante.username.split('@')[0]

    # --- Lógica de Agrupación (Árbol) ---
    detalles_agrupados = {}
    for det in solicitud.detalles.all().select_related('producto', 'proveedor', 'almacen', 'moneda', 'detalle_pedido_origen__producto'):
        if det.detalle_pedido_origen and det.detalle_pedido_origen.producto.tipo_abastecimiento == 'produccion':
            grupo_nombre = f"Producción de: {det.detalle_pedido_origen.producto.nombre}"
        else:
            grupo_nombre = "Compras Directas / Reabastecimiento"
        
        if grupo_nombre not in detalles_agrupados:
            detalles_agrupados[grupo_nombre] = []
        detalles_agrupados[grupo_nombre].append(det)

    grupos_final = []
    nombres_grupos = sorted([g for g in detalles_agrupados.keys() if g != "Compras Directas / Reabastecimiento"])
    if "Compras Directas / Reabastecimiento" in detalles_agrupados:
        nombres_grupos.append("Compras Directas / Reabastecimiento")
        
    for g_nom in nombres_grupos:
        grupos_final.append({
            'nombre': g_nom,
            'items': detalles_agrupados[g_nom],
            'es_directo': g_nom == "Compras Directas / Reabastecimiento"
        })

    context = {
        'solicitud': solicitud,
        'empresa': empresa_actual,
        'solicitante_nombre': solicitante_nombre,
        'grupos': grupos_final,
    }
    return render(request, 'solicitudcompras/imprimir_solicitud.html', context)
